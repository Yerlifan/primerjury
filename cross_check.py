# -*- coding: utf-8 -*-
"""Cross-check, an INDEPENDENT, READ-ONLY audit of a finished run.

Opens no measurement of its own that the pipeline already made; it re-asks the
pipeline's questions using different code and reports where the answers differ.
It NEVER writes to panel files.

Seven modules: identity, internal consistency, membership, literature, error
patterns, database health, taxon coverage. Findings are graded KRITIK / CIDDI /
UYARI / BILGI, and checks that COULD NOT RUN are reported as ATLANDI, never as
passed. That distinction is the point: a check that did not run is not a check
that succeeded, and the exit code reflects it.

"""
from __future__ import print_function

import argparse
import ast
import collections
import glob
import hashlib
import io
import json
import math
import os
import re
import sys
import time
import traceback

VERSIYON = u'1.0 (2026-08-09)'

# UTF-8 output: the Windows console defaults to cp857/cp1254 and raises
# UnicodeEncodeError on non-ASCII characters. Rather than crash half way through
# the script, the stream is rewrapped at the start. (The error is NOT SWALLOWED,
# only the encoding is fixed.)
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8',
                                  errors='replace', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8',
                                  errors='replace', line_buffering=True)


# =========================================================================
# SEVERITY LEVELS
# =========================================================================
# KRITIK : it makes the order, or a number that will be reported, DIRECTLY wrong
# CIDDI  : the basis of a claim collapses, but the order may not change at once
# UYARI  : inconsistent or indefensible, but the measured number does not change
# BILGI  : worth noting, not an error
# ATLANDI: the check COULD NOT RUN. That IS NOT a "pass".
KRITIK, CIDDI, UYARI, BILGI, ATLANDI = u'KRITIK', u'CIDDI', u'UYARI', u'BILGI', u'ATLANDI'
_SIRA = {KRITIK: 0, CIDDI: 1, UYARI: 2, BILGI: 3, ATLANDI: 4}


class Bulgu(object):
    """A single audit finding.

        Every finding MUST answer FOUR questions, or the report becomes unreadable:
          beklenen : what the rule was
          bulunan  : the real value that was measured or read
          dosya    : the source the finding rests on (a path, with a line if possible)
          ciddiyet : one of the five levels above

    """
    __slots__ = ('modul', 'kod', 'ciddiyet', 'beklenen', 'bulunan', 'dosya', 'oneri')

    def __init__(self, modul, kod, ciddiyet, beklenen, bulunan, dosya, oneri=u''):
        self.modul = modul
        self.kod = kod
        self.ciddiyet = ciddiyet
        self.beklenen = beklenen
        self.bulunan = bulunan
        self.dosya = dosya
        self.oneri = oneri


class Rapor(object):
    """Collects the findings of every module, counts them and produces the exit code."""

    def __init__(self):
        self.bulgular = []
        self.modul_durumu = collections.OrderedDict()   # modul -> dict(durum, sure, not)
        self.olcum = collections.OrderedDict()          # olculen sureler / oranlar
        self.tablolar = collections.OrderedDict()       # rapora gomulecek ek tablolar

    def ekle(self, modul, kod, ciddiyet, beklenen, bulunan, dosya, oneri=u''):
        self.bulgular.append(Bulgu(modul, kod, ciddiyet, beklenen, bulunan, dosya, oneri))

    def atla(self, modul, kod, beklenen, sebep, dosya=u'-'):
        """A check could not run. NO SILENT SKIP: it is recorded as a finding."""
        self.bulgular.append(Bulgu(modul, kod, ATLANDI, beklenen,
                                   u'COULD NOT BE RUN: %s' % sebep, dosya,
                                   u'Bu kontrol OLCULMEDI, "gecti" sayilmamalidir.'))

    def say(self, ciddiyet=None, modul=None):
        return len([b for b in self.bulgular
                    if (ciddiyet is None or b.ciddiyet == ciddiyet)
                    and (modul is None or b.modul == modul)])

    def cikis_kodu(self):
        k = 0
        if self.say(KRITIK):
            k |= 1
        if self.say(CIDDI):
            k |= 2
        if self.say(ATLANDI):
            k |= 4
        return k


# ===========================================================================
# THE BASIC HELPERS
# ===========================================================================
def unicode_(x):
    if isinstance(x, bytes):
        return x.decode('utf-8', 'replace')
    return u'%s' % (x,)


def yaz(*a):
    u"""A progress line. Time stamped, and its output is flushed at once."""
    print(u'[%s] %s' % (time.strftime('%H:%M:%S'), u' '.join(unicode_(x) for x in a)))
    try:
        sys.stdout.flush()
    except Exception:
        pass


def vir(x, b=2):
    """A decimal in the local format. None -> '-'. '0' and 'not measured' ARE NEVER CONFUSED."""
    if x is None:
        return u'-'
    try:
        return (u'%.*f' % (b, float(x))).replace(u'.', u',')
    except (TypeError, ValueError):
        return unicode_(x)


def sayi(x):
    """Convert decimal text, in either format, to a float. None if it cannot be converted.

        Returning None MATTERS: counting an unconvertible cell as 0 erases the
        difference between "not measured" and "measured as zero", and that is exactly
        why a separation factor was once taken for 0.00x in this project.

    """
    if x is None:
        return None
    s = unicode_(x).strip().replace(u'%', u'').replace(u'x', u'')
    s = s.replace(u'−', u'-')          # unicode eksi isareti
    if not s or s.lower() in (u'-', u'--', u'yok', u'olculmedi', u'nd', u'nan', u'?'):
        return None
    # Telling "1.234,56" (TR thousands) from "1,234.56" (EN thousands): the LAST
    # separator is the decimal one
    if u',' in s and u'.' in s:
        s = s.replace(u'.', u'') if s.rfind(u',') > s.rfind(u'.') else s.replace(u',', u'')
    s = s.replace(u',', u'.')
    try:
        return float(s)
    except ValueError:
        return None


def md5_metin(*parcalar):
    """The checkpoint key. hash() IS NOT USED.

        Python's built-in hash() is SALTED PER PROCESS (PYTHONHASHSEED); the same
        input produces a different key in two runs, and the checkpoint either never
        matches or, worse, collides. md5 is stable and stays the same from run to run.

    """
    h = hashlib.md5()
    for p in parcalar:
        h.update(unicode_(p).encode('utf-8', 'replace'))
        h.update(b'\x00')
    return h.hexdigest()


def dosya_imzasi(yol):
    """The identity of an input: the absolute path plus the size plus the modification time.

        That goes into the checkpoint key; when an input changes the key changes and
        the old checkpoint invalidates itself.

    """
    try:
        st = os.stat(yol)
        return u'%s|%d|%d' % (os.path.abspath(yol), st.st_size, int(st.st_mtime))
    except OSError:
        return u'%s|YOK' % os.path.abspath(yol)


class KontrolNoktasi(object):
    """An md5 keyed checkpoint store that checks input freshness.

        TWO LAYERS OF INVALIDATION:
          1) The key is derived from the signature of the input files (path, size,
             mtime). If an input changes the key changes and the old record is not
             found.
          2) The stored file's mtime is also compared against the inputs'. If an
             input is NEWER than the checkpoint, the record counts as INVALID.
        The second layer is a seat belt against things like the same key being copied
        by hand. A stale checkpoint has already led to a run being taken for "clean"
        in this project.

    """

    def __init__(self, klasor, etkin=True):
        self.klasor = klasor
        self.etkin = etkin
        self.isabet = 0
        self.iska = 0
        if etkin and not os.path.isdir(klasor):
            os.makedirs(klasor)

    def _yol(self, anahtar):
        return os.path.join(self.klasor, anahtar + '.json')

    def oku(self, anahtar, girdiler=()):
        if not self.etkin:
            return None
        y = self._yol(anahtar)
        if not os.path.exists(y):
            self.iska += 1
            return None
        try:
            kn_zaman = os.path.getmtime(y)
        except OSError:
            self.iska += 1
            return None
        for g in girdiler:                      # IF THE INPUT IS NEWER -> INVALID
            try:
                if os.path.exists(g) and os.path.getmtime(g) > kn_zaman:
                    self.iska += 1
                    return None
            except OSError:
                self.iska += 1
                return None
        try:
            with io.open(y, encoding='utf-8') as fh:
                v = json.load(fh)
            self.isabet += 1
            return v
        except (ValueError, IOError) as e:
            # A corrupt checkpoint is not ignored SILENTLY, it is printed to the screen.
            yaz(u'  WARNING: could not read checkpoint (%s): %s' % (anahtar, e))
            self.iska += 1
            return None

    def yazdir(self, anahtar, veri):
        if not self.etkin:
            return
        try:
            with io.open(self._yol(anahtar), 'w', encoding='utf-8') as fh:
                fh.write(unicode_(json.dumps(veri, ensure_ascii=False, default=str)))
        except IOError as e:
            yaz(u'  WARNING: could not write checkpoint (%s): %s' % (anahtar, e))


class Canlilik(object):
    """A regular sign of life, for long loops.

        A stage that prints nothing to the screen cannot be told apart from a stage
        that has FROZEN. This class prints a line at least every `aralik` seconds and
        estimates the remaining time from the measured speed.

    """

    def __init__(self, etiket, toplam=None, aralik=20.0):
        self.etiket = etiket
        self.toplam = toplam
        self.aralik = aralik
        self.t0 = time.time()
        self.son = self.t0

    def vur(self, n, ek=u''):
        t = time.time()
        if t - self.son < self.aralik:
            return
        self.son = t
        gecen = t - self.t0
        if self.toplam:
            oran = n / float(self.toplam)
            kalan = (gecen / oran - gecen) if oran > 0 else None
            yaz(u'  ... %s %d/%d (%%%s) gecen %s, kalan ~%s %s'
                % (self.etiket, n, self.toplam, vir(100 * oran, 1),
                   sure_metni(gecen), sure_metni(kalan), ek))
        else:
            yaz(u'  ... %s %d, gecen %s %s' % (self.etiket, n, sure_metni(gecen), ek))

    def bitti(self, n):
        d = time.time() - self.t0
        yaz(u'  %s done: %d items, %s' % (self.etiket, n, sure_metni(d)))
        return d


def sure_metni(sn):
    if sn is None:
        return u'?'
    sn = float(sn)
    if sn < 90:
        return u'%d s' % int(round(sn))
    if sn < 5400:
        return u'%d min' % int(round(sn / 60.0))
    return u'%s h' % vir(sn / 3600.0, 1)


# =========================================================================
# READ ONLY SOURCE READERS
# =========================================================================
def gecersiz_isareti(yol):
    u"""Looks for the '# GECERSIZ' mark at the head of a file.

    Returns: the reason text when the mark is there, otherwise None. Only the first
    10 lines are read; the mark has to sit AT THE VERY TOP of the file so that it
    cannot be missed.
    """
    try:
        with io.open(yol, encoding='utf-8', errors='replace') as fh:
            for i, s in enumerate(fh):
                if i >= 10:
                    break
                if not s.startswith(u'#'):
                    break
                if u'GECERSIZ' in s.upper():
                    return s.lstrip(u'#').strip()
    except (IOError, OSError):
        return None
    return None


def en_yeni_kaynak(kok, desen):
    """Returns the NEWEST file matching the pattern that is not marked GECERSIZ.

        2026-08-09: the source path had been fixed to a dated file name
        ('ESIK_VE_OLCUT_2026-08-08.tsv'). When a new version was written, the audit
        carried on reading the old one. Now the files are found by pattern and sorted
        by name; since the name carries a date, the largest name is the newest
        version. If there is none, the pattern itself is returned so that the "no such
        file" finding is still produced.

    """
    adaylar = sorted(glob.glob(os.path.join(kok, desen)))
    for y in reversed(adaylar):
        if not gecersiz_isareti(y):
            return y
    return adaylar[-1] if adaylar else os.path.join(kok, desen)


def tsv_oku(yol, yorum=u'#'):
    """A TSV with a header -> [dict]. Comment lines are skipped.

        Returns None if the file is missing (NOT an empty list). "No such file" and
        "the file is empty" are different things and produce different findings;
        merging them is a silent skip.

    """
    if not os.path.exists(yol):
        return None
    # 2026-08-09: a source marked GECERSIZ IS NOT READ. The reason: in the 2026-08-09
    # 18:09 run, all 22 of the 22 M5-D1-SESSIZ-SIFIR findings came from
    # NIHAI_SIPARIS_LISTESI_2026-08-07.tsv, which had already been superseded by
    # ESIK_VE_OLCUT_2026-08-09.tsv. Producing findings from an invalid decision table
    # fills the audit with noise. The file IS NOT DELETED; a '# GECERSIZ' line is put
    # in its header and the audit stops treating it as a source. The skip IS NOT
    # SILENT: the caller sees None and produces an ATLANDI finding through rap.atla().
    if gecersiz_isareti(yol):
        return None
    satirlar = []
    with io.open(yol, encoding='utf-8', errors='replace') as fh:
        for s in fh:
            if yorum and s.startswith(yorum):
                continue
            if s.strip():
                satirlar.append(s.rstrip(u'\n').rstrip(u'\r'))
    if not satirlar:
        return []
    bas = [b.strip() for b in satirlar[0].split(u'\t')]
    out = []
    for i, s in enumerate(satirlar[1:], 2):
        p = s.split(u'\t')
        p += [u''] * (len(bas) - len(p))
        d = dict(zip(bas, p[:len(bas)]))
        d['_satir'] = i
        out.append(d)
    return out


def metin_oku(yol):
    if not os.path.exists(yol):
        return None
    with io.open(yol, encoding='utf-8', errors='replace') as fh:
        return fh.read()


def _kod_govdesi(metin):
    """Strips comments and docstrings from Python source and returns THE CODE BODY.

        Why it is needed: a directory or API name APPEARING IN A COMMENT does not mean
        that code path is actually used. A plain text search cannot tell the two
        apart and answers the question "is there still unpatched code" wrongly.

        It does THE SAME JOB as screening/yon_kod_taramasi.kod_govdesi(), but the
        criterion is corrected (2026-08-21, measured). The version there rules strings
        out by LINE COUNT:  end_lineno - lineno >= 1.  That is wrong in both
        directions:
          * A SINGLE LINE docstring is not dropped -> a name in a comment is taken
            for code (a false positive; confirmed on a synthetic test).
          * A multi-line string REALLY USED in the code is dropped -> an embedded
            path goes unnoticed (a false negative).
        The right discriminator is not the line count but whether the string IS A
        DOCSTRING: a docstring is an expression statement whose value is a string
        constant (ast.Expr). A string used in code is always the child of some other
        node. So only string constants under an ast.Expr are dropped.

        On a file that cannot be parsed the docstring set stays empty, so at worst it
        falls back to the old (over-broad) behaviour; it never skips the check
        silently.

    """
    try:
        agac = ast.parse(metin)
        ds = set()
        for d in ast.walk(agac):
            if isinstance(d, ast.Expr) and isinstance(d.value, ast.Constant) \
               and isinstance(d.value.value, str):
                b = getattr(d, 'lineno', None)
                s = getattr(d, 'end_lineno', b)
                if b:
                    ds.update(range(b, (s or b) + 1))
    except Exception:
        ds = set()
    # THE LINE NUMBERS ARE PRESERVED: a docstring line is NOT DROPPED, it is EMPTIED.
    # Dropping shifts the numbers, and a shift breaks the alignment with the line sets
    # that come from the AST (message strings, for instance), after which the filter
    # silently does nothing. Measured: screening/run_all.py has 222 lines in the
    # original and 183 in the trimmed body.
    out = []
    for i, l in enumerate(metin.splitlines(), 1):
        out.append(u'' if i in ds else l.split(u'#', 1)[0])
    return u'\n'.join(out)


# Tools that read the mixed orientation directory BY DESIGN. Flagging these is
# wrong: the orientation audit and the canonical builder cannot do their job
# without reading that directory.
# The list must stay SHORT, and every entry carries a reason.
D9_MESRU_OKUYUCULAR = {
    u'orientation_audit.py':    u'the orientation auditor, whose job is to measure the mixed directory',
    u'build_canonical.py':    u'PRODUCES the canonical directory out of that one',
    u'orientation.py':             u'the source of the orientation definitions',
    u'orientation_code_scan.py': u'the sibling tool that scans for the same risk',
    u'orientation_report.py':     u'the report producer that documents the orientation decisions',
}

# Calls that print text to the screen or a log. The string INSIDE them is not A
# CODE PATH but a message shown to the user (measured: screening/run_all.py:183
# only prints the "Kaynak: ..." line, and was being counted as RISKLI for it).
D9_CIKTI_CAGRILARI = frozenset(
    [u'yaz', u'print', u'write', u'log', u'uyar', u'bilgi', u'hata', u'yazdir'])


def d9_karisik_klasor_yollari(metin, ad=u'', karisik=u'consensus sequences'):
    """Returns the code lines that REALLY read the mixed orientation directory.

        [(line_number, source_line), ...]  - an empty list means clean.

        Three filters, and the reason for each was measured (2026-08-21):
          1) docstrings and comments are stripped -> a name in a comment is not code
          2) the arguments of output calls are stripped -> a message printed to the
             screen is not code
          3) tools that read it by design are exempt -> D9_MESRU_OKUYUCULAR

        A plain text search could tell none of the three apart, and in the 2026-08-09
        run it produced five false positives.

    """
    if os.path.basename(ad) in D9_MESRU_OKUYUCULAR:
        return []
    try:
        agac = ast.parse(metin)
    except Exception:
        # If it cannot be parsed we fall back to plain text: far too broad, but not silent.
        return [(i, l) for i, l in enumerate(metin.splitlines(), 1)
                if karisik in l and not l.strip().startswith(u'#')]

    mesaj_satirlari = set()
    for d in ast.walk(agac):
        if not isinstance(d, ast.Call):
            continue
        f = d.func
        adi = getattr(f, 'id', None) or getattr(f, 'attr', None)
        if adi in D9_CIKTI_CAGRILARI:
            for arg in list(d.args) + [k.value for k in d.keywords]:
                for s in ast.walk(arg):
                    if isinstance(s, ast.Constant) and isinstance(s.value, str) \
                       and karisik in s.value:
                        b = getattr(s, 'lineno', None)
                        e = getattr(s, 'end_lineno', b)
                        if b:
                            mesaj_satirlari.update(range(b, (e or b) + 1))

    govde = _kod_govdesi(metin).splitlines()
    return [(i, l) for i, l in enumerate(govde, 1)
            if karisik in l and i not in mesaj_satirlari]


def xlsx_sayfalari(yol):
    """Excel -> {sheet: [[cell,...],...]}. On error it returns text.

        It is opened READ ONLY (read_only=True): the chance of writing to the file is
        zero.

    """
    # The path can be None: when there is no panel xlsx in the root,
    # Kaynaklar.panel_xlsx stays None. os.path.exists(None) was being called
    # directly, TypeError took the whole module down, and that is exactly why
    # 2 INTERNAL CONSISTENCY could not catch its planted error in the self test.
    # A missing file is a missing source, not a crash.
    if not yol or not os.path.exists(yol):
        return None
    try:
        import openpyxl
    except ImportError:
        return u'OPENPYXL_YOK'
    try:
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    except Exception as e:
        return u'ACILAMADI: %s: %s' % (type(e).__name__, e)
    out = collections.OrderedDict()
    try:
        for ad in wb.sheetnames:
            sh = wb[ad]
            out[ad] = [[(u'' if c is None else unicode_(c)) for c in r]
                       for r in sh.iter_rows(values_only=True)]
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def modul_yukle(yol, ad):
    """Load a .py file AS A MODULE (a script, not a package).

        The logic of verification/identity_verification.py and all_bin_identities.py
        IS NOT REWRITTEN here; it is imported from there. Rewriting it would create
        two separate decision logics and leave it unclear which one holds.

    """
    import importlib.util
    if not os.path.exists(yol):
        return None, u'there is no such file: %s' % yol
    try:
        sp = importlib.util.spec_from_file_location(ad, yol)
        m = importlib.util.module_from_spec(sp)
        sys.modules[ad] = m
        sp.loader.exec_module(m)
        return m, None
    except Exception as e:
        return None, u'%s: %s' % (type(e).__name__, e)


# ===========================================================================
# THE PROJECT SOURCES, defined in one place
# ===========================================================================
class Kaynaklar(object):
    """The one list of every file the audit reads.

        If a path changes it changes in ONE place. Every module reads from here, so
        the answer to "which file did you look at" can be traced from the report.

    """

    # NOTHING IS EVER WRITTEN TO THESE PATHS. Other sessions may be running.
    DOKUNULMAZ = (u'install.sh', u'build_index.sh',
                  os.path.join(u'verification', u'one_key.py'), u'KONSENSUS_YENIDEN')

    def __init__(self, kok):
        self.kok = os.path.abspath(kok)
        y = lambda *p: os.path.join(self.kok, *p)
        self.y = y
        # --- the decision tables
        # 2026-08-09 20:05 TRIED AND REVERTED: I tried pointing this path at
        # ESIK_VE_OLCUT_2026-08-09.tsv, because marking the 08-07 file GECERSIZ drops four
        # checks to ATLANDI. But the two files do not have THE SAME COLUMN STRUCTURE:
        # ESIK_VE_OLCUT has no forward and reverse primer columns, so the M2 consistency
        # check read a numeric field from there and produced 22 FALSE CRITICAL
        # contradictions. A false alarm is worse than an honest ATLANDI. The old path was
        # kept; those four checks stay ATLANDI and that is deliberate. The permanent fix is
        # to rewrite the checks against the new table's columns, not to change the path.
        self.nihai_siparis = y('NIHAI_SIPARIS_LISTESI_2026-08-07.tsv')
        # 2026-08-09: instead of a name with a fixed date, THE NEWEST valid version is chosen.
        self.esik_olcut = en_yeni_kaynak(self.kok, 'ESIK_VE_OLCUT_*.tsv')
        self.siparis_listesi = y('SIPARIS_LISTESI.tsv')
        self.hedef_disi = y('HEDEF_DISI_AYRINTI_2026-08-07.tsv')
        # --- panel tanimi
        self.ciftler = y('screening', 'pairs.tsv')
        self.hedef_uyelik = y('screening', 'target_membership.tsv')
        self.hedef_klad = y('screening', 'target_clades.tsv')
        self.hedefler_wsl = y('steps', 'targets.tsv')
        self.takson_esleme = y('screening', 'target_taxon_mapping.py')
        # --- excel
        # 2026-08-11: the delivery xlsx was moved to the archive (six of the pairs inside it
        # carried stale sequences). In its place stands the one file that is GENERATED on
        # every run; since its name carries a date, the newest is picked. If none is found
        # the old name is tried and the audit says "no such file"; it does NOT fall back
        # silently to the old file.
        import glob as _glob
        _ad = sorted(_glob.glob(y('PrimerJury_PANEL_*.xlsx')))
        # If none is found it is LEFT EMPTY and the audit says "no such file". FALLING BACK
        # to the old delivery file is exactly what we want to avoid: six of the pairs in
        # that file carried stale sequences, which is why it was archived.
        self.panel_xlsx = _ad[-1] if _ad else None
        # --- measurement outputs
        self.tek_protokol = y('ONE_PROTOCOL_RESULT', 'panel_tek_protokol.tsv')
        self.uyelik_turetme = y('uyelik_yeniden_turetme_uyelik_20260803.tsv')
        self.uyelik_ciftler = y('uyelik_yeniden_turetme_ciftler_20260803.tsv')
        # --- literatur ve toplanti
        self.literatur = y('LITERATUR_2026-08-07.md')
        self.toplanti = y('TOPLANTI_KARARLARI_SON_DURUM.md')
        # --- veri
        self.refdb = y('REFERANS_DB')
        self.konsensus_indeks = y('konsensus_kanonik', 'INDEKS.tsv')
        self.konsensus_kok = y('konsensus_kanonik')
        self.fastq = y('fastq files')
        self.kraken = y('kraken results')
        self.bracken = y('bracken results')
        # --- code (for the pattern scan). KONSENSUS_YENIDEN IS DELIBERATELY ABSENT.
        self.kod_klasorleri = [y('screening'), y('verification'), y('steps')]
        # --- ice aktarilacak mantik
        self.kimlik_dogrulama = y('verification', 'identity_verification.py')
        self.tum_kutu = y('verification', 'all_bin_identities.py')


# =========================================================================
# MODULE 1 - IDENTITY
# =========================================================================
# THE QUESTION: which organism is really in each bin, and is that what Kraken2 said?
#
# WHY A SEPARATE MEASUREMENT: Kraken2 is a k-mer based CLASSIFIER. It does not
# align, and it labels an organism absent from its database with the name of that
# organism's nearest relative. In this project more than one Kraken label was
# refuted by measurement (Trichoderma asperellum -> Petriella; Dictyostelium
# discoideum, label refuted). So every bin is re-measured independently, by
# alignment, against OFFLINE databases.
#
# SIX OUTPUTS ARE PRODUCED for each bin:
#   1 the best hit at species level (species, database, accession, identity %,
#     aligned bp)
#   2 a TYPE MATERIAL flag (much stronger evidence, in A SEPARATE column)
#   3 the second best hit plus the gap (a small gap means the species assignment is
#     NOT reliable)
#   4 DISCRIMINABILITY (can these two species be told apart with our data quality)
#   5 the N fraction and a mixture indicator
#   6 the Kraken comparison (label, confidence, did the name change)
# -------------------------------------------------------------------------

# --- the identity thresholds: TAKEN FROM identity_verification.py, not rewritten here.
#     (K.TUR_ESIGI, K.CINS_ESIGI, K.AYRIM_PAYI)

# If the N fraction exceeds this threshold the consensus IS NOT USED and the code
# falls back to the raw reads.
# THE REASON (measured): 43.6% of the F2-4_500148 consensus is N. Because Ns always
# count as a mismatch in alignment, that bin looked like an "unnameable lineage";
# looked at read by read, the bin turned out to be Microascaceae, and MIXED. The
# threshold is 20%: that value falls in a wide gap between the worst of the sound
# bins (0.2%) and the broken one (43.6%).
N_ESIGI = 20.0

# The sample sizes for the mixture and purity measurement. The raw read count runs
# into the thousands and aligning all of them is unnecessary. 60 reads catch a
# subpopulation of 10% with probability above 99% (1 - 0.9^60 = 0.998).
KARISIM_OKUMA = 60
KARISIM_PENCERE = 700      # the window used in read to read alignments (bp)
SAF_ESIGI = 90.0           # ESKI KURAL - artik TEK BASINA hukum vermez, bkz asagisi

# -------------------------------------------------------------------------
# THE 2026-08-09 FIX - THE MIXED BIN RULE
# -------------------------------------------------------------------------
# THE SYMPTOM: in the 2026-08-09 18:09 run, 62 of the 99 bins were marked MIXED.
#          Most of that was a false alarm. Take A1-2_1826872: of 60 reads, 53 were
#          in one cluster and the remaining 7 formed seven separate clusters of ONE
#          read each; the dominant share was 88.3% and, because the only criterion
#          was "dominant >= 90%", the bin counted as MIXED.
#          But single read clusters are the noise expected under the Nanopore error
#          rate (measured eps 1.3-2.4%); they are not a separate organism.
#
# THE NEW RULE: for a bin to count as MIXED there has to be at least one cluster of
#          MEANINGFUL size OUTSIDE the dominant one.
#
# THE MEANINGFUL THRESHOLD WAS MEASURED, NOT CHOSEN:
#   Bins whose dominant cluster is very high are single organisms by definition, so
#   every non-dominant cluster there is noise. The size distribution of those
#   clusters was measured (from the cluster data of the 2026-08-09 run, 98 bins):
#
#     clean definition   bins   noise clusters   largest noise cluster
#     dominant >=96.5%      7               10                       2
#     dominant >=95.0%     12               25                       2
#     dominant >=93.0%     19               50                       2
#     dominant >=91.5%     29               96                       4  <- the break
#
#   Under three separate definitions of "clean", the largest noise cluster observed
#   across 50 noise clusters in total is 2 READS; a noise cluster of 3 reads was
#   NEVER seen. By the rule of three: P(a noise cluster >= 3 reads) < 3/50 = 6%.
#   An independent second derivation, Poisson over the same clean subset: lam =
#   1.040, P(a spurious cluster >= 3) = 8.8%, so 17.9 clusters expected across 98
#   bins; 48 observed. The observed excess cannot be explained by noise, which means
#   clusters of >=3 reads carry real signal.
#
#   Hence the threshold: A MEANINGFUL CLUSTER = at least 3 reads AND at least 5% of
#   the sample. (With a sample of 60 reads, 3/60 = 5%; the proportional floor is
#   there so the rule does not drift if the sample size changes.)
#
# THE REPRESENTATION FLOOR: if the dominant cluster is less than HALF the sample,
#   the bin counts as MIXED even when its secondary clusters are small. That is not
#   a noise threshold but a representation requirement: the consensus is derived
#   from the medoid of the dominant cluster, and a consensus derived from a cluster
#   in the minority does not represent most of the reads. The 50% is not a chosen
#   parameter, it is the definition of a majority.
#
# THE EFFECT (measured, same data): MIXED 62 -> 37. Twenty-six bins went back to
#   PURE, and one bin (F1-4_101201, 55/4/1) went from PURE to MIXED. The genuinely
#   mixed ones are still caught; A2-1_1826872, for instance, whose dominant cluster
#   is 46.4% (26/8/3/2/1...), is still MIXED.
ANLAMLI_KUME_OKUMA = 3     # the measured noise ceiling is 2 reads; meaningful = >=3
ANLAMLI_KUME_ORAN = 5.0    # ve ornegin en az %5'i
BASKIN_TEMSIL_TABANI = 50.0  # baskin kume cogunlukta degilse konsensus temsil etmez


def saflik_hukmu(boyut, n_okuma):
    """The purity verdict from the cluster sizes. (purity, meaningful_clusters, reason_suffix)

        To count as MIXED there has to be a cluster outside the dominant one with >=3
        reads and >=5%, OR the dominant cluster has to have lost the majority.

    """
    if not boyut or not n_okuma:
        return u'OLCULEMEDI', [], u''
    baskin = 100.0 * boyut[0] / n_okuma
    anlamli = [b for b in boyut[1:]
               if b >= ANLAMLI_KUME_OKUMA and 100.0 * b / n_okuma >= ANLAMLI_KUME_ORAN]
    temsil_yok = baskin < BASKIN_TEMSIL_TABANI
    if anlamli:
        ek = (u'MEANINGFUL secondary cluster(s): %s reads (the threshold: >=%d reads '
              u'and >=%s per cent)'
              % (u', '.join(str(b) for b in anlamli), ANLAMLI_KUME_OKUMA,
                 vir(ANLAMLI_KUME_ORAN, 0)))
    elif temsil_yok:
        ek = (u'there is NO meaningful secondary cluster, but the dominant cluster is '
              u'not a majority either (%s%% < %s%%): the consensus does not represent '
              u'most of the reads'
              % (vir(baskin, 1), vir(BASKIN_TEMSIL_TABANI, 0)))
    else:
        ek = (u'every cluster apart from the dominant one is the size of noise '
              u'(the largest is %d reads < %d): the measured noise ceiling is 2 reads'
              % (max(boyut[1:]) if len(boyut) > 1 else 0, ANLAMLI_KUME_OKUMA))
    return (u'KARISIK' if (anlamli or temsil_yok) else u'SAF'), anlamli, ek

# The read to read clustering threshold IS NOT FIXED; it is derived from the
# measured error rate.
#
# WHY: even when two reads come from THE SAME organism, each carries its own read
# error, and the difference expected between them is 2*eps. Measured (F2-4_500148):
# eps = 2.208%, so two reads of the same organism look 4.4% apart on average. Under
# a fixed 97% threshold those reads fell into SEPARATE clusters, 60 reads split into
# 34 clusters, and every bin looked "mixed". What was being measured was not biology
# but read error. The threshold is computed to allow a tolerance of up to
# KARISIM_KAT times the expected pairwise difference.
KARISIM_KAT = 2.5
KARISIM_ESIK_TABAN = 88.0   # bundan gevsek olmaz (ayri cinsler birlesmesin)
KARISIM_ESIK_TAVAN = 99.0   # bundan siki olmaz (hatasiz veride bile pay birak)

# The aligned length floor. Matching a very short reference record at 99% is WEAKER
# evidence than matching a long record at 98%: a short record sees only a small
# window of the query. Measured: a 533 bp 28S record at 99.44% was beating a 1719 bp
# record at 99.01%.
ASGARI_HIZ_UZ = 400

# The short list and the alignment budget.
KL_UST = 120               # her veritabanindan kac aday tam degerlendirilsin
ON_PENCERE = 900           # on eleme hizalamasinda kullanilan sorgu penceresi
KESIN_UST = 24             # how many candidates are FULLY aligned after the pre-filter
ADAY_HAVUZU_KONTROL = 800  # the candidate pool kept during the pass (for memory)

# The SEPARABILITY coefficient: how many sigma. 3 sigma = 99.7 per cent confidence.
AYIRT_SIGMA = 3.0

# The small databases scanned in module 1's "hizli" mode. The big ones (SILVA,
# UNITE, PR2, ROD) are scanned in "tam" mode alone, and the reason is the measured
# time, written below.
HIZLI_VTB = (u'RefSeq bakteri 16S', u'RefSeq arke 16S', u'RefSeq mantar ITS',
             u'RefSeq mantar 28S', u'RefSeq mantar 18S', u'RefSeq ref_all2')


def _tip_kaydi(etiket, baslik):
    """Parse the TYPE MATERIAL flag out of a reference header.

        THERE ARE THREE STATES and all three differ:
          EVET      : the header says type material explicitly
          HAYIR     : the database DOES CARRY type information, but not on this record
          BILGI_YOK : the database's headers CARRY NO type information at all

        Counting "BILGI_YOK" as "HAYIR" would be wrong: SILVA, UNITE and ROD headers
        have no type field at all, and we cannot claim a record there is not type
        material. The measured coverage (2026-08-09, a header scan):
            fungi.ITS.fna      20 271 / 20 394 records "from TYPE material"
            fungi.28SrRNA.fna  12 845 / 12 890
            fungi.18SrRNA.fna   4 009 /  4 037
            ref_all2.fna       37 125 / 65 358
            bacteria.16S.fna        0 / 26 877   <- NO field (NR_ records unmarked)
            archaea.16S.fna         0 /  1 160   <- NO field
            SILVA SSU/LSU, UNITE, ROD   0        <- NO field

    """
    b = baslik or u''
    tasiyor = etiket in (u'RefSeq mantar ITS', u'RefSeq mantar 28S',
                         u'RefSeq mantar 18S', u'RefSeq ref_all2')
    if re.search(r'from\s+TYPE\s+material', b, re.I):
        return u'EVET'
    if re.search(r'\btype\s+strain\b', b, re.I):
        return u'EVET'
    if tasiyor:
        return u'HAYIR'
    return u'BILGI_YOK'


def _kayit_no(baslik):
    u"""Pull the record number out of a header (the token up to the first space or pipe)."""
    b = (baslik or u'').strip()
    if not b:
        return u'-'
    return re.split(r'[\s|]', b, 1)[0][:40] or u'-'


def _okuma_ornekle(fastq_yolu, n=KARISIM_OKUMA, en_az_uz=300):
    """Sample n reads at a regular interval from a FASTQ. Returns: (sequences, mean_error_rate).

        Sampling at a regular interval stops the reads at the start of the file
        (usually shorter or earlier) from dominating. The error rate is computed from
        Phred: mean_error = mean(10^(-Q/10)). The error CANNOT BE DERIVED from the
        mean Q; an exponential mean and an arithmetic mean are not the same thing, and
        the error rate would come out smaller than it is.

    """
    if not os.path.exists(fastq_yolu):
        return [], None
    diziler = []
    hata_top = 0.0
    hata_n = 0
    havuz = []
    with io.open(fastq_yolu, encoding='utf-8', errors='replace') as fh:
        d = None
        for i, satir in enumerate(fh):
            m = i % 4
            if m == 1:
                d = satir.strip().upper()
            elif m == 3 and d is not None:
                q = satir.strip()
                if len(d) >= en_az_uz:
                    havuz.append((d, q))
                d = None
    if not havuz:
        return [], None
    adim = max(1, len(havuz) // n)
    secilen = havuz[::adim][:n]
    for d, q in secilen:
        diziler.append(d)
        for c in q:
            hata_top += 10.0 ** (-(ord(c) - 33) / 10.0)
            hata_n += 1
    return diziler, (hata_top / hata_n if hata_n else None)


def _karisim_esigi(hata_orani):
    """Derive the clustering threshold from the measured read error rate.

        The expected difference between two reads from the same organism is 2*eps. The
        threshold is set with a margin of KARISIM_KAT on top of that. If the error
        rate could not be measured the cautious path is taken (a high error is
        assumed), because too tight a threshold makes a PURE bin look mixed, and that
        is the more expensive direction to be wrong in.

    """
    eps = hata_orani if hata_orani is not None else 0.02
    beklenen_fark = 2.0 * eps * 100.0
    esik = 100.0 - KARISIM_KAT * beklenen_fark
    return max(KARISIM_ESIK_TABAN, min(KARISIM_ESIK_TAVAN, esik))


def _kumele(K, diziler, esik, pencere=KARISIM_PENCERE):
    """Cluster the reads by pairwise identity. Returns: [[index,...], ...], largest first.

        Simple single linkage clustering: two reads join the same cluster if they are
        within `esik` of one another. The aim is not species assignment but the
        question HOW MANY SEPARATE ORGANISMS ARE THERE, and single linkage is enough
        for that, and cheap.

        Cost control: only the middle `pencere` bases of each read are used. A full
        length alignment would come to 1770 pairs x ~0.1 s = 3 minutes for 60 reads;
        with the window it drops to ~0.02 s.

    """
    n = len(diziler)
    if n == 0:
        return []
    parca = []
    for d in diziler:
        d = K.temizle(d)
        if len(d) > pencere:
            b = (len(d) - pencere) // 2
            d = d[b:b + pencere]
        parca.append(d)
    ebeveyn = list(range(n))

    def bul(a):
        while ebeveyn[a] != a:
            ebeveyn[a] = ebeveyn[ebeveyn[a]]
            a = ebeveyn[a]
        return a

    for i in range(n):
        for j in range(i + 1, n):
            if bul(i) == bul(j):
                continue                      # already in the same cluster, no alignment needed
            k1, _d1 = K.hizala(parca[i], parca[j])
            k2, _d2 = K.hizala(parca[i], K.rc(parca[j]))   # yon farki kume bozmasin
            if max(k1, k2) >= esik:
                ebeveyn[bul(i)] = bul(j)
    kume = collections.defaultdict(list)
    for i in range(n):
        kume[bul(i)].append(i)
    return sorted(kume.values(), key=len, reverse=True)


def _medoid(K, diziler, indeksler, pencere=KARISIM_PENCERE):
    """The MEDOID of a cluster: the read with the highest total similarity to the rest.

        A real read is chosen rather than an average or a consensus, so that the query
        sequence is not an invented chimera. If the cluster has one element, that
        element comes back.

    """
    if not indeksler:
        return None
    if len(indeksler) == 1:
        return diziler[indeksler[0]]
    parca = {}
    for i in indeksler:
        d = K.temizle(diziler[i])
        if len(d) > pencere:
            b = (len(d) - pencere) // 2
            d = d[b:b + pencere]
        parca[i] = d
    en_iyi, en_iyi_p = None, -1.0
    for i in indeksler:
        p = 0.0
        for j in indeksler:
            if i == j:
                continue
            p += K.hizala(parca[i], parca[j])[0]
        if p > en_iyi_p:
            en_iyi, en_iyi_p = i, p
    return diziler[en_iyi]


def _pencere_sec(dizi, pencere=ON_PENCERE):
    """Pick the `pencere` base stretch of the query with the FEWEST Ns.

        Using a region full of Ns in the pre-filter alignment drowns the candidate
        ranking in noise (an N always counts as a mismatch). A sliding window picks
        the cleanest region.

    """
    if len(dizi) <= pencere:
        return dizi
    adim = max(1, (len(dizi) - pencere) // 40)
    en_iyi, en_az = 0, None
    for b in range(0, len(dizi) - pencere + 1, adim):
        n = dizi.count('N', b, b + pencere)
        if en_az is None or n < en_az:
            en_iyi, en_az = b, n
            if n == 0:
                break
    return dizi[en_iyi:en_iyi + pencere]


def _ayirt_edilebilir(K, en_iyi, ikinci, hata_orani, sorgu_uz):
    """CAN WE TELL THE BEST TWO SPECIES APART WITH OUR DATA QUALITY?

        This rule is IN THE CODE; it is not decided by hand. Two independent tests are
        applied and BOTH have to pass:

        (A) REFERENCE SEPARATION - are the two reference sequences different enough
            from one another? The difference between them is D_ref bases. Our read
            error is eps per read, so over L aligned bases the expected number of
            erroneous bases is E = eps*L with a spread of ~sqrt(E). For two species to
            be separable,
                D_ref >= AYIRT_SIGMA * sqrt(E + 1)   and   D_ref >= 3
            must hold. Otherwise the two species look IDENTICAL at our data quality,
            and giving a species name claims more than the measurement can carry.

        (B) MEASUREMENT SEPARATION - is the gap between our two hits larger than the
            noise? The best hit has m1 mismatched bases and the second m2. The
            counting noise is ~sqrt(m1+m2). For them to be separable,
                |m1 - m2| >= AYIRT_SIGMA * sqrt(m1 + m2 + 1)
            must hold. Otherwise the two candidates are equal within measurement error.

        Returns: dict(ayrilir, sebep, d_ref, esik_a, olcum_farki, esik_b)

    """
    out = dict(ayrilir=None, sebep=u'', d_ref=None, esik_a=None,
               olcum_farki=None, esik_b=None)
    if not en_iyi or not ikinci:
        out['ayrilir'] = None
        out['sebep'] = u'there is no second hit, so the separability COULD NOT BE MEASURED'
        return out
    eps = hata_orani if hata_orani is not None else 0.01   # olculemezse temkinli
    L = float(en_iyi.get('hiz_uz') or sorgu_uz or 1)

    # (A) align the two references TO ONE ANOTHER
    r1 = en_iyi.get('dizi') or u''
    r2 = ikinci.get('dizi') or u''
    if not r1 or not r2:
        out['sebep'] = u'the reference sequences are not at hand, so the separability COULD NOT BE MEASURED'
        return out
    kisa, uzun = (r1, r2) if len(r1) <= len(r2) else (r2, r1)
    ref_kimlik, _u = K.hizala(kisa, uzun)
    d_ref = (100.0 - ref_kimlik) / 100.0 * len(kisa)
    esik_a = max(3.0, AYIRT_SIGMA * math.sqrt(eps * L + 1.0))

    # (B) the difference between our own two measurements
    m1 = en_iyi.get('uzaklik')
    m2 = ikinci.get('uzaklik')
    if m1 is None or m2 is None:
        out['sebep'] = u'there are no mismatch counts, so the separation of the measurements COULD NOT BE MEASURED'
        out['d_ref'] = d_ref
        out['esik_a'] = esik_a
        return out
    olcum_farki = abs(m1 - m2)
    esik_b = AYIRT_SIGMA * math.sqrt(m1 + m2 + 1.0)

    a_gecti = d_ref >= esik_a
    b_gecti = olcum_farki >= esik_b
    out.update(d_ref=d_ref, esik_a=esik_a, olcum_farki=olcum_farki, esik_b=esik_b,
               ayrilir=bool(a_gecti and b_gecti))
    if a_gecti and b_gecti:
        out['sebep'] = (u'the two references differ by %s bases (the threshold is %s) '
                        u'and our two measurements differ by %s bases (the threshold '
                        u'is %s), so they SEPARATE'
                        % (vir(d_ref, 1), vir(esik_a, 1), vir(olcum_farki, 1), vir(esik_b, 1)))
    elif not a_gecti:
        out['sebep'] = (u'the two REFERENCES differ from one another by only %s bases, '
                        u'and with our error rate of %s the needed threshold is %s, so '
                        u'these two species CANNOT BE SEPARATED at our data quality'
                        % (vir(d_ref, 1), vir(100 * eps, 3), vir(esik_a, 1)))
    else:
        out['sebep'] = (u'the measurements differ by %s bases against a noise threshold '
                        u'of %s bases, so the two candidates are EQUAL within the '
                        u'measurement error' % (vir(olcum_farki, 1), vir(esik_b, 1)))
    return out


def _kraken_haritasi(kraken_kok):
    """Map the barcode reports to the bin classes. Returns: {(class,no): report_path}.

        HOW THE MAPPING WAS VERIFIED: the reports in each class directory are sorted
        by barcode number and attached in order to bins 1..4 of that class
        (A1 -> 01..04, A2 -> 05..08, F2 -> 09..12, F1 -> 13..16, B -> 17..20).
        The mapping was verified (2026-08-09): ALL 99 taxids of the 20 bins are
        present in their assigned barcode report, with 0 unmatched. Under a wrong
        mapping most of the taxids would not have been found in the report.

    """
    harita = {}
    if not os.path.isdir(kraken_kok):
        return harita
    for sinif in sorted(os.listdir(kraken_kok)):
        kls = os.path.join(kraken_kok, sinif)
        if not os.path.isdir(kls):
            continue
        raporlar = sorted(glob.glob(os.path.join(kls, '*_kraken2.report')))
        for i, r in enumerate(raporlar, 1):
            harita[(sinif, i)] = r
    return harita


def _kraken_oku(yol):
    """A Kraken2 report -> {taxid: dict(name, rank, percent, assigned, clade)}.

        The columns of a Kraken2 report: percent, clade_reads, direct_reads, rank,
        taxid, name. A "confidence value" IS NOT a separate column in a Kraken2
        report; the nearest equivalent is the percentage of reads assigned directly
        to that taxon, so that is what goes into the "confidence" column, and the
        report says so openly.

    """
    out = {}
    if not os.path.exists(yol):
        return out
    with io.open(yol, encoding='utf-8', errors='replace') as fh:
        for satir in fh:
            p = satir.rstrip(u'\n').split(u'\t')
            if len(p) < 6:
                continue
            try:
                out[p[4].strip()] = dict(
                    yuzde=sayi(p[0]), altagac=int(sayi(p[1]) or 0),
                    atanan=int(sayi(p[2]) or 0), duzey=p[3].strip(),
                    ad=p[5].strip())
            except (ValueError, TypeError):
                continue
    return out


def modul_1_kimlik(kay, rap, kn, kip=u'hizli', yalniz=None, tavan=0):
    """MODULE 1 - measure the identity of every bin and put it beside Kraken2's.

        mode: 'yok'   -> do not run at all (the finding: ATLANDI)
              'hizli' -> the small RefSeq sets only (the measured time is below)
              'tam'   -> every offline database (SILVA SSU/LSU, UNITE, PR2 and ROD included)

    """
    M = u'1 IDENTITY'
    t_basla = time.time()

    if kip == u'yok':
        rap.atla(M, u'M1-KAPALI', u'a species level identity measurement for every bin',
                 u'--m1-mode yok was given, so the module was turned off on purpose', u'-')
        return

    # --- import the decision logic from verification (it IS NOT REWRITTEN here)
    K, hata = modul_yukle(kay.kimlik_dogrulama, 'kimlik_dogrulama')
    if K is None:
        rap.atla(M, u'M1-MOTOR', u'identity_verification.py\'s alignment engine must be loadable',
                 hata, kay.kimlik_dogrulama)
        return
    T, hata2 = modul_yukle(kay.tum_kutu, 'tum_kutu_kimlikleri')
    if T is None:
        rap.atla(M, u'M1-MOTOR2', u'all_bin_identities.py\'s bulk scan must be loadable',
                 hata2, kay.tum_kutu)
        return
    try:
        import numpy  # noqa: F401  (hizala() numpy'siz cok yavas)
    except ImportError:
        rap.atla(M, u'M1-NUMPY', u'numpy must be installed (the alignment engine uses it)',
                 u'numpy ice aktarilamadi', u'-')
        return

    # Shrink the candidate pool for memory. This affects only the copy INSIDE THIS
    # PROCESS; the file verification/identity_verification.py IS UNCHANGED.
    # The reason: by that file's own note, the worst measured pre-filter position is 45;
    # a pool of 800 is 17 times that limit, so the cut is not binding.
    K.ADAY_HAVUZU = ADAY_HAVUZU_KONTROL

    # --- the bin inventory: the CANONICAL consensus index (all of it SENSE)
    ind = tsv_oku(kay.konsensus_indeks, yorum=None)
    if ind is None:
        rap.atla(M, u'M1-ENVANTER', u'konsensus_kanonik/INDEKS.tsv must be readable',
                 u'there is no such file', kay.konsensus_indeks)
        return
    if not ind:
        rap.ekle(M, u'M1-ENVANTER-BOS', KRITIK,
                 u'the canonical consensus index must hold bins',
                 u'the index is EMPTY, not one bin could be read', kay.konsensus_indeks,
                 u'reproduce the index with build_canonical.py')
        return

    kutular = []
    for r in ind:
        yol = os.path.join(kay.konsensus_kok, r.get('dosya', u''))
        if not os.path.exists(yol):
            rap.ekle(M, u'M1-KONSENSUS-YOK', CIDDI,
                     u'the consensus file named in the index must be present on disk',
                     u'%s -> there is no such file' % r.get('dosya'), kay.konsensus_indeks,
                     u'The bin cannot be measured; the index and the directory have drifted apart.')
            continue
        try:
            with io.open(yol, encoding='utf-8', errors='replace') as fh:
                diz = u''.join(s.strip() for s in fh if not s.startswith(u'>')).upper()
        except IOError as e:
            rap.ekle(M, u'M1-KONSENSUS-OKUNAMADI', CIDDI,
                     u'the consensus file must be readable', u'%s: %s' % (r.get('dosya'), e),
                     yol)
            continue
        kutular.append(dict(kutu=r.get('kutu'), sinif=r.get('sinif'),
                            yol=yol, dizi=K.temizle(diz),
                            yon=r.get('eski_yon'), cevrildi=r.get('cevrildi')))

    if yalniz:
        istenen = set(x.strip() for x in yalniz.split(u',') if x.strip())
        kutular = [k for k in kutular if k['kutu'] in istenen]
    if tavan:
        kutular = kutular[:tavan]
    if not kutular:
        rap.atla(M, u'M1-KUTU-YOK', u'at least one bin must be measured',
                 u'no bin was left after the filter', kay.konsensus_indeks)
        return

    yaz(u'M1: %d bins, mode=%s' % (len(kutular), kip))

    # -----------------------------------------------------------------
    # STAGE A - the N fraction, the read error rate, mixture and purity.
    # This stage READS NO DATABASE AT ALL; it measures entirely from the bin's own data.
    # -----------------------------------------------------------------
    yaz(u'M1/A: N fraction, read error rate and mixture measurement...')
    # We count how many bins were REALLY computed. Bins read from a checkpoint take zero
    # seconds, and counting those too gives "0 seconds per bin", which makes the full
    # scan estimate look unrealistically short.
    taze_a = [0]
    can = Canlilik(u'M1/A bin', len(kutular))
    for i, kb in enumerate(kutular, 1):
        can.vur(i, kb['kutu'])
        diz = kb['dizi']
        n_say = diz.count('N')
        kb['n_oran'] = 100.0 * n_say / max(1, len(diz))
        kb['kons_uz'] = len(diz)

        # the FASTQ belonging to the bin: <class-no>/<class-no>-reads_<taxid>.fastq
        grup = kb['kutu'].rsplit(u'_', 1)[0]
        taxid = kb['kutu'].rsplit(u'_', 1)[-1]
        kb['grup'] = grup
        kb['taxid'] = taxid
        fq = os.path.join(kay.fastq, grup, u'%s-reads_%s.fastq' % (grup, taxid))
        if not os.path.exists(fq):
            adaylar = glob.glob(os.path.join(kay.fastq, grup, u'*%s*.fastq*' % taxid))
            fq = adaylar[0] if adaylar else None
        kb['fastq'] = fq

        anahtar = md5_metin(u'M1A', VERSIYON, kb['kutu'], dosya_imzasi(kb['yol']),
                            dosya_imzasi(fq or u'-'), KARISIM_OKUMA, KARISIM_KAT,
                            KARISIM_ESIK_TABAN, KARISIM_ESIK_TAVAN, SAF_ESIGI)
        onbellek = kn.oku(anahtar, [kb['yol']] + ([fq] if fq else []))
        if onbellek:
            kb.update(onbellek)
            continue
        taze_a[0] += 1          # this bin was REALLY computed (for the time estimate)

        if not fq:
            kb['hata_orani'] = None
            kb['okuma_sayisi'] = None
            kb['baskin_oran'] = None
            kb['saflik'] = u'OLCULEMEDI'
            kb['saflik_sebep'] = u'the FASTQ belonging to the bin was not found'
            kb['kume_boyutlari'] = []
            kb['okuma_sorgu'] = None
            rap.ekle(M, u'M1-FASTQ-YOK', UYARI,
                     u'the raw read file of every bin must be present',
                     u'there is no FASTQ for %s, so the mixture and the error rate COULD NOT BE MEASURED' % kb['kutu'],
                     os.path.join(kay.fastq, grup))
        else:
            okumalar, eps = _okuma_ornekle(fq)
            kb['hata_orani'] = eps
            kb['okuma_sayisi'] = len(okumalar)
            if len(okumalar) < 5:
                kb['baskin_oran'] = None
                kb['saflik'] = u'OLCULEMEDI'
                kb['saflik_sebep'] = (u'there is no read long enough to sample (%d)'
                                      % len(okumalar))
                kb['kume_boyutlari'] = []
                kb['okuma_sorgu'] = None
            else:
                k_esik = _karisim_esigi(eps)
                kumeler = _kumele(K, okumalar, k_esik)
                boyut = [len(c) for c in kumeler]
                baskin = 100.0 * boyut[0] / len(okumalar)
                kb['kume_boyutlari'] = boyut
                kb['baskin_oran'] = baskin
                kb['karisim_esigi'] = k_esik
                # 2026-08-09: the verdict now comes not from the dominant share alone but
                # from the presence of a MEANINGFUL secondary cluster. The reasoning, with
                # the measurement, is in the block above saflik_hukmu().
                kb['saflik'], kb['anlamli_kumeler'], _ek = saflik_hukmu(
                    boyut, len(okumalar))
                kb['saflik_sebep'] = (
                    u'the %d sampled reads fell into %d clusters (%s); the dominant '
                    u'cluster is %s%%. The clustering threshold was derived from the '
                    u'measured read error rate: eps=%s%% -> a threshold of %s%%. '
                    u'THE VERDICT: %s'
                    % (len(okumalar), len(boyut),
                       u'/'.join(str(b) for b in boyut[:8]) +
                       (u'/...' if len(boyut) > 8 else u''),
                       vir(baskin, 1), vir(100 * eps, 3) if eps is not None else u'?',
                       vir(k_esik, 1), _ek))
                kb['okuma_sorgu'] = _medoid(K, okumalar, kumeler[0])
        kn.yazdir(anahtar, dict((a, kb.get(a)) for a in (
            'n_oran', 'kons_uz', 'hata_orani', 'okuma_sayisi', 'baskin_oran',
            'saflik', 'saflik_sebep', 'kume_boyutlari', 'okuma_sorgu',
            'karisim_esigi')))
    a_sure = can.bitti(len(kutular))
    # The preparation time per bin is measured ONLY from freshly computed bins.
    if taze_a[0] >= 2:
        rap.olcum[u'_m1a_kutu_sn'] = a_sure / float(taze_a[0])
        rap.olcum[u'M1/A preparation'] = u'%d bins (%d fresh, %d from a checkpoint), ' \
                                         u'%s -> %s s per bin' % (
            len(kutular), taze_a[0], len(kutular) - taze_a[0], sure_metni(a_sure),
            vir(a_sure / float(taze_a[0]), 1))
    else:
        rap.olcum[u'M1/A preparation'] = (
            u'%d of the %d bins were read from a checkpoint, so the preparation time '
            u'per bin WAS NOT MEASURED IN THIS RUN'
            % (len(kutular) - taze_a[0], len(kutular)))

    # --- CHOOSING THE QUERY: when N is high the consensus IS NOT USED and the
    #     reads are used instead
    sorgular = {}
    for kb in kutular:
        if kb['n_oran'] > N_ESIGI and kb.get('okuma_sorgu'):
            kb['sorgu_kaynagi'] = u'RAW READ (the medoid of the dominant cluster)'
            kb['sorgu'] = K.temizle(kb['okuma_sorgu'])
            rap.ekle(M, u'M1-N-YUKSEK', UYARI,
                     u'the N ratio of the consensus must be below %%%s' % vir(N_ESIGI, 0),
                     u'%s: the N ratio is %%%s, so the consensus was not used and the raw reads were used instead (purity: %s)' % (kb['kutu'], vir(kb['n_oran'], 1), kb.get('saflik')),
                     kb['yol'],
                     u'This bin looks "unnameable" through the consensus; the read based measurement is the one to go by.')
        elif kb['n_oran'] > N_ESIGI:
            kb['sorgu_kaynagi'] = u'CONSENSUS (N is high but there are no reads)'
            kb['sorgu'] = kb['dizi']
            rap.ekle(M, u'M1-N-YUKSEK-OKUMASIZ', CIDDI,
                     u'in a bin with a high N ratio it must be possible to fall back on the raw reads',
                     u'%s: the N ratio is %%%s but a read query could not be produced'
                     % (kb['kutu'], vir(kb['n_oran'], 1)), kb['yol'])
        else:
            kb['sorgu_kaynagi'] = u'KONSENSUS'
            kb['sorgu'] = kb['dizi']
        sorgular[kb['kutu']] = kb['sorgu']

    # -----------------------------------------------------------------
    # STAGE B - the database scan. Every database is scanned in ONE pass and every bin
    # is fed from that same pass (a separate pass per bin would come to 100x12=1200 full
    # file passes).
    # -----------------------------------------------------------------
    vtb = [(e, d, t) for e, d, t, kullan, _n in K.VTB if kullan]
    if kip == u'hizli':
        vtb = [v for v in vtb if v[0] in HIZLI_VTB]
    yaz(u'M1/B: %d databases to scan: %s' % (len(vtb), u', '.join(v[0] for v in vtb)))

    bulgular_kutu = collections.defaultdict(list)   # bin -> [hit, ...]
    for ei, (etiket, dosya, lokus) in enumerate(vtb, 1):
        yol = os.path.join(kay.refdb, dosya)
        if not os.path.exists(yol):
            rap.ekle(M, u'M1-VTB-YOK', CIDDI,
                     u'the database file in the list must be present on disk',
                     u'%s (%s) is not there' % (etiket, dosya), kay.refdb,
                     u'This database could not vote for any bin.')
            continue
        anahtar = md5_metin(u'M1B', VERSIYON, etiket, dosya_imzasi(yol), KL_UST,
                            ON_PENCERE, KESIN_UST, ADAY_HAVUZU_KONTROL,
                            md5_metin(*[u'%s=%s' % (k, sorgular[k]) for k in sorted(sorgular)]))
        onbellek = kn.oku(anahtar, [yol])
        if onbellek is not None:
            for kutu, isabetler in onbellek.items():
                bulgular_kutu[kutu].extend(isabetler)
            yaz(u'  [%d/%d] %s - read from checkpoint' % (ei, len(vtb), etiket))
            continue

        t0 = time.time()
        mb = os.path.getsize(yol) / 1e6
        yaz(u'  [%d/%d] %s (%s MB) scanning...' % (ei, len(vtb), etiket, vir(mb, 1)))
        can2 = Canlilik(u'%s kayit' % etiket, None, aralik=20.0)
        try:
            kls, taranan = T.toplu_kisa_liste(K, yol, sorgular, KL_UST,
                                              ilerle=lambda n: can2.vur(n))
        except MemoryError:
            rap.atla(M, u'M1-VTB-BELLEK', u'%s taranabilmeli' % etiket,
                     u'out of memory (MemoryError), so lower the bin count with --m1-cap or run the database on its own', yol)
            continue
        except Exception as e:
            # ERRORS ARE NOT SWALLOWED: which database could not be scanned, and why, is reported.
            rap.atla(M, u'M1-VTB-HATA', u'%s taranabilmeli' % etiket,
                     u'%s: %s' % (type(e).__name__, e), yol)
            continue
        tarama_sn = time.time() - t0
        rap.olcum[u'M1 scan %s' % etiket] = (
            u'%s MB, %d records, %s  (%s MB/s)'
            % (vir(mb, 1), taranan, sure_metni(tarama_sn), vir(mb / max(0.001, tarama_sn), 2)))

        # --- compare against the expected record count (this catches a truncated scan)
        bekl = getattr(T, 'BEKLENEN_KAYIT', {}).get(etiket)
        if bekl and taranan < bekl:
            rap.ekle(M, u'M1-KAPSAM-EKSIK', CIDDI,
                     u'the %s scan must see all %d records' % (etiket, bekl),
                     u'only %d records were scanned (%%%s)'
                     % (taranan, vir(100.0 * taranan / bekl, 1)), yol,
                     u'A pruned scan can miss the real best hit.')

        # --- a two stage alignment for each bin
        t1 = time.time()
        can3 = Canlilik(u'%s hizalama' % etiket, len(kutular))
        vtb_sonuc = {}
        for ki, kb in enumerate(kutular, 1):
            can3.vur(ki, kb['kutu'])
            kl = kls.get(kb['kutu']) or []
            if not kl:
                continue
            q = kb['sorgu']
            qp = _pencere_sec(q, ON_PENCERE)
            qp_rc = K.rc(qp)
            # THE PRE-FILTER: rank every candidate with a short window (cheap).
            #
            # TWO THINGS ARE CRITICAL HERE:
            # (1) hizala() is INFIX: it places the SHORT sequence inside the LONG one.
            #     If the query is longer than the target the alignment stops meaning
            #     anything and the identity falls artificially to the length ratio
            #     (measured: a 3707 bp query against a 1700 bp 18S record gave 47% -
            #     that was not a similarity, it was the length difference itself). So
            #     on every call the shorter one is made the query and the longer one
            #     the target; kimlik_dogrulama.kl_degerlendir() does exactly the same.
            # (2) The ORIENTATION of reference records varies from set to set. A
            #     measurement searching one direction misses part of the records, so
            #     both directions are tried at the window stage and THE WINNING
            #     DIRECTION is used for the full alignment (aligning the full length
            #     twice would cost twice as much).
            on = []
            for a in kl:
                t = a['dizi']
                d, u_ = (qp, t) if len(qp) <= len(t) else (t, qp)
                ileri, _x = K.hizala(d, u_)
                d2, u2 = (qp_rc, t) if len(qp_rc) <= len(t) else (t, qp_rc)
                geri, _y = K.hizala(d2, u2)
                on.append((max(ileri, geri), a, u'+' if ileri >= geri else u'-'))
            on.sort(key=lambda x: -x[0])
            # TYPE MATERIAL must be evaluated without fail: similarity to a type record
            # is far stronger evidence, and losing it in the pre-filter is unacceptable.
            secili = [(a, yon) for _k, a, yon in on[:KESIN_UST]]
            secili_bas = set(id(a) for a, _y in secili)
            for _k, a, yon in on[KESIN_UST:]:
                if _tip_kaydi(etiket, a['baslik']) == u'EVET' and id(a) not in secili_bas:
                    secili.append((a, yon))
                    secili_bas.add(id(a))
                    if len(secili) >= KESIN_UST + 8:
                        break
            # THE DEFINITIVE MEASUREMENT: a full length alignment, in the winning direction,
            # the SHORT one inside the LONG one.
            isabetler = []
            q_rc = K.rc(q)
            for a, yon in secili:
                qq = q if yon == u'+' else q_rc
                t = a['dizi']
                d, u_ = (qq, t) if len(qq) <= len(t) else (t, qq)
                kim, uz = K.hizala(d, u_)
                cins, tur, tam = K.ad_coz(a['baslik'])
                isabetler.append(dict(
                    kimlik=kim, uzaklik=int(uz), baslik=a['baslik'][:200],
                    kayit=_kayit_no(a['baslik']), cins=cins, tur=tur,
                    vtb=etiket, lokus=lokus, sira=a.get('sira'),
                    tip=_tip_kaydi(etiket, a['baslik']),
                    hiz_uz=min(len(q), len(a['dizi'])),
                    dizi=a['dizi'][:6000]))
            isabetler.sort(key=lambda x: -x['kimlik'])
            vtb_sonuc[kb['kutu']] = isabetler[:6]
            bulgular_kutu[kb['kutu']].extend(isabetler[:6])
        hiz_sn = time.time() - t1
        rap.olcum[u'M1 alignment %s' % etiket] = (
            u'%d bins, %s (%s s per bin)'
            % (len(kutular), sure_metni(hiz_sn), vir(hiz_sn / max(1, len(kutular)), 2)))
        kn.yazdir(anahtar, vtb_sonuc)
        yaz(u'  [%d/%d] %s done: scan %s + alignment %s'
            % (ei, len(vtb), etiket, sure_metni(tarama_sn), sure_metni(hiz_sn)))

    # -----------------------------------------------------------------
    # STAGE C - the verdict plus the Kraken comparison
    # -----------------------------------------------------------------
    yaz(u'M1/C: verdict and Kraken comparison...')
    kharita = _kraken_haritasi(kay.kraken)
    if not kharita:
        rap.atla(M, u'M1-KRAKEN-YOK', u'the Kraken2 reports must be readable',
                 u'no report was found under "kraken results"', kay.kraken)
    kraken_onbellek = {}

    satirlar = []
    olculemeyen = []
    for kb in kutular:
        isabetler = sorted(bulgular_kutu.get(kb['kutu']) or [],
                           key=lambda x: -x['kimlik'])
        # --- Kraken tarafi
        grup = kb['grup']
        sinif = grup.split(u'-')[0]
        try:
            no = int(grup.split(u'-')[1])
        except (IndexError, ValueError):
            no = None
        krapor = kharita.get((sinif, no))
        k_etiket = k_guven = k_duzey = u'-'
        if krapor:
            if krapor not in kraken_onbellek:
                kraken_onbellek[krapor] = _kraken_oku(krapor)
            kt = kraken_onbellek[krapor].get(kb['taxid'])
            if kt:
                k_etiket = kt['ad']
                k_duzey = kt['duzey']
                # A Kraken2 report HAS NO separate "confidence" column; the
                # closest counterpart is the percentage of the reads assigned
                # DIRECTLY to this taxon.
                k_guven = u'%s%% (the reads assigned directly; Kraken2 gives no ' \
                          u'separate confidence column)' % vir(kt['yuzde'], 2)
            else:
                k_etiket = u'NOT IN THE REPORT'
                rap.ekle(M, u'M1-KRAKEN-TAXID-YOK', UYARI,
                         u'a bin\'s taxid must be present in its own barcode report',
                         u'%s: taxid %s is not in the %s report'
                         % (kb['kutu'], kb['taxid'], os.path.basename(krapor)), krapor)
        else:
            k_etiket = u'THE REPORT DID NOT MATCH'

        if not isabetler:
            olculemeyen.append(kb['kutu'])
            rap.ekle(M, u'M1-OLCULEMEDI', CIDDI,
                     u'there must be at least one database hit for every bin',
                     u'%s: there is no hit in any database, so it COULD NOT BE MEASURED' % kb['kutu'],
                     kb['yol'],
                     u'It was not skipped silently; there is NO identity claim for this bin.')
            satirlar.append(dict(
                kutu=kb['kutu'], kraken_etiket=k_etiket, kraken_guven=k_guven,
                olculen=u'OLCULEMEDI', duzey=u'OLCULEMEDI', kimlik=None,
                vtb=u'-', kayit=u'-', tip=u'-', ikinci=u'-', fark=None,
                ayrilir=u'-', ayrim_sebep=u'there is no hit', n_oran=kb['n_oran'],
                saflik=kb.get('saflik'), saflik_sebep=kb.get('saflik_sebep'),
                sorgu_kaynagi=kb['sorgu_kaynagi'], uyusan=0, uyusmayan=u'-',
                ad_degisti=u'OLCULEMEDI'))
            continue

        # --- CHOOSING THE BEST HIT: identity ON ITS OWN is not enough.
        #
        # Matching a very short reference record at 99.4% is weaker evidence than
        # matching a long one at 99.0%: the short record sees only a small window of
        # the query, and if that window is a conserved region then nearly every
        # relative gives 99%. Measured (F2-1_500148): a 533 bp 28S record came first
        # at 99.44% and knocked out a 1719 bp record at 99.01%.
        #
        # THE RULE: hits falling inside the NOISE BAND of the highest identity are
        # statistically equal, and among those the one carrying the LONGEST alignment
        # is chosen. The band is derived from counting noise (3 sigma).
        # The choice is made in TWO STAGES and THE SAME rule is used at both:
        #   1) each database picks its own winner,
        #   2) the overall winner is picked from among the database winners.
        # Using different rules at the two stages made the winning genus fail to get a
        # vote from its own database (measured: the overall choice said
        # Pseudallescheria while the vote of the same database went to Lomentospora,
        # so the winner appeared to be "supported by 0 databases").
        def _band_sec(havuz):
            t = max(havuz, key=lambda x: x['kimlik'])
            b = AYIRT_SIGMA * math.sqrt((t.get('uzaklik') or 0) + 1.0) / max(
                1.0, float(t.get('hiz_uz') or 1)) * 100.0
            esit = [h for h in havuz if (t['kimlik'] - h['kimlik']) <= b]
            return max(esit, key=lambda h: (h.get('hiz_uz') or 0, h['kimlik'])), t

        vtb_grup = collections.defaultdict(list)
        for h in isabetler:
            vtb_grup[h['vtb']].append(h)
        vtb_en_iyi = {}
        for v, hh in vtb_grup.items():
            vtb_en_iyi[v] = _band_sec(hh)[0]
        en_iyi, tepe = _band_sec(list(vtb_en_iyi.values()))
        if en_iyi is not tepe:
            rap.ekle(M, u'M1-KISA-HIZALAMA', BILGI,
                     u'the best hit must be chosen by weighing the identity and the aligned length together',
                     u'%s: the highest identity is %s%% but over only %s bp (%s); the %s%% hit over %s bp, which stays inside the noise band, was chosen (%s)'
                     % (kb['kutu'], vir(tepe['kimlik'], 2), tepe.get('hiz_uz'),
                        tepe['kayit'], en_iyi.get('hiz_uz'),
                        vir(en_iyi['kimlik'], 2), en_iyi['kayit']), kb['yol'])
        if (en_iyi.get('hiz_uz') or 0) < ASGARI_HIZ_UZ:
            rap.ekle(M, u'M1-HIZALAMA-KISA', CIDDI,
                     u'the identity verdict must be given over at least %d bp' % ASGARI_HIZ_UZ,
                     u'%s: the best hit was measured over only %s bp (%s%%, %s)'
                     % (kb['kutu'], en_iyi.get('hiz_uz'), vir(en_iyi['kimlik'], 2),
                        en_iyi['kayit']), kb['yol'],
                     u'In a short window a conserved region gives every relative a high identity; that is not enough evidence for a species assignment.')
        # Let the verdict order put the chosen hit first as well.
        isabetler = [en_iyi] + [h for h in isabetler if h is not en_iyi]
        # THE SECOND BEST: a species DIFFERENT from the best one (a second record of
        # the same species answers no question about separation and would be a tautology)
        # The second hit must also be measured over AT LEAST ASGARI_HIZ_UZ; showing a
        # short record as "the competitor" sets the separation question up wrongly.
        def _ikinci_sec(uzunluk_sarti):
            for h in isabetler[1:]:
                if uzunluk_sarti and (h.get('hiz_uz') or 0) < ASGARI_HIZ_UZ:
                    continue
                if (h.get('tur') or h.get('cins')) != (en_iyi.get('tur') or
                                                       en_iyi.get('cins')):
                    return h
            return None
        ikinci = _ikinci_sec(True) or _ikinci_sec(False)
        fark = (en_iyi['kimlik'] - ikinci['kimlik']) if ikinci else None
        if fark is not None and fark < 0:
            # A negative gap: the competitor's raw identity is higher, but over a SHORTER
            # alignment. That is not a contradiction, it is the length against identity
            # trade-off, and the reader should not have to guess it, so it is written out.
            rap.ekle(M, u'M1-RAKIP-DAHA-YUKSEK', UYARI,
                     u'the identity of the best hit chosen must not be lower than its rival\'s',
                     u'%s: the hit chosen is %s%% (%s bp, %s) and the rival %s%% (%s bp, %s), so the rival has a higher identity but over a SHORTER alignment'
                     % (kb['kutu'], vir(en_iyi['kimlik'], 2), en_iyi.get('hiz_uz'),
                        en_iyi['kayit'], vir(ikinci['kimlik'], 2),
                        ikinci.get('hiz_uz'), ikinci['kayit']), kb['yol'],
                     u'A species assignment must not be settled between these two records.')

        # --- the defensible level: identity_verification.py's rule (NOT REWRITTEN)
        sav = K.savunulabilir_duzey(isabetler, en_iyi.get('lokus') or 'SSU')
        duzey = sav['duzey']
        ad = sav['onerilen_ad']
        gerekce = sav['gerekce']

        # --- the DISCRIMINABILITY rule (in code, not by hand)
        ay = _ayirt_edilebilir(K, en_iyi, ikinci, kb.get('hata_orani'), len(kb['sorgu']))
        if ay['ayrilir'] is False and duzey == u'TUR':
            cins = en_iyi.get('cins')
            tur = en_iyi.get('tur') or u''
            if cins and ikinci and ikinci.get('cins') == cins:
                duzey = u'CINS (tur ayrilamiyor)'
                ad = u'%s cf. %s' % (cins, tur.split()[-1]) if tur else u'%s sp.' % cins
            elif cins:
                duzey = u'CINS (tur ayrilamiyor)'
                ad = u'%s sp.' % cins
            else:
                duzey = u'AILE ve USTU (ad VERILEMEZ)'
                ad = u'adlandirilamayan soy'
            gerekce = u'%s | AYIRT EDILEBILIRLIK: %s' % (gerekce, ay['sebep'])
        else:
            gerekce = u'%s | AYIRT EDILEBILIRLIK: %s' % (gerekce, ay['sebep'])

        # --- how many independent databases gave the same GENUS as their best hit
        # vtb_en_iyi above was computed with THE SAME rule as en_iyi.
        cins_oy = collections.Counter(
            (h.get('cins') or u'?') for h in vtb_en_iyi.values())
        kazanan_cins = en_iyi.get('cins') or u'?'
        uyusan = cins_oy.get(kazanan_cins, 0)
        uyusmayan = u'; '.join(u'%s: %s (%%%s)' % (v, h.get('cins') or u'?', vir(h['kimlik'], 1))
                               for v, h in sorted(vtb_en_iyi.items())
                               if (h.get('cins') or u'?') != kazanan_cins) or u'-'
        if uyusan < 2:
            rap.ekle(M, u'M1-TEK-KAYNAK', UYARI,
                     u'an identity claim must be confirmed in at least TWO independent databases',
                     u'%s: "%s" is supported by only %d database (the ones that disagree: %s)'
                     % (kb['kutu'], ad, uyusan, uyusmayan), kb['yol'])

        # --- has the name changed (the main column shown in the report)
        k_ad_kok = re.sub(r'[^a-z ]', u'', (k_etiket or u'').lower()).strip()
        o_ad_kok = re.sub(r'[^a-z ]', u'', (ad or u'').lower()).strip()
        if k_etiket in (u'-', u'NOT IN THE REPORT', u'THE REPORT DID NOT MATCH'):
            ad_degisti = u'KARSILASTIRILAMADI'
        elif not o_ad_kok:
            ad_degisti = u'KARSILASTIRILAMADI'
        elif k_ad_kok.split()[:1] == o_ad_kok.split()[:1]:
            ad_degisti = u'no' if k_ad_kok == o_ad_kok else u'partly (the same genus, a different species)'
        else:
            ad_degisti = u'EVET'
        if ad_degisti == u'EVET':
            rap.ekle(M, u'M1-AD-DEGISTI', CIDDI,
                     u'the measured identity and the Kraken2 label must point at the same genus',
                     u'%s: Kraken "%s" -> olculen "%s" (%%%s, %s%s)'
                     % (kb['kutu'], k_etiket, ad, vir(en_iyi['kimlik'], 2),
                        en_iyi['vtb'],
                        u', TIP KAYDI' if en_iyi['tip'] == u'EVET' else u''),
                     krapor or kb['yol'],
                     u'This is the row to show in the report: the Kraken label changed with the measurement.')

        satirlar.append(dict(
            kutu=kb['kutu'], kraken_etiket=k_etiket, kraken_guven=k_guven,
            kraken_duzey=k_duzey, olculen=ad, duzey=duzey,
            kimlik=en_iyi['kimlik'], vtb=en_iyi['vtb'], kayit=en_iyi['kayit'],
            hiz_uz=en_iyi['hiz_uz'], tip=en_iyi['tip'],
            ikinci=(u'%s (%s, %%%s%s)' % (ikinci.get('tur') or ikinci.get('cins') or ikinci['kayit'],
                                          ikinci['vtb'], vir(ikinci['kimlik'], 2),
                                          u', TIP' if ikinci['tip'] == u'EVET' else u'')
                    ) if ikinci else u'yok',
            fark=fark, ayrilir=(u'EVET' if ay['ayrilir'] else
                                (u'HAYIR' if ay['ayrilir'] is False else u'OLCULEMEDI')),
            ayrim_sebep=ay['sebep'], n_oran=kb['n_oran'],
            saflik=kb.get('saflik'), saflik_sebep=kb.get('saflik_sebep'),
            sorgu_kaynagi=kb['sorgu_kaynagi'], uyusan=uyusan, uyusmayan=uyusmayan,
            ad_degisti=ad_degisti, gerekce=gerekce,
            hata_orani=kb.get('hata_orani')))

        if kb.get('saflik') == u'KARISIK':
            rap.ekle(M, u'M1-KARISIK-KUTU', CIDDI,
                     u'a bin must hold a single organism',
                     u'%s: %s' % (kb['kutu'], kb.get('saflik_sebep')),
                     kb.get('fastq') or kb['yol'],
                     u'The consensus derived from a mixed bin, and every discrimination calculation derived from that, is suspect.')

    rap.tablolar[u'M1 identity table'] = satirlar
    rap.olcum[u'M1 bin count'] = u'%d measured, %d COULD NOT BE MEASURED' % (
        len(satirlar) - len(olculemeyen), len(olculemeyen))
    rap.olcum[u'M1 total time'] = sure_metni(time.time() - t_basla)

    # --- THE FULL SCAN TIME ESTIMATE, FROM THE SPEEDS MEASURED IN THIS RUN
    # The estimate is not a guess but rests on MEASUREMENT: it is scaled by what was
    # really measured in this run, (a) the preparation time per bin, (b) the scan time
    # per MB, (c) the alignment time per bin and database pair. If any component was
    # not measured, NO ESTIMATE IS PRODUCED and it says "olculmedi".
    try:
        tum_vtb = [(e, d) for e, d, _t, kullan, _n in K.VTB if kullan]
        toplam_mb = sum(os.path.getsize(os.path.join(kay.refdb, d)) / 1e6
                        for _e, d in tum_vtb
                        if os.path.exists(os.path.join(kay.refdb, d)))
        n_kutu_tam = len(ind)
        # the parts that were measured
        olc_hazirlik = rap.olcum.get(u'_m1a_kutu_sn')
        tarama_mbsn = []
        hiz_kutu_vtb = []
        for k, v in rap.olcum.items():
            m = re.search(r'\(([\d,]+) MB/s\)', unicode_(v))
            if k.startswith(u'M1 scan') and m:
                tarama_mbsn.append(sayi(m.group(1)))
            m2 = re.search(r'\(([\d,]+) s per bin\)', unicode_(v))
            if k.startswith(u'M1 alignment') and m2:
                hiz_kutu_vtb.append(sayi(m2.group(1)))
        if olc_hazirlik and tarama_mbsn and hiz_kutu_vtb:
            orta_mbsn = sum(tarama_mbsn) / len(tarama_mbsn)
            orta_hiz = sum(hiz_kutu_vtb) / len(hiz_kutu_vtb)
            t_hazirlik = olc_hazirlik * n_kutu_tam
            t_tarama = toplam_mb / max(0.01, orta_mbsn)
            t_hizalama = orta_hiz * n_kutu_tam * len(tum_vtb)
            rap.olcum[u'M1 TAM TARAMA TAHMINI'] = (
                u'~%s  =  preparation %s (%s s per bin x %d bins) + scanning %s '
                u'(%s MB / %s MB/s) + alignment %s (%s s per bin and database x %d '
                u'bins x %d databases). How it scales: preparation with the number '
                u'of bins, scanning with the total bytes, and alignment with bins '
                u'times databases.'
                % (sure_metni(t_hazirlik + t_tarama + t_hizalama),
                   sure_metni(t_hazirlik), vir(olc_hazirlik, 1), n_kutu_tam,
                   sure_metni(t_tarama), vir(toplam_mb, 0), vir(orta_mbsn, 2),
                   sure_metni(t_hizalama), vir(orta_hiz, 2), n_kutu_tam,
                   len(tum_vtb))
                + (u'  CAREFUL: the scan speed was measured with %d queries in this '
                   u'run; because a full run carries %d queries, the real scan comes '
                   u'out SLOWER than this (the scan costs a fixed amount per record '
                   u'plus a small overhead per query).' % (len(kutular), n_kutu_tam)
                   if len(kutular) < n_kutu_tam else u''))
        else:
            rap.olcum[u'M1 TAM TARAMA TAHMINI'] = (
                u'not measured, because one of the parts an estimate needs was not '
                u'measured in this run (preparation, scanning or alignment)')
    except (OSError, TypeError, ZeroDivisionError) as e:
        rap.olcum[u'M1 FULL SCAN ESTIMATE'] = u'not measured (%s)' % e
    if kip == u'hizli':
        rap.ekle(M, u'M1-KISMI-KAPSAM', BILGI,
                 u'the identity verdict must be given with every offline database',
                 u'in "fast" mode only %d small RefSeq sets were scanned; SILVA SSU and LSU, UNITE, PR2 and ROD were not' % len(HIZLI_VTB), kay.refdb,
                 u'For a final verdict run it with --m1-mode tam.')
    yaz(u'M1 done: %d rows, %d bins not measurable, %s'
        % (len(satirlar), len(olculemeyen), sure_metni(time.time() - t_basla)))


# =========================================================================
# MODULE 2 - INTERNAL CONSISTENCY
# =========================================================================
# THE QUESTION: when the same number appears in more than one file, are they all
# the same?
#
# WHY: in this project the decision tables were updated MANY TIMES, by hand and by
# script. When a value corrected in one file stayed at its old value in another, a
# different order decision came out depending on which file you looked at. In the
# past the same bin was counted as a MEMBER in one table and a COMPETITOR in
# another; the M2-UYE-RAKIP check below looks for exactly that pattern.
#
# THE METHOD: (target, field) -> value is extracted from every source, and sources
# giving different values for the same (target, field) are listed as a
# CONTRADICTION. The field mapping is written OUT EXPLICITLY and is not matched
# automatically by column NAME: the column "R", for example, means THE REVERSE
# PRIMER in pairs.tsv and THE ABUNDANCE RATIO in ESIK_VE_OLCUT, and matching those
# two by name would produce a false contradiction.
# -------------------------------------------------------------------------

def _primer_norm(s):
    """Normalise a primer sequence for comparison: upper case, no spaces."""
    return re.sub(r'[^A-Z]', u'', unicode_(s).upper())


def _ad_norm(s):
    """Normalise a target name for comparison (spaces, underscores, case)."""
    return re.sub(r'[^a-z0-9]', u'', unicode_(s).lower())


def modul_2_ic_tutarlilik(kay, rap):
    M = u'2 INTERNAL CONSISTENCY'

    # (source_name, path, key_column, {canonical_field: column_name}, type)
    # the type: 'primer' | 'sayi' | 'metin'
    ESLEME = [
        (u'pairs.tsv', kay.ciftler, u'hedef', {
            u'ileri_primer': u'F', u'geri_primer': u'R'}, ),
        (u'NIHAI_SIPARIS', kay.nihai_siparis, u'hedef', {
            u'ileri_primer': u'F', u'geri_primer': u'R', u'urun_bp': u'urun_bp',
            u'kraken_etiketi': u'kraken_etiketi', u'olculen_kimlik': u'olculen_kimlik',
            u'dCq': u'dCq'}, ),
        (u'SIPARIS_LISTESI', kay.siparis_listesi, u'hedef', {
            u'ileri_primer': u'F', u'geri_primer': u'R', u'urun_bp': u'urun_bp',
            u'kraken_etiketi': u'kraken_etiketi', u'olculen_kimlik': u'olculen_kimlik',
            u'dCq': u'dCq_karsiligi'}, ),
        (u'ESIK_VE_OLCUT', kay.esik_olcut, u'hedef', {
            u'dCq': u'dCq_olculen'}, ),
    ]
    ALAN_TIPI = {u'ileri_primer': u'primer', u'geri_primer': u'primer',
                 u'urun_bp': u'sayi', u'dCq': u'sayi',
                 u'kraken_etiketi': u'metin', u'olculen_kimlik': u'metin'}

    # deger[(target, field)] = [(source, raw_value, line_number)]
    deger = collections.defaultdict(list)
    okunan_kaynak = 0
    for kaynak, yol, anahtar, alanlar in ESLEME:
        satirlar = tsv_oku(yol)
        if satirlar is None:
            _gs = gecersiz_isareti(yol)
            rap.atla(M, u'M2-KAYNAK-YOK', u'%s must be readable' % kaynak,
                     (u'marked GECERSIZ, so it did not count as a source: %s' % _gs) if _gs
                     else u'there is no such file', yol)
            continue
        if not satirlar:
            rap.ekle(M, u'M2-KAYNAK-BOS', CIDDI,
                     u'%s must hold data rows' % kaynak,
                     u'the file is EMPTY (comments and a header only)', yol)
            continue
        okunan_kaynak += 1
        for r in satirlar:
            h = (r.get(anahtar) or u'').strip()
            if not h:
                continue
            for kanonik, sutun in alanlar.items():
                if sutun not in r:
                    rap.ekle(M, u'M2-SUTUN-YOK', UYARI,
                             u'the "%s" column must be present in the %s file' % (sutun, kaynak),
                             u'there is no such column, so this field DID NOT ENTER the comparison',
                             yol)
                    continue
                ham = (r.get(sutun) or u'').strip()
                if ham:
                    deger[(_ad_norm(h), kanonik)].append((kaynak, ham, r.get('_satir')))

    if okunan_kaynak < 2:
        rap.atla(M, u'M2-CAPRAZ', u'at least two sources are needed for a cross comparison',
                 u'only %d source could be read' % okunan_kaynak, u'-')
        return

    # --- EXCEL: panel sayfasindaki primerleri de karsilastirmaya kat
    sayfalar = xlsx_sayfalari(kay.panel_xlsx)
    if sayfalar is None:
        rap.atla(M, u'M2-XLSX-YOK', u'the panel Excel file must be readable',
                 u'there is no such file', kay.panel_xlsx)
    elif isinstance(sayfalar, type(u'')):
        rap.atla(M, u'M2-XLSX', u'the panel Excel file must be readable', sayfalar,
                 kay.panel_xlsx)
    else:
        bulundu = False
        for ad, satirlar in sayfalar.items():
            if not satirlar:
                continue
            bas = [unicode_(c).strip() for c in satirlar[0]]
            bl = [b.lower() for b in bas]
            if u'hedef' not in bl:
                continue
            ih = next((i for i, b in enumerate(bl) if b.startswith(u'ileri primer')), None)
            gh = next((i for i, b in enumerate(bl) if b.startswith(u'geri primer')), None)
            if ih is None and gh is None:
                continue
            hi = bl.index(u'hedef')
            bulundu = True
            for r in satirlar[1:]:
                if len(r) <= hi or not unicode_(r[hi]).strip():
                    continue
                h = _ad_norm(r[hi])
                if ih is not None and ih < len(r) and unicode_(r[ih]).strip():
                    deger[(h, u'ileri_primer')].append((u'Excel/%s' % ad, r[ih], None))
                if gh is not None and gh < len(r) and unicode_(r[gh]).strip():
                    deger[(h, u'geri_primer')].append((u'Excel/%s' % ad, r[gh], None))
            break
        if not bulundu:
            rap.atla(M, u'M2-XLSX-SAYFA',
                     u'the Excel must hold a sheet with a "Hedef" and an "Ileri primer" column',
                     u'no such sheet was found', kay.panel_xlsx)

    # --- CELISKI ARAMA
    celiski = 0
    karsilastirilan = 0
    for (h, alan), kayitlar in sorted(deger.items()):
        if len(kayitlar) < 2:
            continue
        karsilastirilan += 1
        tip = ALAN_TIPI.get(alan, u'metin')
        if tip == u'primer':
            norm = [(k, _primer_norm(v), s) for k, v, s in kayitlar]
            ayri = set(n for _k, n, _s in norm)
        elif tip == u'sayi':
            norm = [(k, sayi(v), s) for k, v, s in kayitlar]
            sayilar = [n for _k, n, _s in norm if n is not None]
            # A rounding difference of 0.01 in a number does not count as a contradiction;
            # anything larger does.
            ayri = set()
            if sayilar and (max(sayilar) - min(sayilar)) > 0.011:
                ayri = set(sayilar)
            if len(sayilar) != len(norm):
                ayri.add(None)
        else:
            norm = [(k, unicode_(v).strip(), s) for k, v, s in kayitlar]
            ayri = set(n.lower() for _k, n, _s in norm)
        if len(ayri) <= 1:
            continue
        celiski += 1
        ayrinti = u' | '.join(
            u'%s%s = %s' % (k, (u' (row %s)' % s) if s else u'', v)
            for k, v, s in kayitlar)
        agir = KRITIK if alan in (u'ileri_primer', u'geri_primer', u'urun_bp') else CIDDI
        rap.ekle(M, u'M2-CELISKI', agir,
                 u'the "%s" value of target "%s" must be THE SAME in every file' % (h, alan),
                 ayrinti,
                 u'; '.join(sorted(set(k for k, _v, _s in kayitlar))),
                 u'The order changes depending on which file was looked at.'
                 if agir == KRITIK else u'')

    rap.olcum[u'M2 fields compared'] = u'%d (target, field) pairs, %d contradictions' % (
        karsilastirilan, celiski)
    if karsilastirilan == 0:
        rap.atla(M, u'M2-BOS', u'en az bir alan capraz karsilastirilmali',
                 u'not one (target, field) appearing in both sources was found', u'-')

    # --- IS THE SAME BIN BOTH A MEMBER AND A COMPETITOR (this happened before)
    uyelik = tsv_oku(kay.hedef_uyelik)
    if uyelik is None:
        rap.atla(M, u'M2-UYE-RAKIP', u'target_membership.tsv must be readable', u'there is no such file',
                 kay.hedef_uyelik)
    else:
        for r in uyelik:
            h = (r.get(u'hedef') or u'').strip()
            uye = set(x.strip() for x in (r.get(u'uye_taxid') or u'').split(u',') if x.strip())
            haric = set(x.strip() for x in (r.get(u'haric') or u'').split(u',') if x.strip())
            kesisim = uye & haric
            if kesisim:
                rap.ekle(M, u'M2-UYE-RAKIP', KRITIK,
                         u'a taxid cannot be both a member and an exclusion on the same target',
                         u'%s: %s is in both the uye_taxid and the haric list'
                         % (h, u', '.join(sorted(kesisim))),
                         u'%s (row %s)' % (kay.hedef_uyelik, r.get('_satir')),
                         u'The discrimination ratio is wrong because this taxid was counted twice.')


# =========================================================================
# MODULE 3 - MEMBERSHIP INTEGRITY
# =========================================================================
# THE QUESTION: is the member set of every pair DEFINED, and was the dCq computed
# for that pair really computed with THAT PAIR'S membership?
#
# WHY: that is exactly what happened with the Petriella_cinsi target. The target had
# NO membership row of its own, the membership had been inherited from another
# target, and the reported dCq was in fact that other target's measurement. The
# number looked filled in, so nobody noticed. An empty membership silently returning
# zero is one of the most expensive patterns in this project.
# -------------------------------------------------------------------------

def modul_3_uyelik(kay, rap):
    M = u'3 MEMBERSHIP'

    ciftler = tsv_oku(kay.ciftler)
    uyelik = tsv_oku(kay.hedef_uyelik)
    esik = tsv_oku(kay.esik_olcut)

    if ciftler is None:
        rap.atla(M, u'M3-CIFTLER', u'screening/pairs.tsv must be readable',
                 u'there is no such file', kay.ciftler)
        return
    if uyelik is None:
        rap.atla(M, u'M3-UYELIK', u'screening/target_membership.tsv must be readable',
                 u'there is no such file', kay.hedef_uyelik)
        return

    uyelik_indeks = {}
    for r in uyelik:
        h = (r.get(u'hedef') or u'').strip()
        if h:
            uyelik_indeks[_ad_norm(h)] = r

    dcq_indeks = {}
    if esik:
        for r in esik:
            h = (r.get(u'hedef') or u'').strip()
            if h:
                dcq_indeks[_ad_norm(h)] = r

    # --- sign the member sets: an identical set on two targets means a tautology
    imza = collections.defaultdict(list)

    for r in ciftler:
        h = (r.get(u'hedef') or u'').strip()
        if not h:
            continue
        hn = _ad_norm(h)
        uye_ham = (r.get(u'uye_taksonlar') or u'').strip()
        uye = set(x.strip() for x in uye_ham.split(u',') if x.strip())
        olcu = (r.get(u'olcu_tipi') or u'').strip()
        durum = (r.get(u'uye_kumesi_durumu') or u'').strip()
        dsat = dcq_indeks.get(hn)
        dcq = sayi(dsat.get(u'dCq_olculen')) if dsat else None

        # 1) is the member set EMPTY
        #    AN IMPORTANT DISTINCTION: on a target that measures COVERAGE (the universal
        #    primers) an empty member set is BY DESIGN, not an error. Such a target is not
        #    trying to separate one lineage from another, it measures coverage across a
        #    domain. An empty membership is an error only (a) on a target that measures
        #    SEPARATION, or (b) when a dCq resting on membership has been reported. An
        #    auditor that does not make this distinction shouts about sound design as if it
        #    were an error, and the real errors get lost among them.
        if not uye:
            kapsam_olcusu = (olcu.lower().startswith(u'kapsam') or
                             durum.upper() == u'KAPSAM_OLCUSU')
            if dcq is not None:
                rap.ekle(M, u'M3-UYELIK-BOS', KRITIK,
                         u'a dCq resting on membership must not be reported for a target whose membership is empty',
                         u'%s: uye_taksonlar is EMPTY but a dCq of %s was reported (measure type: %s)' % (h, vir(dcq), olcu or u'?'),
                         u'%s (row %s)' % (kay.ciftler, r.get('_satir')),
                         u'A discrimination ratio computed with an empty membership IS MEANINGLESS.')
            elif kapsam_olcusu:
                rap.ekle(M, u'M3-UYELIK-BOS-KAPSAM', BILGI,
                         u'in targets that measure coverage the member set may be empty',
                         u'%s: uye_taksonlar bos, olcu tipi "%s" - bu TASARIM '
                         u'geregidir, ayrim hesabina girmez' % (h, olcu),
                         u'%s (row %s)' % (kay.ciftler, r.get('_satir')))
            else:
                rap.ekle(M, u'M3-UYELIK-BOS', KRITIK,
                         u'the member taxa of every pair that measures discrimination must be defined',
                         u'%s: uye_taksonlar BOS, olcu tipi "%s"' % (h, olcu or u'?'),
                         u'%s (row %s)' % (kay.ciftler, r.get('_satir')),
                         u'A discrimination ratio computed with an empty membership IS MEANINGLESS.')
        else:
            imza[u','.join(sorted(uye))].append((h, olcu))

        # 2) does the target have a membership row OF ITS OWN
        if hn not in uyelik_indeks:
            agir = KRITIK if dcq is not None else CIDDI
            rap.ekle(M, u'M3-UYELIK-SATIRI-YOK', agir,
                     u'every target must have ITS OWN row inside target_membership.tsv',
                     u'%s: it has NO membership row of its own%s'
                     % (h, u'; and yet a dCq of %s was reported, so that measurement may have been made with another target\'s membership' % vir(dcq)
                        if dcq is not None else u''),
                     kay.hedef_uyelik,
                     u'This is the Petriella_cinsi pattern: the membership was inherited, the number looks filled in, but it is not that pair\'s measurement.')

        # 3) does the membership status report a problem
        if durum and durum.upper() not in (u'PANELLE_TUTUYOR', u'YENIDEN_KURULDU',
                                           u'KAPSAM_OLCUSU'):
            rap.ekle(M, u'M3-UYELIK-DURUM', CIDDI,
                     u'uye_kumesi_durumu must be a known and valid value',
                     u'%s: uye_kumesi_durumu = "%s"' % (h, durum),
                     u'%s (row %s)' % (kay.ciftler, r.get('_satir')))

        # 4) the THRESHOLD table says the membership is invalid, yet a dCq was still used
        if dsat:
            gecerli = (dsat.get(u'uyelik_gecerli_mi') or u'').strip().upper()
            if gecerli in (u'HAYIR', u'YOK') and dcq is not None:
                rap.ekle(M, u'M3-DCQ-GECERSIZ-UYELIK', KRITIK,
                         u'a dCq must not be reported for a target whose membership is invalid',
                         u'%s: uyelik_gecerli_mi=%s but dCq_olculen=%s and the verdict is "%s"'
                         % (h, gecerli, vir(dcq), (dsat.get(u'YENI_HUKUM') or u'?')),
                         u'%s (row %s)' % (kay.esik_olcut, dsat.get('_satir')),
                         u'This dCq was computed with an undefined membership.')

    # 5) two "separation" targets with an IDENTICAL member set -> A TAUTOLOGY
    for anahtar, hedefler in imza.items():
        ayrim = [h for h, o in hedefler if o.lower().startswith(u'ayrim')]
        if len(hedefler) > 1 and ayrim:
            rap.ekle(M, u'M3-TOTOLOJIK-UYELIK', CIDDI,
                     u'the member sets of two targets that measure discrimination must differ',
                     u'the member sets of the %s targets are EXACTLY THE SAME (%d taxids); the ones among them that measure discrimination: %s'
                     % (u', '.join(h for h, _o in hedefler),
                        len(anahtar.split(u',')), u', '.join(ayrim)),
                     kay.ciftler,
                     u'A measurement trying to separate a set from itself is TAUTOLOGICAL; the ratio it gives is not a real separation.')

    rap.olcum[u'M3 pairs audited'] = u'%d pairs, %d membership rows' % (
        len(ciftler), len(uyelik_indeks))


# =========================================================================
# MODULE 4 - THE LITERATURE RULES
# =========================================================================
# It applies the NUMERIC rules that came out of a literature review of 55 sources to
# the panel. The rules are NOT hard coded here; they are first verified against the
# LITERATURE file. If that file says a different number, the rule is STALE, and that
# is itself a finding.
# -------------------------------------------------------------------------

# The abundance weighted threshold: required_dCq = max(log2(R) + EK, TABAN)
LIT_EK = 4.3
LIT_TABAN = 3.32


def modul_4_literatur(kay, rap):
    M = u'4 LITERATURE'

    lit = metin_oku(kay.literatur)
    if lit is None:
        rap.atla(M, u'M4-LITERATUR-YOK', u'LITERATUR_2026-08-07.md must be readable',
                 u'there is no such file', kay.literatur)
        lit = u''
    else:
        # IS THE RULE ITSELF STALE? Do our constants occur in the literature.
        for sabit, ad in ((LIT_EK, u'the log2(R) addition'), (LIT_TABAN, u'the fixed floor')):
            desen = unicode_(sabit).replace(u'.', u'[.,]')
            if not re.search(desen, lit):
                rap.ekle(M, u'M4-KURAL-BAYAT', CIDDI,
                         u'the %s (%s) value must appear in the literature file' % (ad, vir(sabit)),
                         u'the %s value was not found inside LITERATUR_2026-08-07.md'
                         % vir(sabit), kay.literatur,
                         u'The auditor\'s rule may have drifted away from the literature.')

    esik = tsv_oku(kay.esik_olcut)
    siparis = tsv_oku(kay.nihai_siparis)

    # --- RULE 1: the abundance weighted threshold
    if esik is None:
        rap.atla(M, u'M4-KURAL1', u'gerekli_dCq = max(log2(R)+%s, %s) must be confirmed'
                 % (vir(LIT_EK), vir(LIT_TABAN)), u'there is no ESIK_VE_OLCUT file',
                 kay.esik_olcut)
    else:
        denetlenen = 0
        for r in esik:
            h = (r.get(u'hedef') or u'').strip()
            R = sayi(r.get(u'R'))
            yazan = sayi(r.get(u'gerekli_dCq'))
            olculen = sayi(r.get(u'dCq_olculen'))
            durum = (r.get(u'yeni_kural_durum') or u'').strip().upper()
            if R is None or R <= 0:
                rap.ekle(M, u'M4-R-YOK', CIDDI,
                         u'R (the abundance ratio) must be defined for the abundance weighted threshold',
                         u'%s: R = "%s" (sayiya cevrilemedi ya da <= 0)'
                         % (h, r.get(u'R')),
                         u'%s (row %s)' % (kay.esik_olcut, r.get('_satir')),
                         u'gerekli_dCq CANNOT BE COMPUTED for this target.')
                continue
            hesap = max(math.log(R, 2) + LIT_EK, LIT_TABAN)
            denetlenen += 1
            if yazan is None:
                rap.ekle(M, u'M4-GEREKLI-YOK', CIDDI,
                         u'the gerekli_dCq column must be filled in',
                         u'%s: gerekli_dCq is empty, the computed value is %s' % (h, vir(hesap)),
                         u'%s (row %s)' % (kay.esik_olcut, r.get('_satir')))
            elif abs(yazan - hesap) > 0.02:
                rap.ekle(M, u'M4-ESIK-YANLIS', KRITIK,
                         u'gerekli_dCq = max(log2(R)+%s ; %s) = %s'
                         % (vir(LIT_EK), vir(LIT_TABAN), vir(hesap)),
                         u'%s: the table says %s (R=%s)' % (h, vir(yazan), vir(R, 3)),
                         u'%s (row %s)' % (kay.esik_olcut, r.get('_satir')),
                         u'If the threshold is wrong the GECER or KALIR verdict is wrong too.')
            # does the verdict hold against its own numbers
            if olculen is not None and yazan is not None:
                bekl = u'GECER' if olculen >= yazan else u'KALIR'
                if durum and durum != bekl:
                    rap.ekle(M, u'M4-HUKUM-CELISIK', KRITIK,
                             u'if dCq is %s and the requirement %s, the verdict has to be "%s"'
                             % (vir(olculen), vir(yazan), bekl),
                             u'%s: yeni_kural_durum = "%s"' % (h, durum),
                             u'%s (row %s)' % (kay.esik_olcut, r.get('_satir')))
        rap.olcum[u'M4 threshold rule'] = u'recomputed on %d targets' % denetlenen
        if denetlenen == 0:
            rap.atla(M, u'M4-KURAL1-BOS',
                     u'gerekli_dCq must be recomputed for at least one target',
                     u'no valid R was found for any target', kay.esik_olcut)

    # --- RULE 2: the last two bases at the 3' end must not decide ON THEIR OWN
    if siparis is None:
        rap.atla(M, u'M4-KURAL2', u'the last two bases at the 3\' end must not decide on their own',
                 u'NIHAI_SIPARIS was not read: there is no such file, or it is marked GECERSIZ', kay.nihai_siparis)
    else:
        UC_DESEN = re.compile(r"3'\s*son\s*iki|son\s*iki\s*baz|3'\s*ucundaki?\s*iki", re.I)
        for r in siparis:
            h = (r.get(u'hedef') or u'').strip()
            katman = (r.get(u'hukmu_veren_katman') or u'')
            gerekce = (r.get(u'GEREKCE') or u'')
            if UC_DESEN.search(katman):
                rap.ekle(M, u'M4-UC-BAZ-HUKUM', CIDDI,
                         u'the last two bases at the 3\' end must not decide ON THEIR OWN (the literature: section 3)',
                         u'%s: the layer that gave the verdict is "%s"' % (h, katman.strip()[:160]),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                         u'The literature says the last two bases at the 3\' end cannot predict an off target product on their own.')
            elif UC_DESEN.search(gerekce) and not re.search(
                    r'dCq|ayrim|kapsam|MFE|erime', gerekce, re.I):
                rap.ekle(M, u'M4-UC-BAZ-TEK-GEREKCE', UYARI,
                         u'beside the last two bases at the 3\' end there must also be a measured piece of evidence',
                         u'%s: the reason rests on the 3\' criterion alone: "%s"'
                         % (h, gerekce.strip()[:160]),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')))

    # --- RULE 3: universal primers have to be judged on THREE metrics
    if siparis is not None:
        UCLU = ((u'coverage', re.compile(r'kapsam|coverage', re.I)),
                (u'the phylum spectrum', re.compile(r'filum|phylum|spektrum|spectrum', re.I)),
                (u'the organelle fraction',
                 re.compile(r'organel|mitokondri|kloroplast|plastid|mitochondri|chloroplast', re.I)))
        for r in siparis:
            h = (r.get(u'hedef') or u'').strip()
            if not re.search(r'universal|evrensel', h, re.I):
                continue
            butun = u' '.join(unicode_(v) for k, v in r.items() if k != '_satir')
            eksik = [ad for ad, d in UCLU if not d.search(butun)]
            if eksik:
                rap.ekle(M, u'M4-EVRENSEL-METRIK', CIDDI,
                         u'universal primers must be judged on a TRIPLE metric (coverage, phylum spectrum, organelle ratio), the literature sections 1 and 2',
                         u'%s: the row never mentions these criteria: %s'
                         % (h, u', '.join(eksik)),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                         u'Judging a universal primer by its "outside the clade count" goes against the literature; being outside the clade is what you expect.')

    # --- RULE 4: MIQE - in-silico and experimental validation are SEPARATE requirements
    #
    # In this project the EXPERIMENTAL requirement is held in A SEPARATE file (the
    # LABORATUVARDA_NE_YAPILMALI column inside SIPARIS_LISTESI.tsv). A check looking only
    # at NIHAI_SIPARIS would say "missing" even though the requirement is written in the
    # other file. So the two files are JOINED on the target name, and a finding is
    # produced only when the requirement is absent from both.
    if siparis is not None:
        SIL = re.compile(r'in.?siliko|in.?silico|MFE|hizalama|nt |NCBI', re.I)
        DEN = re.compile(r'jel|gel|erime\s*egrisi|dizileme|amplikon\s*dizile|'
                         r'deneysel|laboratuvar|NTC|qPCR\s*kosu', re.I)
        siparis_l = tsv_oku(kay.siparis_listesi) or []
        ek_metin = collections.defaultdict(list)
        for r in siparis_l:
            ek_metin[_ad_norm(r.get(u'hedef') or u'')].append(
                u' '.join(unicode_(v) for k, v in r.items() if k != '_satir'))
        if not siparis_l:
            rap.atla(M, u'M4-MIQE-IKINCI-KAYNAK',
                     u'the experimental confirmation condition must be audited together with SIPARIS_LISTESI.tsv',
                     u'SIPARIS_LISTESI.tsv could not be read; the MIQE check looked at a single file only and may be incomplete', kay.siparis_listesi)
        for r in siparis:
            h = (r.get(u'hedef') or u'').strip()
            hukum = (r.get(u'HUKUM') or u'').strip().upper()
            if hukum in (u'', u'SIPARIS EDILMEZ'):
                continue
            butun = u' '.join(unicode_(v) for k, v in r.items() if k != '_satir')
            butun += u' ' + u' '.join(ek_metin.get(_ad_norm(h), []))
            var_sil = bool(SIL.search(butun))
            var_den = bool(DEN.search(butun))
            if var_sil and not var_den:
                rap.ekle(M, u'M4-MIQE-DENEYSEL-YOK', CIDDI,
                         u'MIQE in siliko ve DENEYSEL dogrulamayi AYRI zorunlu '
                         u'maddeler sayar (literatur bolum 4)',
                         u'%s (verdict: %s): there is in-silico evidence, but the row never mentions the experimental confirmation condition' % (h, hukum),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                         u'For every pair ordered, the gel, melt curve or sequencing condition must be written plainly.')
            elif not var_sil and not var_den:
                rap.ekle(M, u'M4-MIQE-KANIT-YOK', CIDDI,
                         u'every pair that goes into the order must have at least one confirmation',
                         u'%s (verdict: %s): neither in-silico nor experimental evidence is given'
                         % (h, hukum),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')))


# =========================================================================
# MODULE 5 - KNOWN FAILURE PATTERNS
# =========================================================================
# Nine patterns that earned a module of their own by coming up AGAIN AND AGAIN in
# this project. Each pattern has its own check and each is reported as a "what was
# expected / what was found" pair. If a pattern produces no finding, that means the
# pattern was measured and came out clean; it DOES NOT MEAN the check did not run.
# Checks that did not run are counted separately as ATLANDI.
# -------------------------------------------------------------------------

# The masking patterns looked for in the code scan.
_MASKE_PY = [
    (re.compile(r'except[^\n:]*:\s*\n\s*pass\b'), u'except ... : pass  (the error is swallowed)'),
    (re.compile(r'except[^\n:]*:\s*\n\s*continue\b'), u'except ... : continue  (the error is swallowed)'),
    (re.compile(r'\bsys\.exit\(\s*0\s*\)'), u'sys.exit(0)  (an unconditional successful exit)'),
    (re.compile(r'subprocess\.(call|run)\((?![^)]*check\s*=\s*True)[^)]*\)'),
     u'subprocess without check=True  (the exit code of the child is not read)'),
]
_MASKE_BAT = [
    (re.compile(r'\|\|\s*(true|ver\b)', re.I), u'|| true  (the error is suppressed)'),
    (re.compile(r'>\s*nul\s+2>&1', re.I), u'> nul 2>&1  (the error output is destroyed)'),
    (re.compile(r'\bexit\s*/b\s*0\b', re.I), u'exit /b 0  (an unconditional successful exit)'),
]

# The known caps looked for in the "a cap was taken for a count" pattern.
_TAVANLAR = (500, 1000, 3000, 120001, 200000, 100, 50)


def modul_5_desenler(kay, rap):
    M = u'5 ERROR PATTERNS'

    siparis = tsv_oku(kay.nihai_siparis)
    ciftler = tsv_oku(kay.ciftler)
    uyelik = tsv_oku(kay.hedef_uyelik)
    hdisi = tsv_oku(kay.hedef_disi)

    # ---- PATTERN 1: A SILENT ZERO ------------------------------------
    # The EXACT definition of the pattern: a LAYER did not run, or could not reach a
    # verdict, and yet that layer's numeric field was still written as 0. Such a 0 reads
    # like "no off-target was found", when it means "nobody looked".
    #
    # CAUTION - it is very easy to get the boundary wrong here, and it has been:
    # calling "every 0 suspicious" produces 58 false findings and drowns the real one.
    # So a 0 counts as a finding ONLY IF its own layer says "I did not run". The field
    # to layer mapping is read from the column prefix.
    KATMAN_ONEKI = ((u'NCBI_', u'NCBI'), (u'MFE_', u'MFEprimer'),
                    (u'yerel_', u'yerel'))
    KOSMADI = re.compile(
        r'sinanmad|hukum\s*veremedi|kosulmad|[cç]al[iı][sş]t[iı]r[iı]lmad|'
        r'atland|BA[SŞ]ARISIZ|yap[iı]lamad|olculemedi', re.I)
    if siparis is None:
        rap.atla(M, u'M5-D1', u'the silent zero scan',
                 u'NIHAI_SIPARIS was not read: there is no such file, or it is marked GECERSIZ',
                 kay.nihai_siparis)
    else:
        bakilan = 0
        for r in siparis:
            h = (r.get(u'hedef') or u'').strip()
            baglam = u' '.join([r.get(u'GEREKCE') or u'',
                                r.get(u'hukmu_veren_katman') or u'',
                                r.get(u'NCBI_durumu') or u''])
            # It is examined SENTENCE BY SENTENCE. One long reason can hold TWO SEPARATE
            # statements, such as "MFEprimer found 0. NCBI could not decide.", and a check
            # looking at the whole text at once would blame the MFE fields for NCBI not
            # running as well (measured: 80 of 88 findings were like that).
            cumleler = [c for c in re.split(r'[.;|]\s*', baglam) if c.strip()]
            # ONE reason sentence is kept per layer; when we blamed the same layer in two
            # separate sentences, the report printed the same finding twice.
            kosmayan = collections.OrderedDict()
            for c in cumleler:
                if not KOSMADI.search(c):
                    continue
                for _onek, k in KATMAN_ONEKI:
                    if k.lower() in c.lower() and k not in kosmayan:
                        kosmayan[k] = c.strip()[:140]
            sifir_alan = collections.defaultdict(list)
            for alan, ham in sorted(r.items()):
                if alan == '_satir' or not unicode_(ham).strip():
                    continue
                katman = next((k for onek, k in KATMAN_ONEKI
                               if alan.startswith(onek)), None)
                if katman is None:
                    continue
                bakilan += 1
                v = sayi(ham)
                if v is not None and abs(v) < 1e-9:
                    sifir_alan[katman].append(alan)
            # ONE finding per layer: repeating it field by field bloats the report.
            for katman, cumle in kosmayan.items():
                alanlar = sifir_alan.get(katman)
                if not alanlar:
                    continue
                rap.ekle(M, u'M5-D1-SESSIZ-SIFIR', CIDDI,
                         u'the numeric field of a layer that did not run must be written as "olculmedi" and not as 0',
                         u'%s: the "%s" layer did not run ("%s") but these fields say 0: %s' % (h, katman, cumle, u', '.join(alanlar)),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                         u'0 reads as "no off target was found"; what it means is "it was not looked at".')
        rap.olcum[u'M5 D1 taranan katman alani'] = u'%d' % bakilan
        if bakilan == 0:
            rap.atla(M, u'M5-D1', u'the silent zero scan',
                     u'no numeric field carrying a layer prefix was found',
                     kay.nihai_siparis)

    # ---- PATTERN 2: COUNTING THE TARGET'S OWN MEMBERS AS OFF-TARGET ---
    if uyelik is None or hdisi is None:
        rap.atla(M, u'M5-D2', u'a target must not count its own members as off target',
                 u'there is no target_membership.tsv or HEDEF_DISI_AYRINTI',
                 kay.hedef_disi if uyelik else kay.hedef_uyelik)
    else:
        uye_harita = {}
        for r in uyelik:
            h = _ad_norm(r.get(u'hedef') or u'')
            uye_harita[h] = set(x.strip() for x in
                                (r.get(u'uye_taxid') or u'').split(u',') if x.strip())
        hedef_sut = next((s for s in (hdisi[0].keys() if hdisi else [])
                          if _ad_norm(s) == u'hedef'), None)
        if hedef_sut is None:
            rap.atla(M, u'M5-D2', u'a target must not count its own members as off target',
                     u'the HEDEF_DISI_AYRINTI file has no "hedef" column', kay.hedef_disi)
        else:
            sayac = 0
            for r in hdisi:
                h = _ad_norm(r.get(hedef_sut) or u'')
                uye = uye_harita.get(h)
                if not uye:
                    continue
                butun = u' '.join(unicode_(v) for k, v in r.items() if k != '_satir')
                gecen = [t for t in uye if re.search(r'(?<!\d)%s(?!\d)' % re.escape(t), butun)]
                if gecen:
                    sayac += 1
                    rap.ekle(M, u'M5-D2-UYE-HEDEF-DISI', CIDDI,
                             u'a target\'s OWN member must not appear in that target\'s "off target" list',
                             u'%s: taxid %s, a member of the target itself, appears on an off target row'
                             % (r.get(hedef_sut), u', '.join(sorted(gecen))),
                             u'%s (row %s)' % (kay.hedef_disi, r.get('_satir')),
                             u'The off target count is inflated, so the specificity looks worse than it is.')
            rap.olcum[u'M5 P2 off target rows scanned'] = u'%d rows, %d findings' % (
                len(hdisi), sayac)

    # ---- DESEN 3: MASKELENMIS CIKIS KODU -----------------------------
    tarandi = 0
    for klasor in kay.kod_klasorleri:
        if not os.path.isdir(klasor):
            continue
        for yol in sorted(glob.glob(os.path.join(klasor, u'*.py'))):
            if u'.orig' in yol:
                continue
            icerik = metin_oku(yol)
            if icerik is None:
                continue
            tarandi += 1
            for desen, ad in _MASKE_PY:
                n = len(desen.findall(icerik))
                if n:
                    rap.ekle(M, u'M5-D3-MASKE-PY', UYARI,
                             u'errors and exit codes must not be masked',
                             u'%s: %d kez "%s"' % (os.path.basename(yol), n, ad),
                             yol,
                             u'A swallowed error makes a run look as though it succeeded.')
    for yol in sorted(glob.glob(os.path.join(kay.kok, u'*.bat'))):
        # The .bat files that run are READ but NOT MODIFIED.
        icerik = metin_oku(yol)
        if icerik is None:
            continue
        tarandi += 1
        for desen, ad in _MASKE_BAT:
            n = len(desen.findall(icerik))
            if n:
                rap.ekle(M, u'M5-D3-MASKE-BAT', UYARI,
                         u'bat files must not suppress the exit code of a subprocess',
                         u'%s: %d kez "%s"' % (os.path.basename(yol), n, ad), yol)
    if tarandi == 0:
        rap.atla(M, u'M5-D3', u'the masked exit code scan',
                 u'no code file to scan was found', u'; '.join(kay.kod_klasorleri))
    else:
        rap.olcum[u'M5 D3 taranan kod dosyasi'] = u'%d' % tarandi

    # ---- DESEN 4: BAYAT KONTROL NOKTASI ------------------------------
    kn_klasorleri = [d for d in glob.glob(os.path.join(kay.kok, u'*', u'kontrol'))
                     if os.path.isdir(d)]
    if not kn_klasorleri:
        rap.atla(M, u'M5-D4', u'the stale checkpoint scan',
                 u'not one */kontrol directory was found', kay.kok)
    else:
        # The consensus index and the panel definition are taken as input: if the
        # checkpoint is OLDER than either, the result of that run is stale.
        girdiler = [kay.konsensus_indeks, kay.ciftler, kay.hedef_uyelik]
        en_yeni = 0
        for g in girdiler:
            if os.path.exists(g):
                en_yeni = max(en_yeni, os.path.getmtime(g))
        bayat = 0
        toplam = 0
        for d in kn_klasorleri:
            dosyalar = glob.glob(os.path.join(d, u'*'))
            toplam += len(dosyalar)
            eski = [f for f in dosyalar
                    if os.path.isfile(f) and os.path.getmtime(f) < en_yeni]
            if eski:
                bayat += len(eski)
                rap.ekle(M, u'M5-D4-BAYAT-KN', CIDDI,
                         u'the checkpoints must be NEWER than their inputs (panel, membership, consensus)',
                         u'%s: %d of %d checkpoints are older than their inputs'
                         % (os.path.relpath(d, kay.kok), len(eski), len(dosyalar)), d,
                         u'A stale checkpoint makes an old result be reused with changed input.')
        rap.olcum[u'M5 P4 checkpoints'] = u'%d files, %d stale' % (toplam, bayat)

    # ---- PATTERN 5: MISTAKING A CAP VALUE FOR A COUNT -----------------
    # Only COUNT fields are examined. A product length (urun_bp = 100) is not a count
    # and has nothing to do with a cap; a check that looked at every column was
    # producing false findings like that.
    # And if the row already says "SONUC TAVANI", the project has marked it
    # deliberately; shouting again about a known problem pollutes the report.
    SAYIM_DESENI = re.compile(
        r'(sayi|sayisi|adet|urun$|_urun|hedef_disi|klad_disi|kayit|hit|isabet)', re.I)
    if siparis is None:
        rap.atla(M, u'M5-D5', u'the ceiling value scan',
                 u'NIHAI_SIPARIS was not read: there is no such file, or it is marked GECERSIZ',
                 kay.nihai_siparis)
    else:
        bakilan = 0
        for r in siparis:
            kabul = u' '.join([r.get(u'NCBI_durumu') or u'',
                               r.get(u'GEREKCE') or u''])
            zaten_biliniyor = bool(re.search(r'tavan', kabul, re.I))
            for k, v in sorted(r.items()):
                if k == '_satir' or k.lower() in (u'urun_bp', u'sira'):
                    continue
                if not SAYIM_DESENI.search(k):
                    continue
                n = sayi(v)
                if n is None or int(n) != n or int(n) not in _TAVANLAR:
                    continue
                bakilan += 1
                rap.ekle(M, u'M5-D5-TAVAN',
                         BILGI if zaten_biliniyor else UYARI,
                         u'a COUNT that is reported must not be exactly equal to a tool\'s ceiling',
                         u'%s: %s = %d%s' % (
                             r.get(u'hedef'), k, int(n),
                             u' (the row has already marked this as a "ceiling")'
                             if zaten_biliniyor else
                             u' - a ceiling value, and the row DOES NOT SAY that it is a ceiling'),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                         u'A ceiling value is not "the number found" but "the number it was cut at"; the real number is larger.')
        rap.olcum[u'M5 D5 tavana esit sayim'] = u'%d' % bakilan

    # ---- PATTERN 6: A TAUTOLOGICAL MEASUREMENT -----------------------
    # M3 looks at the membership side; here the question is whether THE EVIDENCE FILE
    # verifies itself: if the file shown as the evidence for a target is the very file
    # the target's verdict was written into, the measurement is proving itself.
    if siparis is not None:
        for r in siparis:
            kanit = (r.get(u'GEREKCE') or u'') + u' ' + (r.get(u'hukmu_veren_katman') or u'')
            if os.path.basename(kay.nihai_siparis) in kanit:
                rap.ekle(M, u'M5-D6-TOTOLOJI', CIDDI,
                         u'the evidence for a verdict must be a measurement file OTHER than the file the verdict is written in',
                         u'%s: its own file (%s) is shown as the evidence'
                         % (r.get(u'hedef'), os.path.basename(kay.nihai_siparis)),
                         u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')))

    # ---- PATTERN 7: TAKING TWO LAYERS THAT SHARE AN ENGINE AS INDEPENDENT EVIDENCE
    # If two layers use the same alignment core they ARE NOT TWO INDEPENDENT PIECES OF
    # EVIDENCE: a shared bug in the core votes twice.
    #
    # THE DEPENDENCE IS NOT GUESSED, IT IS READ FROM THE CODE. Whether a layer is inside
    # the project is determined by looking at whether that layer's script imports
    # engine_gateway.py. Third party tools from outside (the MFEprimer binary, the NCBI
    # web service) do not use the project engine and ARE INDEPENDENT. Counting those as
    # "the same engine" would present a sound cross-validation as an error.
    ici_motor = set()
    for klasor in kay.kod_klasorleri:
        for yol in glob.glob(os.path.join(klasor, u'*.py')):
            if u'.orig' in yol:
                continue
            icerik = metin_oku(yol) or u''
            if re.search(r'^\s*(from\s+\S*\s+)?import\s+.*\bmotor\b', icerik, re.M):
                ici_motor.add(os.path.splitext(os.path.basename(yol))[0])
    # Which script each layer name in the decision table corresponds to.
    KATMAN_BETIK = {u'yerel': u'motor', u'kapsam': u'motor', u'okuma': u'okuma_motoru',
                    u'in siliko': u'motor', u'uyelik': u'uyelik_denetimi'}
    DIS_ARAC = (u'mfeprimer', u'mfe', u'ncbi', u'blast', u'kraken', u'bracken')
    if siparis is None:
        rap.atla(M, u'M5-D7', u'the layer independence scan',
                 u'NIHAI_SIPARIS was not read: there is no such file, or it is marked GECERSIZ',
                 kay.nihai_siparis)
    elif not ici_motor:
        rap.atla(M, u'M5-D7', u'the layer independence scan',
                 u'no script importing engine_gateway.py was found, so the dependency map could not be drawn', u'; '.join(kay.kod_klasorleri))
    else:
        rap.olcum[u'M5 P7 scripts that use engine_gateway.py'] = u'%d' % len(ici_motor)
        for r in siparis:
            katman = (r.get(u'hukmu_veren_katman') or u'').strip()
            if not katman:
                continue
            paylasan = collections.Counter()
            for ad, betik in KATMAN_BETIK.items():
                if betik not in ici_motor and betik != u'motor':
                    continue
                if re.search(re.escape(ad), katman, re.I):
                    # do not count a match that falls inside the name of an outside tool
                    parcalar = re.split(r'[+;,]', katman)
                    for p in parcalar:
                        if re.search(re.escape(ad), p, re.I) and not any(
                                d in p.lower() for d in DIS_ARAC):
                            paylasan[betik] += 1
                            break
            for betik, n in paylasan.items():
                if n >= 2:
                    rap.ekle(M, u'M5-D7-AYNI-MOTOR', CIDDI,
                             u'the two layers supporting a verdict must be INDEPENDENT',
                             u'%s: %d of the layers that gave the verdict use the same core (%s.py), "%s"'
                             % (r.get(u'hedef'), n, betik, katman[:140]),
                             u'%s (row %s)' % (kay.nihai_siparis, r.get('_satir')),
                             u'Two outputs of the same engine are one piece of evidence; a fault they share votes twice.')

    # ---- PATTERN 8: A DEGENERATE BASE LEAK ---------------------------
    # Primers carrying a degenerate base must be shown to have been evaluated by a
    # measurement that HANDLES degenerate bases. Otherwise the tool may have counted the
    # degenerate base as an N and silently discarded it.
    DEJ = set(u'RYSWKMBDHVN')
    if ciftler is None:
        rap.atla(M, u'M5-D8', u'the degenerate base leak scan', u'there is no pairs.tsv',
                 kay.ciftler)
    else:
        dej_hedef = []
        for r in ciftler:
            for alan in (u'F', u'R'):
                s = _primer_norm(r.get(alan))
                if s and set(s) & DEJ:
                    dej_hedef.append((r.get(u'hedef'), alan, s,
                                      sorted(set(s) & DEJ), r.get('_satir')))
        rap.olcum[u'M5 D8 dejenere primer'] = u'%d oligo' % len(dej_hedef)
        for h, alan, s, bazlar, sat in dej_hedef:
            ilgili = None
            if siparis:
                ilgili = next((x for x in siparis
                               if _ad_norm(x.get(u'hedef') or u'') == _ad_norm(h or u'')), None)
            butun = u' '.join(unicode_(v) for k, v in (ilgili or {}).items()
                              if k != '_satir')
            if not re.search(r'dejenere|degenerate|IUPAC|acilim|expand', butun, re.I):
                rap.ekle(M, u'M5-D8-DEJENERE', CIDDI,
                         u'the assessment of an oligo carrying a degenerate base must say how the degenerate bases were handled',
                         u'%s / %s: %s (degenerate base: %s), the decision row says nothing about how degenerate bases were handled'
                         % (h, alan, s, u', '.join(bazlar)),
                         u'%s (row %s)' % (kay.ciftler, sat),
                         u'A tool counting a degenerate base as N eliminates this oligo silently and the coverage comes out lower than it is.')

    # ---- PATTERN 9: AN ORIENTATION FAULT ------------------------------
    ind = tsv_oku(kay.konsensus_indeks, yorum=None)
    if ind is None:
        rap.atla(M, u'M5-D9', u'the orientation error scan',
                 u'there is no konsensus_kanonik/INDEKS.tsv', kay.konsensus_indeks)
    else:
        yon_yok = [r for r in ind if not (r.get(u'eski_yon') or u'').strip()]
        cevrilen = [r for r in ind if (r.get(u'cevrildi') or u'').strip().lower()
                    in (u'evet', u'yes', u'1')]
        rap.olcum[u'M5 P9 orientation'] = u'%d bins, %d turned round, %d with no orientation' % (
            len(ind), len(cevrilen), len(yon_yok))
        if yon_yok:
            rap.ekle(M, u'M5-D9-YON-BILGISI-YOK', CIDDI,
                     u'the orientation of every bin in the canonical index must be recorded',
                     u'eski_yon is empty in %d bins: %s'
                     % (len(yon_yok), u', '.join(r.get(u'kutu') or u'?'
                                                 for r in yon_yok[:8])),
                     kay.konsensus_indeks,
                     u'On a reversed consensus, in-silico PCR SILENTLY gives 0 products.')
        # Still reading the non-canonical, mixed orientation directory is a risk.
        #
        # THE B FIX (2026-08-21): this check used to do a plain text search over THE WHOLE
        # file, comments and docstrings included. Files that merely MENTIONED the directory
        # name in a comment were being marked RISKLI; in the 2026-08-09 run that produced
        # five false positives (steps/generate_primer_candidates.py and
        # design_group_primers.py among them, both of which take the path as a CLI argument
        # and carry no embedded path).
        # The CORRECT version of the same check already existed in the project:
        # screening/yon_kod_taramasi.kod_govdesi(). Two scanners were giving different
        # answers to the same question; both now look at the body.
        for klasor in kay.kod_klasorleri:
            for yol in sorted(glob.glob(os.path.join(klasor, u'*.py'))):
                if u'.orig' in yol:
                    continue
                ham = metin_oku(yol) or u''
                vurus = d9_karisik_klasor_yollari(ham, yol)
                if vurus and u'konsensus_kanonik' not in _kod_govdesi(ham):
                    rap.ekle(M, u'M5-D9-KARISIK-KLASOR', CIDDI,
                             u'the consensus reads must be made from the CANONICAL directory',
                             u'%s: it reads the mixed orientation "consensus sequences" directory IN THE CODE (line %s) and never mentions the canonical directory'
                             % (os.path.relpath(yol, kay.kok),
                                u', '.join(str(i) for i, _ in vurus[:6])), yol,
                             u'That directory is mixed orientation; a reversed consensus silently gives 0 products.')


# =========================================================================
# MODULE 6 - DATABASE HEALTH
# =========================================================================
# THE QUESTION: is every database index REALLY working?
#
# WHY: the SILVA index returned zero SILENTLY for months. The evidence of the broken
# index is still there (SILVA_138.2_SSURef_NR99.fasta.BOZUK_KANIT.txt): the broken
# build said "Sorting 19683 kmers", and 19683 = 3^9, which means one of the four
# bases had dropped out. On a sound index kmer_count must be 4^9 = 262144. That is
# why the k-mer count is checked DIRECTLY.
# -------------------------------------------------------------------------

KMER_BEKLENEN = 262144      # 4^9 - the k-mer count of a sound index
KVALUE_BEKLENEN = 9


def modul_6_veritabani(kay, rap, baglanma_sinamasi=True):
    M = u'6 DATABASE'

    K, hata = modul_yukle(kay.kimlik_dogrulama, 'kimlik_dogrulama_m6')
    if K is None:
        rap.atla(M, u'M6-VTB-LISTESI', u'the database list must be read from identity_verification.py',
                 hata, kay.kimlik_dogrulama)
        return
    T, _h = modul_yukle(kay.tum_kutu, 'tum_kutu_m6')
    beklenen_kayit = getattr(T, 'BEKLENEN_KAYIT', {}) if T else {}

    if not os.path.isdir(kay.refdb):
        rap.atla(M, u'M6-REFDB', u'the REFERANS_DB directory must be present', u'there is no such directory',
                 kay.refdb)
        return

    saglikli = 0
    denetlenen = 0
    for etiket, dosya, lokus, kullan, _not in K.VTB:
        if not kullan:
            continue                     # twin and subset sets do not join the vote
        denetlenen += 1
        yol = os.path.join(kay.refdb, dosya)
        if not os.path.exists(yol):
            rap.ekle(M, u'M6-FASTA-YOK', KRITIK,
                     u'the %s database file must be present' % etiket,
                     u'%s is not there' % dosya, kay.refdb,
                     u'This source cannot vote on any identity verdict.')
            continue
        if os.path.getsize(yol) == 0:
            rap.ekle(M, u'M6-FASTA-BOS', KRITIK,
                     u'the %s database file must not be empty' % etiket,
                     u'%s: 0 bayt' % dosya, yol)
            continue

        # --- 1) the health log of the k-mer index
        log = yol + u'.log'
        icerik = metin_oku(log)
        if icerik is None:
            rap.ekle(M, u'M6-INDEKS-GUNLUK-YOK', CIDDI,
                     u'an index build log must be present for %s' % etiket,
                     u'there is no %s.log, so a healthy index build CANNOT BE CONFIRMED'
                     % dosya, kay.refdb,
                     u'The k-mer count could not be audited for this database.')
        else:
            m = re.findall(r'kvalue=(\d+),\s*kmer_count=(\d+)', icerik)
            if not m:
                rap.ekle(M, u'M6-KMER-SATIRI-YOK', CIDDI,
                         u'the %s log must hold a "kvalue=..., kmer_count=..." line'
                         % etiket,
                         u'there is no k-mer line in the log (the index may have been built by an old version, or in a different format)', log,
                         u'The k-mer count could not be confirmed; the silent zero in the SILVA case was caught by this very line.')
            else:
                kv, kc = int(m[-1][0]), int(m[-1][1])
                bekl = 4 ** kv
                if kc != bekl or kv != KVALUE_BEKLENEN or kc != KMER_BEKLENEN:
                    rap.ekle(M, u'M6-KMER-BOZUK', KRITIK,
                             u'the k-mer count has to be 4^%d = %d'
                             % (KVALUE_BEKLENEN, KMER_BEKLENEN),
                             u'%s: kvalue=%d, kmer_count=%d (4^%d = %d)'
                             % (etiket, kv, kc, kv, bekl), log,
                             u'19683 = 3^9 is the signature of a broken index: it means a base has dropped out, and the index SILENTLY returns zero.')

        # --- 2) is the index file newer than the FASTA (a stale index)
        indeksler = [yol + u'.primerqc.bin', yol + u'.primerqc']
        var = [i for i in indeksler if os.path.exists(i)]
        if not var:
            rap.ekle(M, u'M6-INDEKS-YOK', CIDDI,
                     u'a k-mer index (.primerqc.bin) must be present for %s' % etiket,
                     u'there is no index file, so this database can only be scanned by streaming the FASTA directly, and the layers that rely on an index return ZERO on this source', kay.refdb,
                     u'The SILVA LSU sets are in that state; a measurement relying on the index skips this source silently.')
        else:
            for i in var:
                if os.path.getmtime(i) < os.path.getmtime(yol):
                    rap.ekle(M, u'M6-INDEKS-BAYAT', CIDDI,
                             u'the index has to be NEWER than the FASTA file',
                             u'%s: the index is %s and the FASTA %s (the index is OLDER)'
                             % (etiket,
                                time.strftime('%Y-%m-%d', time.localtime(os.path.getmtime(i))),
                                time.strftime('%Y-%m-%d', time.localtime(os.path.getmtime(yol)))),
                             i, u'The FASTA was updated but the index was not rebuilt.')

        # --- 3) is the record count what was expected
        bekl_n = beklenen_kayit.get(etiket)
        fai = yol + u'.fai'
        gercek_n = None
        if os.path.exists(fai):
            try:
                with io.open(fai, encoding='utf-8', errors='replace') as fh:
                    gercek_n = sum(1 for s in fh if s.strip())
            except IOError:
                gercek_n = None
        if bekl_n and gercek_n is not None and gercek_n != bekl_n:
            rap.ekle(M, u'M6-KAYIT-SAYISI', CIDDI,
                     u'%s must hold %d records' % (etiket, bekl_n),
                     u'the .fai index shows %d records' % gercek_n, fai)
        elif bekl_n and gercek_n is None:
            rap.ekle(M, u'M6-KAYIT-SAYISI-YOK', UYARI,
                     u'the record count must be confirmable for %s' % etiket,
                     u'there is no .fai, so the record count WAS NOT CONFIRMED', kay.refdb)

        # --- 4) A LIVE BINDING TEST: does a known primer come back with more than
        #        zero binding? So that this does not become a check that "always
        #        passes", the test reads the real file.
        if baglanma_sinamasi:
            sonuc = _baglanma_sinamasi(K, yol, lokus)
            if sonuc is None:
                rap.atla(M, u'M6-BAGLANMA', u'a known conserved sequence must be present inside %s' % etiket,
                         u'the test could not be run (the file could not be read)', yol)
            elif sonuc == 0:
                rap.ekle(M, u'M6-BAGLANMA-SIFIR', KRITIK,
                         u'there must be at least one binding for the known conserved sequences inside %s' % etiket,
                         u'ZERO bindings in the first 20 000 records: the database is being read but its content does not match what was expected', yol,
                         u'The silent zero in the SILVA case looked exactly like this.')
            else:
                saglikli += 1

    rap.olcum[u'M6 databases'] = u'%d sources audited, %d passed the live binding test' % (
        denetlenen, saglikli)
    if denetlenen == 0:
        rap.atla(M, u'M6-BOS', u'at least one database must be audited',
                 u'none of the sources in the database list is present', kay.kimlik_dogrulama)


# Known conserved regions. These are NOT our DESIGN primers; they are widely
# accepted universal primer sequences and they MUST be present at the relevant
# locus. If they are not, the file or the index is broken.
_SINAMA_DIZILERI = {
    'SSU': [u'GTGCCAGCMGCCGCGGTAA',    # 515F, SSU'da evrensel
            u'GGATTAGATACCC'],          # 787 bolgesi, cok korunmus
    'LSU': [u'GCATATCAATAAGCGGAGGA',    # LR0R
            u'ACCCGCTGAACTTAAGC'],      # LSU korunmus
    'ITS': [u'GGAAGTAAAAGTCGTAACAAGG',  # ITS1
            u'TCCTCCGCTTATTGATATGC'],   # ITS4
    'OPERON': [u'GTGCCAGCMGCCGCGGTAA', u'GGATTAGATACCC'],
    'KARISIK': [u'GTGCCAGCMGCCGCGGTAA', u'GGAAGTAAAAGTCGTAACAAGG',
                u'GGATTAGATACCC'],
}
_IUPAC = {u'M': u'[AC]', u'R': u'[AG]', u'W': u'[AT]', u'S': u'[GC]', u'Y': u'[CT]',
          u'K': u'[GT]', u'V': u'[ACG]', u'H': u'[ACT]', u'D': u'[AGT]',
          u'B': u'[CGT]', u'N': u'[ACGT]'}


def _dizi_desen(d):
    u"""The regular expression that OPENS the degenerate bases. Counting a degenerate
    base as an N would make this test return zero by mistake, which is pattern 8
    itself."""
    return re.compile(u''.join(_IUPAC.get(c, re.escape(c)) for c in d.upper()))


def _baglanma_sinamasi(K, yol, lokus, tavan=20000):
    """Does at least one of the known conserved sequences occur in the first `tavan` records?

        Returns: the number of bindings found, or None if the file cannot be read.
        It searches both forward and reverse complement; searching one direction would
        miss half the records (the orientation varies from set to set).

    """
    desenler = []
    for d in _SINAMA_DIZILERI.get(lokus, _SINAMA_DIZILERI['KARISIK']):
        desenler.append(_dizi_desen(d))
        desenler.append(_dizi_desen(K.rc(re.sub(r'[^A-Z]', u'', d.upper()))))
    n = 0
    bulunan = 0
    try:
        for _bas, diz in K.fasta_akisi(yol):
            n += 1
            for de in desenler:
                if de.search(diz):
                    bulunan += 1
                    break
            if n >= tavan:
                break
    except (IOError, OSError):
        return None
    return bulunan


# =========================================================================
# MODULE 7 - TAXON COVERAGE
# =========================================================================
# THE QUESTION: does EVERY target requested in the decisions find a counterpart in
# the panel? If not, which category does the reason fall into?
#
# FOUR CATEGORIES (they lead to different conclusions):
#   NO ORGANISM        - the requested taxon is not in the sample at all. No primer
#                        can be designed; that is not a failure, it is a finding.
#   NO DISCRIMINATION  - the organism is there but cannot be separated from its
#                        relatives.
#   NO LOCUS           - the organism is there and can be separated, but there is no
#                        discriminating region at the locus we have (another locus
#                        is needed).
#   MEMBER SET SPLIT   - the target is not a single lineage; its members are far
#                        apart from one another.
# A gap whose category cannot be determined is an UNCLOSED gap.
# -------------------------------------------------------------------------

_KATEGORI = (
    (u'NO ORGANISM', re.compile(
        r'numunede\s*(hic\s*)?yok|t[uü]r\s*numunede\s*yok|cins\s*numunede\s*yok|'
        r'organizma\s*yok|etiket\s*[cç][uü]r[uü]t[uü]ld[uü]'
        r'|not\s*in\s*the\s*sample|no\s*such\s*organism'
        r'|absent\s*from\s*the\s*sample|label\s*refuted', re.I)),
    (u'NO DISCRIMINATION', re.compile(
        r'e[sş]ik\s*alt|ayr[iı]m\s*(kat[iı]\s*)?(yok|tan[iı]ms[iı]z|0[,.]0)|'
        r'ayr[iı]lam[iı]yor|ayr[iı]m\s*e[sş]ik'
        r'|below\s*(the\s*)?threshold|no\s*discrimination'
        r'|discrimination\s*(is\s*)?(undefined|zero)'
        r'|cannot\s*be\s*separated', re.I)),
    (u'NO LOCUS', re.compile(
        r'lokus|ba[sş]ka\s*b[oö]lge|ay[iı]rt\s*edici\s*b[oö]lge\s*yok|'
        r'ITS.*yetersiz|farkl[iı]\s*lokus'
        r'|locus|another\s*region|no\s*discriminating\s*region', re.I)),
    (u'THE MEMBER SET IS SPLIT', re.compile(
        r'heterojen|ayr[iı][sş][iı]k|birbirine\s*%?\s*\d+[,.]?\d*\s*[-, ]\s*\d+|'
        r'kutu\s*heterojen|[uü]ye\s*k[uü]mesi\s*ayr'
        r'|heterogen|split\s*member\s*set', re.I)),
)


def _toplanti_istekleri(metin):
    u"""Pull (request, status_text) out of the tables in the decisions markdown.

    The first column of the tables carries the requested taxon and the last column
    carries its status. Heading and separator rows are skipped.
    """
    out = []
    bolum = u''
    for satir in metin.splitlines():
        s = satir.strip()
        if s.startswith(u'## '):
            bolum = s[3:].strip()
            continue
        if not s.startswith(u'|'):
            continue
        hucre = [h.strip() for h in s.strip(u'|').split(u'|')]
        if len(hucre) < 2:
            continue
        if set(u''.join(hucre)) <= set(u'-: '):
            continue                                  # ayrac satiri
        ilk = re.sub(r'[*_`]', u'', hucre[0]).strip()
        if not ilk or ilk.lower().startswith((u'istenen', u'hedef', u'#', u'toplanti')):
            continue                                  # baslik satiri
        out.append((bolum, ilk, u' | '.join(hucre[1:])))
    return out


def modul_7_kapsam(kay, rap):
    M = u'7 TAXON COVERAGE'

    toplanti = metin_oku(kay.toplanti)
    if toplanti is None:
        rap.atla(M, u'M7-TOPLANTI-YOK', u'TOPLANTI_KARARLARI_SON_DURUM.md must be readable',
                 u'there is no such file', kay.toplanti)
        return
    istekler = _toplanti_istekleri(toplanti)
    if not istekler:
        rap.atla(M, u'M7-ISTEK-YOK',
                 u'the requested targets must be extractable from the meeting decisions',
                 u'no table row was found inside the markdown', kay.toplanti)
        return

    ciftler = tsv_oku(kay.ciftler)
    panel_hedefleri = set(_ad_norm(r.get(u'hedef') or u'') for r in (ciftler or []))
    if ciftler is None:
        rap.atla(M, u'M7-PANEL-YOK', u'the panel pair list must be readable',
                 u'there is no pairs.tsv', kay.ciftler)

    # The link between the panel targets and the meeting decisions is held in the KARAR
    # dictionary inside target_taxon_mapping.py. The file IS NOT EXECUTED, it is parsed
    # with ast; a read only auditor does not run a script.
    esleme_karar = {}
    kaynak = metin_oku(kay.takson_esleme)
    if kaynak is None:
        rap.atla(M, u'M7-ESLEME-YOK',
                 u'the KARAR dictionary inside target_taxon_mapping.py must be readable',
                 u'there is no such file', kay.takson_esleme)
    else:
        try:
            agac = ast.parse(kaynak)
            for d in agac.body:
                if not isinstance(d, ast.Assign):
                    continue
                adlar = [t.id for t in d.targets if isinstance(t, ast.Name)]
                if u'KARAR' not in adlar:
                    continue
                for anahtar, deger in zip(d.value.keys, d.value.values):
                    try:
                        v = ast.literal_eval(deger)
                        esleme_karar[ast.literal_eval(anahtar)] = v
                    except (ValueError, SyntaxError):
                        continue
        except SyntaxError as e:
            rap.atla(M, u'M7-ESLEME-AYRISTIRILAMADI',
                     u'target_taxon_mapping.py ayristirilabilmeli',
                     u'SyntaxError: %s' % e, kay.takson_esleme)

    karar_hedefleri = set(_ad_norm(v[0]) for v in esleme_karar.values()
                          if isinstance(v, (list, tuple)) and v)

    kapatilmamis = 0
    kategori_sayaci = collections.Counter()
    for bolum, istek, durum in istekler:
        # The source document (TOPLANTI_KARARLARI_SON_DURUM.md) is a delivery file
        # written in Turkish, but the same document written in English has to be
        # auditable too. MEASURED: item 7 of the self test planted an English status
        # line ("**Not achieved.** No reason recorded"), the pattern knew Turkish
        # only, so the line never counted as "not achieved" and the module never saw
        # its planted error at all. Both languages are recognised now.
        yapilamadi = bool(re.search(
            r'yap[iı]lamad[iı]|sipari[sş]\s*edilmez|panelden\s*[cç][iı]kar'
            r'|not\s*achieved|could\s*not\s*be\s*done|not\s*ordered'
            r'|removed\s*from\s*the\s*panel', durum, re.I))
        if not yapilamadi:
            continue
        kategoriler = [ad for ad, d in _KATEGORI if d.search(durum)]
        if kategoriler:
            for k in kategoriler:
                kategori_sayaci[k] += 1
        else:
            kapatilmamis += 1
            rap.ekle(M, u'M7-KATEGORISIZ-BOSLUK', CIDDI,
                     u'the reason for every unmet target must fall into one of four categories (no organism / no discrimination / no locus / a split member set)',
                     u'"%s" (%s): "%s", the reason falls into no category'
                     % (istek, bolum, durum.strip()[:200]), kay.toplanti,
                     u'A gap with no category cannot answer the question "why was it not there" in the report.')

    # --- targets that are in the panel but link to no decision
    if ciftler is not None and karar_hedefleri:
        baglanmayan = sorted(h for h in panel_hedefleri if h and h not in karar_hedefleri)
        for h in baglanmayan:
            asil = next((r.get(u'hedef') for r in ciftler
                         if _ad_norm(r.get(u'hedef') or u'') == h), h)
            rap.ekle(M, u'M7-PANELDE-FAZLA', UYARI,
                     u'every target in the panel must be tied to a meeting decision',
                     u'the target "%s" is not in the KARAR table of target_taxon_mapping.py'
                     % asil, kay.takson_esleme,
                     u'If the report is asked "why is this in the panel" there is no ground to show.')
    elif ciftler is not None and not karar_hedefleri:
        rap.atla(M, u'M7-BAGLANTI',
                 u'the targets in the panel must be tied to the meeting decisions',
                 u'the KARAR table could not be read, so the link could not be audited',
                 kay.takson_esleme)

    rap.olcum[u'M7 coverage'] = (
        u'%d requests scanned; how they fall into the categories: %s; gaps with no '
        u'category: %d'
        % (len(istekler),
           u', '.join(u'%s=%d' % (k, v) for k, v in sorted(kategori_sayaci.items()))
           or u'-', kapatilmamis))


# =========================================================================
# WRITING THE REPORT
# =========================================================================
def _tsv_kacis(s):
    return unicode_(s).replace(u'\t', u' ').replace(u'\n', u' ').replace(u'\r', u' ')


def raporla(kay, rap, cikti, kosulan, sureler):
    """One markdown report plus a machine readable TSV. Nothing is written anywhere else."""
    if not os.path.isdir(cikti):
        os.makedirs(cikti)
    damga = time.strftime('%Y-%m-%d_%H%M')
    md_yol = os.path.join(cikti, u'CAPRAZ_KONTROL_%s.md' % damga)
    tsv_yol = os.path.join(cikti, u'CAPRAZ_KONTROL_%s.tsv' % damga)

    sirali = sorted(rap.bulgular,
                    key=lambda b: (_SIRA.get(b.ciddiyet, 9), b.modul, b.kod))

    # ---------------- TSV
    with io.open(tsv_yol, 'w', encoding='utf-8') as fh:
        fh.write(u'\t'.join([u'ciddiyet', u'modul', u'kod', u'ne_bekleniyordu',
                             u'ne_bulundu', u'dosya', u'oneri']) + u'\n')
        for b in sirali:
            fh.write(u'\t'.join(_tsv_kacis(x) for x in
                                [b.ciddiyet, b.modul, b.kod, b.beklenen,
                                 b.bulunan, b.dosya, b.oneri]) + u'\n')

    # ---------------- the module 1 identity table as its own TSV (the main table
    #                  that goes into the report)
    kimlik_yol = None
    satirlar = rap.tablolar.get(u'M1 identity table')
    if satirlar:
        kimlik_yol = os.path.join(cikti, u'KIMLIK_VE_KRAKEN_%s.tsv' % damga)
        sut = [(u'kutu', u'kutu'), (u'Kraken_etiketi', u'kraken_etiket'),
               (u'Kraken_guven', u'kraken_guven'), (u'olculen_kimlik', u'olculen'),
               (u'olculen_duzey', u'duzey'), (u'kimlik_yuzde', u'kimlik'),
               (u'hizalanan_bp', u'hiz_uz'), (u'karar_veren_vtb', u'vtb'),
               (u'kayit_no', u'kayit'), (u'TIP_KAYDI', u'tip'),
               (u'ikinci_isabet', u'ikinci'), (u'fark_yuzde_puan', u'fark'),
               (u'ayirt_edilebilir', u'ayrilir'), (u'ayirt_gerekce', u'ayrim_sebep'),
               (u'N_orani', u'n_oran'), (u'saflik', u'saflik'),
               (u'saflik_gerekce', u'saflik_sebep'), (u'sorgu_kaynagi', u'sorgu_kaynagi'),
               (u'uyusan_vtb_sayisi', u'uyusan'), (u'uyusmayanlar', u'uyusmayan'),
               (u'AD_DEGISTI_MI', u'ad_degisti')]
        with io.open(kimlik_yol, 'w', encoding='utf-8') as fh:
            fh.write(u'\t'.join(a for a, _b in sut) + u'\n')
            for r in satirlar:
                fh.write(u'\t'.join(_tsv_kacis(
                    vir(r.get(b), 2) if isinstance(r.get(b), float) else
                    (u'-' if r.get(b) is None else r.get(b))) for _a, b in sut) + u'\n')

    # ---------------- MARKDOWN
    sayim = collections.OrderedDict(
        (c, rap.say(c)) for c in (KRITIK, CIDDI, UYARI, BILGI, ATLANDI))
    kod = rap.cikis_kodu()
    L = []
    L.append(u'# THE CROSS-CHECK REPORT')
    L.append(u'')
    L.append(u'The PrimerJury qPCR panel: an independent, read only audit.')
    L.append(u'')
    L.append(u'| | |')
    L.append(u'|---|---|')
    L.append(u'| Date | %s |' % time.strftime('%Y-%m-%d %H:%M'))
    L.append(u'| Script | cross_check.py %s |' % VERSIYON)
    L.append(u'| Root directory | `%s` |' % kay.kok)
    L.append(u'| Modules that ran | %s |' % (u', '.join(kosulan) or u'-'))
    L.append(u'| Exit code | **%d** |' % kod)
    L.append(u'')
    L.append(u'## Summary')
    L.append(u'')
    L.append(u'| Severity | Count |')
    L.append(u'|---|---:|')
    for c, n in sayim.items():
        L.append(u'| %s | %d |' % (c, n))
    L.append(u'')
    if sayim[ATLANDI]:
        L.append(u'> **%d checks COULD NOT BE RUN.** These are NOT "passed"; the reason for each is written in the SKIPPED section below. The exit code carries bit 4.' % sayim[ATLANDI])
        L.append(u'')

    L.append(u'## The state of the modules')
    L.append(u'')
    L.append(u'| Module | State | Time | Findings (K/C/U/B/A) |')
    L.append(u'|---|---|---|---|')
    for m in kosulan:
        d = rap.modul_durumu.get(m, {})
        L.append(u'| %s | %s | %s | %d/%d/%d/%d/%d |' % (
            m, d.get(u'durum', u'?'),
            sure_metni(sureler.get(m)) if sureler.get(m) is not None else u'olculmedi',
            rap.say(KRITIK, m), rap.say(CIDDI, m), rap.say(UYARI, m),
            rap.say(BILGI, m), rap.say(ATLANDI, m)))
    L.append(u'')

    if rap.olcum:
        L.append(u'## The measured values')
        L.append(u'')
        L.append(u'Every number in this section WAS MEASURED in this run. Nothing that was not measured is written here.')
        L.append(u'')
        L.append(u'| Measurement | Value |')
        L.append(u'|---|---|')
        for k, v in rap.olcum.items():
            if k.startswith(u'_'):
                continue          # an internal value; it does not go into the report
            L.append(u'| %s | %s |' % (k, v))
        L.append(u'')

    for c in (KRITIK, CIDDI, UYARI, BILGI, ATLANDI):
        grup = [b for b in sirali if b.ciddiyet == c]
        if not grup:
            continue
        L.append(u'## %s (%d)' % (c, len(grup)))
        L.append(u'')
        for b in grup:
            L.append(u'### %s - %s' % (b.kod, b.modul))
            L.append(u'')
            L.append(u'- **What was expected:** %s' % b.beklenen)
            L.append(u'- **What was found:** %s' % b.bulunan)
            L.append(u'- **File:** `%s`' % b.dosya)
            if b.oneri:
                L.append(u'- **Why it matters:** %s' % b.oneri)
            L.append(u'')

    if satirlar:
        L.append(u'## The identity against the Kraken comparison')
        L.append(u'')
        L.append(u'The full table: `%s`' % os.path.basename(kimlik_yol))
        L.append(u'')
        degisen = [r for r in satirlar if r.get(u'ad_degisti') == u'EVET']
        L.append(u'**The bins whose NAME CHANGED (%d)**, the table to show in the report:'
                 % len(degisen))
        L.append(u'')
        if degisen:
            L.append(u'| Bin | Kraken label | Measured identity | % | The database that decided | Type record | Second hit | Difference | Distinguishable |')
            L.append(u'|---|---|---|---:|---|---|---|---:|---|')
            for r in degisen:
                L.append(u'| %s | %s | %s | %s | %s | %s | %s | %s | %s |' % (
                    r.get(u'kutu'), r.get(u'kraken_etiket'), r.get(u'olculen'),
                    vir(r.get(u'kimlik'), 2), r.get(u'vtb'), r.get(u'tip'),
                    r.get(u'ikinci'), vir(r.get(u'fark'), 2), r.get(u'ayrilir')))
        else:
            L.append(u'_No bin changed its name in this run._')
        L.append(u'')

    L.append(u'## What the exit code means')
    L.append(u'')
    L.append(u'A bit mask, added together: 1 = there is a CRITICAL finding, 2 = there is a SERIOUS finding, 4 = at least one check was SKIPPED, 8 = the script itself failed.')
    L.append(u'This run: **%d**.' % kod)
    L.append(u'')

    with io.open(md_yol, 'w', encoding='utf-8') as fh:
        fh.write(u'\n'.join(L))
    return md_yol, tsv_yol, kimlik_yol


# =========================================================================
# THE SELF TEST  -  DELIBERATELY BROKEN INPUT
# =========================================================================
# "A check that always passes is probably measuring nothing." This section gives
# every module a DELIBERATELY BROKEN input and shows whether that module REALLY
# catches the error. A module that does not catch it is useless however green it
# looks, and the self test counts as FAILED.
#
# The test is set up in a temporary directory; the real project files ARE NOT
# TOUCHED.
# -------------------------------------------------------------------------

def _sinama_kok_kur(gecici, kay):
    """Set up a small, complete and SOUND fake project root. Errors are seeded into it after."""
    os.makedirs(os.path.join(gecici, 'screening'))
    os.makedirs(os.path.join(gecici, 'verification'))
    os.makedirs(os.path.join(gecici, 'REFERANS_DB'))
    os.makedirs(os.path.join(gecici, 'konsensus_kanonik'))
    y = lambda *p: os.path.join(gecici, *p)

    def d(yol, icerik):
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(icerik)

    # --- the panel definition: two targets, both sound
    d(y('screening', 'pairs.tsv'),
      u'satir\thedef\tsinif\tF\tR\tuye_taksonlar\tolcu_tipi\tuye_kumesi_durumu\n'
      u'2\tHedef_A\tA\tACGTACGTACGTACGTACGT\tTTGCATGCATGCATGCATGC\t111,222\tayrim\tPANELLE_TUTUYOR\n'
      u'3\tHedef_B\tB\tGGGTACGTACGTACGTACGT\tCCGCATGCATGCATGCATGC\t333\tkapsam\tPANELLE_TUTUYOR\n')
    d(y('screening', 'target_membership.tsv'),
      u'hedef\tuye_taxid\tharic\tkaynak\tnot\n'
      u'Hedef_A\t111,222\t999\tPANEL\t-\n'
      u'Hedef_B\t333\t\tPANEL\t-\n')
    d(y('screening', 'target_taxon_mapping.py'),
      u"KARAR = {\n 2: ('Hedef_A', 'cins', 'Karar 1', 'not'),\n"
      u" 3: ('Hedef_B', 'cins', 'Karar 1', 'not'),\n}\n")
    # --- the decision tables: consistent
    d(y('NIHAI_SIPARIS_LISTESI_2026-08-07.tsv'),
      u'sira\tHUKUM\thedef\turun_bp\tF\tR\thukmu_veren_katman\tGEREKCE\t'
      u'yerel_hedef_disi\tkraken_etiketi\tolculen_kimlik\tdCq\n'
      u'1\tSIPARIS\tHedef_A\t150\tACGTACGTACGTACGTACGT\tTTGCATGCATGCATGCATGC\t'
      u'yerel tarama\tolculdu: in siliko hizalama ve jel dogrulamasi gerekli; '
      u'kapsam, filum spektrumu ve organel orani olculdu\t3\tTaxon A\tTaxon A\t5,00\n'
      u'2\tSIPARIS\tHedef_B\t120\tGGGTACGTACGTACGTACGT\tCCGCATGCATGCATGCATGC\t'
      u'MFE katmani\tolculdu: in siliko MFE ve erime egrisi dogrulamasi gerekli; '
      u'kapsam, filum spektrumu ve organel orani olculdu\t2\tTaxon B\tTaxon B\t4,00\n')
    d(y('SIPARIS_LISTESI.tsv'),
      u'sira\thedef\tF\tR\turun_bp\tkraken_etiketi\tolculen_kimlik\tdCq_karsiligi\n'
      u'1\tHedef_A\tACGTACGTACGTACGTACGT\tTTGCATGCATGCATGCATGC\t150\tTaxon A\tTaxon A\t5,00\n'
      u'2\tHedef_B\tGGGTACGTACGTACGTACGT\tCCGCATGCATGCATGCATGC\t120\tTaxon B\tTaxon B\t4,00\n')
    # R=1 -> log2(1)+4.3 = 4.3 ; floor 3.32 -> required 4.30
    d(y('ESIK_VE_OLCUT_2026-08-08.tsv'),
      u'hedef\tESKI_HUKUM\tYENI_HUKUM\tdCq_olculen\tR\tgerekli_dCq\t'
      u'yeni_kural_durum\tuyelik_gecerli_mi\n'
      u'Hedef_A\tGECER\tGECER\t5,00\t1\t4,30\tGECER\tEVET\n'
      u'Hedef_B\tKALIR\tKALIR\t4,00\t1\t4,30\tKALIR\tEVET\n')
    d(y('HEDEF_DISI_AYRINTI_2026-08-07.tsv'),
      u'hedef\ttaxid\tnot\nHedef_A\t777\tbaska organizma\n')
    # --- literatur ve toplanti
    d(y('LITERATUR_2026-08-07.md'),
      u'# Literatur\n\ngerekli dCq >= log2(R) + 4,3 ve taban 3,32 olmali.\n')
    d(y('TOPLANTI_KARARLARI_SON_DURUM.md'),
      u'## Karar 1\n\n| Istenen tur | Durum |\n|---|---|\n'
      u'| *Taxon A* | Var, siparis edilir |\n'
      u'| *Taxon C* | **Yapilamadi.** Tur numunede yok |\n')
    # --- the consensus index and its file
    d(y('konsensus_kanonik', 'A-1_111.kanonik.fa'), u'>x\n%s\n' % (u'ACGT' * 200))
    d(y('konsensus_kanonik', 'INDEKS.tsv'),
      u'kutu\tsinif\tdosya\tkaynak\teski_yon\tcevrildi\tuzunluk\n'
      u'A-1_111\tA\tA-1_111.kanonik.fa\tkons\tSENSE\thayir\t800\n')
    # --- copy the identity engine (the VTB list is needed for M6)
    if os.path.exists(kay.kimlik_dogrulama):
        with io.open(kay.kimlik_dogrulama, encoding='utf-8', errors='replace') as fh:
            d(y('verification', 'identity_verification.py'), fh.read())
    if os.path.exists(kay.tum_kutu):
        with io.open(kay.tum_kutu, encoding='utf-8', errors='replace') as fh:
            d(y('verification', 'all_bin_identities.py'), fh.read())
    return gecici


def _sinama_vtb_dosyalari(gecici, kay, kmer_satiri=u'kvalue=9, kmer_count=262144'):
    """Produce a small but REALISTIC copy of every file in the VTB list.

        For the binding test to pass, the files have to hold the known conserved
        sequences; that way the "sound" state really looks sound and the seeded error
        can be told apart.

    """
    K, _h = modul_yukle(kay.kimlik_dogrulama, 'kd_sinama')
    if K is None:
        return []
    govde = (u'GTGCCAGCAGCCGCGGTAA' + u'ACGT' * 60 + u'GGATTAGATACCC' +
             u'ACGT' * 60 + u'GGAAGTAAAAGTCGTAACAAGG' + u'ACGT' * 60 +
             u'GCATATCAATAAGCGGAGGA' + u'ACGT' * 60 + u'TCCTCCGCTTATTGATATGC' +
             u'ACGT' * 60 + u'ACCCGCTGAACTTAAGC' + u'ACGT' * 60)
    yazilan = []
    for etiket, dosya, _lokus, kullan, _n in K.VTB:
        if not kullan:
            continue
        yol = os.path.join(gecici, 'REFERANS_DB', dosya)
        with io.open(yol, 'w', encoding='utf-8') as fh:
            for i in range(3):
                fh.write(u'>RECORD_%d Example organism %d; from TYPE material\n%s\n'
                         % (i, i, govde))
        with io.open(yol + u'.log', 'w', encoding='utf-8') as fh:
            fh.write(u'Binary index v5-single-strand-64 created: 1 MB, %s\n' % kmer_satiri)
        with io.open(yol + u'.primerqc.bin', 'w', encoding='utf-8') as fh:
            fh.write(u'a fake index')
        yazilan.append((etiket, yol))
    return yazilan


def kendini_sina(kay, cikti):
    """Give every module a deliberately broken input and measure whether it catches the error.

        Returns: (modules_passed, total_modules, detail_list)

    """
    import shutil
    import tempfile

    def kur_m1(g):
        # THE ERROR: the consensus file named in the index IS NOT on disk.
        os.remove(os.path.join(g, 'konsensus_kanonik', 'A-1_111.kanonik.fa'))
        return u'INDEKS.tsv names a consensus file but the file was deleted'

    def kur_m2(g):
        # THE ERROR: the forward primer of the same target DIFFERS between two files.
        yol = os.path.join(g, 'SIPARIS_LISTESI.tsv')
        s = metin_oku(yol).replace(u'ACGTACGTACGTACGTACGT', u'AAAAACGTACGTACGTACGT', 1)
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(s)
        return u'the forward primer of Hedef_A was written differently in SIPARIS_LISTESI'

    def kur_m3(g):
        # THE ERROR: a pair's member set was EMPTIED but its dCq is still there.
        yol = os.path.join(g, 'screening', 'pairs.tsv')
        s = metin_oku(yol).replace(u'\t111,222\t', u'\t\t', 1)
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(s)
        return u'the uye_taksonlar column of Hedef_A was emptied, the dCq was kept'

    def kur_m4(g):
        # THE ERROR: required_dCq was pulled to a wrong value by hand (4.30 -> 2.00).
        yol = os.path.join(g, 'ESIK_VE_OLCUT_2026-08-08.tsv')
        s = metin_oku(yol).replace(u'\t4,30\tGECER', u'\t2,00\tGECER', 1)
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(s)
        return u'gerekli_dCq was written as 2,00 instead of 4,30, against the log2(R)+4.3 rule'

    def kur_m5(g):
        # THE ERROR: the target's OWN member (111) was put on the off-target list.
        yol = os.path.join(g, 'HEDEF_DISI_AYRINTI_2026-08-07.tsv')
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(u'hedef\ttaxid\tnot\nHedef_A\t111\tits own member was counted as off-target\n')
        return u'taxid 111, a member of Hedef_A itself, was added to its off target list'

    def kur_m6(g):
        # THE ERROR: a database's k-mer count is 3^9 = 19683 (the broken index signature).
        K, _h = modul_yukle(kay.kimlik_dogrulama, 'kd_sinama6')
        ilk = next(d for _e, d, _t, k, _n in K.VTB if k)
        yol = os.path.join(g, 'REFERANS_DB', ilk + u'.log')
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(u'Binary index v5 created: kvalue=9, kmer_count=19683\n')
        return u'the %s index log says kmer_count=19683 (3^9, the signature of a broken index)' % ilk

    def kur_m7(g):
        # THE ERROR: the reason for an unmet target falls into no category.
        yol = os.path.join(g, 'TOPLANTI_KARARLARI_SON_DURUM.md')
        with io.open(yol, 'w', encoding='utf-8') as fh:
            fh.write(u'## Decision 1\n\n| Requested species | Status |\n|---|---|\n| *Taxon C* | **Not achieved.** No reason recorded |\n')
        return u'the reason for an unmet target falls into none of the four categories'

    SINAVLAR = [
        (u'1 IDENTITY', kur_m1, u'M1-KONSENSUS-YOK',
         lambda kk, rr: modul_1_kimlik(kk, rr, KontrolNoktasi(u'', etkin=False),
                                       kip=u'hizli')),
        (u'2 INTERNAL CONSISTENCY', kur_m2, u'M2-CELISKI', lambda kk, rr: modul_2_ic_tutarlilik(kk, rr)),
        (u'3 MEMBERSHIP', kur_m3, u'M3-UYELIK-BOS', lambda kk, rr: modul_3_uyelik(kk, rr)),
        (u'4 LITERATURE', kur_m4, u'M4-ESIK-YANLIS', lambda kk, rr: modul_4_literatur(kk, rr)),
        (u'5 ERROR PATTERNS', kur_m5, u'M5-D2-UYE-HEDEF-DISI',
         lambda kk, rr: modul_5_desenler(kk, rr)),
        (u'6 DATABASE', kur_m6, u'M6-KMER-BOZUK',
         lambda kk, rr: modul_6_veritabani(kk, rr, baglanma_sinamasi=True)),
        (u'7 TAXON COVERAGE', kur_m7, u'M7-KATEGORISIZ-BOSLUK',
         lambda kk, rr: modul_7_kapsam(kk, rr)),
    ]

    yaz(u'')
    yaz(u'=== SELF-TEST: each module is given DELIBERATELY BROKEN input ===')
    ayrinti = []
    gecen = 0
    for ad, kur, beklenen_kod, kos in SINAVLAR:
        gecici = tempfile.mkdtemp(prefix='capraz_sinama_')
        try:
            _sinama_kok_kur(gecici, kay)
            _sinama_vtb_dosyalari(gecici, kay)
            ekilen = kur(gecici)
            kk = Kaynaklar(gecici)
            rr = Rapor()
            t0 = time.time()
            try:
                kos(kk, rr)
                cokme = None
            except Exception as e:
                cokme = u'%s: %s' % (type(e).__name__, e)
            kodlar = set(b.kod for b in rr.bulgular)
            yakalandi = beklenen_kod in kodlar
            if yakalandi:
                gecen += 1
            durum = u'CAUGHT IT' if yakalandi else (
                u'CRASHED (%s)' % cokme if cokme else u'DID NOT CATCH IT')
            yaz(u'  [%s] %-22s  planted error: %s' % (
                u'OK ' if yakalandi else u'ERROR', ad, ekilen))
            yaz(u'        expected finding code: %s  ->  %s  (%s)'
                % (beklenen_kod, durum, sure_metni(time.time() - t0)))
            ayrinti.append(dict(modul=ad, ekilen=ekilen, beklenen=beklenen_kod,
                                durum=durum, yakalandi=yakalandi,
                                uretilen_kod_sayisi=len(kodlar)))
        finally:
            shutil.rmtree(gecici, ignore_errors=True)

    yaz(u'')
    yaz(u'=== SELF-TEST RESULT: %d/%d modules caught the planted error ==='
        % (gecen, len(SINAVLAR)))
    if not os.path.isdir(cikti):
        os.makedirs(cikti)
    yol = os.path.join(cikti, u'KENDINI_SINAMA_%s.tsv' % time.strftime('%Y-%m-%d_%H%M'))
    with io.open(yol, 'w', encoding='utf-8') as fh:
        fh.write(u'modul\tekilen_hata\tbeklenen_bulgu_kodu\tsonuc\n')
        for a in ayrinti:
            fh.write(u'%s\t%s\t%s\t%s\n' % (a['modul'], _tsv_kacis(a['ekilen']),
                                            a['beklenen'], a['durum']))
    yaz(u'Self-test detail: %s' % yol)
    return gecen, len(SINAVLAR), ayrinti


# ===========================================================================
# SURUCU
# ===========================================================================
MODUL_ADLARI = collections.OrderedDict([
    (u'1', u'1 IDENTITY'), (u'2', u'2 INTERNAL CONSISTENCY'), (u'3', u'3 MEMBERSHIP'),
    (u'4', u'4 LITERATURE'), (u'5', u'5 ERROR PATTERNS'), (u'6', u'6 DATABASE'),
    (u'7', u'7 TAXON COVERAGE'),
])



# --- CLI value normalisation ------------------------------------------------
# English option values are accepted alongside the original Turkish ones and
# mapped back here. The internal values are unchanged on purpose: they are
# compared in dozens of places and, in some cases, written to output files.
# Translating the interface must not translate the data.
_DEGER = {'auto': 'oto', 'manual': 'elle', 'none': 'yok', 'quick': 'hizli',
          'full': 'tam', 'member': 'uye', 'competitor': 'rakip',
          'exclude': 'disla'}


def _ing_deger(a):
    for _ad in ('nt', 'literatur', 'ncbi', 'karisik', 'moduller', 'mod'):
        _v = getattr(a, _ad, None)
        if isinstance(_v, str) and _v in _DEGER:
            setattr(a, _ad, _DEGER[_v])
    return a

def main():
    p = argparse.ArgumentParser(
        description=u'The PrimerJury panel: an independent, read only cross-check')
    p.add_argument('--root', dest='kok', default='.', help=u'project root directory')
    p.add_argument('--output', dest='cikti', default=None, help=u'report directory (default CROSSCHECK_RESULT)')
    p.add_argument('--modules', dest='moduller', default='hepsi',
                   help=u'modules to run, comma-separated: 1,2,3 ... or "all"')
    p.add_argument('--m1-mode', dest='m1_kip', default='hizli',
                   choices=['none', 'quick', 'full', 'yok', 'hizli', 'tam'],
                   help=u'M1 scope of the identity scan (default: quick)')
    p.add_argument('--m1-only', dest='m1_yalniz', default=None,
                   help=u'these bins only (comma-separated), example: F2-1_500148,F2-4_500148')
    p.add_argument('--m1-cap', dest='m1_tavan', type=int, default=0,
                   help=u'M1 for maximum number of bins (0 = all)')
    p.add_argument('--reset', dest='sifirla', action='store_true',
                   help=u'ignore checkpoints and recompute everything')
    p.add_argument('--no-checkpoint', dest='kn_yok', action='store_true',
                   help=u'never use a checkpoint')
    p.add_argument('--self-test', dest='kendini_sina', action='store_true',
                   help=u'feed each module deliberately broken input and show that it catches the error')
    a = p.parse_args()
    a = _ing_deger(a)

    kok = os.path.abspath(a.kok)
    cikti = a.cikti or os.path.join(kok, u'CROSSCHECK_RESULT')
    kay = Kaynaklar(kok)

    yaz(u'CROSS-CHECK %s' % VERSIYON)
    yaz(u'Root directory : %s' % kok)
    yaz(u'Output         : %s' % cikti)
    yaz(u'READ-ONLY: this script writes to nothing outside the output directory.')
    if not os.path.isdir(kok):
        yaz(u'ERROR: root directory does not exist: %s' % kok)
        return 8

    if a.kendini_sina:
        gecen, toplam, _ay = kendini_sina(kay, cikti)
        # If a module cannot catch the fault seeded into it, the test FAILS.
        return 0 if gecen == toplam else 8

    if a.moduller.strip().lower() in (u'hepsi', u'all', u'*'):
        secili = list(MODUL_ADLARI.keys())
    else:
        secili = [x.strip() for x in a.moduller.split(u',') if x.strip()]
        bilinmeyen = [x for x in secili if x not in MODUL_ADLARI]
        if bilinmeyen:
            yaz(u'ERROR: unknown module: %s (valid: %s)'
                % (u', '.join(bilinmeyen), u', '.join(MODUL_ADLARI)))
            return 8

    kn_klasor = os.path.join(cikti, u'kontrol')
    if a.sifirla and os.path.isdir(kn_klasor):
        import shutil
        shutil.rmtree(kn_klasor, ignore_errors=True)
        yaz(u'Checkpoints deleted (--reset).')
    kn = KontrolNoktasi(kn_klasor, etkin=not a.kn_yok)

    rap = Rapor()
    kosulan = []
    sureler = {}
    t_hepsi = time.time()

    ISLER = {
        u'1': lambda: modul_1_kimlik(kay, rap, kn, kip=a.m1_kip,
                                     yalniz=a.m1_yalniz, tavan=a.m1_tavan),
        u'2': lambda: modul_2_ic_tutarlilik(kay, rap),
        u'3': lambda: modul_3_uyelik(kay, rap),
        u'4': lambda: modul_4_literatur(kay, rap),
        u'5': lambda: modul_5_desenler(kay, rap),
        u'6': lambda: modul_6_veritabani(kay, rap),
        u'7': lambda: modul_7_kapsam(kay, rap),
    }

    for no in secili:
        ad = MODUL_ADLARI[no]
        kosulan.append(ad)
        yaz(u'')
        yaz(u'===== MODUL %s =====' % ad)
        t0 = time.time()
        try:
            ISLER[no]()
            rap.modul_durumu[ad] = dict(durum=u'kostu')
        except KeyboardInterrupt:
            # AN INTERRUPTION IS NOT SWALLOWED: because the checkpoints are written it can be
            # resumed where it stopped, but the run is said openly to be INCOMPLETE.
            yaz(u'INTERRUPTED (Ctrl+C). The checkpoints were kept, so run it again.')
            rap.modul_durumu[ad] = dict(durum=u'KESILDI')
            rap.atla(ad, u'KESINTI', u'modul bastan sona kosmali',
                     u'kullanici kesti (Ctrl+C)', u'-')
            sureler[ad] = time.time() - t0
            break
        except Exception as e:
            # ERRORS ARE NOT MASKED: if a module crashes that is an ATLANDI finding and it
            # shows in the exit code.
            iz = traceback.format_exc()
            yaz(u'MODUL COKTU: %s: %s' % (type(e).__name__, e))
            rap.modul_durumu[ad] = dict(durum=u'COKTU')
            rap.atla(ad, u'MODUL-COKTU', u'the module must run without an error',
                     u'%s: %s | %s' % (type(e).__name__, e,
                                       iz.strip().splitlines()[-1]), u'-')
        sureler[ad] = time.time() - t0
        yaz(u'--- %s done: %s, findings C%d/S%d/W%d/I%d/K%d' % (
            ad, sure_metni(sureler[ad]), rap.say(KRITIK, ad), rap.say(CIDDI, ad),
            rap.say(UYARI, ad), rap.say(BILGI, ad), rap.say(ATLANDI, ad)))

    rap.olcum[u'The total run time'] = sure_metni(time.time() - t_hepsi)
    rap.olcum[u'Kontrol noktasi'] = u'%d isabet, %d iska' % (kn.isabet, kn.iska)
    for no in MODUL_ADLARI:
        if MODUL_ADLARI[no] not in kosulan:
            rap.atla(MODUL_ADLARI[no], u'MODUL-KOSULMADI',
                     u'every module must run',
                     u'it was left out by the --modules selection', u'-')

    md, tsv, kimlik = raporla(kay, rap, cikti, kosulan, sureler)
    kod = rap.cikis_kodu()
    yaz(u'')
    yaz(u'================= SUMMARY =================')
    for c in (KRITIK, CIDDI, UYARI, BILGI, ATLANDI):
        yaz(u'  %-8s : %d' % (c, rap.say(c)))
    yaz(u'Report         : %s' % md)
    yaz(u'Machine TSV    : %s' % tsv)
    if kimlik:
        yaz(u'Identity TSV   : %s' % kimlik)
    yaz(u'Exit code      : %d  (1=CRITICAL, 2=SERIOUS, 4=SKIPPED present; these are summed)' % kod)
    return kod


if __name__ == '__main__':
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        yaz(u'INTERRUPTED.')
        sys.exit(8)
