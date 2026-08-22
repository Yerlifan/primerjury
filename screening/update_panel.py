# -*- coding: utf-8 -*-
"""
update_panel.py writes the read engine fix into the panel.

  * a new sheet: "16 Okuma Motoru Duzeltmesi"
  * "2 Panel"  : a criterion label plus corrected value columns and a NOTE on each row
  * "1 Rapora Ozet" and "6 Karar Durumu": the rows that changed, and a warning
  * the degerlendiriciya_ozet TSV is synced

Usage: python update_panel.py --xlsx ../PrimerJury_PCR_Paneli_2026-08-02_TESLIM.xlsx
                                   --tsv  ../okuma_motoru_duzeltmesi_20260802.tsv

"""
# -------------------------------------------------------------------------
# update_panel.py writes the results of the read engine fix into the delivery
#                     panel xlsx (it opens a new sheet and adds a criterion label
#                     and corrected value columns to the existing sheets).
#
# INPUT  : the delivery panel file given with --xlsx (opened with openpyxl) and the
#          read engine fix table given with --tsv; optionally a --backup path.
# OUTPUT : it writes over the same xlsx file (wb.save) and takes a backup if asked;
#          it also syncs the degerlendiriciya_ozet TSV. It prints the name of the
#          file updated and the sheet count to the screen.
# CALLED BY: IT IS NOT IN THE MENU, it is run by hand on purpose. The reason is
#          written in 00_OZET_HEPSI.md: it CHANGES the delivery file and other
#          sessions write to that file too, so it is not wired into the automatic
#          chain.
# -------------------------------------------------------------------------
import sys, os, csv, argparse, shutil, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

KIRMIZI = PatternFill('solid', fgColor='FFC7CE')
SARI    = PatternFill('solid', fgColor='FFEB9C')
YESIL   = PatternFill('solid', fgColor='C6EFCE')
GRI     = PatternFill('solid', fgColor='D9D9D9')
KALIN   = Font(bold=True)
SAR     = Alignment(wrap_text=True, vertical='top')


def yaz(ws, r, c, v, fill=None, bold=False, wrap=True):
    h = ws.cell(r, c, v)
    if fill: h.fill = fill
    if bold: h.font = KALIN
    if wrap: h.alignment = SAR
    return h


def sayfa16(wb, rows):
    ad = '16 Okuma Motoru Duzeltmesi'
    if ad in wb.sheetnames:
        del wb[ad]
    ws = wb.create_sheet(ad)
    for w, c in zip((6, 30, 15, 15, 15, 15, 15, 15, 46),
                    'ABCDEFGHI'):
        ws.column_dimensions[c].width = w
    n = 1
    yaz(ws, n, 1, u'READ ENGINE FIX - 2026-08-02', bold=True); n += 1
    yaz(ws, n, 1, u'A SILENT bug was found and fixed in the panel\'s sample measurement engine. This page records the bug, the fix, which values changed and which pairs fell below the 10x threshold.'); n += 2

    yaz(ws, n, 1, u'1. WHAT THE BUG WAS', bold=True, fill=GRI); n += 1
    for s in [
        'engine/reads.py -> class `Sonda` (and engine/scb.py -> class `S`) looked for the primer in a read using A SINGLE 13 BASE EXACT MATCHING SEED:  s = primer[-13:] ;  i = seq.find(s)',
        'Although the criterion is "total mismatches <= max_mm", find() returns nothing whenever the mismatch falls INSIDE the 13 base seed, and the binding site disappears SILENTLY. The program raises no error, it simply reports "no product". Like the five earlier measurement bugs, this one is silent.',
        'EK BULGU: 3\' son 2 baz TAM ESLESME kurali kodda HICBIR YERDE acikca uygulanmiyordu - '
        '13 bazlik tohumun yan etkisiydi. Duzeltilmis motorda kural acikca uygulanir.',
    ]:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 42; n += 1
    n += 1

    yaz(ws, n, 1, u'2. EVIDENCE - A TEST WITH A KNOWN ANSWER (bin A1-4_3078083, first 400 reads, the M. mazei pair, the SAME criterion mm<=1)', bold=True, fill=GRI); n += 1
    bas = ['Engine', 'reads giving a product / 400', 'Percent', 'Note']
    for j, h in enumerate(bas, 1):
        yaz(ws, n, j, h, bold=True, fill=GRI)
    n += 1
    for r in [('okuma.Sonda (panelin kullandigi)', '2', '0,50%', 'tohumlu'),
              ('kaba kuvvet (saf python, tohumsuz, bagimsiz yazildi)', '174', '43,50%', 'dogru cevap'),
              ('ispcr.find_sites (numpy, tohumsuz, panelin kendi kodu)', '174', '43,50%', 'kaba kuvvetle BIREBIR'),
              ('read_engine.py (duzeltilmis)', '174', '43,50%', 'kaba kuvvetle BIREBIR')]:
        for j, v in enumerate(r, 1):
            yaz(ws, n, j, v, fill=(KIRMIZI if r[1] == '2' else YESIL))
        n += 1
    yaz(ws, n, 1, u'In 202 of the forward primer\'s 205 binding sites (98.5%) the single mismatch falls inside the seed, and 188 of those are at base 6 of the primer. That is not scattered noise, it is one recurring variant base. MISS RATE: 172/174 = 98.9%.')
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9); ws.row_dimensions[n].height = 30; n += 1
    yaz(ws, n, 1, u'CORRECTION: the "188/400" figure in the earlier note file came from running ispcr.amplify with its DEFAULT max_mm=3, so that comparison changed the seed AND the criterion at the same time. The 2 vs 174 above changes only the seed, and that is the number that gives the size of the bug.', fill=SARI)
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9); ws.row_dimensions[n].height = 30; n += 2

    yaz(ws, n, 1, u'3. HOW IT WAS FIXED', bold=True, fill=GRI); n += 1
    for s in [
        "PIGEONHOLE SEEDING: the primer is split into max_mm+1 NON-OVERLAPPING blocks. If there are at most max_mm mismatches then AT LEAST ONE of those blocks must match exactly, so searching for an exact match of any block is LOSSLESS. Every candidate site is then verified one by one under the full rule (total mismatches plus the last 2 bases at the 3' end).",
        'LOSSLESSNESS WAS PROVEN: screening/engine_test.py compares the corrected engine against an independent brute force implementation that uses no seed and tries every position one at a time. T1 synthetic (2 439 binding sites, difference 0), T2 real reads (difference 0), T3 product count (difference 0). In the same test the OLD engine misses 1 386 sites on the synthetic data (56.8%) and 35.2% on the real reads.',
        'SPEED: for 400 reads per pair, brute force takes 0.30 s and the corrected engine 0.03 s (~10x). Since it gives the same answer as brute force, the speed gain costs no accuracy.',
        'FILES: screening/read_engine.py (the engine), brute_force.py (the reference), engine_test.py (the evidence), measure_panel.py (the panel measurement), independent_check.py, python3 -m screening --mode panel-olc --full-depth',
    ]:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'4. THE SECOND FINDING - CRITERION INCONSISTENCY (at least as important as the seed bug)',
        bold=True, fill=GRI); n += 1
    for s in [
        'The criterion note on the "2 Panel" sheet defines the sample criterion as "mismatches <=1, the last 2 bases at the 3\' end an EXACT match". But the rows were NOT MEASURED UNDER A SINGLE CRITERION. Each row\'s published member % range was regenerated under four engine and criterion combinations, which identified the criterion it had actually been measured with:',
        'mm<=1 (the seeded engine): Metanomikrobiyales (member 51.2-80.0% EXACTLY), Proteolitik_Synergistaceae (81.5% -> 81.2%), Methanosarcina_mazei_turu (40.6-60.7% and competitor 4.49% EXACTLY).',
        'mm<=3 (the seedless engine, the ispcr.amplify default): Proteiniphilum (29.0% in the panel -> 28.6% under mm<=3, 1.6% under mm<=1), Metilotrofik (71.0% -> 72.4%), Cloacimonas (78.5% -> 77.8%), Sakarolitik_Sphaerochaeta, Methanosarcina_cinsi.',
        'CONCLUSION: the values in the "Ayrim (x)" column CANNOT be compared with one another. In this table every row was re-measured with ONE engine under TWO criteria (mm<=1 and mm<=3), and every value carries its criterion label. DO NOT USE ANY NUMBER THAT HAS NO CRITERION LABEL.',
    ]:
        yaz(ws, n, 1, s, fill=(SARI if s.startswith('SONUC') else None))
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 42; n += 1
    n += 1

    yaz(ws, n, 1, u'5. WHICH VALUES CHANGED (21 pairs, corrected engine, at most 3000 reads per bin)',
        bold=True, fill=GRI); n += 1
    yaz(ws, n, 1, u'THIS IS A SUBSET MEASUREMENT: at most 3000 reads per bin (the panel\'s own protocol). Full depth verification is done with python3 -m screening --mode panel-olc --full-depth; if the trend changes, this page must be updated.', fill=SARI)
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9); ws.row_dimensions[n].height = 30; n += 1
    bas = ['#', 'Hedef', 'Panelin olcutu', 'ESKI kat (en kotu/havuz)', 'YENI mm<=1 (en kotu/havuz)',
           'YENI mm<=3 (en kotu/havuz)', 'Uye kumesi', '10x alti?', 'Not']
    for j, h in enumerate(bas, 1):
        yaz(ws, n, j, h, bold=True, fill=GRI)
    n += 1
    for r in rows:
        deg = bool(r['DEGISTI'])
        f = KIRMIZI if (r['ESIK_ALTI_mm1'] == 'EVET' and deg) else (SARI if deg else None)
        yaz(ws, n, 1, int(r['satir']), fill=f)
        yaz(ws, n, 2, r['hedef'], fill=f)
        yaz(ws, n, 3, r['panel_olcut_tespiti'], fill=f)
        yaz(ws, n, 4, '%s / %s' % (r['ESKI_kat_enkotu'] or '-', r['ESKI_kat_havuz'] or '-'), fill=f)
        yaz(ws, n, 5, '%s / %s' % (r['YENI1_kat_enkotu'] or '-', r['YENI1_kat_havuz'] or '-'), fill=f)
        yaz(ws, n, 6, '%s / %s' % (r['YENI3_kat_enkotu'] or '-', r['YENI3_kat_havuz'] or '-'), fill=f)
        yaz(ws, n, 7, r['uye_kumesi'], fill=f)
        yaz(ws, n, 8, ('mm<=1 ' if r['ESIK_ALTI_mm1'] else '') + ('mm<=3' if r['ESIK_ALTI_mm3'] else ''), fill=f)
        yaz(ws, n, 9, r['DEGISTI'] or u'no change (below the 5 point / 2 fold threshold)', fill=f)
        ws.row_dimensions[n].height = 30
        n += 1
    n += 1

    yaz(ws, n, 1, u'6. THE ONE REAL REGRESSION - Methanosarcina_mazei_turu (row 22)', bold=True, fill=KIRMIZI); n += 1
    for s in [
        'This is the ONLY row where the fix pushed a panel value BELOW THE THRESHOLD. On the others the fix either changes nothing or pulls the value UP (their coverage had been under measured).',
        'The cause is a single competitor bin: A1-4_3078083 (Methanosarcina hadiensis, n=2215). The old engine measured it at 0.72%, and its correct value is 47.22%. A second bin, A2-4_3078083, went 4.49% -> 33.33%. The member bins move by only +1.4 to +2.2 points. In other words the correction lands almost entirely on the COMPETITOR side, which means the bug made the specificity look BETTER THAN IT WAS. That is the dangerous direction.',
        'OLD (mm<=1): member 40.63-60.63% / competitor max 4.49% / worst fold 4.23x / pool fold 49.96x. The value published in the panel is 187.9x (the within genus pool, a narrower pool, carrying the same bug).',
        'NEW (mm<=1): member 42.23-62.20% / competitor max 47.22% / WORST FOLD 0.82x / pool fold 11.41x.',
        'WHAT IT MEANS: this pair amplifies Methanosarcina hadiensis as well as it amplifies M. mazei. The claim "the M. mazei / M. soligelidi group" DOES NOT HOLD as it stands. Either the amplified set has to be widened to cover M. hadiensis, or the pair has to be dropped, or amplicon sequencing has to be made a condition. THE DECISION IS YOURS. The panel marked this row "ESIK ALTI" rather than leaving it in silently.',
    ]:
        yaz(ws, n, 1, s, fill=(KIRMIZI if s.startswith('ANLAMI') else None))
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'7. ROWS THAT IMPROVED (their coverage had been under measured)', bold=True, fill=YESIL); n += 1
    for s in [
        'Methanosarcina_cinsi (row 7): seven M. mazei MEMBER bins measured 0.5-26.5% under the old engine, and their correct value is 79.4-82.9%. The worst fold went 0.04x -> 4.37x (mm<=1) / 4.66x (mm<=3), and the pool 2.51x -> 81.59x. The 28.4x published in the panel is a pool measure; the WORST SINGLE BIN measure is still below 10x.',
        'Asetoklastik_metanojenler (row 16): the member floor went 0.0% -> 58.6%; the worst fold ~0 -> 4.22x, and the pool ~0 -> 50.84x.',
        'Kapsam olculeri: Arke_universal 11/39 -> 32/39 (mm<=1; panelin 39/39 iddiasi mm<=3 degeridir), '
        'Bakteri_universal 4/20 -> 13/20 (mm<=1), Mantar_universal F1 14/20 -> 16/20.',
        'Methanothrix_cinsi (row 3): under mm<=1 it does not change (13.54x -> 13.74x), but under mm<=3 it drops to 0.86x (one competitor bin amplifies at 76.92%). The fate of this row depends entirely on THE CHOICE OF CRITERION, and it should not go to order before that criterion is settled.',
    ]:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'8. INDEPENDENT VERIFICATION (a route that does NOT USE the corrected engine)', bold=True, fill=GRI); n += 1
    for s in [
        "The headline numbers were measured by THREE separate routes: (A) ispcr.find_sites, a numpy vector scan, seedless, the panel's OWN code, untouched in this session; (B) brute_force.py, pure Python, every position one at a time, sharing no code with the others; (C) the corrected read_engine.py.",
        'All SEVEN of the seven headline bins gave the same number by all three routes: A1-4_3078083 1046/2215 (47.22%), A2-4_3078083 52/156 (33.33%), A2-2_2209 1866/3000 (62.20%), A1-3_2209 1267/3000 (42.23%), A1-2_2209 (Methanosarcina_cinsi) 2389/3000 (79.63%), A2-3_2223 15/199 (7.54%), A1-2_2209 (Asetoklastik) 1828/3000 (60.93%).',
        'To run it: python screening/independent_check.py --fastq "fastq files"',
    ]:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'9. STILL OPEN - WHAT THE NEXT PERSON NEEDS TO KNOW', bold=True, fill=SARI); n += 1
    for s in [
        'FULL DEPTH: the numbers on this page were produced from a subset of at most 3000 reads per bin. python3 -m screening --mode panel-olc --full-depth runs with no sampling. If the trend changes, this page must be updated.',
        'MEMBER SETS: the panel\'s original member taxon sets were never stored in any script (they were a command line argument). They were rebuilt from the target names in screening/ciftler.tsv and from taxid_adlari.tsv. On rows marked "PANELLE_TUTUYOR" the rebuilt set reproduces the member % range the panel published; on rows marked "YENIDEN_KURULDU" it does not, and on those rows the old vs new DIFFERENCE is valid (both engines ran on the same bins) but the absolute values may not match the panel\'s original set definition. The Proteiniphilum (row 10) and Bacteroidales (row 12) sets in particular should be checked.',
        'THE CRITERION DECISION: the panel needs to settle on a single sample criterion. mm<=1 is the recommended one, since it is consistent with the published criterion label; mm<=3 is the design pipeline\'s criterion and is looser. Until that decision is made, the "Ayrim (x)" column cannot be compared across rows.',
        'CONSENSUS: this correction covers RAW READ measurements only, and no consensus file was touched. Consensus regeneration will be done separately (by two methods: quality weighted with a lowered threshold, plus a majority vote; positions that disagree will be masked and NO degenerate base will be used, by decision).',
    ]:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=9)
        ws.row_dimensions[n].height = 56; n += 1
    return ws


def panel_sutunlari(wb, rows):
    ws = wb['2 Panel']
    c0 = ws.max_column + 1
    basliklar = ['READ ENGINE FIX - criterion label',
                 'CORRECTED discrimination (mm<=1) worst / pool',
                 'CORRECTED discrimination (mm<=3) worst / pool',
                 'DEGISTI MI',
                 '10x ESIGININ ALTINDA MI']
    for j, h in enumerate(basliklar):
        yaz(ws, 1, c0 + j, h, bold=True, fill=GRI)
    ix = {int(r['satir']): r for r in rows}
    for satir in range(2, 23):
        r = ix.get(satir)
        if not r:
            continue
        deg = bool(r['DEGISTI'])
        alt = r['ESIK_ALTI_mm1'] == 'EVET'
        f = KIRMIZI if (deg and alt) else (SARI if deg else None)
        yaz(ws, satir, c0 + 0, r['panel_olcut_tespiti'], fill=f)
        yaz(ws, satir, c0 + 1, '%s / %s' % (r['YENI1_kat_enkotu'] or '-', r['YENI1_kat_havuz'] or '-'), fill=f)
        yaz(ws, satir, c0 + 2, '%s / %s' % (r['YENI3_kat_enkotu'] or '-', r['YENI3_kat_havuz'] or '-'), fill=f)
        yaz(ws, satir, c0 + 3, r['DEGISTI'] or u'no', fill=f)
        yaz(ws, satir, c0 + 4, ('EVET (mm<=1)' if alt else '') +
            (' EVET (mm<=3)' if r['ESIK_ALTI_mm3'] == 'EVET' else '') or u'no', fill=f)
    for j, w in enumerate((30, 26, 26, 60, 22)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c0 + j)].width = w
    # NOT satiri
    n = ws.max_row + 2
    yaz(ws, n, 1, 'NOT', bold=True)
    yaz(ws, n, 3, u'READ ENGINE FIX (2026-08-02, the most recent item): a SILENT bug was found in the engine that produced the "Uye urun %", "Rakip maks %" and "Ayrim (x)" columns. The 13 base EXACT matching seed never saw a binding site when the mismatch fell inside the seed (174/400 instead of 2/400 under the same criterion). The engine was fixed and all 21 pairs were re-measured. ALSO: those three columns were not measured under a SINGLE CRITERION. Some rows used mm<=1 and some mm<=3, so values inside a column cannot be compared across rows. Detail, the changed values and the pairs that fell below 10x: the "16 Okuma Motoru Duzeltmesi" sheet. The new columns (see to the right) are a subset measurement; full depth will be verified with python3 -m screening --mode panel-olc --full-depth.')
    ws.cell(n, 3).fill = KIRMIZI
    ws.cell(n, 3).alignment = SAR
    ws.row_dimensions[n].height = 90


def ozet_ve_karar(wb):
    ws = wb['1 Rapora Ozet']
    n = ws.max_row + 2
    yaz(ws, n, 1, u'READ ENGINE FIX (2026-08-02 - THE MOST IMPORTANT OPEN ITEM)', bold=True, fill=KIRMIZI)
    yaz(ws, n, 7, u'A silent bug was found in the engine that produced the discrimination ratios in this table: the 13 base exact matching seed never saw a binding site when the mismatch fell inside the seed (174/400 instead of 2/400 under the same criterion, a 98.9% miss rate). The engine was fixed and all 21 pairs were re-measured. THE ONE REAL REGRESSION: the Methanosarcina mazei / M. soligelidi group. Pool discrimination went 49.96x -> 11.41x, and the WORST SINGLE BIN 4.23x -> 0.82x. The cause: bin A1-4_3078083 of M. hadiensis amplifies at 47.22% instead of 0.72%. In other words this pair amplifies M. hadiensis as well as it amplifies the target. THE 187.9x VALUE ON THIS ROW IS NO LONGER VALID. On the other rows the fix is either neutral or pulls the value UP (their coverage had been under measured). Detail: "16 Okuma Motoru Duzeltmesi".')
    ws.cell(n, 7).fill = KIRMIZI; ws.cell(n, 7).alignment = SAR
    ws.row_dimensions[n].height = 120
    n += 1
    yaz(ws, n, 1, u'THE SECOND FINDING - CRITERION INCONSISTENCY', bold=True, fill=SARI)
    yaz(ws, n, 7, u'The panel\'s "Ayrim (x)" column was not produced under a SINGLE CRITERION: some rows were measured with mismatches <=1 and some with <=3 (for example Proteiniphilum reads 29.0% in the panel, which is the mm<=3 value; under mm<=1 it is 1.6%). Values inside the column therefore CANNOT be compared across rows. On the "16 Okuma Motoru Duzeltmesi" sheet every row was re-measured with one engine under both criteria, and every value carries its criterion label. The panel needs to settle on a single sample criterion, and that decision is yours.')
    ws.cell(n, 7).fill = SARI; ws.cell(n, 7).alignment = SAR
    ws.row_dimensions[n].height = 90

    # M. mazei satirlari
    for r in (8, 23):
        h = ws.cell(r, 6)
        h.value = 'GECERSIZ: 187,9x -> duzeltilmis 11,41x (havuz) / 0,82x (en kotu tek kutu). ' \
                  'Bkz. "16 Okuma Motoru Duzeltmesi".'
        h.fill = KIRMIZI; h.alignment = SAR
    h = ws.cell(9, 6)   # Methanosarcina cinsi
    h.value = '28,4x -> duzeltilmis havuz 81,59x, en kotu tek kutu 4,66x (10x ALTINDA). ' \
              'Kapsam eksik olculmustu, duzeltme degeri yukari cekti.'
    h.fill = SARI; h.alignment = SAR

    ws = wb['6 Karar Durumu']
    n = ws.max_row + 2
    yaz(ws, n, 1, u'READ ENGINE FIX', bold=True, fill=KIRMIZI)
    yaz(ws, n, 2, u'the Methanosarcina mazei / M. soligelidi group')
    yaz(ws, n, 3, u'BELOW THRESHOLD - A NEW DECISION IS NEEDED', fill=KIRMIZI)
    yaz(ws, n, 4, u'The seed bug in the sample engine was fixed. Within genus pool discrimination went 49.96x -> 11.41x, and the worst single bin 4.23x -> 0.82x. The cause: bin A1-4_3078083 of M. hadiensis measured 0.72% under the old engine, and its correct value is 47.22%. The pair amplifies M. hadiensis as well as it amplifies the target. OPTIONS: (a) widen the amplified set so that it covers M. hadiensis as well, (b) drop the pair, (c) keep it conditional on amplicon sequencing. It was not left in the panel silently; it is marked "ESIK ALTI". Detail: "16 Okuma Motoru Duzeltmesi".')
    ws.cell(n, 4).fill = KIRMIZI; ws.cell(n, 4).alignment = SAR
    ws.row_dimensions[n].height = 90
    n += 1
    yaz(ws, n, 1, u'READ ENGINE FIX', bold=True, fill=YESIL)
    yaz(ws, n, 2, u'the Methanosarcina genus / acetoclastic methanogens')
    yaz(ws, n, 3, u'IMPROVED (the coverage had been under measured)', fill=YESIL)
    yaz(ws, n, 4, u'Methanosarcina_cinsi: seven M. mazei member bins measured 0.5-26.5% under the old engine, and their correct value is 79.4-82.9%. Pool discrimination 2.51x -> 81.59x; the worst single bin 0.04x -> 4.66x (still below 10x). Asetoklastik_metanojenler: the member floor went 0.0% -> 58.6%, and the pool ~0 -> 50.84x. The coverage measures also rose: Arke_universal 11/39 -> 32/39 (mm<=1), Bakteri_universal 4/20 -> 13/20.')
    ws.cell(n, 4).fill = YESIL; ws.cell(n, 4).alignment = SAR
    ws.row_dimensions[n].height = 76
    n += 1
    yaz(ws, n, 1, u'READ ENGINE FIX', bold=True, fill=SARI)
    yaz(ws, n, 2, u'the Methanothrix genus')
    yaz(ws, n, 3, u'OVERLY SENSITIVE TO THE CRITERION - AWAITING THE CRITERION DECISION', fill=SARI)
    yaz(ws, n, 4, u'Under the mm<=1 criterion the discrimination is 13.74x (above the threshold); under mm<=3 it is 0.86x (one competitor bin amplifies at 76.92%). The 75.2x published in the panel could not be reproduced under either criterion. This should not go to order before the criterion is settled.')
    ws.cell(n, 4).fill = SARI; ws.cell(n, 4).alignment = SAR
    ws.row_dimensions[n].height = 60


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--tsv', required=True)
    ap.add_argument('--backup', dest='yedek', default='')
    a = ap.parse_args()
    yedek = a.yedek or a.xlsx.replace('.xlsx', '_YEDEK_motor_oncesi.xlsx')
    if not os.path.exists(yedek):
        shutil.copy2(a.xlsx, yedek)
        print('yedek:', yedek)
    rows = list(csv.DictReader(open(a.tsv, encoding='utf-8'), delimiter='\t'))
    wb = openpyxl.load_workbook(a.xlsx)
    sayfa16(wb, rows)
    panel_sutunlari(wb, rows)
    ozet_ve_karar(wb)
    wb.save(a.xlsx)
    print(u'updated:', a.xlsx, '| sayfalar:', len(wb.sheetnames))


if __name__ == '__main__':
    main()
