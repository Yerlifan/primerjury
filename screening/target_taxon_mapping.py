# -*- coding: utf-8 -*-
"""target_taxon_mapping.py produces two tables:
  Table A  the 21 targets in the panel -> which meeting decision each comes from
  Table B  the 44 taxa in the sample -> which target covers each (measured)

Output: hedef_takson_eslemesi_20260802.md plus "18 Hedef-Takson Eslemesi" in the
panel

"""
# -------------------------------------------------------------------------
# target_taxon_mapping.py ties two questions to a table: which meeting decision
#                          each target in the panel comes from, and which target
#                          covers each taxon in the sample.
#
# INPUT  : the measurement result pattern given with --r2, the json cross_coverage.py
#          produces given with --cross, the pair table with --pairs,
#          taxid_names.tsv with --taxid and the delivery panel with --xlsx. The
#          target to decision mapping (the KARAR dictionary) is fixed inside the
#          file; the rows that do not come from a meeting decision are marked there
#          plainly as "Karar 5 - derived from the measurement".
# OUTPUT : the Markdown file given with --md and the sheet
#          "18 Hedef-Takson Eslemesi" added to the --xlsx file (wb.save).
# CALLED BY: IT IS NOT IN THE MENU, it is run by hand; it changes the delivery
#          xlsx. cross_coverage.py, one of its inputs, is run by hand as well.
#
# The value of Table B is that it makes visible the taxa no specific target covers:
# if a taxon amplifies only with a universal or control primer, there is no specific
# measurement in the panel for that organism.
# -------------------------------------------------------------------------
import sys, os, json, glob, csv, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

KIRMIZI = PatternFill('solid', fgColor='FFC7CE')
SARI    = PatternFill('solid', fgColor='FFEB9C')
GRI     = PatternFill('solid', fgColor='D9D9D9')
YESIL   = PatternFill('solid', fgColor='C6EFCE')
KALIN   = Font(bold=True)
SAR     = Alignment(wrap_text=True, vertical='top')

# the panel row -> (target name, level, decision, an explanation of the source)
KARAR = {
 2:  ('Metanojen_universal',              'grup (islevsel)', 'Karar 4 - alan/evrensel',
      'Toplanti: "universal metanojen" kontrol primeri.'),
 3:  ('Methanothrix_cinsi',               'cins',            'Karar 5 - olcumden turetilen',
      'Toplanti kararlarinda YOK. Bu oturumda tasarlandi; M. soehngenii tur iddiasi dusurulunce cins duzeyi kondu.'),
 4:  ('Petrimonas_cinsi',                 'cins',            'Karar 2 - dort cins ozgul',
      'Toplanti: istenen dort cinsten biri.'),
 5:  ('Metanomikrobiyales_hidrojenotrof', 'takim',           'Karar 3 - islev/ekolojik grup',
      'Toplanti: "hidrojenotrofik metanojenler". Adi olculen kapsama gore Metanomikrobiyales takimina daraltildi.'),
 6:  ('Mantar_universal (F1)',            'alan',            'Karar 4 - alan/evrensel',
      'Toplanti: mantar evrensel kontrol primeri (F1 barkod sinifi).'),
 7:  ('Methanosarcina_cinsi',             'cins',            'Karar 5 - olcumden turetilen',
      'Toplanti kararlarinda YOK dogrudan. Karar 1\'in "M. barkeri tur ozgul" hedefinin YERINE GECER: '
      'tur numunede yok (2208 kutusu M. vacuolata\'ya %97,4-97,9), cins duzeyi verildi.'),
 8:  ('Metilotrofik_metanojen',           'grup (islevsel)', 'Karar 3 - islev/ekolojik grup',
      'Toplanti: metilotrofik metanojenler.'),
 9:  ('Nitrosocosmicus_AOA',              'grup (ekolojik)', 'Karar 3 - islev/ekolojik grup',
      'Toplanti: amonyak oksitleyen arke (AOA). Sonradan eklenmedi.'),
 10: ('Proteiniphilum_cinsi',             'cins',            'Karar 2 - dort cins ozgul',
      'Toplanti: istenen dort cinsten biri.'),
 11: ('Proteolitik_Synergistaceae',       'soy',             'Karar 5 - olcumden turetilen (Karar 3 altinda)',
      'Toplanti "proteolitik/sintrofik bakteriler" dedi; olculen soy Synergistaceae cikti. Eski ad: Proteolitik_Cloacibacillus.'),
 12: ('Bacteroidales_kumesi',             'soy',             'Karar 5 - olcumden turetilen',
      'Karar 2\'nin "Bacteroides" ve "Alistipes" hedeflerinin YERINE GECER: iki cins de numunede yok, '
      'kutular birbirine %95,3-96,8 benziyor, adlandirilamayan tek bir Bacteroidales soyu.'),
 13: ('Mantar_universal (F2)',            'alan',            'Karar 4 - alan/evrensel',
      'Toplanti: mantar evrensel kontrol primeri (F2 barkod sinifi).'),
 14: ('Microascaceae_askomikot',          'aile',            'Karar 5 - olcumden turetilen',
      'AILE DUZEYI HEDEF. Karar 1\'in "Trichoderma asperellum tur ozgul" hedefinin YERINE GECER: '
      '101201 kutusundaki organizma Trichoderma degil, ITS\'te Petriella/Microascaceae. PANELDEN CIKARILDI (ayrim 0,7x).'),
 15: ('Sakarolitik_Sphaerochaeta',        'cins',            'Karar 5 - olcumden turetilen (Karar 3 altinda)',
      'Toplanti "sakarolitik bakteriler" dedi; olculen uye Sphaerochaeta associata cikti.'),
 16: ('Asetoklastik_metanojenler',        'grup (islevsel)', 'Karar 3 - islev/ekolojik grup',
      'Toplanti: asetoklastik metanojenler.'),
 17: ('Bakteri_universal',                'alan',            'Karar 4 - alan/evrensel',
      'Toplanti: bakteri evrensel kontrol primeri.'),
 18: ('Petriella_musispora',              'cins -> sinif',   'Karar 5 - olcumden turetilen',
      'Karar 1\'in "Podospora pseudopauciseta tur ozgul" hedefinin yerine olcumden cikti. PANELDEN CIKARILDI (ayrim 0,7x).'),
 19: ('Proteolitik_Cloacimonas',          'cins',            'Karar 5 - olcumden turetilen (Karar 3 altinda)',
      'Toplanti "proteolitik/sintrofik bakteriler" dedi; olculen uye Ca. Cloacimonas acidaminovorans cikti.'),
 20: ('Arke_universal',                   'alan',            'Karar 4 - alan/evrensel',
      'Toplanti: arke evrensel kontrol primeri.'),
 21: ('Methanothrix_soehngenii_turu',     'TUR (kosullu)',   'Karar 1 - alti tur ozgul',
      'Toplanti: istenen alti turden biri. TUR DUZEYI VERILDI (kosullu).'),
 22: ('Methanosarcina_mazei_turu',        'TUR GRUBU',       'Karar 1 - alti tur ozgul',
      'Toplanti: istenen alti turden biri. TUR GRUBU verildi (M. mazei / M. soligelidi ayrilamiyor). '
      'OKUMA MOTORU DUZELTMESI SONRASI ESIK ALTI - bkz. "16 Okuma Motoru Duzeltmesi".'),
}

# requests asked for under Decisions 1 and 2 that COULD NOT ENTER THE PANEL AS A TARGET
KARSILANMAYAN = [
 ('Karar 1 - tur ozgul', 'Methanosarcina barkeri',
  'Organizma numunede yok (2208 kutusu M. vacuolata %97,4-97,9). Yerine satir 7 Methanosarcina_cinsi.'),
 ('Karar 1 - tur ozgul', 'Podospora pseudopauciseta',
  'Organizma numunede yok. Yerine satir 18 Petriella_musispora (o da panelden cikarildi).'),
 ('Karar 1 - tur ozgul', 'Dictyostelium discoideum (44689)',
  'Kraken2 etiketi curutuldu; kutu heterojen. HICBIR PANEL HEDEFI YOK - yalniz Mantar_universal F1 cogaltiyor.'),
 ('Karar 1 - tur ozgul', 'Trichoderma asperellum (101201)',
  'Kutudaki organizma Trichoderma degil (ITS: Petriella/Microascaceae). Yerine satir 14, o da cikarildi.'),
 ('Karar 2 - cins ozgul', 'Bacteroides',
  'Cins numunede yok (en iyi eslesme Alistipes putredinis %85,5). Yerine satir 12 Bacteroidales_kumesi.'),
 ('Karar 2 - cins ozgul', 'Alistipes',
  'Bacteroides ile ayni organizma; satir 12 altinda birlestirildi.'),
]

GENIS_AD = {2: 'Metanojen_universal', 6: 'Mantar_universal (F1)', 13: 'Mantar_universal (F2)',
            17: 'Bakteri_universal', 20: 'Arke_universal'}


def yukle(r2_desen, capraz, ciftler_tsv, taxid_tsv):
    R = {}
    for f in glob.glob(r2_desen):
        R.update(json.load(open(f)))
    CP = json.load(open(capraz))
    ad = {}
    for l in open(taxid_tsv, encoding='utf-8'):
        p = l.rstrip('\n').split('\t')
        if len(p) >= 2:
            ad[p[0]] = p[1]
    uye = {}
    for c in csv.DictReader(open(ciftler_tsv, encoding='utf-8'), delimiter='\t'):
        uye[int(c['satir'])] = set(x for x in c['uye_taksonlar'].split(',') if x)
    return R, CP, ad, uye


def takson_tablosu(R, CP, ad, uye):
    tax = {}
    for kaynak, veri in (('sinif', R), ('capraz', CP)):
        for k, v in veri.items():
            s, kutu = k.split('|')
            s = int(s)
            tx = kutu.split('_')[1]
            if kaynak == 'sinif':
                y1, n = v[1], v[4]
                y3 = v[3]
            else:
                y1, y3, n = v[0], v[1], v[2]
            if not n:
                continue
            d = tax.setdefault(tx, {})
            a = d.get(s, (0.0, 0.0))
            d[s] = (max(a[0], 100.0 * y1 / n), max(a[1], 100.0 * y3 / n))
    satirlar = []
    for tx in sorted(tax, key=lambda x: ad.get(x, 'zz')):
        d = tax[tx]
        u = sorted(KARAR[s][0] for s in d if tx in uye.get(s, set()))
        c1 = sorted(KARAR[s][0] for s in d if d[s][0] >= 10)
        c3 = sorted(KARAR[s][0] for s in d if d[s][1] >= 10 and d[s][0] < 10)
        ozgul1 = [x for x in c1 if x not in GENIS_AD.values()]
        if c1:
            durum = 'kapsanan (ozgul hedef)' if ozgul1 else 'yalniz evrensel/kontrol'
        elif c3:
            durum = 'yalniz mm<=3 ile'
        else:
            durum = 'BOSLUK - hicbir hedef cogaltmiyor'
        satirlar.append(dict(taxid=tx, ad=ad.get(tx, '?'),
                             uye='; '.join(u) or '-',
                             cog1='; '.join(c1) or '-',
                             cog3='; '.join(c3) or '-', durum=durum))
    return satirlar


def md_yaz(yol, satirlar):
    L = []
    L.append(u'# The target to taxon mapping, 2 August 2026\n')
    L.append(u'The 21 targets in the panel are not the 44 taxa in the sample themselves. The targets are derived from the meeting decisions and from the measurement.\n')
    L.append(u'\n## Table A, the 21 targets in the panel by their source\n')
    L.append(u'| # | Target | Level | Source | Explanation |')
    L.append('|---|---|---|---|---|')
    for s in sorted(KARAR):
        h, dz, kr, ac = KARAR[s]
        L.append('| %d | %s | %s | %s | %s |' % (s, h, dz, kr, ac))
    L.append(u'\n### The count per decision\n')
    say = {}
    for s in KARAR:
        k = KARAR[s][2].split(' (')[0]
        say[k] = say.get(k, 0) + 1
    L.append(u'| Source | Targets | Which ones |')
    L.append('|---|---|---|')
    for k in sorted(say):
        h = [KARAR[s][0] for s in sorted(KARAR) if KARAR[s][2].split(' (')[0] == k]
        L.append('| %s | %d | %s |' % (k, say[k], ', '.join(h)))
    L.append(u'\n### Asked for in the meeting decisions but not admitted to the panel as a target\n')
    L.append(u'| Decision | Asked for | What happened |')
    L.append('|---|---|---|')
    for a, b, c in KARSILANMAYAN:
        L.append('| %s | %s | %s |' % (a, b, c))
    L.append(u'\n### The family level target\n')
    L.append(u'The one family level target: **Microascaceae_askomikot** (row 14, Decision 5). It was removed from the panel (discrimination 0,7x).\n')
    L.append(u'\n## Table B, the 44 taxa in the sample and which target covers them\n')
    L.append(u'The measurement: the corrected read engine, at most 3000 reads per bin, a threshold of >=%10 product. The five universal and broad targets were measured in all 99 bins with no class boundary.\n')
    L.append(u'| taxid | Taxon | The target(s) it is a member of | The targets that amplify it (mm<=1) | Extra (mm<=3 only) | State |')
    L.append('|---|---|---|---|---|---|')
    for r in satirlar:
        L.append('| %s | %s | %s | %s | %s | %s |' % (r['taxid'], r['ad'], r['uye'],
                                                      r['cog1'], r['cog3'], r['durum']))
    bos = [r for r in satirlar if r['durum'].startswith('BOSLUK')]
    yev = [r for r in satirlar if r['durum'] == 'yalniz evrensel/kontrol']
    m3 = [r for r in satirlar if r['durum'] == 'yalniz mm<=3 ile']
    L.append(u'\n### The gaps\n')
    L.append(u'| State | Count | Taxa |')
    L.append('|---|---|---|')
    L.append(u'| No target amplifies it | %d | %s |' % (len(bos), ', '.join(r['ad'] for r in bos) or '-'))
    L.append(u'| It amplifies only under the mm<=3 criterion | %d | %s |' % (len(m3), ', '.join(r['ad'] for r in m3) or '-'))
    L.append(u'| Only a universal or control primer amplifies it, it has no specific target | %d | %s |'
             % (len(yev), ', '.join(r['ad'] for r in yev) or '-'))
    L.append(u'| It is covered by a specific target | %d | - |'
             % len([r for r in satirlar if r['durum'].startswith('kapsanan')]))
    L.append(u'\nTaxa in total: **%d**\n' % len(satirlar))
    open(yol, 'w', encoding='utf-8').write('\n'.join(L) + '\n')


def xlsx_yaz(xlsx, satirlar):
    wb = openpyxl.load_workbook(xlsx)
    adx = '18 Hedef-Takson Eslemesi'
    if adx in wb.sheetnames:
        del wb[adx]
    ws = wb.create_sheet(adx)
    for w, c in zip((8, 34, 20, 30, 44, 44, 26), 'ABCDEFG'):
        ws.column_dimensions[c].width = w
    n = 1

    def yaz(r, c, v, fill=None, bold=False):
        h = ws.cell(r, c, v)
        if fill: h.fill = fill
        if bold: h.font = KALIN
        h.alignment = SAR
        return h

    yaz(n, 1, u'TARGET TO TAXON MAPPING', bold=True); n += 1
    yaz(n, 1, u'The 21 targets in the panel are NOT the same thing as the 44 taxa in the sample. The targets are derived from the meeting decisions and from measurement.'); n += 2

    yaz(n, 1, u'TABLE A - the 21 targets in the panel, by where each came from', bold=True, fill=GRI); n += 1
    for j, h in enumerate(['#', 'Hedef', 'Duzey', 'Kaynak', 'Aciklama'], 1):
        yaz(n, j, h, bold=True, fill=GRI)
    n += 1
    for s in sorted(KARAR):
        h, dz, kr, ac = KARAR[s]
        f = SARI if kr.startswith('Karar 5') else None
        yaz(n, 1, s, fill=f); yaz(n, 2, h, fill=f); yaz(n, 3, dz, fill=f)
        yaz(n, 4, kr, fill=f); yaz(n, 5, ac, fill=f)
        ws.row_dimensions[n].height = 30
        n += 1
    n += 1

    yaz(n, 1, u'THE COUNT PER DECISION', bold=True, fill=GRI); n += 1
    say = {}
    for s in KARAR:
        k = KARAR[s][2].split(' (')[0]
        say.setdefault(k, []).append(KARAR[s][0])
    for k in sorted(say):
        yaz(n, 1, len(say[k]), bold=True); yaz(n, 2, k, bold=True)
        yaz(n, 4, ', '.join(say[k])); ws.merge_cells(start_row=n, start_column=4, end_row=n, end_column=5)
        ws.row_dimensions[n].height = 30
        n += 1
    n += 1

    yaz(n, 1, u'REQUESTED IN THE DECISIONS BUT COULD NOT ENTER THE PANEL AS A TARGET', bold=True, fill=KIRMIZI); n += 1
    for j, h in enumerate(['Karar', 'Istenen', 'Ne oldu'], 1):
        yaz(n, j + 1, h, bold=True, fill=GRI)
    n += 1
    for a, b, c in KARSILANMAYAN:
        yaz(n, 2, a); yaz(n, 3, b); yaz(n, 4, c)
        ws.merge_cells(start_row=n, start_column=4, end_row=n, end_column=5)
        ws.row_dimensions[n].height = 30
        n += 1
    n += 1

    yaz(n, 1, u'TABLE B - the 44 taxa in the sample, and which target covers each', bold=True, fill=GRI); n += 1
    yaz(n, 1, u'Measurement: the corrected read engine, <=3000 reads per bin, threshold >=%10 product. The five universal or broad targets were measured across all 99 bins WITH NO CLASS BOUNDARY (the panel\'s own measurements were class based).')
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 30
    n += 1
    for j, h in enumerate(['taxid', 'Takson', 'Uyesi oldugu hedef(ler)', 'Cogaltan hedefler (mm<=1)',
                           'Ek (yalniz mm<=3)', 'Durum'], 1):
        yaz(n, j, h, bold=True, fill=GRI)
    n += 1
    for r in satirlar:
        f = (KIRMIZI if r['durum'].startswith('BOSLUK') else
             SARI if r['durum'] in ('yalniz evrensel/kontrol', 'yalniz mm<=3 ile') else None)
        yaz(n, 1, r['taxid'], fill=f); yaz(n, 2, r['ad'], fill=f); yaz(n, 3, r['uye'], fill=f)
        yaz(n, 4, r['cog1'], fill=f); yaz(n, 5, r['cog3'], fill=f); yaz(n, 6, r['durum'], fill=f)
        ws.row_dimensions[n].height = 26
        n += 1
    n += 1

    yaz(n, 1, 'BOSLUKLAR', bold=True, fill=KIRMIZI); n += 1
    gr = [('Hicbir hedef cogaltmiyor', [r for r in satirlar if r['durum'].startswith('BOSLUK')], KIRMIZI),
          ('Yalniz mm<=3 olcutuyle cogaliyor', [r for r in satirlar if r['durum'] == 'yalniz mm<=3 ile'], SARI),
          ('Yalniz evrensel/kontrol primeri cogaltiyor, ozgul hedefi yok',
           [r for r in satirlar if r['durum'] == 'yalniz evrensel/kontrol'], SARI),
          ('Ozgul bir hedefin kapsaminda',
           [r for r in satirlar if r['durum'].startswith('kapsanan')], YESIL)]
    for ad2, lst, f in gr:
        yaz(n, 1, len(lst), bold=True, fill=f); yaz(n, 2, ad2, fill=f)
        yaz(n, 4, ', '.join(r['ad'] for r in lst) or '-', fill=f)
        ws.merge_cells(start_row=n, start_column=4, end_row=n, end_column=6)
        ws.row_dimensions[n].height = 30
        n += 1
    yaz(n, 1, u'TOTAL', bold=True); yaz(n, 2, len(satirlar), bold=True)
    wb.save(xlsx)
    return len(satirlar)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--r2', required=True)
    ap.add_argument('--cross', dest='capraz', required=True)
    ap.add_argument('--pairs', dest='ciftler', required=True)
    ap.add_argument('--taxid', required=True)
    ap.add_argument('--md', required=True)
    ap.add_argument('--xlsx', required=True)
    a = ap.parse_args()
    R, CP, ad, uye = yukle(a.r2, a.capraz, a.ciftler, a.taxid)
    satirlar = takson_tablosu(R, CP, ad, uye)
    md_yaz(a.md, satirlar)
    n = xlsx_yaz(a.xlsx, satirlar)
    print('md:', a.md)
    print(u'xlsx sheet added, taxon:', n)
    for r in satirlar:
        if not r['durum'].startswith('kapsanan'):
            print(' ', r['durum'], '|', r['ad'])


if __name__ == '__main__':
    main()
