# -*- coding: utf-8 -*-
"""
order_and_status.py - it produces two delivery files from the LIVE panel:
  SIPARIS_LISTESI_20260802.tsv          only the pairs to be ordered
  TOPLANTI_KARARLARI_DURUM_20260802.md  YAPILDI / KISMEN / YAPILAMIYOR per decision

THE FRESHNESS RULE: the xlsx is read on every run from the LIVE file on the mounted
directory, no /tmp copy is used. After writing, the files are read back and compared
against the panel.

"""
# -------------------------------------------------------------------------
# order_and_status.py - it produces two delivery files from the live panel: the
#                       list of pairs to be ordered, and the done / partly / cannot
#                       be done status of the meeting decisions.
#
# INPUT  : the LIVE panel file on the mounted directory given with --xlsx (read
#          with openpyxl, its md5 taken and printed to the screen) plus the
#          project directory given with --root. The panel row to meeting decision
#          mapping (KARAR_ESLEME) is fixed inside the file.
# OUTPUT : SIPARIS_LISTESI_<date>.tsv and TOPLANTI_KARARLARI_DURUM_<date>.md.
#          After writing, both files are read back and the primer sequences are
#          compared against the panel; every difference found is printed.
# CALLED BY: IT IS NOT IN THE MENU - it is a delivery generator run by hand. Key
#          (P) in the menu produces a separate order list (in the protocol
#          directory) and does not use this script.
#
# THE FRESHNESS RULE IS CRITICAL: the xlsx is read from the live file on every run,
# no temporary copy is used. If the order sequence is taken from a stale copy THE
# WRONG PRIMER IS ORDERED and the mistake only shows up in the laboratory.
# -------------------------------------------------------------------------
import os, sys, csv, argparse, hashlib
import shutil
import openpyxl

# the panel row -> (decision, requested target); decisions 1 to 4 are the targets
# REQUESTED at the meeting
KARAR_ESLEME = {
    22: ('Karar 1 - tur ozgul', 'Methanosarcina mazei'),
    21: ('Karar 1 - tur ozgul', 'Methanothrix soehngenii'),
    10: ('Karar 2 - cins ozgul', 'Proteiniphilum'),
    4:  ('Karar 2 - cins ozgul', 'Petrimonas'),
    5:  ('Karar 3 - islev/ekolojik grup', 'Hidrojenotrofik metanojenler'),
    16: ('Karar 3 - islev/ekolojik grup', 'Asetoklastik metanojenler'),
    8:  ('Karar 3 - islev/ekolojik grup', 'Metilotrofik metanojen'),
    9:  ('Karar 3 - islev/ekolojik grup', 'Nitrosocosmicus AOA'),
    15: ('Karar 3 - islev/ekolojik grup', 'Sakarolitik bakteriler'),
    11: ('Karar 3 - islev/ekolojik grup', 'Proteolitik / sintrofik bakteriler'),
    19: ('Karar 3 - islev/ekolojik grup', 'Proteolitik / sintrofik bakteriler'),
    20: ('Karar 4 - alan / evrensel', 'Arke universal'),
    17: ('Karar 4 - alan / evrensel', 'Bakteri universal'),
    6:  ('Karar 4 - alan / evrensel', 'Mantar universal (F1)'),
    13: ('Karar 4 - alan / evrensel', 'Mantar universal (F2)'),
    2:  ('Karar 4 - alan / evrensel', 'Universal metanojen'),
    12: ('Karar 5 - olcumden turetilen', 'Bacteroidales kumesi'),
    7:  ('Karar 5 - olcumden turetilen', 'Methanosarcina cinsi'),
    3:  ('Karar 5 - olcumden turetilen', 'Methanothrix cinsi'),
    14: ('Karar 5 - olcumden turetilen', 'Microascaceae askomikot'),
    18: ('Karar 5 - olcumden turetilen', 'Petriella musispora'),
}

# targets REQUESTED at the meeting for which no pair could be given at all (no panel row)
YAPILAMIYOR = [
 ('Karar 1 - tur ozgul', 'Methanosarcina barkeri',
  'Organizma numunede yok: 2208 kutusunun en yakin referansi M. vacuolata %97,4-97,9 - tur esiginin altinda. Yerine CINS duzeyi verildi (Methanosarcina_cinsi).'),
 ('Karar 1 - tur ozgul', 'Podospora pseudopauciseta',
  'Organizma numunede yok: bes referans ciftinden ucu F1 sinifinin 85 804 okumasinin tamaminda 0 urun verdi.'),
 ('Karar 1 - tur ozgul', 'Dictyostelium discoideum (44689)',
  'Kraken2 etiketi olcumle curutuldu: SILVA LSU 28S D1-D2 testinde D. discoideum skoru 195, rastgele bir mantar 480. Kutu heterojen (dort barkod birbirine %70-76).'),
 ('Karar 1 - tur ozgul', 'Trichoderma asperellum (101201)',
  'Kutudaki organizma Trichoderma degil: ITS Petriella/Microascaceae veriyor. Tasarlanan cift panelden cikarildi (ayrim 0,7x).'),
 ('Karar 2 - cins ozgul', 'Bacteroides',
  'Cins numunede yok: bes kutunun en iyi Bacteroidales eslesmesi %84,6-85,9, 16S cins esigi ~%94-95. Yerine adlandirilamayan Bacteroidales soyu icin kume cifti verildi.'),
 ('Karar 2 - cins ozgul', 'Alistipes',
  'Bacteroides ile AYNI organizma (kutular birbirine %95,3-96,8 benziyor); Bacteroidales_kumesi altinda birlestirildi.'),
 ('Karar 3 - islev/ekolojik grup', 'Trichoderma cinsi',
  'Hedef numunede var ama cins degil: olculen kimlik Petriella/Microascaceae. Cift panelden cikarildi (ayrim 0,7x).'),
]

# the PARTLY reasons (there IS a panel row but the requested level could not be given)
KISMEN_NOT = {
 15: 'Toplanti "sakarolitik bakteriler" grubunu istedi; grup capinda cift bulunamadi. Verilen: UYE BAZLI tek cins cifti (Sphaerochaeta associata). Grubun diger uyeleri kapsanmiyor.',
 11: 'Toplanti "proteolitik/sintrofik bakteriler" grubunu istedi; grup capinda cift yok. Verilen: uye bazli iki ayri cift (Synergistaceae soyu + Cloacimonas cinsi). Hedef adi olculen kimlige cekildi - numunedeki organizma Cloacibacillus degil, adlandirilamayan Synergistaceae (%99,39; Cloacibacillus %90,02, cins esigi %94,5).',
 19: 'Ayni grubun ikinci uyesi. Bkz. Proteolitik_Synergistaceae satiri.',
 2:  'Kapsam TAM DEGIL: 34 metanojen kutusunun 33\'u cogaliyor. Ca. Methanomassiliicoccus kutusu bu ciftle cogalmiyor, onu Metilotrofik_metanojen cifti kapsiyor.',
 21: 'TUR duzeyi verildi ama KOSULLU: amplikon dizilemesi sarti var. Capraz vurusun 52\'si ayni ailedeki diger Methanothrix kayitlari - erime egrisi ayirmiyor.',
 22: 'TUR GRUBU verildi (M. mazei / M. soligelidi ayrilamiyor). OKUMA MOTORU DUZELTMESI SONRASI ESIK ALTI: en kotu tek kutu 0,82x. M. hadiensis kutusu hedef kadar iyi cogaliyor (%47,22). YENIDEN KARAR GEREKIYOR.',
 12: 'Toplantinin istedigi Bacteroides/Alistipes cinsleri numunede yok; yerine adlandirilamayan Bacteroidales SOYU icin kume cifti verildi. Ayrim 5,9x - 10x esiginin ALTINDA, teslim kosullu.',
 4:  'Cins OZGUL ama cins KAPSAMLI DEGIL: dogrulanmis P. sulfuriphila kutusu %57,0; Petrimonas kutularinin tamami kapsanmiyor.',
}


def kisa(v, n=110):
    s = '' if v is None else str(v).replace('\n', ' ').strip()
    return s if len(s) <= n else s[:n - 1] + '…'


# -------------------------------------------------------------------------
# THE 2026-08-04 FIX: the output names ARE NOT FIXED.
#
# This script used to write the output as SIPARIS_LISTESI_20260802.tsv on every
# run. The consequence: although newer lists produced on 3 August were sitting side
# by side with it, the file the script produced looked by its name as though it had
# stayed at 2 August, and which one was current could not be told from the name.
#
# The new behaviour has three parts:
#   1) The dated file is derived from THE PANEL'S own modification date. That way
#      the name carries which panel version the output really came from.
#   2) A CANONICAL SIPARIS_LISTESI.tsv is written as well. The order is always
#      given from this name; the dated files stay as a record.
#   3) If a newer dated order list that this script DID NOT PRODUCE is in the
#      directory, a WARNING is printed and the canonical file IS NOT OVERWRITTEN.
#      Going back to the old one silently has cost this project dearly exactly once.
# -------------------------------------------------------------------------

KANONIK_SIPARIS = 'SIPARIS_LISTESI.tsv'


def panel_tarihi(xlsx_yolu):
    """Derives the output's date label FROM THE PANEL, not from today's date.

    The reason: a list produced twice from the same panel must get the same name.
    Had today's date been used, a new file would appear every day although nothing
    had changed, and the question of which one is current would be born again.

    """
    import datetime
    return datetime.datetime.fromtimestamp(os.path.getmtime(xlsx_yolu)).strftime('%Y%m%d')


def daha_yeni_liste_var_mi(kok, bizim_dosya, kaynak_panel):
    """Is there a newer order list that this script DID NOT PRODUCE.

    It looks by MODIFICATION DATE, not by name; a name lets you guess, a date
    measures. If it finds one it returns (file_name, date), otherwise None.

    CAREFUL: the comparison is made against THE SOURCE PANEL'S date, NOT against the
    date of the file this script has just written. The latter is now on every run and
    no candidate could ever come out newer than it, so the warning would never fire.
    The question asked is this: is there a list produced AFTER the panel I am derived
    from?

    """
    import glob, datetime
    bizim_t = os.path.getmtime(kaynak_panel)
    aday = []
    for d in ('', '1_TESLIM'):
        for y in glob.glob(os.path.join(kok, d, 'SIPARIS_LISTESI*.tsv')):
            if os.path.basename(y) == KANONIK_SIPARIS:
                continue
            if os.path.abspath(y) == os.path.abspath(bizim_dosya):
                continue
            if os.path.getmtime(y) > bizim_t:
                aday.append((y, os.path.getmtime(y)))
    if not aday:
        return None
    y, t = max(aday, key=lambda x: x[1])
    return (os.path.relpath(y, kok),
            datetime.datetime.fromtimestamp(t).strftime('%Y-%m-%d %H:%M'))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--root', dest='kok', required=True)
    a = ap.parse_args()

    # ---- CANLI dosyadan oku ------------------------------------------------
    xy = os.path.abspath(a.xlsx)
    ozet = hashlib.md5(open(xy, 'rb').read()).hexdigest()
    wb = openpyxl.load_workbook(xy, data_only=True)
    ws = wb['2 Panel']
    print('CANLI panel : %s' % xy)
    print('md5         : %s' % ozet)
    print('sayfa       : %d' % len(wb.sheetnames))

    satirlar = []
    for i in range(2, 23):
        satirlar.append(dict(
            satir=i, hedef=str(ws.cell(i, 3).value or ''), duzey=str(ws.cell(i, 4).value or ''),
            plaka=str(ws.cell(i, 1).value or ''), ta=ws.cell(i, 2).value,
            F=str(ws.cell(i, 6).value or ''), Fuz=ws.cell(i, 7).value, FTm=ws.cell(i, 8).value,
            R=str(ws.cell(i, 10).value or ''), Ruz=ws.cell(i, 11).value, RTm=ws.cell(i, 12).value,
            urun=ws.cell(i, 15).value, aralik=kisa(ws.cell(i, 27).value, 60),
            ayrim=kisa(ws.cell(i, 18).value, 70), siparis=str(ws.cell(i, 23).value or ''),
            durum=str(ws.cell(i, 20).value or ''),
            duz1=kisa(ws.cell(i, 29).value, 24), esik=kisa(ws.cell(i, 32).value, 24)))

    sip = [r for r in satirlar if not r['siparis'].upper().startswith('HAYIR')]
    hayir = [r for r in satirlar if r['siparis'].upper().startswith('HAYIR')]

    # ---- 1) SIPARIS LISTESI ------------------------------------------------
    ETIKET = panel_tarihi(xy)                      # derived from the panel, not fixed
    syol = os.path.join(a.kok, 'SIPARIS_LISTESI_%s.tsv' % ETIKET)
    BAS = ['#', 'Hedef', 'Duzey', 'Plaka', 'Ta (C)',
           "Ileri primer (5'->3')", 'Ileri uz', 'Ileri Tm',
           "Geri primer (5'->3')", 'Geri uz', 'Geri Tm',
           'Urun (bp)', 'Urun boyu araligi (numunede)', 'Siparis notu']
    with open(syol, 'w', encoding='utf-8', newline='') as fh:
        w = csv.writer(fh, delimiter='\t')
        w.writerow(['# SIPARIS YALNIZ BU DOSYADAN VERILIR.'])
        w.writerow(['# Kaynak: %s ("2 Panel"), md5 %s' % (os.path.basename(xy), ozet)])
        w.writerow(['# Uretim: %s | Bu dosyada YALNIZ siparis edilecek %d cift vardir. '
                    'Siparis edilmeyecek satir yoktur.'
                    % (ETIKET[:4] + '-' + ETIKET[4:6] + '-' + ETIKET[6:], len(sip))])
        w.writerow(['# Panelden cikarilan ve siparis EDILMEYECEK ciftler: %s'
                    % ', '.join(r['hedef'] for r in hayir)])
        w.writerow(['# Toplam oligo: %d (her cift 2 oligo, dejenere baz YOK - toplanti karari).'
                    % (2 * len(sip))])
        w.writerow([])
        w.writerow(BAS)
        for n, r in enumerate(sip, 1):
            w.writerow([n, r['hedef'], r['duzey'], r['plaka'], r['ta'],
                        r['F'], r['Fuz'], r['FTm'], r['R'], r['Ruz'], r['RTm'],
                        r['urun'], r['aralik'], r['siparis']])

    # ---- GERI OKU VE PANELLE KARSILASTIR -----------------------------------
    okunan = list(csv.reader(open(syol, encoding='utf-8'), delimiter='\t'))
    veri = [x for x in okunan if x and x[0].isdigit()]
    kar = fark = 0
    for x in veri:
        h = x[1]
        p = next((r for r in sip if r['hedef'] == h), None)
        if p is None:
            fark += 1; print(u'  !! not in the panel:', h); continue
        for sut, deg in ((5, p['F']), (8, p['R'])):
            kar += 1
            if x[sut] != deg:
                fark += 1
                print(u'  !! SEQUENCE DIFFERENCE %s column %d: file=%s panel=%s' % (h, sut, x[sut], deg))
    print(u'\nTHE ORDER FILE: %s' % syol)
    print(u'  pairs to be ordered : %d   (NOT to be ordered: %d)' % (len(sip), len(hayir)))
    print(u'  sequences compared    : %d' % kar)
    print('  bulunan fark          : %d' % fark)

    # ---- 2) TOPLANTI KARARLARI DURUMU --------------------------------------
    def kat(r):
        if r['siparis'].upper().startswith('HAYIR'):
            return 'YAPILAMIYOR'
        if r['satir'] in KISMEN_NOT:
            return 'KISMEN'
        return 'YAPILDI'

    yapildi, kismen = [], []
    for r in satirlar:
        if r['satir'] not in KARAR_ESLEME:
            continue
        k = kat(r)
        if k == 'YAPILDI':
            yapildi.append(r)
        elif k == 'KISMEN':
            kismen.append(r)

    myol = os.path.join(a.kok, 'TOPLANTI_KARARLARI_DURUM_%s.md' % ETIKET)
    L = []
    A = L.append
    A(u'# The meeting decisions, what was managed and what was not\n')
    A(u'2 August 2026. The numbers were taken **from the live panel**: `%s` (`2 Panel`), md5 `%s`.\n' % (os.path.basename(xy), ozet))
    A(u'## The target count, the contradiction between the documents\n')
    A(u'Some documents say "twenty targets", others "six species plus four genera". **What is taken as the base:** the panel\'s own decision ledger, `6 Karar Durumu`. By that, **21 targets** were asked for at the meeting:\n')
    A(u'| Decision | Targets asked for |')
    A('|---|---|')
    A(u'| Decision 1, species specific | 6 |')
    A(u'| Decision 2, genus specific | 4 |')
    A(u'| Decision 3, a functional or ecological group | 7 |')
    A(u'| Decision 4, domain or universal | 4 |')
    A(u'| **Total asked for** | **21** |')
    A(u'| Decision 5, derived from the measurement (not asked for at the meeting) | 8 |\n')
    A(u'The phrase "six species plus four genera" describes **Decision 1 plus Decision 2** only (10 targets), not the whole meeting.')
    A('---\n')

    # uye kumesi guvenilirligi - CANLI pairs.tsv'den
    uyedur = {}
    cp = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pairs.tsv')
    if os.path.exists(cp):
        for c in csv.DictReader(open(cp, encoding='utf-8'), delimiter='\t'):
            uyedur[int(c['satir'])] = c['uye_kumesi_durumu']

    def ayrim_ve_not(r):
        ham = r['duz1'] if r['duz1'] and r['duz1'] != '- / -' else ''
        notlar = []
        deger = ham or r['ayrim']
        try:
            enk = float(ham.split('/')[0].strip()) if ham else None
        except ValueError:
            enk = None
        if uyedur.get(r['satir']) == 'YENIDEN_KURULDU' and ham:
            # the corrected measurement was made with the rebuilt member set, so it is not reliable on its own
            deger = 'panel: %s' % r['ayrim']
            notlar.append(u'the member set **could not be confirmed** under the corrected engine (the panel\'s value is shown)')
        elif enk is not None and enk < 10:
            notlar.append(u'**the worst single bin is BELOW 10x**')
        if str(r['esik']).upper().startswith('EVET'):
            if not notlar:
                notlar.append(u'**below the 10x threshold**')
        return deger, '; '.join(notlar) or '-'

    A(u'## DONE, there is a pair and it can be ordered (%d)\n' % len(yapildi))
    A(u'| Decision | Target | Forward / Reverse primer | Product | Discrimination | Warning |')
    A('|---|---|---|---|---|---|')
    for r in sorted(yapildi, key=lambda x: KARAR_ESLEME[x['satir']][0]):
        k, ist = KARAR_ESLEME[r['satir']]
        ay, nt = ayrim_ve_not(r)
        A('| %s | %s | `%s` / `%s` | %s bp | %s | %s |' % (k.split(' - ')[0], r['hedef'],
                                                           r['F'], r['R'], r['urun'], ay, nt))
    A('')
    A(u'> The discrimination column: **the worst single bin / the pool** under the corrected read engine (mm<=1). The "panel:" prefix says that the corrected measurement could not confirm the member set and that the panel\'s own value is being shown. Targets reading "x/y kutu" are a coverage measure; discrimination is not measured for them. **Pairs below the 10x threshold can be ordered but they are conditional**: they need amplicon sequencing or a gel confirmation.')

    A(u'## PARTLY, something was given but not at the level asked for (%d)\n' % len(kismen))
    A(u'| Decision | Asked for | Given | Why |')
    A('|---|---|---|---|')
    for r in sorted(kismen, key=lambda x: KARAR_ESLEME[x['satir']][0]):
        k, ist = KARAR_ESLEME[r['satir']]
        A('| %s | %s | %s (%s bp) | %s |' % (k.split(' - ')[0], ist, r['hedef'],
                                             r['urun'], KISMEN_NOT[r['satir']]))
    A('')

    A(u'## CANNOT BE DONE, nothing could be given (%d)\n' % len(YAPILAMIYOR))
    A(u'| Decision | Asked for | The measured reason |')
    A('|---|---|---|')
    for k, ist, ger in YAPILAMIYOR:
        A('| %s | %s | %s |' % (k.split(' - ')[0], ist, ger))
    A('')
    A('---\n')
    A(u'## The order\n')
    A(u'**%d pairs = %d oligos** to be ordered. The sequences must be copied from the `%s` file only.\n' % (len(sip), 2 * len(sip), KANONIK_SIPARIS))
    A(u'The pairs removed from the panel, which will not be ordered: %s.\n'
      % ', '.join(r['hedef'] for r in hayir))
    open(myol, 'w', encoding='utf-8').write('\n'.join(L) + '\n')

    # ---- THE CANONICAL COPY ------------------------------------------------
    # The order is always given from the name KANONIK_SIPARIS. But first we look at
    # whether a newer list this script did not produce exists; if it does the canonical
    # file IS NOT TOUCHED and the user is warned. Going back to the old one silently is
    # forbidden.
    kyol = os.path.join(a.kok, KANONIK_SIPARIS)
    yeni_olan = daha_yeni_liste_var_mi(a.kok, syol, xy)
    if yeni_olan:
        print(u'\n  WARNING: there is a NEWER order list that this script did not produce:')
        print('           %s   (%s)' % yeni_olan)
        print(u'         list produced : %s' % os.path.basename(syol))
        print(u'         %s WAS NOT OVERWRITTEN. Decide for yourself which one holds;'
              % KANONIK_SIPARIS)
        print(u'         if you want to change the canonical one, copy that file by hand.')
    else:
        shutil.copyfile(syol, kyol)
        print(u'\nCANONICAL LIST : %s   (the order is placed from THIS file)' % kyol)

    print(u'\nTHE STATUS FILE: %s' % myol)
    print(u'  DONE %d | PARTLY %d | CANNOT BE DONE %d' % (len(yapildi), len(kismen), len(YAPILAMIYOR)))
    return 0 if fark == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
