# -*- coding: utf-8 -*-
"""orientation_report.py - panele "19 Yon Normalizasyonu" sayfasini yazar."""
# -------------------------------------------------------------------------
# orientation_report.py writes all the evidence of the orientation normalisation
#                  into the delivery panel xlsx as a sheet named "19 Yon
#                  Normalizasyonu".
#
# INPUT  : the delivery panel given with --xlsx; the per file orientation table
#          orientation_audit.py produces, given with --orientation; the code route
#          classification orientation_code_scan.py produces, given with --code; and
#          konsensus_kanonik/INDEKS.tsv given with --index. The notes on how each
#          file was corrected (KOD_NOT) are fixed inside this file.
# OUTPUT : it adds a new sheet to the given xlsx and saves it (wb.save); it prints
#          the sheet name written and the row count to the screen.
# CALLED BY: IT IS NOT IN THE MENU, it is run by hand, because it changes the
#          delivery file. Both scripts that produce its inputs
#          (orientation_audit.py and orientation_code_scan.py) are run by hand too.
#
# Why the sheet exists: the orientation fault was found and patched in three
# separate places on the same night. This sheet keeps the answers to "which file
# was corrected, which consensus was flipped, which code route is still at risk"
# inside the panel, in a form that can be followed.
# -------------------------------------------------------------------------
import os, sys, csv, argparse, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

KIRMIZI = PatternFill('solid', fgColor='FFC7CE')
SARI = PatternFill('solid', fgColor='FFEB9C')
YESIL = PatternFill('solid', fgColor='C6EFCE')
GRI = PatternFill('solid', fgColor='D9D9D9')
KALIN = Font(bold=True)
SAR = Alignment(wrap_text=True, vertical='top')

KOD_NOT = {
 'screening/config.py':
   'FIXED. The consensus path is now the canonical directory. The raw '
   'directory stands as KONSENSUS_HAM alone and ONLY the canonical generation '
   'reads it.',
 'screening/targets.py':
   'FIXED. The consensus loader now reads the canonical index. Without an '
   'index it RAISES AN ERROR; it DOES NOT fall back to the mixed directory '
   'silently.',
 'screening/build_consensus.py':
   'FIXED in three places: (a) the template is converted to canonical, (b) '
   'the orphan bin template, which is a raw read, is converted too, and (c) '
   'the OUTPUT is measured once more and converted before it is written. A '
   'GATE was also put at the head of the run: generation does not start until '
   'the orientation test passes.',
 'screening/self_test.py':
   "An orientation test was added: orientation.py's own test, whether any "
   'file in the canonical set is reversed, and a known answer impact test. It '
   'DOES NOT depend on an optional dependency such as primer3.',
 'screening/generator.py':
   'Orientation independent: it tries the sequence both forward and reverse '
   'complemented. It works correctly with the canonical source too.',
 'screening/panel_measurement.py':
   'It measures RAW READS and is unaffected by the consensus orientation, '
   'since reads are scanned in both directions anyway.',
 'screening/membership_check.py':
   'Membership and measurement; it takes the consensus path through '
   'targets.py, so it follows the canonical directory.',
 'screening/orientation.py':
   'NEW. The DEFINITION of the canonical orientation and its normalisation. '
   'Two independent criteria, plus a self test.',
 'screening/build_canonical.py':
   'NEW. It produces the canonical directory, writes the manifest, the index '
   'and the undecided list, and confirms its own work.',
 'screening/orientation_audit.py':
   'NEW, an audit tool. It reads the mixed directories DELIBERATELY, because '
   'measuring the orientation is its job.',
 'screening/orientation_impact_test.py':
   'NEW. A known answer test: the product count in the right orientation '
   'against the reversed one.',
 'steps/split_clusters.py':
   'THE OLD LINE. It reads the mixed directory assuming one orientation. It '
   "DOES NOT produce the panel's current numbers, since that line is out of "
   'use, but rerunning it would give a wrong result, so it has to be moved '
   'onto the canonical directory or archived.',
}


def yaz(ws, r, c, v, fill=None, bold=False):
    h = ws.cell(r, c, v)
    if fill: h.fill = fill
    if bold: h.font = KALIN
    h.alignment = SAR
    return h


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--orientation', dest='yon', required=True)
    ap.add_argument('--code', dest='kod', required=True)
    ap.add_argument('--index', dest='indeks', required=True)
    a = ap.parse_args()

    yr = list(csv.DictReader(open(a.yon, encoding='utf-8'), delimiter='\t'))
    kr = list(csv.DictReader(open(a.kod, encoding='utf-8'), delimiter='\t'))
    ix = list(csv.DictReader(open(a.indeks, encoding='utf-8'), delimiter='\t'))

    say = collections.defaultdict(collections.Counter)
    for x in yr:
        say[x['klasor']][x['karar'].split(' (')[0]] += 1

    wb = openpyxl.load_workbook(a.xlsx)
    ad = '19 Yon Normalizasyonu'
    if ad in wb.sheetnames:
        del wb[ad]
    ws = wb.create_sheet(ad)
    for w, c in zip((44, 14, 14, 14, 14, 16, 60), 'ABCDEFG'):
        ws.column_dimensions[c].width = w
    n = 1

    yaz(ws, n, 1, 'YON NORMALIZASYONU - 2026-08-02', bold=True); n += 1
    yaz(ws, n, 1, u'ANSWER: the orientation IS NOW CANONICAL. It was not before: there were four separate consensus sets and two of them were mixed orientation. A single', fill=YESIL)
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 32; n += 2

    yaz(ws, n, 1, u'1. WHY IT WAS ASKED - three separate patches, no single canonical fix', bold=True, fill=GRI); n += 1
    for s in ['The orientation bug was found and patched separately in AT LEAST THREE SEPARATE PLACES over the course of one night: (a) the strand choice of 85 on the source study side, (b) "the consensuses are the reverse complement of SILVA" on the design side, (c) "39 of 58 consensuses are in the reverse orientation" during the B/F re-measurement. Three separate patches means there is no single canonical fix, which means it escapes again on the next change. This page ties "I think it is fixed" to a measurement.']:
        yaz(ws, n, 1, s); ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'2. MEASUREMENT - the orientation of every consensus file BEFORE the fix', bold=True, fill=GRI); n += 1
    yaz(ws, n, 1, u'Method: two INDEPENDENT criteria. (1) the panel\'s own universal pairs: in the sense orientation F and rc(R) are found directly' % len(yr))
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 46; n += 1
    for j, h in enumerate(['Klasor', 'SENSE', 'ANTISENSE', 'BELIRSIZ/bos', 'Toplam', 'Karisik mi?', 'Not'], 1):
        yaz(ws, n, j, h, bold=True, fill=GRI)
    n += 1
    NOT = {
     'consensus sequences': 'the original output of the source study. MIXED. '
                            'The cause: the samtools and minimap2 line builds '
                            'the template around a READ chosen from the data, '
                            'and because nanopore reads arrive about half in '
                            'each direction the output orientation is random.',
     'referans_konsensus/konsensus': 'The set normalised overnight. Clean.',
     'referans_konsensus/baskin/konsensus': 'The dominant allele set. Clean; '
                                            '6 files could not be measured '
                                            'because their N fraction is '
                                            'high.',
     'referans_konsensus/self/konsensus': 'The self set. Clean; 1 file is '
                                          'empty.',
     'SCREENING_RESULT/konsensus_yeni': 'THE OUTPUT OF THE GENERATION THAT '
                                        'WAS TO RUN. MIXED and mostly '
                                        'ANTISENSE. The root cause: it took '
                                        'its template from the mixed '
                                        'directory, as set out below.',
    }
    for k in sorted(say):
        c = say[k]
        se, an = c.get('SENSE', 0), c.get('ANTISENSE', 0)
        bl = sum(v for kk, v in c.items() if kk not in ('SENSE', 'ANTISENSE'))
        kar = 'EVET - KARISIK' if se and an else 'hayir'
        f = KIRMIZI if (se and an) else None
        yaz(ws, n, 1, k, fill=f); yaz(ws, n, 2, se, fill=f); yaz(ws, n, 3, an, fill=f)
        yaz(ws, n, 4, bl, fill=f); yaz(ws, n, 5, se + an + bl, fill=f)
        yaz(ws, n, 6, kar, fill=f); yaz(ws, n, 7, NOT.get(k, ''), fill=f)
        ws.row_dimensions[n].height = 44
        n += 1
    n += 1

    yaz(ws, n, 1, u'3. THE ROOT CAUSE (the evidence in the code)', bold=True, fill=KIRMIZI); n += 1
    for s in ['screening/config.py pointed the consensus path at the raw '
              'directory, so the package took the MIXED directory as its one '
              'source.',
              'screening/build_consensus.py -> _sablon_sec(): the template was taken from the "current consensus". The reads were anchored to the template in both directions and NORMALISED (the code called this "the orientation is normalised"), but THE TEMPLATE\'S OWN ORIENTATION was normalised nowhere. The result: the output inherits the template\'s orientation exactly. Measured evidence: konsensus_yeni was 28 antisense / 7 sense.',
              "For orphan bins the template was chosen straight from a RAW READ, and a read's orientation is random, so the output orientation was random too."]:
        yaz(ws, n, 1, s, fill=KIRMIZI if 'konsensus_uret' in s else None)
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
        ws.row_dimensions[n].height = 46; n += 1
    n += 1

    yaz(ws, n, 1, u'4. A TEST WITH A KNOWN ANSWER - how far each number moves when the orientation is wrong',
        bold=True, fill=GRI); n += 1
    yaz(ws, n, 1, u'Setup: the same primer pair, the same consensus, THE ONLY DIFFERENCE BEING ORIENTATION. Source: referans_konsensus/konsensus (99 files')
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 46; n += 1
    for j, h in enumerate(['Cift', 'Sinif', 'Dogru yon', 'TERS yon', 'Kayip', 'Sonuc', ''], 1):
        yaz(ws, n, j, h, bold=True, fill=GRI)
    n += 1
    for ad2, sn, d, t, top in [('Arke_universal', 'A', 38, 0, 39),
                               ('Bakteri_universal', 'B', 20, 0, 20),
                               ('Mantar_universal (F1)', 'F1', 18, 0, 20),
                               ('Mantar_universal (F2)', 'F2', 15, 0, 20),
                               ('Methanosarcina_mazei_turu', 'A', 12, 0, 39),
                               ('Methanosarcina_cinsi', 'A', 13, 0, 39),
                               ('Proteolitik_Cloacimonas', 'B', 1, 0, 20)]:
        yaz(ws, n, 1, ad2, fill=KIRMIZI); yaz(ws, n, 2, sn, fill=KIRMIZI)
        yaz(ws, n, 3, '%d/%d' % (d, top), fill=KIRMIZI)
        yaz(ws, n, 4, '%d/%d' % (t, top), fill=KIRMIZI)
        yaz(ws, n, 5, d - t, fill=KIRMIZI)
        yaz(ws, n, 6, u'the reverse orientation ZEROES the product', fill=KIRMIZI)
        n += 1
    yaz(ws, n, 1, u'TOTAL: 117 products in the correct orientation, 0 in the reverse orientation. LOSS 100%.', bold=True, fill=KIRMIZI)
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7); n += 1
    yaz(ws, n, 1, u'CROSS-CHECK: the same test was repeated with the panel\'s OWN engine (ispcr.amplify). Arke_universal gave 38 correct')
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 30; n += 2

    yaz(ws, n, 1, u'5. THE CANONICAL FIX - one source, one definition', bold=True, fill=YESIL); n += 1
    for s in ['THE CANONICAL ORIENTATION = SENSE (the reference, or plus, strand). The definition lives in ONE PLACE: screening/orientation.py. Two independent criteria, and if they disagree the file counts as UNCERTAIN, is NOT normalised, and is flagged instead (the project rule: no decision is left to a single code path).',
              'ONE SOURCE: the canonical directory, %d bins and all of them '
              'SENSE. Produced by screening/build_canonical.py, with a '
              "manifest beside it holding each file's source, its old "
              'orientation and whether it was converted, an index of the '
              'valid files, and a list of the undecided ones. In this run %d '
              'files were converted from ANTISENSE to SENSE and none were '
              'left undecided.'
              % (len(ix), sum(1 for r in ix if r['cevrildi'] == 'EVET')),
              'EVERY SCRIPT READS THIS PLACE: targets.py -> konsensusler() now reads INDEKS.tsv and RAISES AN ERROR when the index is missing; it does NOT fall back SILENTLY to the mixed directory. No script carries its own orientation patch any more.',
              'CAUTION - files CANNOT BE DELETED on a mounted drive. konsensus_kanonik/ still holds the misnamed "*_kanonik.fasta" leftovers of the first run. The valid files match the pattern "*.kanonik.fa" and are listed in INDEKS.tsv. Consumers read the INDEX, NOT a glob. The README.txt in the directory repeats this.']:
        yaz(ws, n, 1, s, fill=SARI if s.startswith('DIKKAT') else None)
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
        ws.row_dimensions[n].height = 52; n += 1
    n += 1

    yaz(ws, n, 1, u'6. CODE PATHS - every script that reads a consensus, and how it handles orientation',
        bold=True, fill=GRI); n += 1
    yaz(ws, n, 1, u'Automatic scan (orientation_code_scan.py, with comments and docstrings stripped) plus a manual note. Full list: yon_kod_tara' % len(kr))
    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
    ws.row_dimensions[n].height = 26; n += 1
    for j, h in enumerate(['Betik', 'Sinif', 'Karisik kaynak', 'Normalize kaynak',
                           'Iki yon deniyor', 'Kendi yamasi', 'Not'], 1):
        yaz(ws, n, j, h, bold=True, fill=GRI)
    n += 1
    onem = {'HAM_KAYNAK_VARSAYIM': 0, 'KAYNAK_BELIRSIZ': 1, 'NORMALIZE_KAYNAK': 2,
            'IKI_YON_DENIYOR': 3, 'KENDI_YAMASI': 4, 'YON_ILGISIZ': 5}
    for r in sorted(kr, key=lambda x: (onem.get(x['sinif'], 9), x['betik'])):
        if r['betik'].startswith('screening/eski/'):
            continue
        f = (KIRMIZI if r['sinif'] == 'HAM_KAYNAK_VARSAYIM' and 'yapilandirma' not in r['betik']
             and 'yon_denetimi' not in r['betik'] else
             YESIL if r['betik'] in KOD_NOT and KOD_NOT[r['betik']].startswith(('DUZELTILDI', 'YENI')) else None)
        yaz(ws, n, 1, r['betik'], fill=f); yaz(ws, n, 2, r['sinif'], fill=f)
        yaz(ws, n, 3, r['karisik_kaynak'], fill=f); yaz(ws, n, 4, r['normalize_kaynak'], fill=f)
        yaz(ws, n, 5, r['iki_yon'], fill=f); yaz(ws, n, 6, r['kendi_yamasi'], fill=f)
        yaz(ws, n, 7, KOD_NOT.get(r['betik'], ''), fill=f)
        if KOD_NOT.get(r['betik']):
            ws.row_dimensions[n].height = 40
        n += 1
    n += 1

    yaz(ws, n, 1, u'7. THE PRODUCTION RUN TONIGHT - checked specifically', bold=True, fill=YESIL); n += 1
    for s in ['build_consensus.py was fixed in THREE places: (a) the template is converted to canonical, (b) the orphan bin template (a raw read) is converted too, (c) the OUTPUT is measured once more and converted before it is written (the last seat belt). The output fasta header carries kanonik=... cevrildi=..., and the columns cikti_yon and cikti_cevrildi were added to the TSV.',
              'A GATE WAS PUT IN: konsensus_uret.calistir() runs the orientation test first, and if it does not pass, GENERATION IS NOT STARTED and what to do is printed. The reason: if the output of this step comes out in the wrong orientation, the whole night is wasted.',
              "yon_sinamasi() was added inside self_test.py (3 items: orientation.py's own test, whether the canonical set still holds a reversed file, and a known-answer impact test). It does NOT TRIP over optional dependencies such as primer3, and it can be called on its own. It was run in this session: 3/3 PASSED (correct 39/40, reversed 0/40).",
              'THE ORDER: (1) python screening/build_canonical.py --root .   (2) consensus regeneration from the verification/full_chain.py menu   (3) once generation finishes, run kanonik_uret AGAIN with --priority yeni so that the new set becomes the canonical source.']:
        yaz(ws, n, 1, s, fill=SARI if s.startswith('SIRA') else None)
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
        ws.row_dimensions[n].height = 56; n += 1
    n += 1

    yaz(ws, n, 1, '8. ACIK KALAN', bold=True, fill=SARI); n += 1
    for s in ['konsensus_kanonik currently comes from referans_konsensus/konsensus (99 bins) plus 1 orphan bin (A1-1_2209, from the original directory, which was ANTISENSE and was converted). konsensus_yeni was NOT taken into the canonical set because it is INCOMPLETE (35/99); mixing a half set with a full one creates a heterogeneous base. Once the overnight generation finishes it should be re-run with --priority yeni.',
              "steps/split_clusters.py reads the mixed directory assuming a single orientation. It does not produce the panel's CURRENT numbers (it is a disabled line), but if it is re-run it gives a wrong answer. It should either be converted to canonical or archived.",
              'The scripts flagged KAYNAK_BELIRSIZ in yon_kod_taramasi (most of them under steps and engine) take the consensus path from the command line and leave the orientation to the caller. Those are the old line. If they are to be re-run, the canonical directory must be given.',
              'The orientation measurement on this page applies to the consensus files. RAW READ measurements are unaffected by orientation (reads are scanned in both directions anyway), so the numbers on the "16 Okuma Motoru Duzeltmesi" sheet are NOT AFFECTED by this finding.']:
        yaz(ws, n, 1, s)
        ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=7)
        ws.row_dimensions[n].height = 56; n += 1

    wb.save(a.xlsx)
    print('sayfa yazildi:', ad, '| satir', n)


if __name__ == '__main__':
    main()
