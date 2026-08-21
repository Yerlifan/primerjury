# -*- coding: utf-8 -*-
"""
tsv_senkron.py - TESLIM TSV ikizlerini panel xlsx'inden YENIDEN URETIR.

Neden: TSV'ler panelden elle turetildigi icin kayabiliyor. 2026-08-02 denetimi
Metanojen_universal ileri primerinin TSV'de hala 21 nt'lik ESKI dizi oldugunu buldu
(panelde bir G eklenmisti, kapsam %19 -> %97). Siparis TSV'den kopyalanirsa YANLIS
PRIMER SIPARIS EDILIR.

Once FARKLARI bildirir, sonra --yaz ile uretir. Primer dizisi sutunlarindaki her fark
ayrica KRITIK olarak isaretlenir.

Kullanim:
  python tsv_senkron.py --xlsx ..\PrimerJury_..._TESLIM.xlsx --kok ..          (yalniz rapor)
  python tsv_senkron.py --xlsx ... --kok .. --yaz                                  (uret)
"""
# ---------------------------------------------------------------------------
# tsv_senkron.py — teslim TSV ikizlerini panel xlsx'inden yeniden uretir ve
#                  aradaki her hucre farkini, ozellikle primer dizisi
#                  farklarini, KRITIK olarak bildirir.
#
# GIRDI  : --xlsx teslim paneli (openpyxl) ve --kok altindaki mevcut TSV
#          dosyalari. Sayfa -> TSV yolu eslemesi (ESLEME) dosyanin icinde
#          sabittir.
# CIKTI  : --yaz verilirse eslemedeki TSV dosyalarini panelden yeniden yazar;
#          verilmezse hicbir seyi degistirmeden yalniz farklari ekrana basar.
# CAGRAN : MENUDE DEGILDIR - elle calistirilir.
#
# NEDEN VAR: TSV'ler panelden elle turetildigi icin kayabiliyor. Olculen ornek,
# Metanojen_universal ileri primerinin TSV'de hala 21 nt'lik eski dizi olarak
# durmasidir; panelde bir G eklenmis ve kapsam %19'dan %97'ye cikmisti. Siparis
# TSV'den kopyalanirsa yanlis primer siparis edilir - bu yuzden primer dizisi
# sutunlarindaki farklar diger farklardan AYRI, KRITIK basligi altinda listelenir.
# ---------------------------------------------------------------------------
import os, sys, csv, argparse
import openpyxl

ESLEME = [
    ('1 Rapora Ozet',            'degerlendiriciya_ozet_20260802_TESLIM.tsv'),
    ('2 Panel',                  'primer_final/devir_ciftleri_20260802_sonrotus_TESLIM.tsv'),
    ('3 Triyaj (matris ilgisi)', 'primer_final/triyaj_20260802_TESLIM.tsv'),
    ('4 Degisiklikler',          'primer_final/degisiklikler_20260802_TESLIM.tsv'),
    ('5 Dusen ciftler',          'primer_final/devir_dusenler_20260802_sonrotus_TESLIM.tsv'),
    ('6 Karar Durumu',           'primer_final/karar_durumu_20260802_TESLIM.tsv'),
    ('7 Ayrilik Tablosu',        'primer_final/ayrilik_tablosu_20260802_TESLIM.tsv'),
    ('8 Geometri Denetimi',      'primer_final/geometri_denetimi_20260802_TESLIM.tsv'),
    ('9 Kurtarma ve Onarim',     'primer_final/kurtarma_ve_onarim_20260802_TESLIM.tsv'),
    ('10 Olcum Hatalari',        'primer_final/olcum_hatalari_20260802_TESLIM.tsv'),
    ('11 B-F Yeniden Olcum',     'primer_final/bf_yeniden_olcum_20260802_TESLIM.tsv'),
    ('13 Oksuz Kutular',         'primer_final/oksuz_kutular_20260802_TESLIM.tsv'),
    ('14 Plaka ve Jel',          'primer_final/plaka_ve_jel_20260802_TESLIM.tsv'),
    ('15 On Kararlar',           'primer_final/on_kararlar_20260802_TESLIM.tsv'),
]
# primer dizisi tasiyan sutun basliklari (kritik fark denetimi icin)
DIZI_BASLIK = ("Ileri primer", "Geri primer")


def sayfa_satirlari(ws):
    out = []
    for r in ws.iter_rows(values_only=True):
        out.append(['' if c is None else str(c) for c in r])
    while out and not any(x.strip() for x in out[-1]):
        out.pop()
    return out


def tsv_satirlari(yol):
    if not os.path.exists(yol):
        return None
    with open(yol, encoding='utf-8', newline='') as fh:
        return [list(r) for r in csv.reader(fh, delimiter='\t')]


def dizi_sutunlari(basliklar):
    ix = []
    for i, b in enumerate(basliklar):
        if any(b.startswith(k) for k in DIZI_BASLIK):
            ix.append(i)
    return ix


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--kok', required=True)
    ap.add_argument('--yaz', action='store_true')
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.xlsx, data_only=True)

    kritik, toplam_fark = [], 0
    for sayfa, rel in ESLEME:
        yol = os.path.join(a.kok, rel)
        if sayfa not in wb.sheetnames:
            print('SAYFA YOK, atlandi:', sayfa)
            continue
        yeni = sayfa_satirlari(wb[sayfa])
        eski = tsv_satirlari(yol)
        if eski is None:
            print('%-26s TSV YOK -> uretilecek (%d satir)' % (sayfa, len(yeni)))
            toplam_fark += len(yeni)
        else:
            dx = dizi_sutunlari(yeni[0] if yeni else [])
            n = 0
            for i in range(max(len(yeni), len(eski))):
                sy = yeni[i] if i < len(yeni) else []
                se = eski[i] if i < len(eski) else []
                for j in range(max(len(sy), len(se))):
                    vy = sy[j].strip() if j < len(sy) else ''
                    ve = se[j].strip() if j < len(se) else ''
                    if vy != ve:
                        n += 1
                        if j in dx and (vy or ve):
                            kritik.append((rel, i + 1, (yeni[0][j] if j < len(yeni[0]) else '?'),
                                           ve, vy))
            toplam_fark += n
            print('%-26s %-58s farkli hucre: %d' % (sayfa, os.path.basename(rel), n))
        if a.yaz:
            os.makedirs(os.path.dirname(yol) or '.', exist_ok=True)
            with open(yol, 'w', encoding='utf-8', newline='') as fh:
                w = csv.writer(fh, delimiter='\t')
                for r in yeni:
                    w.writerow(r)

    print('\nTOPLAM farkli hucre:', toplam_fark)
    print('\n=== KRITIK: PRIMER DIZISI FARKLARI ===')
    if not kritik:
        print('  yok')
    for rel, sat, bas, ve, vy in kritik:
        print('  %s satir %d | %s' % (os.path.basename(rel), sat, bas))
        print('     TSV (eski) : %s  (%d nt)' % (ve, len(ve)))
        print('     PANEL(dogru): %s  (%d nt)' % (vy, len(vy)))
    print('\n%s' % ('YAZILDI - TSV\'ler panelden yeniden uretildi.' if a.yaz
                    else 'YALNIZ RAPOR. Uretmek icin --yaz ekleyin.'))


if __name__ == '__main__':
    main()
