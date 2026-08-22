# -*- coding: utf-8 -*-
"""
tsv_sync.py REPRODUCES the DELIVERY TSV twins from the panel xlsx.

Why: because the TSVs are derived from the panel by hand they can drift. The
2026-08-02 audit found that the Metanojen_universal forward primer was still the OLD
21 nt sequence in the TSV (a G had been added in the panel and the coverage went
from 19 percent to 97). If the order is copied from the TSV, THE WRONG PRIMER IS
ORDERED.

It reports the DIFFERENCES first and only produces with --write. Every difference in a
primer sequence column is also marked CRITICAL.

Usage:
  python tsv_sync.py --xlsx ../PrimerJury_..._TESLIM.xlsx --root ..   (report only)
  python tsv_sync.py --xlsx ... --root .. --write                       (produce)

"""
# -------------------------------------------------------------------------
# tsv_sync.py reproduces the delivery TSV twins from the panel xlsx and reports
#                  every cell difference between them, and the primer sequence
#                  differences in particular, as CRITICAL.
#
# INPUT  : the delivery panel given with --xlsx (openpyxl) and the existing TSV
#          files under --root. The sheet to TSV path mapping (ESLEME) is fixed inside
#          the file.
# OUTPUT : with --write it rewrites the TSV files in the mapping from the panel;
#          without it, it changes nothing and only prints the differences.
# CALLED BY: IT IS NOT IN THE MENU, it is run by hand.
#
# WHY IT EXISTS: because the TSVs are derived from the panel by hand they can drift.
# The measured example is the Metanojen_universal forward primer still sitting in
# the TSV as the old 21 nt sequence; a G had been added in the panel and the
# coverage had gone from 19 percent to 97. If the order is copied from the TSV the
# wrong primer is ordered, and that is why differences in the primer sequence
# columns are listed SEPARATELY from the rest, under a CRITICAL heading.
# -------------------------------------------------------------------------
import os, sys, csv, argparse
import openpyxl

ESLEME = [
    ('1 Rapora Ozet',            'degerlendiriciya_ozet_20260802_TESLIM.tsv'),
    ('2 Panel',                  'primer_final/devir_ciftleri_20260802_sonrotus_TESLIM.tsv'),
    ('3 Triyaj (matris ilgisi)', 'primer_final/triyaj_20260802_TESLIM.tsv'),
    ('4 Degisiklikler',          'primer_final/degisiklikler_20260802_TESLIM.tsv'),
    ('5 Dusen ciftler',          'primer_final/devir_dusenler_20260802_sonrotus_TESLIM.tsv'),
    ('6 Karar Durumu',           'primer_final/karar_durumu_20260802_TESLIM.tsv'),
    ('7 Ayrilik Tablosu',        'primer_final/ayrilik_tablosu_20260802_TESLIM.tsv'),
    ('8 Geometri Denetimi',      'primer_final/geometri_denetimi_20260802_TESLIM.tsv'),
    ('9 Kurtarma ve Onarim',     'primer_final/kurtarma_ve_onarim_20260802_TESLIM.tsv'),
    ('10 Olcum Hatalari',        'primer_final/olcum_hatalari_20260802_TESLIM.tsv'),
    ('11 B-F Yeniden Olcum',     'primer_final/bf_yeniden_olcum_20260802_TESLIM.tsv'),
    ('13 Oksuz Kutular',         'primer_final/oksuz_kutular_20260802_TESLIM.tsv'),
    ('14 Plaka ve Jel',          'primer_final/plaka_ve_jel_20260802_TESLIM.tsv'),
    ('15 On Kararlar',           'primer_final/on_kararlar_20260802_TESLIM.tsv'),
]
# the column headers carrying a primer sequence (for the critical difference check)
DIZI_BASLIK = ("Ileri primer", "Geri primer")


def sayfa_satirlari(ws):
    out = []
    for r in ws.iter_rows(values_only=True):
        out.append(['' if c is None else str(c) for c in r])
    while out and not any(x.strip() for x in out[-1]):
        out.pop()
    return out


def tsv_satirlari(yol):
    if not os.path.exists(yol):
        return None
    with open(yol, encoding='utf-8', newline='') as fh:
        return [list(r) for r in csv.reader(fh, delimiter='\t')]


def dizi_sutunlari(basliklar):
    ix = []
    for i, b in enumerate(basliklar):
        if any(b.startswith(k) for k in DIZI_BASLIK):
            ix.append(i)
    return ix


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--root', dest='kok', required=True)
    ap.add_argument('--write', dest='yaz', action='store_true')
    a = ap.parse_args()
    wb = openpyxl.load_workbook(a.xlsx, data_only=True)

    kritik, toplam_fark = [], 0
    for sayfa, rel in ESLEME:
        yol = os.path.join(a.kok, rel)
        if sayfa not in wb.sheetnames:
            print(u'NO SHEET, skipped:', sayfa)
            continue
        yeni = sayfa_satirlari(wb[sayfa])
        eski = tsv_satirlari(yol)
        if eski is None:
            print(u'%-26s NO TSV -> will be generated (%d rows)' % (sayfa, len(yeni)))
            toplam_fark += len(yeni)
        else:
            dx = dizi_sutunlari(yeni[0] if yeni else [])
            n = 0
            for i in range(max(len(yeni), len(eski))):
                sy = yeni[i] if i < len(yeni) else []
                se = eski[i] if i < len(eski) else []
                for j in range(max(len(sy), len(se))):
                    vy = sy[j].strip() if j < len(sy) else ''
                    ve = se[j].strip() if j < len(se) else ''
                    if vy != ve:
                        n += 1
                        if j in dx and (vy or ve):
                            kritik.append((rel, i + 1, (yeni[0][j] if j < len(yeni[0]) else '?'),
                                           ve, vy))
            toplam_fark += n
            print(u'%-26s %-58s differing cells: %d' % (sayfa, os.path.basename(rel), n))
        if a.yaz:
            os.makedirs(os.path.dirname(yol) or '.', exist_ok=True)
            with open(yol, 'w', encoding='utf-8', newline='') as fh:
                w = csv.writer(fh, delimiter='\t')
                for r in yeni:
                    w.writerow(r)

    print(u'\nTOTAL differing cells:', toplam_fark)
    print(u'\n=== CRITICAL: DIFFERENCES IN THE PRIMER SEQUENCES ===')
    if not kritik:
        print(u'  none')
    for rel, sat, bas, ve, vy in kritik:
        print(u'  %s row %d | %s' % (os.path.basename(rel), sat, bas))
        print('     TSV (eski) : %s  (%d nt)' % (ve, len(ve)))
        print('     PANEL(dogru): %s  (%d nt)' % (vy, len(vy)))
    print('\n%s' % (u'WRITTEN - the TSVs were regenerated from the panel.' if a.yaz
                    else u'REPORT ONLY. Add --write to generate them.'))


if __name__ == '__main__':
    main()
