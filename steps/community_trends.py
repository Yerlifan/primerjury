#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
community_trends.py
THE COMMUNITY TREND WORKBOOK, reproduced from the raw Bracken output.

Why again: the previous workbook presented the species level counts with the same
weight as the genus level ones. But two findings we measured make the species level
unreadable for some taxa:
  1. Indistinguishable bins: the sequences of two species bins do not separate from
     one another (ayirt_edilemez.tsv, measured).
  2. Bin identity: a bin's raw reads do not prefer the species it was assigned to,
     they go to another reference (kimlik_*.tsv, measured).
Both files are measurements; this script DERIVES the reliability mark from them and
keeps no hand written list. If the files are not given, no mark is placed and that
is written plainly on the cover sheet.

The group and year mapping is built from the barcode NUMBER rather than from the
directory name: F1 and F2 had been swapped in the source study's directory names.

Usage:
  python3 community_trends.py --bracken "bracken results"       --distinguishable primer_candidates/ayirt_edilemez.tsv       --identity t_kimlik/kimlik_A.tsv t_kimlik/kimlik_B.tsv       --names taxid_names.tsv --out Topluluk_Trend.xlsx

"""
import argparse, csv, datetime, glob, math, os, re, sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

YAZI = "Arial"
BASLIK_DOLGU = PatternFill("solid", fgColor="1F4E78")
BASLIK_YAZI = Font(name=YAZI, bold=True, color="FFFFFF", size=10)
NORMAL = Font(name=YAZI, size=10)
KALIN = Font(name=YAZI, bold=True, size=10)
BUYUK = Font(name=YAZI, bold=True, size=14, color="1F4E78")
INCE = Side(style="thin", color="BFBFBF")
KENAR = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)
YESIL = PatternFill("solid", fgColor="C6EFCE")
SARI = PatternFill("solid", fgColor="FFEB9C")
KIRMIZI = PatternFill("solid", fgColor="FFC7CE")

# Otoriter esleme: ornekleme haritasi. Barkod numarasi belirleyicidir.
GRUP_ARALIK = [("A1", 1, "Arke, kisa amplikon"), ("A2", 5, "Arke, uzun amplikon"),
               ("F2", 9, "Mantar, uzun amplikon"), ("F1", 13, "Mantar, kisa amplikon"),
               ("B", 17, "Bakteri")]
YILLAR = (2021, 2023, 2024, 2025)
BARKOD_GRUP, BARKOD_YIL, GRUP_ACIKLAMA = {}, {}, {}
for _g, _b0, _ac in GRUP_ARALIK:
    GRUP_ACIKLAMA[_g] = _ac
    for _i, _y in enumerate(YILLAR):
        BARKOD_GRUP[_b0 + _i] = _g
        BARKOD_YIL[_b0 + _i] = _y
GRUPLAR = [g for g, _, _ in GRUP_ARALIK]
RUTBE_ADI = {"D": 'domain', "K": 'kingdom', "P": 'phylum', "C": 'class',
             "O": 'order', "F": 'family', "G": 'genus', "S": 'species'}


def get_args():
    p = argparse.ArgumentParser()
    p.add_argument("--bracken", required=True, help="'bracken results' directory")
    p.add_argument("--distinguishable", default=None, help="ayirt_edilemez.tsv")
    p.add_argument("--identity", nargs="*", default=[], help='the identity tables produced '
                                                             'by the identity verification')
    p.add_argument("--names", default=None, help="taxid_names.tsv")
    p.add_argument("--out", required=True)
    p.add_argument("--top", type=int, default=10, help="taxon shown on the page")
    p.add_argument("--rank", default=None,
                   help='the output directory of abundance_rank.py; when it '
                        'is given, the rank coverage and reliable abundance '
                        'sheets are added')
    p.add_argument("--identity-threshold", type=float, default=50.0,
                   help="bin identity is suspect when the dominant fraction is below this value")
    return p.parse_args()


# ------------------------------------------------------------------ veri
def bracken_oku(kok, duzey):
    'Returns: {barcode: {taxon: per cent}}'
    out = {}
    desen = os.path.join(kok, duzey, "*", "*_bracken_%s.txt" % duzey)
    for p in sorted(glob.glob(desen)):
        if p.endswith("_report.txt"):
            continue
        m = re.search(r"barcode(\d+)", os.path.basename(p))
        if not m:
            continue
        bc = int(m.group(1))
        if bc not in BARKOD_GRUP:
            continue
        d = {}
        with open(p, encoding="utf-8") as fh:
            basliklar = fh.readline().rstrip("\n").split("\t")
            try:
                i_ad = basliklar.index("name")
                i_or = basliklar.index("fraction_total_reads")
            except ValueError:
                continue
            for line in fh:
                q = line.rstrip("\n").split("\t")
                if len(q) > max(i_ad, i_or):
                    d[q[i_ad]] = float(q[i_or]) * 100.0
        out[bc] = d
    return out


def shannon(v):
    s = sum(v)
    if s <= 0:
        return 0.0
    return -sum((x / s) * math.log(x / s) for x in v if x > 0)


def simpson(v):
    s = sum(v)
    if s <= 0:
        return 0.0
    return 1.0 - sum((x / s) ** 2 for x in v if x > 0)


def bray_curtis(a, b):
    tk = set(a) | set(b)
    pay = sum(abs(a.get(t, 0.0) - b.get(t, 0.0)) for t in tk)
    payda = sum(a.get(t, 0.0) for t in tk) + sum(b.get(t, 0.0) for t in tk)
    return pay / payda if payda else 0.0


def cins_adi(tur_adi):
    """Extracts the genus name from a species name. 'Candidatus' and its abbreviation
    'Ca.' are not genus names; they are skipped, otherwise every Candidatus taxon is
    gathered into a single 'Ca.' genus.

    """
    if not tur_adi:
        return ""
    p = tur_adi.split()
    i = 0
    while i < len(p) and p[i].rstrip(".").lower() in ("candidatus", "ca"):
        i += 1
    return p[i] if i < len(p) else ""


def guvenilirlik_kur(a):
    """Derives the suspect species and genus sets from the measured files.
        Returns: (suspect species -> reason, suspect genus -> reason)"""
    ad = {}
    if a.names and os.path.exists(a.names):
        for l in open(a.names, encoding="utf-8"):
            q = l.rstrip("\n").split("\t")
            if len(q) > 1:
                ad[q[0]] = q[1]
    tur, cins = {}, {}

    def ekle(sozluk, anahtar, metin):
        if not anahtar:
            return
        sozluk.setdefault(anahtar, [])
        if metin not in sozluk[anahtar]:
            sozluk[anahtar].append(metin)

    if a.distinguishable and os.path.exists(a.distinguishable):
        for r in csv.DictReader(open(a.distinguishable, encoding="utf-8"), delimiter="\t"):
            t1, t2 = r.get("taxid1", ""), r.get("taxid2", "")
            a1, a2 = ad.get(t1, t1), ad.get(t2, t2)
            oz = r.get("kati_ozdeslik", r.get("ozdeslik_yuzde", ""))
            g = ('does not separate from %s at sequence level (strict '
                 'identity %%%s, class %s)' % (a2, oz, r.get("sinif", "")))
            ekle(tur, a1, g)
            ekle(tur, a2, (u'it does not separate from %s at sequence level (strict identity %%%s, class %s)' % (a1, oz, r.get("sinif", ""))))
            for x in (a1, a2):
                ekle(cins, cins_adi(x),
                     u'there is an indistinguishable species pair in this genus')

    for yol in a.identity:
        if not os.path.exists(yol):
            continue
        for r in csv.DictReader(open(yol, encoding="utf-8"), delimiter="\t"):
            m = re.search(r"reads[-_](\d+)\.fastq", r.get("dosya", ""))
            if not m:
                continue
            tx = m.group(1)
            adi = ad.get(tx, tx)
            try:
                oran = float(r.get("baskin_oran", "0"))
            except ValueError:
                continue
            bref = r.get("baskin_referans", "")
            kendi = cins_adi(adi)
            # TWO SEPARATE SIGNALS, reported separately:
            #  (a) the reads go to another GENUS   -> the bin identity is wrong
            #  (b) no reference reaches a majority -> the bin is undefined
            yanlis_cins = bool(kendi) and kendi not in bref
            cogunluk_yok = oran < a.identity_threshold
            if yanlis_cins:
                ekle(tur, adi,
                     u'the bin\'s raw reads go to another genus (in bin %s the dominant reference is %s, %%%.1f)'
                     % (r.get("grup", ""), bref[:52], oran))
                ekle(cins, kendi,
                     u'in the bins of this genus the reads go to other genera')
            elif cogunluk_yok:
                ekle(tur, adi,
                     u'no reference takes the majority of the reads, so the bin is undefined '
                     u'at species level (bin %s, the best reference %s, only %%%.1f)'
                     % (r.get("grup", ""), bref[:44], oran))
                ekle(cins, kendi,
                     u'in the bins of this genus no reference reaches a majority')
    return tur, cins


# ------------------------------------------------------------------ yazim
def yaz_baslik(ws, basliklar, satir=1, dondur=True):
    for j, b in enumerate(basliklar, 1):
        c = ws.cell(row=satir, column=j, value=b)
        c.font = BASLIK_YAZI
        c.fill = BASLIK_DOLGU
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = KENAR
    if dondur:
        ws.freeze_panes = ws.cell(row=satir + 1, column=1)


def genislik(ws, gs):
    for j, g in enumerate(gs, 1):
        ws.column_dimensions[get_column_letter(j)].width = g


def bolum_yaz(ws, satir, veri, ust, supheli, duzey):
    """Writes a taxon by year table for one group. Returns: the next row."""
    basliklar = ["Takson"] + ["%d" % y for y in YILLAR] + ['The four year mean']
    if duzey == "tur":
        basliklar.append(u'Reliability')
    yaz_baslik(ws, basliklar, satir=satir, dondur=False)
    ilk = satir + 1
    for i, (tk, vals) in enumerate(veri[:ust]):
        r = ilk + i
        c = ws.cell(row=r, column=1, value=tk); c.font = NORMAL; c.border = KENAR
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + j, value=round(v, 4))
            c.font = NORMAL; c.number_format = "0.00"; c.border = KENAR
        c = ws.cell(row=r, column=6,
                    value="=AVERAGE(%s%d:%s%d)" % ("B", r, "E", r))
        c.font = NORMAL; c.number_format = "0.00"; c.border = KENAR
        if duzey == "tur":
            g = supheli.get(tk)
            c = ws.cell(row=r, column=7,
                        value=('READ AT GENUS LEVEL:' + "; ".join(g)) if g
                        else 'the measurement did not show this species as '
                             'suspect')
            c.font = NORMAL; c.border = KENAR
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.fill = KIRMIZI if g else YESIL
    son = ilk + min(len(veri), ust) - 1
    # the total check: the taxa shown plus the others
    r = son + 1
    c = ws.cell(row=r, column=1, value='The total of what is shown')
    c.font = KALIN; c.border = KENAR
    for j in range(4):
        col = get_column_letter(2 + j)
        c = ws.cell(row=r, column=2 + j,
                    value="=SUM(%s%d:%s%d)" % (col, ilk, col, son))
        c.font = KALIN; c.number_format = "0.00"; c.border = KENAR
    return r + 2, ilk, son


def main():
    a = get_args()
    cins = bracken_oku(a.bracken, "genus")
    tur = bracken_oku(a.bracken, "species")
    if not cins:
        sys.exit(u'no bracken file at genus level was found: %s' % a.bracken)
    print("cins duzeyi barkod: %d, tur duzeyi barkod: %d" % (len(cins), len(tur)))
    eksik = [b for b in sorted(BARKOD_GRUP) if b not in cins]
    if eksik:
        print(u'WARNING: a barcode missing at genus level: %s' % eksik)

    s_tur, s_cins = guvenilirlik_kur(a)
    print(u'suspect species from the measured files: %d, suspect genera: %d'
          % (len(s_tur), len(s_cins)))

    wb = Workbook()

    # ---------------- Kapak ----------------
    ws = wb.active; ws.title = 'Cover and method'
    genislik(ws, [30, 104])
    ws["A1"] = "PrimerJury topluluk trend analizi"
    ws["A1"].font = BUYUK; ws.merge_cells("A1:B1")
    satirlar = [
        ('Produced at', datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Kaynak", os.path.abspath(a.bracken)),
        ("", ""),
        ("OKUMA SIRASI", ""),
        ("0", 'Read the rank coverage sheet first: it says at which rank the '
              'abundance of each sample can be read. Measured: in the '
              'archaeal samples 86 to 97 per cent of the reads reach genus '
              'level, in the fungal samples 0.1 to 2.5 per cent, and in the '
              'bacterial samples 8 to 10 per cent. The abundance numbers have '
              'to be taken from the reliable abundance sheet.'),
        ("0b", 'Bracken WAS NOT RUN. Bracken pushes the reads that stayed at '
               "a higher rank down according to the database's own priors, "
               'which rests on the assumption that the real organism is in '
               'the database, and on the fungal and bacterial side of this '
               'sample that assumption does not hold.'),
        ("1", 'The dominant genus and dominant species sheets are the '
              'ORIGINAL Bracken output with no confidence threshold applied; '
              'they are kept for comparison.'),
        ("2", 'The species level sheets can be read at species level only on '
              'the rows whose reliability column is green. Red rows have to '
              'be read gathered at genus level.'),
        ("", ""),
        ("NEDEN", ""),
        ('Bins that cannot be told apart',
         'The sequences of some species bins do not separate from one '
         'another, so how many reads fell into which bin does not rest on '
         'sequence evidence. The source: ayirt_edilemez.tsv'),
        ('Bin identity',
         'The raw reads of some bins do not prefer the species they were '
         'assigned to and go to another reference. At its default setting '
         'Kraken2 does not abstain: when the real species is not in the '
         'database, the read falls onto the nearest sibling species. The '
         'source: the identity tables.'),
        ('This mark was not placed by hand',
         'The list of suspect species and genera is derived from the two '
         'MEASUREMENT files above. Without them no mark is placed, and that '
         'is stated on this row.'),
        ("", ""),
        ('The sample mapping', 'The group and the year are built FROM THE '
                                 'BARCODE NUMBER, not from the directory '
                                 'name.'),
    ]
    r = 3
    for k, v in satirlar:
        ws.cell(row=r, column=1, value=k).font = KALIN
        c = ws.cell(row=r, column=2, value=v)
        c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
        if len(v) > 90:
            ws.row_dimensions[r].height = 15 * (1 + len(v) // 90)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Grup").font = BASLIK_YAZI
    ws.cell(row=r, column=1).fill = BASLIK_DOLGU
    ws.cell(row=r, column=2, value='Barcodes and years').font = BASLIK_YAZI
    ws.cell(row=r, column=2).fill = BASLIK_DOLGU
    r += 1
    for g, b0, ac in GRUP_ARALIK:
        ws.cell(row=r, column=1, value="%s  %s" % (g, ac)).font = NORMAL
        ws.cell(row=r, column=2, value=", ".join(
            "barcode%02d=%d" % (b0 + i, YILLAR[i]) for i in range(4))).font = NORMAL
        r += 1
    ws.cell(row=r + 1, column=1, value='The number of source files').font = KALIN
    ws.cell(row=r + 1, column=2,
            value='genus %d barcodes, species %d barcodes' % (len(cins), len(tur))).font = NORMAL
    ws.cell(row=r + 2, column=1, value='The suspect mark').font = KALIN
    ws.cell(row=r + 2, column=2,
            value=('%d species and %d genera were marked' % (len(s_tur), len(s_cins)))
            if (s_tur or s_cins) else
            'NO MEASUREMENT FILE WAS GIVEN, so species level could not be '
            'marked').font = NORMAL
    if not (s_tur or s_cins):
        ws.cell(row=r + 2, column=2).fill = SARI

    # ---------------- Metrik sozlugu ----------------
    ws = wb.create_sheet('The metric glossary')
    genislik(ws, [26, 100])
    yaz_baslik(ws, ["Metrik", 'What it measures and how to read it'])
    for i, (k, v) in enumerate([
        ("Bolluk (%)", "Bracken's fraction_total_reads column turned into a "
                       'percentage. It is not a read count but the fraction '
                       'of the reads in that sample.'),
        ("Shannon", 'It measures how many taxa there are and how evenly they '
                    'are spread at once. The larger it is the more diverse. '
                    'It stays low when a few taxa dominate.'),
        ("Zenginlik", 'The number of taxa with an abundance above zero. It '
                      'takes no account of evenness and only counts how many '
                      'there are.'),
        ("Simpson (1-D)", 'The probability that two random reads belong to '
                          'DIFFERENT taxa. Near zero the community is in the '
                          'hands of one taxon; near one it is even.'),
        ("Bray-Curtis", 'The dissimilarity between two samples. 0 is '
                        'identical and 1 is completely separate. It gives the '
                        'size of the change between years.'),
        ('Reliability', 'Whether this row can be read at species level. It '
                          'is derived from the measurement files and is not '
                          'placed by hand.'),
    ], start=2):
        ws.cell(row=i, column=1, value=k).font = KALIN
        c = ws.cell(row=i, column=2, value=v)
        c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30

    # ---------------- Baskin cins / tur ----------------
    grafik_yeri = {}
    for duzey, veri, ad_sayfa in (('genus', cins, 'Dominant genus'),
                                  ("tur", tur, 'Dominant species')):
        if not veri:
            continue
        ws = wb.create_sheet(ad_sayfa)
        genislik(ws, [46, 12, 12, 12, 12, 18] + ([74] if duzey == "tur" else []))
        r = 1
        # These sheets come from the original Bracken output; the confidence
        # threshold WAS NOT APPLIED. If the rank coverage was measured, which
        # sheet to look at is written plainly; otherwise two different genus
        # tables sit side by side and the reader cannot tell which one holds.
        if a.rank and os.path.exists(os.path.join(a.rank, "ozet.tsv")):
            c = ws.cell(row=1, column=1,
                        value='THIS SHEET COMES FROM Bracken output WITH NO '
                              'CONFIDENCE THRESHOLD applied and is kept for '
                              'comparison. To read abundance, use the '
                              'reliable abundance sheet; which sample can be '
                              'read at which rank is on the rank coverage '
                              'sheet. Most of the species and genus '
                              'assignments on this sheet do not stay standing '
                              'once a confidence threshold is applied.')
            c.font = Font(name=YAZI, bold=True, size=10, color="9C0006")
            c.fill = KIRMIZI
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1,
                           end_column=7 if duzey == "tur" else 6)
            ws.row_dimensions[1].height = 48
            r = 3
        if duzey == "tur":
            c = ws.cell(row=r, column=1,
                        value='CAREFUL: the rows marked red must not be read '
                              'at species level; the bin of that taxon was '
                              'found suspect by measurement. The reason is in '
                              'the reliability column. The genus level sheet '
                              'is the one to go by.')
            c.font = Font(name=YAZI, bold=True, size=10, color="9C0006")
            c.fill = KIRMIZI
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 34
            r += 2
        for g in GRUPLAR:
            bcs = [b for b in sorted(BARKOD_GRUP) if BARKOD_GRUP[b] == g
                   and b in veri]
            if not bcs:
                continue
            c = ws.cell(row=r, column=1,
                        value="%s  %s  (%s)" % (g, GRUP_ACIKLAMA[g],
                                                ", ".join("barcode%02d" % b for b in bcs)))
            c.font = BUYUK
            r += 1
            tumu = set()
            for b in bcs:
                tumu |= set(veri[b])
            sirali = sorted(tumu,
                            key=lambda t: -sum(veri[b].get(t, 0.0) for b in bcs))
            tablo = [(t, [veri[b].get(t, 0.0) for b in bcs]) for t in sirali]
            supheli = s_tur if duzey == "tur" else s_cins
            r, ilk, son = bolum_yaz(ws, r, tablo, a.top, supheli, duzey)
            grafik_yeri.setdefault(ad_sayfa, []).append((g, ilk, son))
        # grafikler
        for g, ilk, son in grafik_yeri.get(ad_sayfa, []):
            ch = BarChart(); ch.type = "col"; ch.grouping = "clustered"
            ch.title = '%s  the dominant %s, by year' % (g, duzey)
            ch.y_axis.title = "bolluk (%)"; ch.x_axis.title = "takson"
            veri_ref = Reference(ws, min_col=2, max_col=5, min_row=ilk - 1,
                                 max_row=son)
            kat = Reference(ws, min_col=1, min_row=ilk, max_row=son)
            ch.add_data(veri_ref, titles_from_data=True)
            ch.set_categories(kat)
            ch.height, ch.width = 8, 24
            ws.add_chart(ch, "I%d" % ilk)

    # ---------------- Alfa cesitlilik ----------------
    for duzey, veri, ad_sayfa in (('genus', cins, 'Alpha diversity, genus'),
                                  ("tur", tur, 'Alpha diversity, species')):
        if not veri:
            continue
        ws = wb.create_sheet(ad_sayfa)
        genislik(ws, [10, 10, 10, 14, 14, 14])
        yaz_baslik(ws, ["Grup", 'Year', "Barkod", "Shannon", "Zenginlik",
                        "Simpson (1-D)"])
        r = 2
        for g in GRUPLAR:
            for b in sorted(BARKOD_GRUP):
                if BARKOD_GRUP[b] != g or b not in veri:
                    continue
                v = [x for x in veri[b].values() if x > 0]
                for j, val in enumerate([g, BARKOD_YIL[b], "barcode%02d" % b,
                                         round(shannon(v), 4), len(v),
                                         round(simpson(v), 4)], start=1):
                    c = ws.cell(row=r, column=j, value=val)
                    c.font = NORMAL; c.border = KENAR
                    if j in (4, 6):
                        c.number_format = "0.000"
                r += 1
        ch = LineChart()
        ch.title = 'Shannon diversity by year (at %s level)' % duzey
        ch.y_axis.title = "Shannon"; ch.x_axis.title = 'sample'
        ch.add_data(Reference(ws, min_col=4, min_row=1, max_row=r - 1),
                    titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=3, min_row=2, max_row=r - 1))
        ch.height, ch.width = 9, 26
        ws.add_chart(ch, "H2")

    # ---------------- Bray-Curtis ----------------
    for duzey, veri, ad_sayfa in (('genus', cins, "Bray-Curtis Cins"),
                                  ("tur", tur, 'Bray-Curtis, species')):
        if not veri:
            continue
        ws = wb.create_sheet(ad_sayfa)
        genislik(ws, [10, 14, 14, 14])
        c = ws.cell(row=1, column=1,
                    value='The Bray-Curtis dissimilarity between consecutive '
                          'years inside each group. 0 is identical and 1 is '
                          'completely separate.')
        c.font = NORMAL
        ws.merge_cells("A1:D1")
        yaz_baslik(ws, ["Grup", 'The comparison', "Bray-Curtis", "Yorum"], satir=3)
        r = 4
        for g in GRUPLAR:
            bcs = [b for b in sorted(BARKOD_GRUP) if BARKOD_GRUP[b] == g
                   and b in veri]
            for i in range(len(bcs) - 1):
                d = bray_curtis(veri[bcs[i]], veri[bcs[i + 1]])
                yorum = ('a small change' if d < 0.3 else
                         'a moderate change' if d < 0.6 else 'a large change')
                for j, val in enumerate([g, "%d - %d" % (BARKOD_YIL[bcs[i]],
                                                         BARKOD_YIL[bcs[i + 1]]),
                                         round(d, 4), yorum], start=1):
                    c = ws.cell(row=r, column=j, value=val)
                    c.font = NORMAL; c.border = KENAR
                    if j == 3:
                        c.number_format = "0.000"
                        c.fill = YESIL if d < 0.3 else (SARI if d < 0.6 else KIRMIZI)
                r += 1

    # ---------------- Rank coverage and reliable abundance ----------------
    # Bracken WAS NOT RUN. The reason was measured: after the confidence
    # correction, 86 to 97 percent of the archaeal reads can reach genus
    # level and 0.1 to 1.5 percent of the fungal reads. Bracken distributes
    # what stays at an upper rank downward by the database's priors; if the
    # real organism is not in the database that is manufacturing numbers
    # rather than measuring.
    rk = os.path.join(a.rank, "rutbe_kapsamasi.tsv") if a.rank else None
    ro = os.path.join(a.rank, "ozet.tsv") if a.rank else None
    rb = os.path.join(a.rank, "bolluk.tsv") if a.rank else None
    if rk and os.path.exists(ro):
        ozetler = list(csv.DictReader(open(ro, encoding="utf-8"), delimiter="\t"))
        ws = wb.create_sheet('Rank coverage')
        genislik(ws, [12, 8, 8, 14, 14, 14, 18, 16, 14, 14])
        c = ws.cell(row=1, column=1,
                    value='At which rank the abundance can be read. The genus '
                          'fraction shows what percentage of the reads in '
                          'that sample could settle at genus level or '
                          'narrower. A low fraction means the organisms of '
                          'that group are not represented in the Kraken2 '
                          'database. That is why Bracken was not run: pushing '
                          'what stayed at a higher rank down to genus would '
                          'have been inventing numbers.')
        c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A1:J1"); ws.row_dimensions[1].height = 46
        yaz_baslik(ws, ["Barkod", "Grup", 'Year', "Toplam okuma", 'Classified',
                        'Unclassified', 'The rank chosen', 'Settling at that rank',
                        'Genus fraction, per cent', 'Species fraction, per cent'], satir=3)
        r = 4
        for x in ozetler:
            vals = [x["barkod"], x["grup"], int(x["yil"]), int(x["toplam_okuma"]),
                    int(x["siniflanan"]), int(x["sinifsiz"]),
                    x["secilen_rutbe_adi"], int(x["bu_rutbede_yerlesen"]),
                    float(x["cins_orani"]), float(x["tur_orani"])]
            for j, v in enumerate(vals, 1):
                cc = ws.cell(row=r, column=j, value=v)
                cc.font = NORMAL; cc.border = KENAR
                if j in (9, 10):
                    cc.number_format = "0.0"
            co = float(x["cins_orani"])
            ws.cell(row=r, column=9).fill = (YESIL if co >= 70 else
                                             SARI if co >= 30 else KIRMIZI)
            ws.cell(row=r, column=7).fill = (YESIL if x["secilen_rutbe"] in ("G", "S")
                                             else SARI if x["secilen_rutbe"] == "F"
                                             else KIRMIZI)
            r += 1
        if rb and os.path.exists(rb):
            ws = wb.create_sheet('Reliable abundance')
            genislik(ws, [12, 8, 8, 12, 52, 12, 14, 12])
            c = ws.cell(row=1, column=1,
                        value='Every sample is given at the narrowest rank '
                              'the data supports in that sample. The rank was '
                              'not chosen by hand; the narrowest rank where '
                              'more than half of the classified reads could '
                              'settle was taken. The reads that could not '
                              'settle are not hidden and stand on their own '
                              'row.')
            c.font = NORMAL; c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells("A1:H1"); ws.row_dimensions[1].height = 34
            yaz_baslik(ws, ["Barkod", "Grup", 'Year', 'Rank', "Takson",
                            "taxid", "Okuma", 'Per cent'], satir=3)
            r = 4
            for x in csv.DictReader(open(rb, encoding="utf-8"), delimiter="\t"):
                vals = [x["barkod"], x["grup"], int(x["yil"]),
                        RUTBE_ADI.get(x["rutbe"], x["rutbe"]), x["takson"],
                        x["taxid"], int(x["okuma"]), float(x["yuzde"])]
                for j, v in enumerate(vals, 1):
                    cc = ws.cell(row=r, column=j, value=v)
                    cc.font = NORMAL; cc.border = KENAR
                    if j == 8:
                        cc.number_format = "0.00"
                if x["takson"].startswith("["):
                    for j in range(1, 9):
                        ws.cell(row=r, column=j).fill = SARI
                r += 1
    elif a.rank:
        print(u'WARNING: --rank was given but there is no ozet.tsv: %s' % ro)

    # ---------------- Guvenilirlik kaniti ----------------
    ws = wb.create_sheet('Species level reliability')
    genislik(ws, [40, 16, 96])
    c = ws.cell(row=1, column=1,
                value='Every mark below is derived from the measurement '
                      'files; there is no hand written list.')
    c.font = NORMAL; ws.merge_cells("A1:C1")
    yaz_baslik(ws, ["Takson", 'Level', 'The measured reason'], satir=3)
    r = 4
    for tk in sorted(s_tur):
        for g in s_tur[tk]:
            for j, v in enumerate([tk, 'species', g], start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = NORMAL; c.border = KENAR
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=1).fill = KIRMIZI
            r += 1
    for tk in sorted(s_cins):
        for g in s_cins[tk]:
            for j, v in enumerate([tk, 'genus', g], start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = NORMAL; c.border = KENAR
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=1).fill = SARI
            r += 1
    if r == 4:
        ws.cell(row=4, column=1,
                value='No measurement file was given, so no mark could be '
                      'placed.').font = NORMAL
        ws.cell(row=4, column=1).fill = SARI

    d = os.path.dirname(os.path.abspath(a.out))
    if d:
        os.makedirs(d, exist_ok=True)
    wb.save(a.out)
    print(u'written: %s' % a.out)
    print("sayfa: %s" % ", ".join(w.title for w in wb.worksheets))
    print(u'suspect species marks: %d, suspect genus marks: %d'
          % (len(s_tur), len(s_cins)))


if __name__ == "__main__":
    main()
