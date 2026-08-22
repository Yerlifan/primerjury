#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_excel.py
Primer tasarımının bütün çıktılarını tek bir Excel dosyasında toplar.

Girdi klasörleri:
  --aday   08'in çıktısı (ozet.tsv, ayirt_edilemez.tsv, *.log)
  --final  09'un çıktısı (primer_final.tsv)
  --bol    05'in çıktısı (isteğe bağlı, *_bolme.json)
  --adlar  taxid ad tablosu
"""
import argparse, csv, datetime, glob, json, os, re, sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import field_audit

YAZI = "Arial"
BASLIK_DOLGU = PatternFill("solid", fgColor="1F4E78")
BASLIK_YAZI = Font(name=YAZI, bold=True, color="FFFFFF", size=10)
NORMAL = Font(name=YAZI, size=10)
KALIN = Font(name=YAZI, bold=True, size=10)
BUYUK = Font(name=YAZI, bold=True, size=14, color="1F4E78")
INCE = Side(style="thin", color="BFBFBF")
KENAR = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)
YESIL = PatternFill("solid", fgColor="C6EFCE")
SARI = PatternFill("solid", fgColor="FFEB9C")
KIRMIZI = PatternFill("solid", fgColor="FFC7CE")


def tsv(p):
    if not p or not os.path.exists(p):
        return []
    with open(p, encoding="utf-8") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def yaz_baslik(ws, basliklar, satir=1):
    for j, b in enumerate(basliklar, 1):
        c = ws.cell(row=satir, column=j, value=b)
        c.font = BASLIK_YAZI
        c.fill = BASLIK_DOLGU
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = KENAR
    ws.freeze_panes = ws.cell(row=satir + 1, column=1)


def genislik(ws, genislikler):
    for j, g in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(j)].width = g


def get_args():
    p = argparse.ArgumentParser()
    p.add_argument("--aday", required=True)
    p.add_argument("--final", required=True)
    p.add_argument("--bol", default=None)
    p.add_argument("--dis", default=None,
                   help="output of the external-databases step (dis_veritabani.tsv)")
    p.add_argument("--referans", default=None,
                   help="output of the design-from-reference step (primer_referans.tsv)")
    p.add_argument("--adlar", default=None)
    p.add_argument("--hedefler", default="hedefler.tsv")
    p.add_argument("--kimlik", default=None,
                   help="output of the target-identity step (hedef_kimlik.tsv); if given, every "
                        "satira olculen kimlik sutunu eklenir")
    p.add_argument("--kons", default=None,
                   help="baskin alel consensus directory; alan tutarliligi "
                        "denetimi icin gerekir")
    p.add_argument("--out", required=True)
    p.add_argument("--not-metni", default="")
    return p.parse_args()


def main():
    a = get_args()
    ad = {}
    if a.adlar and os.path.exists(a.adlar):
        for l in open(a.adlar, encoding="utf-8"):
            q = l.rstrip("\n").split("\t")
            if len(q) > 1:
                ad[q[0]] = q[1]
    # Girdi eksikse sessizce bos ama gecerli gorunen bir Excel uretmek
    # yerine acikca durulur; bos kitap "hicbir aday gecmedi" gibi okunuyordu.
    eksik = [y for y in (os.path.join(a.aday, "ozet.tsv"),
                         os.path.join(a.final, "primer_final.tsv"))
             if not os.path.exists(y)]
    if eksik:
        sys.exit(u'The input file was not found, so no Excel was produced:\n   %s'
                 % "\n   ".join(eksik))
    ozet = tsv(os.path.join(a.aday, "ozet.tsv"))
    ayirt = tsv(os.path.join(a.aday, "ayirt_edilemez.tsv"))
    final = tsv(os.path.join(a.final, "primer_final.tsv"))
    hedefler = tsv(a.hedefler) if os.path.exists(a.hedefler) else []
    ref_ham = tsv(a.referans) if a.referans else []
    dis_yolu = a.dis or os.path.join(a.final, "dis_veritabani.tsv")
    dis_ham = tsv(dis_yolu)
    # (hedef, sinif, ileri, geri) -> toplam hedef disi urun
    dis = {}
    dis_detay = {}
    for x in dis_ham:
        k = (x["hedef"], x["sinif"], x["ileri_dizi"], x["geri_dizi"])
        dis[k] = dis.get(k, 0) + int(x.get("hedef_disi_urun", 0) or 0)
        if x.get("ornekler"):
            dis_detay.setdefault(k, []).append("%s: %s"
                                               % (x["veritabani"], x["ornekler"]))

    # Olculen kimlik: hedef adi Kraken2 atamasindan gelir, dizinin gercekte
    # ne oldugu ayri bir olcumdur. Ikisi ayrilabilir; tabloda yan yana
    # durmazsa okuyan, adin dizi kanitiyla dogrulandigini sanir.
    kimlik = {}
    kyol = a.kimlik or os.path.join(a.final, "hedef_kimlik.tsv")
    for r in tsv(kyol):
        kimlik[r["hedef"]] = r
    if kimlik:
        print(u'measured identity table read: %s (%d targets)' % (kyol, len(kimlik)))
    else:
        print(u'WARNING: there is no measured identity table (%s), so the identity column will stay empty'
              % kyol)

    UYUM_METIN = {
        "tur_uyusuyor": "tür düzeyinde doğrulandı",
        "cins_uyusuyor_tur_farkli": "cins doğru, tür farklı",
        "CINS_FARKLI": "CİNS FARKLI",
        "YAKIN_AKRABA_YOK": "veritabanında yakın akrabası yok",
        "vurus_yok": "veritabanında vuruş yok",
    }

    def kimlik_destek(hedef):
        """Doner: (kimlik, destek, kutu, cogunluk_var_mi)

        Grup hedeflerinde tek bir 'olculen kimlik' yaniltici olabilir: en
        yaygin ad yalnizca COGUNLUK degil COKLUK olabilir. Olculdu:
        Asetoklastik metanojenler hedefinin 19 kutusundan yalnizca 6'si
        Methanothrix soehngenii'ye gidiyor, gerisi Methanosarcina turlerine
        dagiliyor. Bu yuzden destek orani her zaman gosterilir ve yarinin
        altindaysa 'belirgin cogunluk yok' denir."""
        r = kimlik.get(hedef)
        if not r:
            return None, 0, 0, False
        try:
            d = int(r.get("destekleyen_kutu", 0))
            n = int(r.get("kutu_sayisi", 0))
        except (TypeError, ValueError):
            d = n = 0
        return r, d, n, (n > 0 and d * 2 > n)

    def kimlik_hucre(hedef):
        r, d, n, cog = kimlik_destek(hedef)
        if not r:
            return "", "", None
        u = r.get("uyum", "")
        ad_metni = r.get("olculen_kimlik", "")
        if n:
            ad_metni = "%s (%d/%d kutu)" % (ad_metni, d, n)
        if not cog:
            ad_metni += "  [belirgin çoğunluk yok]"
            if r.get("diger"):
                ad_metni += "  diğerleri: " + r["diger"][:70]
            return ad_metni, "kutular tek kimlikte birleşmiyor", SARI
        dolgu = (YESIL if u == "tur_uyusuyor" else
                 SARI if u == "cins_uyusuyor_tur_farkli" else KIRMIZI)
        return ad_metni, UYUM_METIN.get(u, u), dolgu

    wb = Workbook()

    # ---------------- 1. Kapak ve yöntem ----------------
    ws = wb.active
    ws.title = "Kapak ve Yöntem"
    genislik(ws, [42, 96])
    ws["A1"] = "PrimerJury primer tasarımı, toplantı kararlarının uygulanması"
    ws["A1"].font = BUYUK
    ws.merge_cells("A1:B1")
    r = 3
    gecen = [x for x in final if x.get("ozgulluk_durum") == "GECTI"]
    # ALAN TUTARLILIGI: hedefin ait olmadigi lokus kitapligindan tasarlanmis
    # ciftler teslimden CIKARILIR. Kural ihlali degildirler, ama hedefi temsil
    # etmezler; tabloda kalirlarsa o hedef kapsanmis gorunur, oysa kapsanmaz.
    _ta = field_audit.taxid_alanlari(a.kons)
    _ht = field_audit.hedef_taxidleri(a.hedefler)
    alan_disi = []
    temiz = []
    for x in gecen:
        uyumsuz, dag, baskin = field_audit.alan_dagilimi(
            x.get("hedef", ""), x.get("sinif", ""),
            taxid_alan=_ta, hedef_taxid=_ht)
        if uyumsuz:
            x["alan_gerekce"] = field_audit.aciklama(dag, x.get("sinif"), baskin)
            alan_disi.append(x)
        else:
            temiz.append(x)
    if alan_disi:
        print(u'pairs removed from delivery because of a DOMAIN MIX: %d (%s)'
              % (len(alan_disi),
                 ", ".join(sorted(set("%s/%s" % (z["hedef"], z["sinif"])
                                      for z in alan_disi)))))
    gecen = temiz
    for x in gecen:
        k = (x["hedef"], x["sinif"], x["ileri_dizi"], x["geri_dizi"])
        x["dis_hedef_disi"] = dis.get(k, "")
        x["dis_ornek"] = ";".join(dis_detay.get(k, []))[:180]
    kapsanan = sorted(set(x["hedef"] for x in gecen))
    satirlar = [
        ("Üretim zamanı", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Aday klasörü", a.aday),
        ("Doğrulama klasörü", a.final),
        ("Toplam hedef", str(len(set(x["hedef"] for x in ozet)))),
        ("Aday üreten hedef", str(len(set(x["hedef"] for x in ozet
                                          if x.get("durum", "").startswith("TAMAM"))))),
        ("Bütün kurallardan geçen hedef", str(len(kapsanan))),
        ("Bütün kurallardan geçen primer çifti", str(len(gecen))),
        ("", ""),
        ("YÖNTEM", ""),
        ("Oligo kuralı",
         "Yalnız A, C, G, T. Uzunluk 18-25 baz. GC %40-60, sert sınır %35-65. "
         "3' uç G ya da C. Son beş bazda en fazla üç G/C. En fazla dört aynı "
         "ardışık baz."),
        ("Termodinamik",
         "Tm 58-62 C, sert sınır 57-63. Hairpin dG en az -3000, self-dimer ve "
         "hetero-dimer en az -6000. Tm iki bağımsız kütüphaneyle ölçülür "
         "(primer3 ve Biopython); ikisi arasındaki sistematik kayma veriden "
         "hesaplanır, kaymadan sapan oligo elenir."),
        ("Ürün", "70-250 baz, üst sınır 300. 90-150 en yüksek puanı alır. "
                 "Çift Tm farkı 1,5 C altında."),
        ("Bağlanma kuralı",
         "Son iki baz birebir uymalı, son beş bazda en fazla bir uyumsuzluk, "
         "primerin tamamında en fazla üç uyumsuzluk. 5' tarafta sarkma "
         "serbest. İki primer ters zincirlerde ve 3' uçları birbirine bakacak."),
        ("Özgüllük",
         "Ürün hedeflenen HER üyede oluşmalı, rakiplerin HİÇBİRİNDE "
         "oluşmamalı. Primerlerden en az biri rakiplerde hiç bağlanma yeri "
         "bulamamalı; iki primerin de zayıf bağlanmasıyla oluşan temizlik "
         "kabul edilmez."),
        ("Ham okuma doğrulaması",
         "Konsensüste geçen her çift, taksonun ham okumalarında yeniden "
         "sınanır. Rakip okumalarındaki ürün oranı Wilson alt sınırıyla "
         "değerlendirilir; tek okumalık gürültü yüksek oran gibi görünmez."),
        ("Kutu çapraz bulaşması",
         "Kraken kutuları birbirine sızabildiği için, rakip kutusundaki "
         "okumaların ne kadarının aslında hedefe ait olduğu ölçülür. Rakipte "
         "görülen ürün oranı bu sızıntının Wilson üst sınırının altındaysa "
         "rakibin çoğaldığına dair kanıt sayılmaz."),
        ("Konsensüs",
         "Tasarım, ham okumalardan pozisyon başına BASKIN ALEL çağrılarak "
         "üretilen belirsizliksiz konsensüs üzerinde yapılır. IUPAC kodlu "
         "konsensüs ikinci ve daha temkinli ölçüm olarak saklanır; kodların "
         "küme kesişimiyle değerlendirilmesi gerçek farkları örtüyordu."),
        ("Çift ölçüm ilkesi",
         "Hiçbir karar tek bir kod yoluna bırakılmaz. Tm iki kütüphaneyle, "
         "takson ayrımı hem hizalama hem k-mer ile, özgüllük hem konsensüs "
         "hem ham okuma ile ölçülür. Ölçümler ayrışırsa aday reddedilir ve "
         "ayrışma log'lanır."),
    ]
    if a.not_metni:
        satirlar.append(("Not", a.not_metni))
    for k, v in satirlar:
        ws.cell(row=r, column=1, value=k).font = KALIN if k else NORMAL
        c = ws.cell(row=r, column=2, value=v)
        c.font = NORMAL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (1 + len(v) // 95))
        r += 1

    # ---------------- 2. Primer tablosu ----------------
    ws = wb.create_sheet("Primer Tablosu")
    sut = ["karar", "hedef", "sinif", "ileri_dizi", "ileri_tm", "ileri_gc",
           "geri_dizi", "geri_tm", "geri_gc", "tm_farki", "urun_min",
           "urun_maks", "uye_dogrulanan", "uye_toplam", "rakip_wilson",
           "ileri_baglanma_min", "geri_baglanma_min", "sus_ici_fark",
           "yetim_primer", "heterodimer_dg", "dis_hedef_disi", "ceza"]
    basliklar = ["Karar", "Hedef", "Sınıf", "İleri primer (5'-3')", "İleri Tm",
                 "İleri GC", "Geri primer (5'-3')", "Geri Tm", "Geri GC",
                 "Tm farkı", "Ürün en az", "Ürün en çok", "Doğrulanan üye",
                 "Toplam üye", "Rakip Wilson", "İleri bağlanma", "Geri bağlanma",
                 "Suş içi fark", "Yetim primer", "Hetero-dimer dG",
                 "Dış VT hedef dışı ürün", "Ceza"]
    yaz_baslik(ws, basliklar)
    genislik(ws, [7, 34, 7, 30, 9, 9, 30, 9, 9, 9, 11, 11, 13, 11, 13, 13, 13,
                  11, 13, 15, 20, 8])
    i = 2
    for x in sorted(gecen, key=lambda z: (z.get("karar", ""), z["hedef"],
                                          z["sinif"], float(z.get("ceza", 9e9)))):
        for j, k in enumerate(sut, 1):
            v = x.get(k, "")
            try:
                v = float(v) if k not in ("hedef", "sinif", "ileri_dizi",
                                          "geri_dizi", "yetim_primer") else v
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Consolas", size=10) if k in ("ileri_dizi", "geri_dizi") else NORMAL
            c.border = KENAR
        ws.cell(row=i, column=2).fill = YESIL
        i += 1
    if i == 2:
        ws.cell(row=2, column=1, value="Bütün kurallardan geçen aday yok").font = NORMAL

    # ---------------- 2b. Önerilen çift ----------------
    ws = wb.create_sheet("Önerilen Çiftler")
    ws["A1"] = ("Hedef başına bir çift. Sıralama ölçütü sırasıyla: dış "
                "veritabanında hedef dışı ürün sayısı, rakip Wilson alt "
                "sınırı, tasarım cezası. Dış veritabanı ölçümü yapılmamışsa "
                "o sütun boş kalır ve sıralama diğer iki ölçüte göre yapılır.")
    ws["A1"].font = NORMAL
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 45
    basliklar = ["Karar", "Hedef (toplantı adı)", "Sınıf", "İleri primer (5'-3')",
                 "İleri Tm", "Geri primer (5'-3')", "Geri Tm", "Ürün (bp)",
                 "Doğrulanan üye", "Rakip Wilson", "Dış VT hedef dışı",
                 "Yetim primer", "Ölçülen kimlik", "Ad ile veri uyumu"]
    yaz_baslik(ws, basliklar, satir=3)
    genislik(ws, [7, 34, 7, 30, 9, 30, 9, 12, 14, 13, 18, 14, 30, 24])

    def _sira(z):
        d = z.get("dis_hedef_disi", "")
        try:
            d = int(d)
        except (TypeError, ValueError):
            d = 10 ** 9
        try:
            w = float(z.get("rakip_wilson", 1))
        except (TypeError, ValueError):
            w = 1.0
        try:
            c = float(z.get("ceza", 9e9))
        except (TypeError, ValueError):
            c = 9e9
        return (d, w, c)

    en_iyi = {}
    for x in gecen:
        k = (x["hedef"], x["sinif"])
        if k not in en_iyi or _sira(x) < _sira(en_iyi[k]):
            en_iyi[k] = x
    i = 4
    for k in sorted(en_iyi, key=lambda z: (en_iyi[z].get("karar", ""), z[0], z[1])):
        x = en_iyi[k]
        urun = ("%s-%s" % (x.get("urun_min", ""), x.get("urun_maks", ""))
                if x.get("urun_min") != x.get("urun_maks")
                else str(x.get("urun_min", "")))
        satir = [x.get("karar", ""), x["hedef"], x["sinif"], x["ileri_dizi"],
                 x.get("ileri_tm", ""), x["geri_dizi"], x.get("geri_tm", ""),
                 urun,
                 "%s/%s" % (x.get("uye_dogrulanan", ""), x.get("uye_toplam", "")),
                 x.get("rakip_wilson", ""), x.get("dis_hedef_disi", ""),
                 x.get("yetim_primer", "")]
        kad, kuy, kdolgu = kimlik_hucre(x["hedef"])
        satir += [kad, kuy]
        for j, v in enumerate(satir, 1):
            if j in (5, 7, 10, 11):
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    pass
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Consolas", size=10) if j in (4, 6) else NORMAL
            c.border = KENAR
        ws.cell(row=i, column=2).fill = YESIL
        if kdolgu is not None:
            ws.cell(row=i, column=14).fill = kdolgu
        ws.cell(row=i, column=13).alignment = Alignment(wrap_text=True,
                                                        vertical="top")
        i += 1
    if i == 4:
        ws.cell(row=4, column=1, value="Önerilecek çift yok").font = NORMAL
    # Ayni olculen kimlige sahip birden cok hedef: dizi duzeyinde AYNI
    # organizmayi hedefliyorlar demektir. Bu, tabloya bakan birinin
    # kendiliginden goremeyecegi bir sey; acikca yazilir.
    ayni_kimlik = {}
    for k in en_iyi:
        r, d, n, cog = kimlik_destek(k[0])
        # Yalnizca kutularin YARIDAN COGU ayni kimlige gidiyorsa "ayni
        # organizma" denir. Coklukla yetinilirse, uyeleri farkli turlere
        # dagilan grup hedefleri de birlesmis gorunur.
        if r and cog and r.get("olculen_kimlik") \
                and "vurus yok" not in r["olculen_kimlik"] \
                and "esik alti" not in r["olculen_kimlik"]:
            ayni_kimlik.setdefault(r["olculen_kimlik"], set()).add(k[0])
    cakisan = {kk: vv for kk, vv in ayni_kimlik.items() if len(vv) > 1}
    if cakisan:
        i += 1
        c = ws.cell(row=i, column=1,
                    value="DİKKAT: aşağıdaki hedefler dizi düzeyinde AYNI "
                          "organizmayı hedefliyor. Ayrı satırlarda görünmeleri "
                          "toplantı adlarının korunmasındandır, iki farklı "
                          "organizma çoğaltıldığı anlamına gelmez.")
        c.font = Font(name=YAZI, bold=True, size=10, color="9C0006")
        c.fill = KIRMIZI
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)
        ws.row_dimensions[i].height = 32
        for kk, vv in sorted(cakisan.items()):
            i += 1
            c = ws.cell(row=i, column=1, value="%s  ->  %s" % (", ".join(sorted(vv)), kk))
            c.font = NORMAL
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=14)

    # ---------------- 3. Hedef durumu ----------------
    ws = wb.create_sheet("Hedef Durumu")
    basliklar = ["Karar", "Hedef", "Düzey", "Sınıf", "Üye", "Rakip",
                 "Tasarımda çift", "Tasarım durumu", "Kademe",
                 "Doğrulamada geçen", "Açıklama"]
    yaz_baslik(ws, basliklar)
    genislik(ws, [7, 34, 9, 7, 7, 8, 14, 16, 13, 17, 60])
    gecen_say = {}
    for x in gecen:
        gecen_say[(x["hedef"], x["sinif"])] = gecen_say.get((x["hedef"], x["sinif"]), 0) + 1
    i = 2
    for x in ozet:
        anahtar = (x["hedef"], x["sinif"])
        g = gecen_say.get(anahtar, 0)
        satir = [x.get("karar", ""), x["hedef"], x.get("duzey", ""), x["sinif"],
                 x.get("uye", ""), x.get("rakip", ""), x.get("cift", ""),
                 x.get("durum", ""), x.get("kademe", ""), g, x.get("note", "")]
        for j, v in enumerate(satir, 1):
            try:
                v = int(v)
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = KENAR
            c.alignment = Alignment(wrap_text=(j == 11), vertical="top")
        ws.cell(row=i, column=10).fill = YESIL if g else (
            SARI if str(x.get("durum", "")).startswith("TAMAM") else KIRMIZI)
        i += 1
    son = i - 1
    ws.cell(row=i + 1, column=1, value="Toplam doğrulamadan geçen çift").font = KALIN
    ws.cell(row=i + 1, column=10, value="=SUM(J2:J%d)" % son).font = KALIN

    ch = BarChart()
    ch.type = "bar"
    ch.title = "Hedef başına doğrulamadan geçen primer çifti"
    ch.y_axis.title = "çift sayısı"
    ch.x_axis.title = "hedef ve sınıf"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    veri = Reference(ws, min_col=10, min_row=1, max_row=son)
    kat = Reference(ws, min_col=2, min_row=2, max_row=son)
    ch.add_data(veri, titles_from_data=True)
    ch.set_categories(kat)
    ch.height, ch.width = 14, 26
    ch.legend = None
    ws.add_chart(ch, "M2")

    # ---------------- 4. Ayırt edilemez kutular ----------------
    ws = wb.create_sheet("Ayırt Edilemez Kutular")
    ws["A1"] = ("Aynı amplikon sınıfı içinde farklı taksona atanmış ama dizi "
                "düzeyinde ayrılamayan konsensüsler. Bunlar rakip listesinden "
                "çıkarılır, çünkü hedefin kendi dizisi rakip sayılırsa "
                "'rakipte ürün oluşmasın' kuralı mantıken sağlanamaz.")
    ws["A1"].font = NORMAL
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 45
    basliklar = ["Sınıf", "Taxid 1", "Takson 1", "Taxid 2", "Takson 2",
                 "Hizalanan bp", "Kesişimli özdeşlik %", "Katı özdeşlik %",
                 "Gerekçe"]
    yaz_baslik(ws, basliklar, satir=3)
    genislik(ws, [7, 11, 34, 11, 34, 13, 20, 18, 20])
    i = 4
    for x in ayirt:
        t1, t2 = x.get("taxid1", ""), x.get("taxid2", "")
        satir = [x.get("sinif", ""), t1, ad.get(t1, x.get("ad1", "")), t2,
                 ad.get(t2, x.get("ad2", "")), x.get("hizalanan_bp", ""),
                 x.get("kesisimli_ozdeslik", x.get("ozdeslik_yuzde", "")),
                 x.get("kati_ozdeslik", ""), x.get("gerekce", "")]
        for j, v in enumerate(satir, 1):
            try:
                v = float(v)
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = KENAR
        i += 1
    if i == 4:
        ws.cell(row=4, column=1, value="Ayırt edilemez çift bulunmadı").font = NORMAL

    # ---------------- 5. Alt küme bölmesi ----------------
    if a.bol and os.path.isdir(a.bol):
        ws = wb.create_sheet("Alt Küme Bölmesi")
        ws["A1"] = ("Tek primer çiftiyle kapsanamayan işlev grupları, "
                    "tasarlanabilirliğe göre alt kümelere bölündü. Bölme dizi "
                    "benzerliğine değil, hangi üyenin çift oluşumunu "
                    "engellediğine bakılarak yapılır.")
        ws["A1"].font = NORMAL
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 45
        yaz_baslik(ws, ["Grup", "Alt küme", "Üye sayısı", "Geçerli çift",
                        "Üyeler"], satir=3)
        genislik(ws, [34, 11, 12, 14, 80])
        i = 4
        for p in sorted(glob.glob(os.path.join(a.bol, "*_bolme.json"))):
            d = json.load(open(p, encoding="utf-8"))
            etiket = os.path.basename(p).replace("_bolme.json", "")
            for s in d.get("sets", []):
                satir = [etiket, "SET%d" % s["set"], len(s["members"]),
                         s["pairs"], ", ".join(s["members"])]
                for j, v in enumerate(satir, 1):
                    c = ws.cell(row=i, column=j, value=v)
                    c.font = NORMAL
                    c.border = KENAR
                    c.alignment = Alignment(wrap_text=(j == 5), vertical="top")
                i += 1
            if d.get("kalan"):
                c = ws.cell(row=i, column=1, value=etiket)
                c.font = NORMAL
                ws.cell(row=i, column=2, value="kapsanamayan").font = NORMAL
                ws.cell(row=i, column=5, value=", ".join(d["kalan"])).font = NORMAL
                ws.cell(row=i, column=2).fill = SARI
                i += 1

    # ---------------- 5b. Dış veritabanı ----------------
    if dis_ham:
        ws = wb.create_sheet("Dış Veritabanı")
        ws["A1"] = ("Her primer blastn (task blastn-short) ile referans "
                    "veritabanına arandı. Aynı referans dizide ters "
                    "zincirlerde ve 3' uçları birbirine bakan iki vuruş ürün "
                    "aralığında buluşuyorsa hedef dışı ürün sayıldı. Her vuruş "
                    "ayrıca toplantı kararındaki bağlanma kuralından geçirildi. "
                    "Evrensel primerlerde yüksek sayı beklenen sonuçtur.")
        ws["A1"].font = NORMAL
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 60
        yaz_baslik(ws, ["Hedef", "Sınıf", "Veritabanı", "İleri primer",
                        "Geri primer", "Hedef dışı ürün", "Örnekler"], satir=3)
        genislik(ws, [34, 7, 22, 30, 30, 16, 70])
        i = 4
        for x in sorted(dis_ham, key=lambda z: (z["hedef"], z["sinif"],
                                                -int(z.get("hedef_disi_urun", 0) or 0))):
            satir = [x["hedef"], x["sinif"], x["veritabani"], x["ileri_dizi"],
                     x["geri_dizi"], int(x.get("hedef_disi_urun", 0) or 0),
                     x.get("ornekler", "")]
            for j, v in enumerate(satir, 1):
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(name="Consolas", size=10) if j in (4, 5) else NORMAL
                c.border = KENAR
            if satir[5] == 0:
                ws.cell(row=i, column=6).fill = YESIL
            i += 1

    # ---------------- 5c. Uyarılar ----------------
    ws = wb.create_sheet("Uyarılar")
    ws["A1"] = ("Otomatik denetimler. Bu satırlar sonucu geçersiz kılmaz, ama "
                "primerlerin nasıl yorumlanması gerektiğini belirler.")
    ws["A1"].font = NORMAL
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    yaz_baslik(ws, ["Tür", "Konu", "Ayrıntı", "Yorum"], satir=3)
    genislik(ws, [26, 42, 70, 70])
    uyarilar = []

    # 1. Ayni primer cifti birden cok hedefte gecerse hedefler ayrilmiyor
    paylasim = {}
    for x in gecen:
        paylasim.setdefault((x["ileri_dizi"], x["geri_dizi"]), set()).add(
            "%s (%s)" % (x["hedef"], x["sinif"]))
    for k, v in sorted(paylasim.items(), key=lambda z: -len(z[1])):
        if len(v) > 1:
            hedef_adlari = set(h.split(" (")[0] for h in v)
            if len(hedef_adlari) < 2:
                yorum = ("Aynı hedefin farklı amplikon sınıflarında aynı çift "
                         "seçilmiş. Sorun değil, çiftin iki sınıfta da "
                         "çalıştığını gösterir.")
                tur = "Bilgi"
            else:
                yorum = ("Farklı hedefler aynı çiftle karşılanıyor. Bu çift o "
                         "hedefleri birbirinden AYIRMAZ; hedef tanımları "
                         "veride örtüşüyor demektir.")
                tur = "Dikkat"
            uyarilar.append((tur, u'The same primer pair on more than one target',
                             "%s / %s" % k, yorum + "  Hedefler: " + ", ".join(sorted(v))))

    # 1a. Hedef adi ile olculen kimlik uyusmazligi
    for x in sorted(set(z["hedef"] for z in gecen)):
        r = kimlik.get(x)
        if not r:
            continue
        u = r.get("uyum", "")
        if u in ("tur_uyusuyor", ""):
            continue
        tur = "Dikkat" if u == "cins_uyusuyor_tur_farkli" else "Eksik"
        uyarilar.append((tur, u'The target name and the measured identity disagree',
                         u'%s: Kraken2 says %s, the sequence shows %s (%s/%s bins)'
                         % (x, r.get("kraken_etiketi", "")[:40],
                            r.get("olculen_kimlik", ""),
                            r.get("destekleyen_kutu", "?"),
                            r.get("kutu_sayisi", "?")),
                         u'Evidence: %s. The target name comes from the meeting decision and was kept; which organism the primer really amplifies is in the Measured identity column.'
                         % (r.get("kanit", "")[:150] or "yok")))

    # 1b. Alan karisimi nedeniyle teslimden cikarilan ciftler
    for x in alan_disi:
        uyarilar.append(("Eksik", u'A mixture of domains, taken out of the delivery',
                         "%s (%s): F=%s R=%s"
                         % (x["hedef"], x["sinif"], x["ileri_dizi"],
                            x["geri_dizi"]),
                         x.get("alan_gerekce", "")))
    if alan_disi:
        for h in sorted(set(x["hedef"] for x in alan_disi)):
            uyarilar.append(("Eksik", u'The target was not met',
                             h,
                             u'The design in this target\'s own domain gave no valid pair; the single pair that passed came from the wrong locus library and was removed. The target has to count as NOT MET.'))

    # 2. Ayirt edilemez kutular
    for x in ayirt:
        t1, t2 = x.get("taxid1", ""), x.get("taxid2", "")
        uyarilar.append(("Dikkat", u'An indistinguishable taxon bin',
                         u'%s: %s ~ %s (strict identity %%%s)'
                         % (x.get("sinif", ""), ad.get(t1, t1), ad.get(t2, t2),
                            x.get("kati_ozdeslik", x.get("ozdeslik_yuzde", ""))),
                         u'These two bins do not separate at sequence level. It was taken out of the competitor list; the specificity of that target is not claimed for this taxon.'))

    # 3. Aday uretemeyen hedefler
    for x in ozet:
        if not str(x.get("durum", "")).startswith("TAMAM"):
            uyarilar.append(("Eksik", u'No candidate could be produced',
                             u'%s (%s), members=%s competitors=%s'
                             % (x["hedef"], x["sinif"], x.get("uye", ""),
                                x.get("rakip", "")),
                             u'The member that blocks it: %s. See the subset split sheet.' % (x.get("engelleyen", "") or "kayıt yok")))

    # 4. Gevsetilmis kademeyle gecen hedefler
    for x in ozet:
        if x.get("kademe") and x.get("kademe") != "kati":
            uyarilar.append(("Dikkat", u'A candidate was produced with a relaxed step',
                             "%s (%s), kademe=%s" % (x["hedef"], x["sinif"],
                                                     x["kademe"]),
                             u'The strict rule (the orphan primer must not bind anywhere in the competitors) could not be met; at its best placement in the competitors the orphan primer carries at least three mismatches.'))

    # 5. Dis veritabaninda cok hedef disi urun veren ozgul hedefler
    for x in gecen:
        try:
            n = int(x.get("dis_hedef_disi", 0) or 0)
        except (TypeError, ValueError):
            continue
        if n > 100 and not x["hedef"].endswith("universal"):
            uyarilar.append(("Dikkat", u'Many off target products in an external database',
                             u'%s (%s): %d products' % (x["hedef"], x["sinif"], n),
                             u'Clean in the sample but it gives a product in the relatives in the reference database. Candidates of the same target with a lower count should be preferred.'))

    i = 4
    for t, konu, ayrinti, yorum in uyarilar:
        for j, v in enumerate((t, konu, ayrinti, yorum), 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = KENAR
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=1).fill = (KIRMIZI if t == "Eksik"
                                         else SARI if t == "Dikkat" else YESIL)
        i += 1
    if i == 4:
        ws.cell(row=4, column=1, value="Uyarı yok").font = NORMAL

    # ---------------- 5d. Referans tasarım ----------------
    if ref_ham:
        ws = wb.create_sheet("Referans Tasarım")
        ws["A1"] = ("Numunede karşılanamayan hedefler için REFERANS "
                    "veritabanı dizilerinden tasarlanan çiftler. Bu çiftler "
                    "numuneyle DOĞRULANMAZ; sağdaki sütunlar yalnızca "
                    "numunede böyle bir kalıbın bulunup bulunmadığını "
                    "gösterir. Özgüllük iddiası referans veritabanı "
                    "kapsamındadır.")
        ws["A1"].font = NORMAL
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A1:K1")
        ws.row_dimensions[1].height = 60
        yaz_baslik(ws, ["Hedef", "Sınıf", "Veritabanı", "İleri primer",
                        "İleri Tm", "Geri primer", "Geri Tm", "Ürün (bp)",
                        "Numunede ürün veren okuma", "Wilson alt", "Durum"],
                   satir=3)
        genislik(ws, [36, 7, 20, 28, 9, 28, 9, 11, 24, 12, 24])
        i = 4
        for x in ref_ham:
            urun = ("%s-%s" % (x.get("urun_min", ""), x.get("urun_maks", ""))
                    if x.get("urun_min") != x.get("urun_maks")
                    else x.get("urun_min", ""))
            satir = [x.get("hedef", ""), x.get("sinif", ""),
                     x.get("veritabani", ""), x.get("ileri_dizi", ""),
                     x.get("ileri_tm", ""), x.get("geri_dizi", ""),
                     x.get("geri_tm", ""), urun,
                     x.get("numune_urun_okuma", ""),
                     x.get("numune_wilson_alt", ""), x.get("durum", "")]
            for j, v in enumerate(satir, 1):
                if j in (5, 7, 10):
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(name="Consolas", size=10) if j in (4, 6) else NORMAL
                c.border = KENAR
            ws.cell(row=i, column=11).fill = (
                YESIL if x.get("durum") == "numunede_destekli" else SARI)
            i += 1

    # ---------------- 5e. Laboratuvar protokolü ----------------
    ws = wb.create_sheet("Laboratuvar Protokolü")
    ws["A1"] = ("Sipariş ve ilk PCR için hazırlık tablosu. Tavlama sıcaklığı "
                "önerisi çiftin düşük Tm'inden 5 C aşağıdır; gradyan aralığı "
                "bu değerin 4 C altı ve üstüdür. Değerler primer3 ile "
                "hesaplanan Tm'e dayanır, ilk koşuda gradyanla "
                "doğrulanmalıdır.")
    ws["A1"].font = NORMAL
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 50
    yaz_baslik(ws, ["Karar", "Hedef", "Sınıf", "Oligo adı", "Dizi (5'-3')",
                    "Uzunluk", "GC %", "Tm (C)", "Beklenen ürün (bp)",
                    "Önerilen tavlama (C)", "Gradyan aralığı (C)", "Kaynak"],
               satir=3)
    genislik(ws, [7, 34, 7, 26, 30, 9, 8, 9, 18, 20, 20, 20])
    i = 4
    protokol_kayit = []
    for k in sorted(en_iyi, key=lambda z: (en_iyi[z].get("karar", ""), z[0], z[1])):
        protokol_kayit.append((en_iyi[k], "numune"))
    for x in ref_ham:
        if x.get("durum") == "numunede_destekli":
            protokol_kayit.append((x, "referans"))
    for x, kaynak in protokol_kayit:
        try:
            tf, tr = float(x.get("ileri_tm", 0)), float(x.get("geri_tm", 0))
        except (TypeError, ValueError):
            tf = tr = 0.0
        dusuk = min(tf, tr) if tf and tr else max(tf, tr)
        ta = round(dusuk - 5.0, 1) if dusuk else ""
        aralik = ("%.1f - %.1f" % (ta - 4, ta + 4)) if ta else ""
        urun = ("%s-%s" % (x.get("urun_min", ""), x.get("urun_maks", ""))
                if x.get("urun_min") != x.get("urun_maks")
                else x.get("urun_min", ""))
        for yon, dizi_a, tm_a, gc_a in (
                ("F", "ileri_dizi", "ileri_tm", "ileri_gc"),
                ("R", "geri_dizi", "geri_tm", "geri_gc")):
            dizi = x.get(dizi_a, "")
            if not dizi:
                continue
            satir = [x.get("karar", ""), x.get("hedef", ""), x.get("sinif", ""),
                     "%s_%s_%s" % (x.get("hedef", "")[:22], x.get("sinif", ""), yon),
                     dizi, len(dizi), x.get(gc_a, ""), x.get(tm_a, ""),
                     urun, ta, aralik, kaynak]
            for j, v in enumerate(satir, 1):
                if j in (6, 7, 8, 10):
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                c = ws.cell(row=i, column=j, value=v)
                c.font = Font(name="Consolas", size=10) if j == 5 else NORMAL
                c.border = KENAR
            ws.cell(row=i, column=12).fill = (
                YESIL if kaynak == "numune" else SARI)
            i += 1
    son_p = i - 1
    i += 1
    ws.cell(row=i, column=1, value="Toplam oligo").font = KALIN
    ws.cell(row=i, column=6, value="=COUNTA(E4:E%d)" % son_p).font = KALIN
    i += 2
    for baslik, metin in [
        ("Reaksiyon", "25 uL: 12,5 uL 2x master mix, her primer 0,2-0,5 uM, "
                      "1-5 ng kalip DNA, kalan hacim nukleaz iceriksiz su."),
        ("Dongu", "95 C 3 dk; 35 dongu (95 C 15 sn, tavlama 20 sn, 72 C 20 sn); "
                  "72 C 5 dk. Urunler 300 bp altinda oldugu icin uzatma kisa "
                  "tutulmustur."),
        ("Ilk kosu", "Tavlama sicakligi gradyanla sinanmali. Gradyan araligi "
                     "her cift icin yukaridaki sutunda verilmistir."),
        ("Kontroller", "Pozitif kontrol: hedefin en yuksek okuma sayisina "
                       "sahip oldugu ornek. Negatif kontrol: kalipsiz. "
                       "Ozgulluk kontrolu: rakip taksonun baskin oldugu ornek."),
        ("Jel", "%2 agaroz, 100 bp merdiven. Tek bant beklenir; ikinci bant "
                "capraz cogalma isaretidir ve o cift elenmelidir."),
        ("Kaynak sutunu", "numune: primer numunenin kendi okumalarindan "
                          "tasarlandi ve ham okumalarla dogrulandi. "
                          "referans: primer referans veritabanindan "
                          "tasarlandi, numunede yalnizca destek olcumu var."),
    ]:
        ws.cell(row=i, column=1, value=baslik).font = KALIN
        c = ws.cell(row=i, column=2, value=metin)
        c.font = NORMAL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
        ws.row_dimensions[i].height = max(15, 13 * (1 + len(metin) // 110))
        i += 1

    # ---------------- 6. Eleme gerekçeleri ----------------
    ws = wb.create_sheet("Eleme Gerekçeleri")
    yaz_baslik(ws, ["Hedef ve sınıf", "Kompozisyon sonrası oligo",
                    "Termodinamik sonrası", "Her üyeye bağlanan",
                    "Rakipte hiç bağlanmayan", "Çift Tm farkı elemesi",
                    "Bir üyede ürün yok", "Rakipte ürün", "Yetim primer yok",
                    "Hetero-dimer", "Geçerli çift"])
    genislik(ws, [40] + [18] * 10)
    i = 2
    kal = [
        (r"kompozisyon(?: suzgeci)? sonrasi oligo:\s*(\d+)", 2),
        (r"termodinamik sonrasi oligo:\s*(\d+)", 3),
        (r"her hedef uyeye baglanan oligo:\s*(\d+)", 4),
        (r"rakiplerde hic baglanmayan oligo:\s*(\d+)", 5),
        (r"elenen, cift Tm farki\s*:\s*(\d+)", 6),
        (r"elenen, bir uyede urun yok\s*:\s*(\d+)", 7),
        (r"elenen, rakipte urun olusuyor\s*:\s*(\d+)", 8),
        (r"elenen, yetim primer yok\s*:\s*(\d+)", 9),
        (r"elenen, hetero-dimer dG\s*:\s*(\d+)", 10),
        (r"gecerli cift sayisi\s*:\s*(\d+)", 11),
    ]
    for p in sorted(glob.glob(os.path.join(a.aday, "*__*.log"))):
        t = open(p, encoding="utf-8", errors="replace").read()
        ws.cell(row=i, column=1,
                value=os.path.basename(p)[:-4]).font = NORMAL
        ws.cell(row=i, column=1).border = KENAR
        for kalip, sutun in kal:
            m = re.search(kalip, t)
            c = ws.cell(row=i, column=sutun, value=int(m.group(1)) if m else "")
            c.font = NORMAL
            c.border = KENAR
        i += 1

    wb.save(a.out)
    print(u'written: %s' % a.out)
    print("sayfa: %s" % ", ".join(wb.sheetnames))
    print(u'primer pairs passing: %d, targets covered: %d' % (len(gecen), len(kapsanan)))


if __name__ == "__main__":
    main()
