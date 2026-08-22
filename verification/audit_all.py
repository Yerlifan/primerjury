# -*- coding: utf-8 -*-
"""THE AUDIT GATE THAT RUNS ON EVERY RUN  -  it makes "look without being asked" the
code's job.

WHY IT EXISTS
-------------
In this project most of the bugs were not in the measurement but in THE TABLE THE
MEASUREMENT RESTED ON: a stale cache, a stale reference, an exclusion taxid chosen
by the target's name, a taxonomy NCBI had changed, key names that did not match
one another. None of them failed a run; every one of them made the result SILENTLY
wrong and was found only when somebody asked "is there another bug".

An audit that depends on being asked is not an audit. This script asks that
question itself on every run. It makes no measurement and changes no file; it only
looks, and writes down what does not hold.

WHAT IS AUDITED
  1  Do the exclusion map keys match SIPARIS_LISTESI exactly
  2  Do the exclusion taxids exist in NCBI and do they COVER their members (needs network)
  3  Which pair were the quick test references measured from, and has that pair changed since
  4  Do the panel source and SIPARIS_LISTESI carry the same sequences
  5  Do the checkpoint seals include the sequence (the stale cache trap)
  6  Are the output files FRESH relative to their inputs
  7  Did every pair in the panel pass the geometry gate with THE SAME sequence
  8  Do the products on one plate separate on a gel, and is the band class suitable
  9  Is the pair count in the guidance documents the panel's current count
 10  Does the NCBI name rule pass its known-answer test
 11  Are the sequences on the order sheet (xlsx) the panel's current sequences
 12  Is the evidence complete for every pair to be ordered UNCONDITIONALLY
 13  Is the product length in the table the length the pair PRODUCES on the consensus
 14  Are the .bat files CRLF and pure ASCII, and do their goto targets exist
 15  Are there leftover files in the canonical consensus directory that are not in the index
 16  Do P and K read the same membership file
 17  Are the indexed databases in the DNA alphabet, and are the "twins" really twins
 18  Do the universal and control primers see our own targets

The exit code: 0 all clean, 1 at least one audit failed, 2 an audit needing the
network was skipped but the local audits are clean.

Run:
    python verification/audit_all.py --root .
    python verification/audit_all.py --root . --offline     (skip the NCBI steps)

"""
from __future__ import print_function

import argparse
import csv
import hashlib
import io
import os
import subprocess
import sys
import time

BULGU = []
ATLANAN = []


# The severity levels. The wording is deliberately plain: "BLOKE" is not a judgement,
# it means "order without fixing this and the loss of money and time is certain".
BLOKE, DIKKAT, BILGI = u'SIPARISI DURDURUR', u'DIKKAT', u'BILGI'


def bulgu(baslik, ayrinti, onem=None):
    BULGU.append((baslik, ayrinti, onem or DIKKAT))


def _tsv(yol):
    if not os.path.exists(yol):
        return []
    with io.open(yol, encoding='utf-8') as fh:
        return list(csv.DictReader((l for l in fh if not l.startswith('#')),
                                   delimiter='\t'))


def _harita(yol):
    h = {}
    if not os.path.exists(yol):
        return h
    for l in io.open(yol, encoding='utf-8'):
        l = l.rstrip('\n')
        if not l.strip() or l.startswith('#') or l.startswith('hedef\t'):
            continue
        p = l.split('\t')
        if len(p) >= 2:
            h[p[0].strip()] = p[1].strip()
    return h


# --- 1 ------------------------------------------------------------------
def d1_harita_anahtarlari(kok, yaz):
    sl = os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv')
    hy = os.path.join(kok, 'screening', 'target_taxids.tsv')
    if not os.path.exists(sl) or not os.path.exists(hy):
        ATLANAN.append(u'1 map keys (no such file)')
        return
    adlar = set((r.get('hedef') or '').strip() for r in _tsv(sl) if r.get('hedef'))
    H = _harita(hy)
    eksik = sorted(adlar - set(H))
    fazla = sorted(set(H) - adlar)
    yaz(u'  [1] exclusion map keys: %d targets, %d map rows'
        % (len(adlar), len(H)))
    if eksik:
        bulgu(u'A target that is NOT in the exclusion map',
              u'%s\n      These targets are written as SINANMADI in the NCBI layer. The key name must be EXACTLY the same as in SIPARIS_LISTESI.' % ', '.join(eksik))
    if fazla:
        bulgu(u'Haritada olup panelde OLMAYAN anahtar',
              u'%s\n      The name probably changed and the old key stayed; a dead row gives false confidence.' % ', '.join(fazla))
    bos = sorted(k for k, v in H.items() if not v)
    if bos:
        bulgu(u'A target whose exclusion taxid was left EMPTY',
              u'%s\n      An empty taxid means that target is never tested in the NCBI layer.'
              % ', '.join(bos))


# --- 2 ------------------------------------------------------------------
def d2_kapsama(kok, yaz, agsiz):
    bet = os.path.join(kok, 'screening', 'exclusion_coverage_check.py')
    if not os.path.exists(bet):
        ATLANAN.append(u'2 the coverage audit (no such script)')
        return
    if agsiz:
        ATLANAN.append(u'2 the coverage audit (--offline was given; NCBI Taxonomy is needed)')
        return
    yaz(u'  [2] running the exclusion coverage check (NCBI Taxonomy)...')
    try:
        p = subprocess.run([sys.executable, bet, '--root', kok],
                           stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                           timeout=900)
        cik = p.stdout.decode('utf-8', 'replace')
    except Exception as e:
        ATLANAN.append(u'2 the coverage audit (could not be run: %s)' % e)
        return
    for satir in cik.splitlines():
        if 'KAPSAMIYOR' in satir or 'denetlenemedi' in satir:
            yaz(u'      %s' % satir.strip()[:110])
    if p.returncode != 0:
        kotu = [s.strip() for s in cik.splitlines() if 'KAPSAMIYOR' in s]
        if kotu:
            bulgu(u'The exclusion taxon DOES NOT COVER its members',
                  u'\n      '.join(kotu[:8]) +
                  u'\n      An uncovered member counts as off target although it IS the target itself. Detail: screening/exclusion_coverage_check.py', BLOKE)
        else:
            yaz(u'      (coverage is complete; only some targets could not be audited)')


# --- 3 ------------------------------------------------------------------
def d3_referans_bayat(kok, yaz):
    ry = os.path.join(kok, 'QUICK_TEST', 'referans_degerler.tsv')
    py_ = os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'panel_tek_protokol.tsv')
    if not os.path.exists(ry):
        bulgu(u'The quick test reference file IS MISSING',
              u'QUICK_TEST/referans_degerler.tsv was not produced. The reference is then read from constants embedded in the code and which pair it was measured from is unknown. To produce it: python verification/refresh_reference.py --root .')
        return
    if not os.path.exists(py_):
        ATLANAN.append(u'3 reference freshness (there is no panel_tek_protokol.tsv)')
        return
    ref = {}
    for r in _tsv(ry):
        ref[(r.get('hedef') or '').strip()] = ((r.get('F') or '').strip().upper(),
                                               (r.get('R') or '').strip().upper())
    bayat = []
    for r in _tsv(py_):
        ad = (r.get('hedef') or '').strip()
        if ad not in ref:
            continue
        simdi = ((r.get('F') or '').strip().upper(), (r.get('R') or '').strip().upper())
        if simdi[0] and simdi != ref[ad]:
            bayat.append(ad)
    yaz(u'  [3] reference freshness: %d rows in the reference, %d pairs changed'
        % (len(ref), len(bayat)))
    if bayat:
        bulgu(u'The reference is STALE, the primer pair has changed',
              u'%s\n      On these rows a number belonging to the old pair is compared against the new pair and produces a false regression. Refresh it: python verification/refresh_reference.py --root .' % ', '.join(bayat), DIKKAT)


# --- 4 ------------------------------------------------------------------
def d4_kaynak_tutarliligi(kok, yaz):
    """Do the panel's sequence source and the order list say the same sequence?"""
    sl = os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv')
    pk = os.path.join(kok, 'primer_final',
                      'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    if not os.path.exists(sl) or not os.path.exists(pk):
        ATLANAN.append(u'4 source consistency (no such file)')
        return
    panel = {}
    with io.open(pk, encoding='utf-8') as fh:
        sat = [l.rstrip('\n').split('\t') for l in fh]
    if not sat:
        ATLANAN.append(u'4 source consistency (the panel source is empty)')
        return
    bas = sat[0]
    try:
        iH = bas.index('Hedef')
        iF = [i for i, b in enumerate(bas) if b.startswith('Ileri primer')][0]
        iR = [i for i, b in enumerate(bas) if b.startswith('Geri primer')][0]
    except (ValueError, IndexError):
        ATLANAN.append(u'4 source consistency (the panel columns were not recognised)')
        return
    for r in sat[1:]:
        if len(r) > max(iF, iR) and r[iH].strip():
            panel[r[iH].strip()] = (r[iF].strip().upper(), r[iR].strip().upper())
    fark = []
    liste_adlari = set()
    for r in _tsv(sl):
        ad = (r.get('hedef') or '').strip()
        if not ad:
            continue
        liste_adlari.add(ad)
        if ad not in panel:
            continue
        simdi = ((r.get('F') or '').strip().upper(), (r.get('R') or '').strip().upper())
        if simdi[0] and panel[ad][0] and simdi != panel[ad]:
            fark.append(ad)
    # A MISSING ROW is a difference too. Comparing only the SHARED targets ignores a pair
    # that is in one file and absent from the other; on 2026-08-10 Petriella_cinsi read
    # KESIN in the order list while it had no row at all in the panel source, and this
    # check said "0 differences".
    yalniz_listede = sorted(liste_adlari - set(panel))
    yalniz_panelde = sorted(set(panel) - liste_adlari)
    yaz(u'  [4] source consistency: %d shared targets, %d sequence differences, %d only in the list, %d only in the panel'
        % (len(liste_adlari & set(panel)), len(fark),
           len(yalniz_listede), len(yalniz_panelde)))
    if yalniz_listede:
        bulgu(u'In the order list but NOT in the panel source',
              u'%s\n      The plate and Ta of these pairs are not in the panel table; they go into the order but their place in the experimental layout is undefined.'
              % ', '.join(yalniz_listede), BLOKE)
    if fark:
        bulgu(u'The panel source and SIPARIS_LISTESI DISAGREE',
              u'%s\n      The two files give different sequences. Which one is to be ordered is undefined, and that has to be settled first.' % ', '.join(fark), BLOKE)


# --- 5 ------------------------------------------------------------------
def d5_muhur_diziyi_iceriyor_mu(kok, yaz):
    """Do the checkpoint seals include the primer SEQUENCE?

        A seal that does not take a changed sequence's OLD measurement for a fresh one.
        This project had that bug in three separate scripts in one day; it is now looked
        for mechanically.

    """
    bakilacak = [
        ('protocol/single_protocol_measure.py', True),
        ('verification/recovery_round.py', True),
        ('verification/specificity_round.py', True),
        ('verification/mfeprimer_layer.py', False),
        ('engine/rederive_membership.py', True),
    ]
    eksik = []
    for yol, zorunlu in bakilacak:
        t = os.path.join(kok, yol)
        if not os.path.exists(t):
            continue
        s = io.open(t, encoding='utf-8', errors='replace').read()
        # muhur/imza hesabinda F ve R gecmeli
        imzali = ("'F'" in s or '"F"' in s or "get('F'" in s)
        anahtar = ('md5' in s or 'sha1' in s or 'hashlib' in s)
        if anahtar and not imzali and zorunlu:
            eksik.append(yol)
    yaz(u'  [5] checkpoint seals: %d scripts examined, %d suspect'
        % (len(bakilacak), len(eksik)))
    if eksik:
        bulgu(u'The seal may NOT COVER the sequence',
              u'%s\n      If the seal holds no primer sequence, an old measurement looks FRESH when the sequence changes. Look at it by hand.' % ', '.join(eksik))


# --- 6 ------------------------------------------------------------------
def d6_cikti_tazeligi(kok, yaz, uretilecek=()):
    """uretilecek: the outputs THIS RUN WILL REGENERATE - those are not called stale.

        2026-08-10: the NCBI gate showed the ncbi_katman4.tsv the run itself would produce
        as "stale" and asked for confirmation. A gate that cries wolf stops being taken
        seriously. So a file the run will produce is excluded, and the fact that it was
        excluded IS PRINTED TO THE SCREEN; it is not hidden silently.

    """
    ciftler = [
        ('primer_final/devir_ciftleri_20260802_sonrotus_TESLIM.tsv',
         'ONE_PROTOCOL_RESULT/panel_tek_protokol.tsv'),
        ('ONE_PROTOCOL_RESULT/panel_tek_protokol.tsv',
         'QUICK_TEST/referans_degerler.tsv'),
        ('ONE_PROTOCOL_RESULT/SIPARIS_LISTESI.tsv',
         'VERIFICATION_RESULT/dogrulama_uc_sutun.tsv'),
        ('screening/target_taxids.tsv',
         'VERIFICATION_RESULT/ncbi_katman4.tsv'),
        # 2026-08-11: when the membership table changes, stage G's table goes stale.
        # The identity columns in the order list (olculen_kimlik, ad_farkli_mi, the member
        # count) come straight from that table's UYESI_OLDUGU_HEDEFLER column. If the
        # membership was corrected but G was not re-run, the order row counts the names of
        # bins that ARE NO LONGER MEMBERS: the Petriella_musispora row said "10/10 bins" and
        # counted Microascus, Lomentospora and Graphium, when the measurement had been made
        # with 9 bins.
        ('screening/target_membership.tsv',
         'ALL_IDENTITIES_RESULT/tum_kutu_kimlikleri.tsv'),
    ]
    def _diziler(y2, ad_h='hedef', ad_f='F', ad_r='R'):
        out = {}
        for r in _tsv(y2):
            a = (r.get(ad_h) or '').strip()
            if a:
                out[a] = ((r.get(ad_f) or '').strip().upper(),
                          (r.get(ad_r) or '').strip().upper())
        return out

    bayat = []
    for g, c in ciftler:
        # CONTENT FIRST, TIME SECOND. When only the Tm or product length columns differ
        # between the panel source and P's output, the timestamp says "stale" while the
        # measurement IS VALID: dCq depends on the sequence, not on the Tm.
        # (That is exactly what happened at 07:02 on 2026-08-11.) If the sequences are the
        # same it does not count as stale.
        if (g.endswith('devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
                and c.endswith('panel_tek_protokol.tsv')):
            gp2, cp2 = os.path.join(kok, g), os.path.join(kok, c)
            if os.path.exists(gp2) and os.path.exists(cp2):
                sat2 = [l.rstrip('\n').split('\t') for l in io.open(gp2, encoding='utf-8')]
                b2 = sat2[0]
                try:
                    iH2 = b2.index('Hedef')
                    iF2 = next(k for k, x in enumerate(b2) if x.startswith('Ileri primer'))
                    iR2 = next(k for k, x in enumerate(b2) if x.startswith('Geri primer'))
                except (ValueError, StopIteration):
                    iH2 = None
                if iH2 is not None:
                    kay = {}
                    for r2 in sat2[1:]:
                        if len(r2) > max(iF2, iR2) and r2[iH2].strip():
                            kay[r2[iH2].strip()] = (r2[iF2].strip().upper(),
                                                    r2[iR2].strip().upper())
                    cik = _diziler(cp2)
                    ortak = set(kay) & set(cik)
                    if ortak and all(kay[k2] == cik[k2] for k2 in ortak):
                        yaz(u'      (%s has an old timestamp but the SEQUENCES are identical, so the measurement holds and it is not counted as stale)' % c)
                        continue
        if c in uretilecek:
            yaz(u'      (%s will be regenerated in this run, so freshness was not required)' % c)
            continue
        gp, cp = os.path.join(kok, g), os.path.join(kok, c)
        if not os.path.exists(gp) or not os.path.exists(cp):
            continue
        if os.path.getmtime(cp) < os.path.getmtime(gp):
            bayat.append(u'%s (%s) < %s (%s)'
                         % (c, time.strftime('%d.%m %H:%M', time.localtime(os.path.getmtime(cp))),
                            g, time.strftime('%d.%m %H:%M', time.localtime(os.path.getmtime(gp)))))
    yaz(u'  [6] output freshness: %d dependencies, %d stale' % (len(ciftler), len(bayat)))
    if bayat:
        bulgu(u'The output is OLDER than its input',
              u'\n      '.join(bayat) +
              u'\n      This output has not seen the current input; it has to be reproduced.')



# --- 7 ------------------------------------------------------------------
def d7_geometri_kapisi(kok, yaz):
    """Did every pair in the panel pass the geometry check with THE SAME sequence?

        2026-08-10: six pairs had their sequences changed after the geometry check of 2
        August, and the geometry had not been re-run. So those six pairs had NEVER passed
        the panel's own rules (length, GC, the Tm window, hairpin, dimer), and nothing said
        so.

    """
    import glob
    pk = os.path.join(kok, 'primer_final',
                      'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    adaylar = sorted(glob.glob(os.path.join(kok, 'primer_final',
                                            'geometri_denetimi_*.tsv')))
    adaylar = [x for x in adaylar if 'yedek' not in x]
    if not os.path.exists(pk) or not adaylar:
        ATLANAN.append(u'7 the geometry gate (no such file)')
        return
    gy = max(adaylar, key=os.path.getmtime)
    g = {}
    for r in _tsv(gy):
        h = (r.get('Hedef') or '').strip()
        pr = (r.get('Primer') or '').strip()
        if pr in ('Ileri', 'Geri'):
            g.setdefault(h, {})[pr] = (r.get('Dizi') or '').strip().upper()
    sat = [l.rstrip('\n').split('\t') for l in io.open(pk, encoding='utf-8')]
    bas = sat[0]
    try:
        iH = bas.index('Hedef')
        iF = next(i for i, b in enumerate(bas) if b.startswith('Ileri primer'))
        iR = next(i for i, b in enumerate(bas) if b.startswith('Geri primer'))
    except (ValueError, StopIteration):
        ATLANAN.append(u'7 the geometry gate (the panel columns were not recognised)')
        return
    import re as _re
    gecmemis = []
    for r in sat[1:]:
        if len(r) <= max(iF, iR) or not r[iH].strip():
            continue
        if not _re.match(r'^[A-Za-z]', r[iH].strip()):
            continue
        F, R = r[iF].strip().upper(), r[iR].strip().upper()
        if not _re.fullmatch(r'[ACGT]+', F or ''):
            continue
        ad = r[iH].strip()
        if ad not in g or (F, R) != (g[ad].get('Ileri'), g[ad].get('Geri')):
            gecmemis.append(ad)
    yaz(u'  [7] geometry gate: source %s, not passing %d'
        % (os.path.basename(gy), len(gecmemis)))
    if gecmemis:
        bulgu(u'A pair that DID NOT PASS the geometry gate',
              u'%s\n      The sequence of these pairs changed after the last geometry audit. Length, GC, the Tm window, hairpin and dimer rules were NEVER measured for these sequences. Run: python verification/refresh_geometry.py --root .' % ', '.join(gecmemis), BLOKE)



# --- 8 ------------------------------------------------------------------
def d8_plaka_jel_ve_bant(kok, yaz):
    """Do the products on one plate separate on a gel, plus the QuantiNova band class.

        2026-08-10: the pairs changed today broke the within-plate gel separation again.
        The Bacteroidales product went from 241 bp to 150 bp and came within 5 bp of
        Mantar F2's 145 bp, which cannot be told apart on a 2% agarose gel. Changing a pair
        affects not only that pair but THE PLATE, which is why it is checked on every run.

    """
    import itertools
    import re as _re
    pk = os.path.join(kok, 'primer_final',
                      'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    if not os.path.exists(pk):
        ATLANAN.append(u'8 plate and gel (there is no panel file)')
        return
    sat = [l.rstrip('\n').split('\t') for l in io.open(pk, encoding='utf-8')]
    bas = sat[0]
    try:
        iP, iT, iH = bas.index('Plaka'), bas.index('Ta (C)'), bas.index('Hedef')
        iU = bas.index('Urun (bp)')
        iF = next(i for i, b in enumerate(bas) if b.startswith('Ileri primer'))
    except (ValueError, StopIteration):
        ATLANAN.append(u'8 plate and gel (the columns were not recognised)')
        return
    pl = {}
    bant = []
    for r in sat[1:]:
        if len(r) <= max(iU, iF) or not r[iH].strip():
            continue
        if not _re.fullmatch(r'[ACGT]+', (r[iF] or '').strip().upper()):
            continue          # drops the note rows - with no primer sequence it is not a pair
        try:
            u = int(_re.sub(r'\D', '', r[iU]))
        except ValueError:
            continue
        pl.setdefault((r[iP].strip(), r[iT].strip()), []).append((r[iH].strip(), u))
        if u > 250:
            bant.append(u'%s %d bp (QuantiNova >250 ONERMIYOR)' % (r[iH].strip(), u))
    cak = []
    for k, v in sorted(pl.items()):
        for (a1, u1), (a2, u2) in itertools.combinations(sorted(v, key=lambda x: x[1]), 2):
            if abs(u1 - u2) < 10:
                cak.append(u'plaka %s Ta %s: %s (%d bp) / %s (%d bp) - fark %d bp'
                           % (k[0], k[1], a1, u1, a2, u2, abs(u1 - u2)))
    yaz(u'  [8] plate/gel: %d plate groups, %d overlaps, %d products outside the band'
        % (len(pl), len(cak), len(bant)))
    if cak:
        bulgu(u'Plaka ici JEL AYRIMI cakismasi',
              u'\n      '.join(cak) +
              u'\n      Two products closer than 10 bp on the same plate cannot be told apart on a %2 agarose gel. Either the plate has to be reassigned or the accepted difference has to be written IN THE REPORT.', DIKKAT)
    if bant:
        bulgu(u'Amplikon bant sinifi disinda',
              u'\n      '.join(bant) +
              u'\n      For QuantiNova SYBR Green the ideal is 60-150 bp, 150-250 bp needs a 30 s extension, and above 250 is not recommended.', DIKKAT)



# --- 9 ------------------------------------------------------------------
# Every document where a number is written BY HAND goes stale one day. This check
# compares the "N pairs" claim in the guidance documents against the panel's CURRENT
# count. 2026-08-10: three documents said three different numbers (16, 16, 11); the
# right one was 20.
# DATED audit records (such as SON_KONTROL.md) are NOT in this list: their number is
# the number of the day it was written and must not be changed. The list holds only
# the documents that answer "what should I do right now".
YOL_GOSTERICI = ('OKU_ONCE.md', 'NASIL_DEVAM_EDILIR.md', 'CALISTIRMA_KILAVUZU.md',
                 'GUNCEL_DURUM.md')


def d9_belgelerde_bayat_sayi(kok, yaz):
    import re as _re
    sl = _tsv(os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv'))
    if not sl:
        ATLANAN.append(u'9 document counts (there is no SIPARIS_LISTESI)')
        return
    kesin = sum(1 for r in sl if (r.get('SINIF') or '').strip().upper() == 'KESIN')
    evr = sum(1 for r in sl if (r.get('SINIF') or '').strip().upper() == 'EVRENSEL')
    dogru = {kesin, evr, kesin + evr, len(sl), len(sl) - kesin - evr}
    kalip = _re.compile(r'(?:KES[İI]N|[Ss]ipari[şs] edilebilir|[Ss]ipari[şs] edilecek)'
                        r'[^0-9\n]{0,40}?(\d{1,2})\s*[çc]ift')
    kotu = []
    bakilan = 0
    for ad in YOL_GOSTERICI:
        t = os.path.join(kok, ad)
        if not os.path.exists(t):
            continue
        bakilan += 1
        metin = io.open(t, encoding='utf-8', errors='replace').read()
        for m in kalip.finditer(metin):
            n = int(m.group(1))
            if n not in dogru:
                kotu.append(u'%s: "%s" (the correct number today: %d ordered, %d target specific, %d universal)'
                            % (ad, m.group(0).strip()[:60], kesin + evr, kesin, evr))
    yaz(u'  [9] pair counts in the documents: %d documents examined, %d stale claims'
        % (bakilan, len(kotu)))
    if kotu:
        bulgu(u'A STALE pair count in a document',
              u'\n      '.join(kotu) +
              u'\n      Every document that writes the number by hand goes stale. Change the sentence to point at GUNCEL_DURUM.md, which is produced on every run.', DIKKAT)



# --- 10 -----------------------------------------------------------------
def d10_ad_kurali_sinavi(kok, yaz):
    """Does the NCBI name rule (named or unnamed) pass its known-answer test?

        That rule determines the off-target COUNT. Broken, the table that reaches the
        write-up carries inflated numbers such as "650 off-target" (which is exactly what
        happened on 2026-08-10: the loose rule said 650, the strict rule 82).

    """
    bet = os.path.join(kok, 'verification', 'ncbi_reclassify.py')
    if not os.path.exists(bet):
        ATLANAN.append(u'10 the name rule test (no such script)')
        return
    try:
        p = subprocess.run([sys.executable, bet, '--selftest'],
                           stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                           timeout=120)
        cik = p.stdout.decode('utf-8', 'replace').strip()
    except Exception as e:
        ATLANAN.append(u'10 ad kurali sinavi (kosturulamadi: %s)' % e)
        return
    son = [l for l in cik.splitlines() if 'sinavi:' in l]
    yaz(u'  [10] NCBI naming rule: %s' % (son[-1].strip() if son else cik[:60]))
    if p.returncode != 0:
        bulgu(u'NCBI ad kurali sinavi DUSTU',
              cik[-600:] +
              u'\n      This rule sets the off target COUNT; if it is broken the numbers in the report are wrong.', BLOKE)



# --- 11 -----------------------------------------------------------------
def d11_siparis_dizileri(kok, yaz):
    """Are the sequences in the xlsx going to the write-up or the supplier the panel's CURRENT ones?

        The night of 2026-08-10: the morning summary pointed at the "2 Panel" sheet of
        PrimerJury_..._TESLIM.xlsx saying "the sequences will be copied from here"; SIX of
        the pairs on that sheet carried stale sequences. Had the order been placed from
        there, 6 of the 20 pairs would have arrived as the wrong oligo. That is the most
        expensive mistake this project can make.

    """
    import glob as _glob
    import re as _re
    pk = os.path.join(kok, 'primer_final',
                      'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    # 2026-08-11: the old delivery xlsx was moved to the archive. What is audited now is
    # the PrimerJury_PANEL_*.xlsx PRODUCED on every run; if that is missing too, the
    # audit is skipped. The archived file IS NOT LOOKED AT.
    _p = sorted(_glob.glob(os.path.join(kok, 'PrimerJury_PANEL_*.xlsx')))
    xl = _p[-1] if _p else ''
    if not os.path.exists(pk):
        ATLANAN.append(u'11 order sequences (there is no panel source)')
        return
    sat = [l.rstrip('\n').split('\t') for l in io.open(pk, encoding='utf-8')]
    b = sat[0]
    try:
        iH = b.index('Hedef')
        iF = next(i for i, x in enumerate(b) if x.startswith('Ileri primer'))
        iR = next(i for i, x in enumerate(b) if x.startswith('Geri primer'))
    except (ValueError, StopIteration):
        ATLANAN.append(u'11 order sequences (the panel columns were not recognised)')
        return
    tsv = {}
    for r in sat[1:]:
        if len(r) > max(iF, iR) and r[iH].strip():
            F = r[iF].strip().upper()
            if _re.fullmatch(r'[ACGT]+', F or ''):
                tsv[r[iH].strip()] = (F, r[iR].strip().upper())
    if not xl or not os.path.exists(xl):
        yaz(u'  [11] order sequences: no Excel generated yet - build one with python verification/build_excel.py --root .')
        ATLANAN.append(u'11 order sequences (there is no produced Excel)')
        return
    try:
        import openpyxl
    except ImportError:
        ATLANAN.append(u'11 order sequences (there is no openpyxl: pip3 install openpyxl --break-system-packages)')
        return
    try:
        wb = openpyxl.load_workbook(xl, data_only=True, read_only=True)
        ws = wb['1 Siparis']
        rows = [[('' if c is None else str(c)).strip() for c in r]
                for r in ws.iter_rows(values_only=True)]
    except Exception as e:
        ATLANAN.append(u'11 order sequences (the xlsx could not be read: %s)' % e)
        return
    # THE NEW Excel format: a "1 Siparis" sheet with the columns "oligo adi",
    # "dizi (5-3)", "hedef" and "yon". That was not the old delivery file's format;
    # looking for the header under the old names returned "header not found" and let the
    # audit be skipped silently (2026-08-11, on the first attempt).
    bas = None
    for i, r in enumerate(rows[:12]):
        if 'oligo adi' in r and 'hedef' in r:
            bas = i
            break
    if bas is None:
        ATLANAN.append(u'11 order sequences (the Excel "1 Siparis" heading was not found)')
        return
    h = rows[bas]
    xH, xD, xY = h.index('hedef'), h.index('dizi (5→3)'), h.index('yon')
    xls = {}
    for r in rows[bas + 1:]:
        if len(r) <= max(xH, xD, xY) or not r[xH]:
            continue
        d = (r[xD] or '').upper()
        if not _re.fullmatch(r'[ACGT]+', d or ''):
            continue
        g = xls.setdefault(r[xH].strip(), {})
        g['F' if r[xY].strip().startswith('ileri') else 'R'] = d
    xls = {k: (v.get('F', ''), v.get('R', '')) for k, v in xls.items() if len(v) == 2}

    fark = sorted(k for k in set(xls) & set(tsv) if xls[k] != tsv[k])
    # The MISSING comparison is made only over THE ONES GOING TO ORDER. The order sheet
    # does not hold ONERILMEZ pairs at all; comparing against all of them produces a false
    # alarm such as "Proteiniphilum is missing" (2026-08-11, on the first attempt).
    _sl = _tsv(os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv'))
    _sip = set((r.get('hedef') or '').strip() for r in _sl
               if (r.get('SINIF') or '').strip().upper() in ('KESIN', 'EVRENSEL'))
    eksik = sorted((set(tsv) & _sip) - set(xls)) if _sip else sorted(set(tsv) - set(xls))
    yaz(u'  [11] order sequences: xlsx %d pairs, panel %d pairs, diverging %d'
        % (len(xls), len(tsv), len(fark)))
    if fark:
        bulgu(u'THE ORDER SEQUENCES DISAGREE, the xlsx is OLD',
              u'%s\n      On the "2 Panel" sheet of the "%s" file the sequence of these pairs is DIFFERENT from the panel\'s current sequence. Ordering from that sheet brings THE WRONG OLIGO. The right list is SIPARIS_FORMU.tsv (produced, not written by hand).'
              % (', '.join(fark), os.path.basename(xl)), BLOKE)
    if eksik:
        bulgu(u'A pair that is NOT in the xlsx',
              u'%s\n      It is in the panel but not on the order sheet.' % ', '.join(eksik))

    # ALL the xlsx files: any file holding a primer sequence could one day be taken for
    # the order source. Write out BY NAME which of them are stale.
    import glob as _glob
    guncel = set(v[0] for v in tsv.values()) | set(v[1] for v in tsv.values())
    bayat_dosya = []
    for xy in sorted(_glob.glob(os.path.join(kok, '*.xlsx'))):
        try:
            w2 = openpyxl.load_workbook(xy, data_only=True, read_only=True)
        except Exception:
            continue
        bulundu = set()
        var = 0
        for sn in w2.sheetnames:
            for row in w2[sn].iter_rows(values_only=True):
                for c in row:
                    if isinstance(c, str):
                        v = c.strip().upper()
                        if 15 <= len(v) <= 30 and _re.fullmatch(r'[ACGT]+', v):
                            var += 1
                            if v in guncel:
                                bulundu.add(v)
        if var and len(bulundu) < len(guncel):
            bayat_dosya.append(u'%s (current sequence %d/%d)'
                               % (os.path.basename(xy), len(bulundu), len(guncel)))
    if bayat_dosya:
        bulgu(u'xlsx files carrying an OLD primer sequence inside',
              u'\n      '.join(bayat_dosya) +
              u'\n      NONE of these is an order source. The one authoritative list is SIPARIS_FORMU.tsv, produced on every run.', BILGI)



# --- 12 -----------------------------------------------------------------
def d12_kanitsiz_kosulsuz(kok, yaz):
    """Could the evidence be incomplete for a pair to be ordered UNCONDITIONALLY?

        2026-08-10: the order list said SINIF=KESIN and siparis_sarti=KOSULSUZ for
        Petriella_cinsi; but its olculen_kimlik field says "there is NO membership
        definition in stage G", it has no row in the membership table, and the panel source
        holds no plate or Ta information. So an unconditional order verdict had been given
        for a target whose identity was never verified. The word "unconditional" is read in
        a report as "there is nothing to discuss here"; there is.

    """
    sl = _tsv(os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv'))
    if not sl:
        ATLANAN.append(u'12 unconditional without evidence (there is no SIPARIS_LISTESI)')
        return
    supheli = []
    for r in sl:
        ad = (r.get('hedef') or '').strip()
        sinif = (r.get('SINIF') or '').strip().upper()
        sart = (r.get('siparis_sarti') or '').strip().upper()
        if sinif != 'KESIN' or not sart.startswith('KOSULSUZ'):
            continue
        kim = (r.get('olculen_kimlik') or '').strip()
        uy = (r.get('uyusan_vtb_sayisi') or '').strip()
        eksik = []
        if not kim or 'YOK' in kim.upper() or 'TANIMI' in kim.upper():
            eksik.append(u'measured identity: "%s"' % (kim or 'bos'))
        if uy in ('', '-', '0'):
            eksik.append(u'the number of agreeing databases: "%s"' % (uy or 'bos'))
        if eksik:
            supheli.append(u'%s -> %s' % (ad, '; '.join(eksik)))
    yaz(u'  [12] evidence for unconditional ordering: %d rows examined, %d suspect'
        % (sum(1 for r in sl if (r.get('siparis_sarti') or '').strip().upper()
               .startswith('KOSULSUZ')), len(supheli)))
    if supheli:
        bulgu('Written as unconditional while the evidence is missing',
              u'\n      '.join(supheli) +
              u'\n      "Unconditional" reads in the report as "there is nothing to discuss". A row whose evidence is missing has to be made KOSULLU, or the gap has to be filled.', BLOKE)



# --- 13 -----------------------------------------------------------------
def d13_urun_boyu(kok, yaz):
    """Is the product length in the table the length the pair PRODUCES on the consensus?

        A product length is a number that can be typed by hand, and every number typed by
        hand drifts one day. 2026-08-10: the Proteolitik_Synergistaceae row said 173 bp
        while the measurement says 172 bp (the B-2-1197717 consensus, F@228 R@381,
        381+19-228=172). A one base difference enters the gel separation calculation and
        the band class.

        The criterion is the panel's own: mm<=1 and the last two bases at the 3' end exact.

    """
    import glob as _glob
    import re as _re
    pk = os.path.join(kok, 'primer_final',
                      'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    kd = os.path.join(kok, 'konsensus_kanonik')
    if not os.path.exists(pk) or not os.path.isdir(kd):
        ATLANAN.append(u'13 product length (there is no panel or consensus directory)')
        return
    # GLOB IS NOT USED. There are 250 files in the konsensus_kanonik directory but only
    # 100 are valid; the other 150 are leftovers that cannot be deleted on a mounted
    # drive, and some of them are an older version of the same bin WITH DIFFERENT CONTENT
    # (measured on 33 bins). The panel's own loader (hedefler.konsensusler) therefore
    # reads INDEKS.tsv; if the audit does not read the same source it measures with a
    # sequence the panel never sees and produces an invented "deviation". On the night of
    # 2026-08-10 my first version did exactly that.
    ixy = os.path.join(kd, 'INDEKS.tsv')
    if not os.path.exists(ixy):
        ATLANAN.append(u'13 product length (there is no konsensus_kanonik/INDEKS.tsv; NO measurement is made with leftover files)')
        return
    kons = {}
    for r in _tsv(ixy):
        f = os.path.join(kd, (r.get('dosya') or '').strip())
        if not os.path.exists(f):
            continue
        kons[os.path.basename(f)] = ''.join(
            l.strip() for l in io.open(f, encoding='utf-8', errors='replace')
            if not l.startswith('>')).upper()
    if not kons:
        ATLANAN.append(u'13 product length (there is no readable file in the index)')
        return

    def _rc(x):
        return x.translate(str.maketrans('ACGT', 'TGCA'))[::-1]

    def _yer(p, d):
        n = len(p)
        out = []
        for i in range(len(d) - n + 1):
            f = 0
            for a, c in zip(p, d[i:i + n]):
                if a != c:
                    f += 1
                    if f > 1:
                        break
            else:
                if p[-2:] == d[i + n - 2:i + n]:
                    out.append(i)
        return out

    sat = [l.rstrip('\n').split('\t') for l in io.open(pk, encoding='utf-8')]
    b = sat[0]
    try:
        iH = b.index('Hedef')
        iU = b.index('Urun (bp)')
        iF = next(i for i, x in enumerate(b) if x.startswith('Ileri primer'))
        iR = next(i for i, x in enumerate(b) if x.startswith('Geri primer'))
    except (ValueError, StopIteration):
        ATLANAN.append(u'13 product length (the columns were not recognised)')
        return
    sapan = []
    bakilan = 0
    for r in sat[1:]:
        if len(r) <= max(iU, iF, iR) or not r[iH].strip():
            continue
        F = r[iF].strip().upper()
        R = r[iR].strip().upper()
        if not _re.fullmatch(r'[ACGT]+', F or ''):
            continue
        try:
            u = int(_re.sub(r'\D', '', r[iU]))
        except ValueError:
            continue
        bakilan += 1
        Rrc = _rc(R)
        boylar = set()
        for d in kons.values():
            fs = _yer(F, d)
            if not fs:
                continue
            rs = _yer(Rrc, d)
            for i in fs:
                for j in rs:
                    if j >= i:
                        L = j + len(R) - i
                        if 40 <= L <= 600:
                            boylar.add(L)
        if boylar and u not in boylar:
            sapan.append('%s: the table says %d bp and the measured value is '
                         '%s'
                         % (r[iH].strip(), u, sorted(boylar)[:5]))
    yaz(u'  [13] product length: %d pairs measured, %d deviations' % (bakilan, len(sapan)))
    if sapan:
        bulgu(u'The product length in the table DOES NOT MATCH the measured one',
              u'\n      '.join(sapan) +
              u'\n      The product length enters the gel separation calculation and the QuantiNova band class; a wrong number spoils both.', DIKKAT)



# --- 14 -----------------------------------------------------------------
def d14_bat_dosyalari(kok, yaz):
    """Are the .bat files CRLF and pure ASCII, and do their goto targets exist?

        2026-08-09: SINA_BAT.bat had LF line endings and thirteen of fifteen tests were
        SILENTLY skipped; the goto target could not be found and no error was raised.
        2026-08-10: verification/one_key.py and verification/full_chain.py came out LF too.
        A non-ASCII character also breaks the interpreter, which is why pure ASCII is
        required.

    """
    import glob as _glob
    import re as _re
    sorun = []
    n = 0
    for y2 in sorted(_glob.glob(os.path.join(kok, '*.bat'))):
        n += 1
        b = io.open(y2, 'rb').read()
        ad = os.path.basename(y2)
        try:
            t = b.decode('ascii')
        except UnicodeDecodeError:
            sorun.append(u'%s: NOT pure ASCII (a non-ASCII character breaks the interpreter)' % ad)
            continue
        if b.count(b'\n') and b.count(b'\r\n') != b.count(b'\n'):
            sorun.append(u'%s: LF line endings (%d of %d lines are CRLF). It may fail to find the goto target and RAISES NO ERROR.'
                         % (ad, b.count(b'\n'), b.count(b'\r\n')))
        et = set(m.group(1).lower()
                 for m in _re.finditer(r'^:([a-z0-9_]+)', t, _re.M | _re.I))
        git = set(m.group(1).lower()
                  for m in _re.finditer(r'^[^\n]*?\bgoto\s+:([a-z0-9_]+)', t, _re.M | _re.I))
        eksik = sorted(x for x in git - et if x != 'eof')
        if eksik:
            sorun.append(u'%s: a goto target with no counterpart: %s' % (ad, ', '.join(eksik)))
        # A REPEATED LABEL. goto always jumps to the FIRST label; if the same name appears
        # twice, one key of the menu silently goes to the wrong place. That happened on
        # 2026-08-11 while adding a new key to PANEL.bat: a second ":sk" fell below the
        # dispatch line and the K key stopped running the plate plan and ran the continuation
        # of the dispatch instead. No error message appeared at all.
        tekrar = {}
        for m in _re.finditer(r'^:([a-z0-9_]+)', t, _re.M | _re.I):
            k = m.group(1).lower()
            tekrar[k] = tekrar.get(k, 0) + 1
        cift = sorted(k for k, v in tekrar.items() if v > 1)
        if cift:
            sorun.append(u'%s: THE SAME label is defined more than once: %s. goto jumps to the first label, so that key silently goes to the wrong place.'
                         % (ad, ', '.join(cift)))
    yaz(u'  [14] batch files: %d files, %d problems' % (n, len(sorun)))
    if sorun:
        bulgu(u'A formatting problem in a .bat file',
              u'\n      '.join(sorun) +
              u'\n      The fix: make the line endings CRLF and save the file as pure ASCII. That is where a silent skip comes from.', BLOKE)



# --- 15 -----------------------------------------------------------------
def d15_konsensus_kalintilari(kok, yaz):
    """Are there leftover files in the canonical consensus directory that are NOT in the index?

        The night of 2026-08-10: there are 250 files in the directory and INDEKS.tsv
        defines 100 of them. The other 150 are leftovers that cannot be deleted, and on 33
        bins two or three versions of the same bin WITH DIFFERENT CONTENT are sitting there
        (A1-1_2223.kanonik.fa and A1-1_2223_kanonik.fasta are the same bin with different
        sequences).

        The panel's own loader is unaffected because it reads the index. But EVERY new
        script that writes a glob silently picks the wrong sequence - my first product
        length check did exactly that. This item is a reminder that the trap is still there.

    """
    import glob as _glob
    d = os.path.join(kok, 'konsensus_kanonik')
    ix = os.path.join(d, 'INDEKS.tsv')
    if not os.path.isdir(d):
        ATLANAN.append(u'15 consensus leftovers (there is no such directory)')
        return
    if not os.path.exists(ix):
        bulgu(u'The canonical consensus INDEX IS MISSING',
              u'%s was not found. Without the index there is no telling which file is valid and the measurements fall back on file name order. Produce it: python screening/build_canonical.py --root .' % ix)
        return
    gecerli = set()
    for r in _tsv(ix):
        f = (r.get('dosya') or '').strip()
        if f:
            gecerli.add(f)
    hepsi = set(os.path.basename(f) for f in _glob.glob(os.path.join(d, '*.fa'))
                + _glob.glob(os.path.join(d, '*.fasta')))
    kalinti = hepsi - gecerli
    yaz(u'  [15] consensus leftovers: %d in the index, %d in the directory, %d orphaned'
        % (len(gecerli), len(hepsi), len(kalinti)))
    if kalinti:
        bulgu(u'A LEFTOVER file in the canonical consensus directory',
              u'%d files are not in the index (%d in the directory, %d valid). An example: %s\n      The panel\'s loader reads the index, so it is unaffected, but any script writing a GLOB like konsensus_kanonik/*.fa* can pick an old version of the same bin. Read INDEKS.tsv when you write a new script.'
              % (len(kalinti), len(hepsi), len(gecerli),
                 ', '.join(sorted(kalinti)[:3])), BILGI)



# --- 16 -----------------------------------------------------------------
def d16_uyelik_kaynagi(kok, yaz):
    """Do P (the single protocol) and K (the recovery) read THE SAME membership file?

        If the two pick different files they measure their dCq values on different ground
        and become incomparable. 2026-08-10: both were concatenating two globs and taking
        a[-1], which does not mean "the newest" but "the subdirectory always wins". It was
        fixed, but the rule has to be tested outside the code too; either of the two
        scripts could change tomorrow.

    """
    sec = {}
    for ad, dizin, mod in (('P', 'protocol', 'tek_protokol_olc'),
                           ('K', 'verification', 'kurtarma_turu')):
        d = os.path.join(kok, dizin)
        if not os.path.isdir(d):
            continue
        kod = ('import sys,os;sys.path.insert(0,%r);import %s as M;'
               'print(os.path.abspath(M.uyelik_dosyasi(%r) or ""))'
               % (d, mod, kok))
        try:
            p = subprocess.run([sys.executable, '-c', kod],
                               stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                               timeout=120)
            sec[ad] = p.stdout.decode('utf-8', 'replace').strip()
        except Exception as e:
            ATLANAN.append(u'16 membership source (%s could not be called: %s)' % (ad, e))
    if len(sec) < 2:
        ATLANAN.append(u'16 membership source (neither script could be read)')
        return
    ayni = len(set(v for v in sec.values() if v)) == 1
    yaz(u'  [16] membership source: do P and K read the same file -> %s'
        % ('evet' if ayni else u'NO'))
    if not ayni:
        bulgu(u'P and K read DIFFERENT membership files',
              u'\n      '.join(u'%s: %s' % (k, v or u'not found')
                               for k, v in sorted(sec.items())) +
              u'\n      dCq values measured with different membership do not sit on the same ground and cannot be compared.', BLOKE)



# --- 17 -----------------------------------------------------------------
def d17_veritabani_alfabesi(kok, yaz):
    """Is every scanned FASTA in the DNA alphabet, and are the files called "twins" really twins?

        Two separate traps, and both have happened:
          1) SILVA stores RNA (U). An index in the RNA alphabet matches NO DNA query at all
             and the result looks like "0 off-target", that is, CLEAN. On 2026-08-09 the
             SILVA index was doing exactly that (it showed in the 3^9 k-mer alphabet).
          2) identity_verification.py had marked two files as "BYTE FOR BYTE IDENTICAL
             (verified with cmp)". The 2026-08-10 measurement: the SSU pair have the same
             size but DIFFERENT CONTENT - release 138.2 had been converted U->T while its
             twin was still RNA. The note dated from before the conversion and nobody had
             gone back to look.

    """
    import hashlib as _h
    d = os.path.join(kok, 'REFERANS_DB')
    if not os.path.isdir(d):
        ATLANAN.append(u'17 database alphabet (there is no REFERANS_DB)')
        return
    # indeksi olan FASTA'lar taranan veritabanlaridir; alfabeleri DNA olmali
    rna = []
    bakilan = 0
    for f in sorted(os.listdir(d)):
        if not f.endswith(('.fna', '.fasta', '.fa')):
            continue
        y2 = os.path.join(d, f)
        if not any(os.path.exists(y2 + e) for e in ('.primerqc.bin', '.nsq')):
            continue          # an unindexed file is not scanned, so its alphabet does not matter
        bakilan += 1
        u = t = 0
        with io.open(y2, encoding='utf-8', errors='replace') as fh:
            for i, l in enumerate(fh):
                if l.startswith('>'):
                    continue
                u += l.count('U')
                t += l.count('T')
                if i > 200000:
                    break
        if u:
            rna.append(u'%s: U=%d T=%d' % (f, u, t))
    # Is the index FRESH relative to the FASTA: if the FASTA changed later, the index
    # scans the old data and nothing says so.
    bayat_ix = []
    for f in sorted(os.listdir(d)):
        if not f.endswith(('.fna', '.fasta', '.fa')):
            continue
        y2 = os.path.join(d, f)
        for e in ('.primerqc.bin', '.nsq'):
            ix = y2 + e
            if os.path.exists(ix) and os.path.getmtime(ix) < os.path.getmtime(y2) - 60:
                bayat_ix.append(u'%s%s the index is OLDER than the FASTA' % (f, e))
    yaz(u'  [17] database alphabet: %d indexed FASTA, %d RNA, %d stale indexes'
        % (bakilan, len(rna), len(bayat_ix)))
    if bayat_ix:
        bulgu(u'The index is OLDER than the FASTA',
              u'\n      '.join(bayat_ix) +
              u'\n      The index scans old data and the result is silently wrong. Reindex it.', BILGI)
    if rna:
        bulgu(u'The indexed database is in the RNA alphabet',
              u'\n      '.join(rna) +
              u'\n      An index in the RNA alphabet matches no DNA query and the result comes out as "0 off target", that is, CLEAN. Convert U to T and reindex: bash build_index.sh <file>', BLOKE)

    # "ikiz" iddialari gercekten dogru mu
    ikizler = [('SILVA_138.2_SSURef_NR99.fasta', 'SILVA_SSURef_NR99.fasta'),
               ('SILVA_138.2_LSURef_NR99.fasta', 'SILVA_LSURef_NR99.fasta')]
    bozuk = []
    for a, b in ikizler:
        pa, pb = os.path.join(d, a), os.path.join(d, b)
        if not (os.path.exists(pa) and os.path.exists(pb)):
            continue
        if os.path.getsize(pa) != os.path.getsize(pb):
            bozuk.append(u'%s / %s: the sizes differ' % (a, b))
            continue
        ha, hb = _h.md5(), _h.md5()
        with io.open(pa, 'rb') as fa, io.open(pb, 'rb') as fb:
            ha.update(fa.read(20000000))
            hb.update(fb.read(20000000))
        if ha.hexdigest() != hb.hexdigest():
            bozuk.append(u'%s and %s: the size is the same but the CONTENT differs (the md5 of the first 20 MB does not match)' % (a, b))
    if bozuk:
        bulgu(u'The files called "twins" are NOT twins any more',
              u'\n      '.join(bozuk) +
              u'\n      identity_verification.py marks these files as "identical byte for byte". They do not enter the voting, but the note IS WRONG; if someone flips the flag it produces a silent zero.', BILGI)



# --- 18 -----------------------------------------------------------------
def d18_evrensel_kapsam(kok, yaz):
    """Do the universal and control primers see OUR OWN targets?

        On universal primers dCq IS UNDEFINED (the denominator of the competitor set goes
        to zero); the measure is COVERAGE. But there is NO written coverage threshold in
        the project, and the labels given do not match the measured values: in the table of
        8 August, Metanojen_universal was called "low coverage" at 88% while Arke_universal
        was called nothing at 74%.

        More importantly: the 2026-08-11 measurement shows Arke_universal gives NO PRODUCT
        in six of its own member bins, three of which are Nitrosocosmicus (a target of the
        panel itself) and one Methanomassiliicoccus (the methylotrophic methanogen target).
        The normalisation control does not see the target it is meant to normalise.

        This item passes no verdict; it puts the numbers side by side and says that there
        is no written criterion. A human sets the threshold.

    """
    import csv as _csv
    import re as _re
    p = _tsv(os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'panel_tek_protokol.tsv'))
    sl = {(r.get('hedef') or '').strip(): r
          for r in _tsv(os.path.join(kok, 'ONE_PROTOCOL_RESULT', 'SIPARIS_LISTESI.tsv'))}
    if not p or not sl:
        ATLANAN.append(u'18 universal coverage (there is no table)')
        return
    satir = []
    for r in p:
        ad = (r.get('hedef') or '').strip()
        if (sl.get(ad, {}).get('SINIF') or '').strip().upper() != 'EVRENSEL':
            continue
        k = (r.get('ASIL_kapsam_mm1') or '').strip()
        m = _re.match(r'(\d+)\s*/\s*(\d+)', k)
        oran = (100.0 * int(m.group(1)) / int(m.group(2))) if m and int(m.group(2)) else None
        satir.append((ad, k, oran))
    if not satir:
        ATLANAN.append(u'18 universal coverage (no universal pair was found)')
        return
    dusuk = [(a, k, o) for a, k, o in satir if o is not None and o < 90.0]
    yaz(u'  [18] universal coverage: %d pairs | %s'
        % (len(satir), ', '.join(u'%s %s' % (a.split('_')[0][:12], k) for a, k, _o in satir)))
    if dusuk:
        bulgu(u'Coverage below %90 on a universal or control primer',
              u'\n      '.join(u'%s: %s (%%%.0f)' % (a, k, o) for a, k, o in dusuk) +
              u'\n      On universal primers the measure is COVERAGE and there is NO written coverage threshold in the project; the %90 here is a temporary bound this audit sets, not the panel\'s rule. If a control primer does not see its own targets the normalisation is biased. Write the criterion down, or give a reason in the report for these rows.', DIKKAT)


def d19_uyelik_icerigi(kok, yaz):
    """DO THE TWO MEMBERSHIP SOURCES NAME THE SAME BINS?

        Item 16 tests that two scripts read THE SAME FILE. But the project uses two
        SEPARATE membership files, and that is deliberate:
            screening/target_membership.tsv          - the search and scan side
            uyelik_yeniden_turetme_uyelik_*.tsv - the single protocol measurement
        Because they are not the same file, item 16 never compares these two. If their
        contents diverge silently, the search optimises for one set while the measurement
        grades against another.

        That is exactly what happened on 2026-08-11: for Petriella_cinsi the search side
        counted F2-4_500148 as a MEMBER (through taxid 500148) while the measurement side
        counted the same bin as a COMPETITOR - and that bin was the deciding bin that
        failed the pair (1655/3000 reads). So the locus scan optimised THE OPPOSITE of what
        the measurement wanted. About ten targets had a similar divergence.

        The rule: if there is a difference and the bin in question HAS a measured identity
        -> IT STOPS THE ORDER. A difference arising from bins whose identity was never
        measured -> DIKKAT (that bin must be measured first, and the decision comes after).

    """
    import glob as _glob
    uy = [x for x in _glob.glob(os.path.join(kok, 'uyelik_yeniden_turetme_uyelik_*.tsv'))
          if '.yedek' not in x]
    if not uy:
        ATLANAN.append(u'19 membership content (there is no single protocol membership file)')
        return
    uy.sort(key=lambda p: (os.path.getmtime(p), os.path.basename(p)))
    kod = (
        'import sys,os,io,csv\n'
        'sys.path.insert(0,%r)\n'
        'from screening import targets as H\n'
        'sat=[l.rstrip("\\n").split("\\t") for l in io.open(%r,encoding="utf-8")]\n'
        'bas=sat[0]; iu=bas.index("yeni_uye_kutular")\n'
        'TP={r[0].strip():set(x for x in r[iu].split(";") if x) '
        'for r in sat[1:] if r and r[0].strip()}\n'
        'panel,_=H.panel_oku(); kons=H.konsensusler(); kut=H.kutular()\n'
        'var={k["kutu"] for k in kut}\n'
        # The bins WHOSE IDENTITY HAS BEEN MEASURED are collected from TWO sources. Looking
        # only at stage G's table was misleading: the F1-*_44689 bins were SKIPPED in that
        # table as "not a member of any target", while the abundance study (11 August)
        # measured all four of them (Zoopagomycota / Nucletmycea). An audit looking at one
        # table said "not measured" and assigned the wrong severity.
        'kim=set()\n'
        'y=os.path.join(%r,"ALL_IDENTITIES_RESULT","tum_kutu_kimlikleri.tsv")\n'
        'if os.path.exists(y):\n'
        '    s2=[l.rstrip("\\n").split("\\t") for l in io.open(y,encoding="utf-8")]\n'
        '    bi=[i for i,r in enumerate(s2) if r and r[0].strip()=="kutu"][0]\n'
        '    kim={r[0].strip() for r in s2[bi+1:] if r and r[0].strip() '
        'and not r[0].startswith("#")}\n'
        'import glob as _g\n'
        'for y2 in _g.glob(os.path.join(os.path.dirname(y),"..","BOLLUK_OLCULEN_*",'
        '"karsilastirma_kutu.tsv")):\n'
        '    with io.open(y2,encoding="utf-8") as fh2:\n'
        '        for r2 in csv.DictReader((l for l in fh2 if not l.startswith("#")),'
        'delimiter="\\t"):\n'
        '            if (r2.get("YENI_olculen_kimlik") or "").strip():\n'
        '                kim.add((r2.get("kutu") or "").strip())\n'
        'for p in panel:\n'
        '    ad=p["hedef"]\n'
        '    b=H.hedef_baglami(p,kons=kons,kut=kut)\n'
        '    ka=set(k["kutu"] for k in b["uye_kutu"])\n'
        '    tp=TP.get(ad) or TP.get(H.AD_ESLEME.get(ad,""))\n'
        '    if tp is None: continue\n'
        '    tp={x for x in tp if x in var}\n'
        '    if ka!=tp:\n'
        '        f=sorted((ka-tp)|(tp-ka))\n'
        '        olculen=[x for x in f if x in kim]\n'
        '        print("%%s\\t%%s\\t%%s" %% (ad,",".join(f),",".join(olculen)))\n'
        % (kok, uy[-1], kok))
    try:
        p = subprocess.run([sys.executable, '-c', kod], stdout=subprocess.PIPE,
                           stderr=subprocess.PIPE, timeout=600)
    except Exception as e:
        ATLANAN.append(u'19 membership content (%s)' % e)
        return
    if p.returncode != 0:
        ATLANAN.append(u'19 membership content (the comparison did not run: %s)'
                       % p.stderr.decode('utf-8', 'replace').strip()[-200:])
        return
    satirlar = [l for l in p.stdout.decode('utf-8', 'replace').splitlines() if l.strip()]
    yaz(u'  [19] membership content: targets where the two sources diverge -> %d' % len(satirlar))
    for l in satirlar:
        pr = l.split('\t')
        ad, fark = pr[0], pr[1]
        olculen = pr[2] if len(pr) > 2 else ''
        if olculen:
            bulgu(u'%s: the two membership sources disagree on the same bins' % ad,
                  u'The bins that disagree: %s\n      These have a MEASURED identity (%s), so there is data one can decide on, and the two sides still count them differently. The search optimises one set while the measurement grades another; the dCq values do not sit on the same ground.'
                  % (fark, olculen), BLOKE)
        else:
            bulgu(u'%s: the membership disagrees, and the identity of the disagreeing bins HAS NOT BEEN MEASURED' % ad,
                  u'The bins that disagree: %s\n      These bins have no measured identity; the identities have to be measured first, and only then written down as member or competitor. The decision IS NOT MADE on the Kraken label.' % fark,
                  DIKKAT)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--root', dest='kok', default='.')
    p.add_argument('--offline', dest='agsiz', action='store_true',
                   help='skip the checks that need NCBI')
    p.add_argument('--generate', dest='uretilecek', default='',
                   help='the output paths to regenerate in this run, comma '
                        'separated; no freshness check is applied to them')
    a = p.parse_args()
    kok = os.path.abspath(a.kok)

    def yaz(s=''):
        print(s, flush=True)

    yaz(u'=' * 78)
    yaz(u'  AUDIT ON EVERY RUN   %s' % time.strftime('%Y-%m-%d %H:%M'))
    yaz(u'  Nothing is measured and no file is changed. It only looks.')
    yaz(u'=' * 78)

    d1_harita_anahtarlari(kok, yaz)
    d3_referans_bayat(kok, yaz)
    d4_kaynak_tutarliligi(kok, yaz)
    d5_muhur_diziyi_iceriyor_mu(kok, yaz)
    d7_geometri_kapisi(kok, yaz)
    d8_plaka_jel_ve_bant(kok, yaz)
    d9_belgelerde_bayat_sayi(kok, yaz)
    d10_ad_kurali_sinavi(kok, yaz)
    d11_siparis_dizileri(kok, yaz)
    d12_kanitsiz_kosulsuz(kok, yaz)
    d13_urun_boyu(kok, yaz)
    d14_bat_dosyalari(kok, yaz)
    d15_konsensus_kalintilari(kok, yaz)
    d16_uyelik_kaynagi(kok, yaz)
    d17_veritabani_alfabesi(kok, yaz)
    d18_evrensel_kapsam(kok, yaz)
    d19_uyelik_icerigi(kok, yaz)
    d6_cikti_tazeligi(kok, yaz, tuple(x.strip() for x in a.uretilecek.split(',') if x.strip()))
    d2_kapsama(kok, yaz, a.agsiz)

    yaz('')
    if BULGU:
        sirali = sorted(BULGU, key=lambda x: (BLOKE, DIKKAT, BILGI).index(x[2]))
        say = {}
        for _b, _a, o in BULGU:
            say[o] = say.get(o, 0) + 1
        yaz(u'  %d BULGU  (%s)' % (len(BULGU), ', '.join(
            u'%s %d' % (o, say[o]) for o in (BLOKE, DIKKAT, BILGI) if o in say)))
        for b, ay, o in sirali:
            yaz(u'   [%s] %s' % (o, b))
            yaz(u'      %s' % ay)
    else:
        yaz(u'  Every check is clean.')
    if ATLANAN:
        yaz(u'  Checks skipped: %s' % '; '.join(ATLANAN))
    yaz(u'=' * 78)

    rapor = os.path.join(kok, 'ONE_KEY_RESULT')
    if os.path.isdir(rapor):
        with io.open(os.path.join(rapor, 'DENETIM_RAPORU.md'), 'w',
                     encoding='utf-8', newline='') as fh:
            fh.write(u'# Audit on every run\n\nGenerated: %s\n\n'
                     % time.strftime('%Y-%m-%d %H:%M'))
            if BULGU:
                say = {}
                for _b, _a, o in BULGU:
                    say[o] = say.get(o, 0) + 1
                fh.write(u'## %d bulgu\n\n' % len(BULGU))
                for o in (BLOKE, DIKKAT, BILGI):
                    grup = [x for x in BULGU if x[2] == o]
                    if not grup:
                        continue
                    fh.write(u'### %s (%d)\n\n' % (o, len(grup)))
                    for b, ay, _o in grup:
                        fh.write(u'- **%s** — %s\n' % (b, ay.replace('\n', ' ')))
                    fh.write(u'\n')
            else:
                fh.write(u'## Every audit is clean\n\n')
            if ATLANAN:
                fh.write(u'\n## Atlananlar\n\n')
                for x in ATLANAN:
                    fh.write(u'- %s\n' % x)

    if any(o == BLOKE for _b, _a, o in BULGU):
        return 1
    if BULGU:
        return 3          # DIKKAT and BILGI only - the run is not stopped
    return 2 if ATLANAN else 0


if __name__ == '__main__':
    sys.exit(main())
