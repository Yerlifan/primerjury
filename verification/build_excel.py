# -*- coding: utf-8 -*-
"""TEK EXCEL - panelin bugunku butun dogru verisi, URETILEREK.

NEDEN
-----
Klasorde alti ayri xlsx vardi ve hicbiri guncel degildi; besinde eski primer
dizisi duruyordu. Hangisinin dogru oldugunu soyleyen bir sey yoktu ve teslim
dosyasindan siparis verilseydi yirmi ciftin altisi yanlis oligo gelecekti.

Bu betik TEK bir Excel uretir ve icindeki her sayi projenin kendi cikti
dosyalarindan okunur. Elle yazilan tek sey aciklama metinleridir. Kaynak
dosyalarin md5 ozeti kapak sayfasina yazilir; kaynak degisirse Excel de
degisir.

Kosum:
    python verification/build_excel.py --kok .
Cikti:
    PrimerJury_PANEL_<tarih>.xlsx
"""
from __future__ import print_function

import argparse
import csv
import hashlib
import io
import os
import re
import sys
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit('HATA: openpyxl yok.  pip3 install openpyxl --break-system-packages')

BASLIK = PatternFill('solid', fgColor='1F3864')
IYI = PatternFill('solid', fgColor='E2EFDA')
UYARI = PatternFill('solid', fgColor='FFF2CC')
KOTU = PatternFill('solid', fgColor='FCE4EC')
GRI = PatternFill('solid', fgColor='F2F2F2')
KENAR = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def _tsv(yol):
    if not os.path.exists(yol):
        return []
    with io.open(yol, encoding='utf-8') as fh:
        return list(csv.DictReader((l for l in fh if not l.startswith('#')),
                                   delimiter='\t'))


def _f(x):
    try:
        return float(str(x).replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _md5(yol):
    if not os.path.exists(yol):
        return 'YOK'
    return hashlib.md5(io.open(yol, 'rb').read()).hexdigest()[:12]

def _terminal_ad(baslik):
    """Soy dizgesinin en dar GERCEK takson halkasi.

    Iki tuzak var, ikisi de olculdu (2026-08-11):
      1) Kaynak alan 120 karakterde KESILIYOR (82 satirda tam 120). Son halka
         yarim kalmis oluyor: "...;Sphaerochaeta;bacterium enrichment cu".
         Yarim parcayi ad diye yazmak "Candidatus Nitrosoc" gibi uydurma
         adlar uretir.
      2) Kesilmemis olsa bile son halka cogu kayitta takson degil, kaydin
         tanimi oluyor: "uncultured bacterium", "bacterium enrichment culture".
    Bu yuzden sondan basa yurunur; yarim ya da takson olmayan halkalar
    atlanir ve ilk GERCEK takson adi dondurulur.
    """
    b = (baslik or '').strip()
    if not b:
        return ''
    kesik = len(b) >= 120          # kaynak alanin tavani
    b2 = re.sub(r'\(%[^)]*\)', ' ', b)
    parca = [x.strip() for x in re.split(r'[;|]', b2) if x.strip()]
    if not parca:
        return ''
    # ilk halka erisim numarasi + alan olabilir: "AB854355.1.910 Archaea"
    parca[0] = re.sub(r'^[A-Z]{1,2}[_A-Z]*\d{5,}(\.\d+)*\s*', '', parca[0]).strip()
    if kesik and len(parca) > 1:
        parca = parca[:-1]         # son halka yarim, at
    kotu = ('uncultured', 'unidentified', 'unclassified', 'metagenome',
            'environmental', 'enrichment', 'clone', 'bacterium', 'archaeon',
            'prokaryote', 'incertae', 'sedis', 'unknown', 'sp', 'sp.')
    for x in reversed(parca):
        y2 = re.sub(r'^[a-z]__', '', x).strip()
        if not y2 or len(y2) < 4:
            continue
        ilk = y2.split()[0].lower().strip('.,')
        if ilk in kotu:
            continue
        if any(y2.lower().startswith(k) for k in ('uncultured', 'unidentified')):
            continue
        return y2[:60]
    return ''

def sayfa(wb, ad, basliklar, satirlar, genislik=None, notlar=None, renk=None):
    """Bir sayfa yazar. renk(satir_sozlugu) -> PatternFill ya da None."""
    ws = wb.create_sheet(ad[:31])
    r = 1
    if notlar:
        for n in notlar:
            ws.cell(row=r, column=1, value=n).font = Font(italic=True, size=9,
                                                          color='555555')
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=max(len(basliklar), 2))
            ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True,
                                                           vertical='top')
            ws.row_dimensions[r].height = 13 * (1 + len(n) // 130)
            r += 1
        r += 1
    bas_satir = r
    for j, b in enumerate(basliklar, start=1):
        c = ws.cell(row=r, column=j, value=b)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = BASLIK
        c.alignment = Alignment(wrap_text=True, vertical='center')
        c.border = KENAR
    ws.row_dimensions[r].height = 30
    for s in satirlar:
        r += 1
        f = renk(s) if renk else None
        for j, b in enumerate(basliklar, start=1):
            v = s.get(b, '')
            c = ws.cell(row=r, column=j, value=v)
            c.border = KENAR
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if f:
                c.fill = f
    for j, b in enumerate(basliklar, start=1):
        g = (genislik or {}).get(b)
        if not g:
            uzunluklar = [len(str(s.get(b, ''))) for s in satirlar] + [len(b)]
            g = min(46, max(9, int(sorted(uzunluklar)[len(uzunluklar) // 2] * 1.25) + 2))
        ws.column_dimensions[get_column_letter(j)].width = g
    ws.freeze_panes = ws.cell(row=bas_satir + 1, column=1)
    return ws


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--root', '--kok', dest='kok', default='.')
    a = p.parse_args()
    kok = os.path.abspath(a.kok)
    T = lambda *y: os.path.join(kok, *y)

    # ---------------- kaynaklar ----------------
    panel_yolu = T('primer_final', 'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    kaynaklar = {
        'panel (diziler, plaka, Ta)': panel_yolu,
        'siparis listesi (hukum)': T('TEK_PROTOKOL_SONUC', 'SIPARIS_LISTESI.tsv'),
        'panel olcumu (dCq, kapsam)': T('TEK_PROTOKOL_SONUC', 'panel_tek_protokol.tsv'),
        'geometri (primer3)': None,     # asagida en yenisi secilir
        'kutu kimlikleri': T('TUM_KIMLIK_SONUC', 'tum_kutu_kimlikleri.tsv'),
        'dogrulama (4 katman)': T('DOGRULAMA_SONUC', 'dogrulama_uc_sutun.tsv'),
        'NCBI 4. katman (siki)': T('DOGRULAMA_SONUC', 'ncbi_katman4_siki.tsv'),
        'iki esik kurali': T('ESIK_IKI_KURAL.tsv'),
        'toplanti durumu': T('toplanti_durumu.tsv'),
    }
    import glob
    gl = sorted(glob.glob(T('primer_final', 'geometri_denetimi_2026*.tsv')),
                key=os.path.getmtime)
    kaynaklar['geometri (primer3)'] = gl[-1] if gl else None

    SL = {(r.get('hedef') or '').strip(): r for r in _tsv(kaynaklar['siparis listesi (hukum)'])}
    PM = {(r.get('hedef') or '').strip(): r for r in _tsv(kaynaklar['panel olcumu (dCq, kapsam)'])}
    DG = {(r.get('hedef') or '').strip(): r for r in _tsv(kaynaklar['dogrulama (4 katman)'])}
    NC = {(r.get('hedef') or '').strip(): r for r in _tsv(kaynaklar['NCBI 4. katman (siki)'])}
    ES = {(r.get('hedef') or '').strip(): r for r in _tsv(kaynaklar['iki esik kurali'])}
    TP = _tsv(kaynaklar['toplanti durumu'])
    GEO = {}
    for r in _tsv(kaynaklar['geometri (primer3)'] or ''):
        if (r.get('Primer') or '') in ('Ileri', 'Geri'):
            GEO.setdefault((r.get('Hedef') or '').strip(), {})[r['Primer']] = r

    # panel kaynagi
    sat = [l.rstrip('\n').split('\t') for l in io.open(panel_yolu, encoding='utf-8')]
    b = sat[0]
    iP, iT, iH = b.index('Plaka'), b.index('Ta (C)'), b.index('Hedef')
    iU = b.index('Urun (bp)')
    iF = next(i for i, x in enumerate(b) if x.startswith('Ileri primer'))
    iR = next(i for i, x in enumerate(b) if x.startswith('Geri primer'))
    ciftler = []
    for r in sat[1:]:
        if len(r) <= max(iU, iF, iR) or not r[iH].strip():
            continue
        F, R = r[iF].strip().upper(), r[iR].strip().upper()
        if not re.fullmatch(r'[ACGT]+', F or ''):
            continue
        ciftler.append(dict(hedef=r[iH].strip(), plaka=r[iP].strip(), ta=r[iT].strip(),
                            F=F, R=R, urun=r[iU].strip()))

    # PANEL KAYNAGINDA SATIRI OLMAYAN ama SIPARISE GIREN cift sessizce
    # dusmesin: 2, 5 ve 6. sayfalar bu birlesik liste uzerinden kurulur.
    # (Ilk surumde Petriella_cinsi 5 ve 6. sayfalarda YOKTU - panel kaynagini
    # dolasip siparis listesini gormuyordum.)
    panelde = {c['hedef'] for c in ciftler}
    for ad, sl in SL.items():
        if ad in panelde:
            continue
        F = (sl.get('F') or '').strip().upper()
        R = (sl.get('R') or '').strip().upper()
        if not re.fullmatch(r'[ACGT]+', F or ''):
            continue
        ciftler.append(dict(hedef=ad, plaka='BELIRSIZ', ta='BELIRSIZ', F=F, R=R,
                            urun=(sl.get('urun_bp') or '').strip(),
                            panelde_yok=True))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ================= 0 OKU =================
    ws = wb.create_sheet('0 OKU')
    metin = [
        ('PrimerJury qPCR paneli - tek dogru dosya', 16, True),
        ('Uretim: %s' % time.strftime('%d.%m.%Y %H:%M'), 11, False),
        ('', 11, False),
        ('Bu dosyadaki HER SAYI projenin kendi cikti dosyalarindan okunarak '
         'uretildi. Elle yazilan tek sey aciklama metinleridir.', 11, False),
        ('', 11, False),
        ('SIPARIS BU DOSYADAN VERILIR. Klasordeki eski Excel dosyalarinda alti '
         'ciftin dizisi eskidir; oralardan kopyalanirsa yanlis oligo gelir. '
         'Onlar _SILINECEKLER klasorune tasindi.', 11, True),
        ('', 11, False),
        ('OLCUM ile TERCIH ayrimi: "1 Siparis" ve "2 Panel" sayfalarindaki '
         'diziler, Tm, GC, urun boyu ve dCq OLCUMDUR. "5 Esik" sayfasindaki iki '
         'kural arasindaki secim, "8 Bulgular" sayfasindaki acik maddeler ve '
         'plaka yerlesimi TERCIHTIR - onlari insan karara baglar.', 11, False),
        ('', 11, False),
        ('KAYNAK DOSYALAR (md5 ilk 12 hane)', 12, True),
    ]
    r = 1
    for t, boy, kalin in metin:
        c = ws.cell(row=r, column=1, value=t)
        c.font = Font(size=boy, bold=kalin)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 15 * (1 + len(t) // 110)
        r += 1
    for ad, yol in kaynaklar.items():
        ws.cell(row=r, column=1, value=ad).font = Font(bold=True)
        ws.cell(row=r, column=2, value=os.path.relpath(yol, kok) if yol else 'YOK')
        ws.cell(row=r, column=3, value=_md5(yol) if yol else '-')
        r += 1
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 16

    # ================= 1 SIPARIS =================
    oligo = []
    for c in ciftler:
        sl = SL.get(c['hedef'], {})
        sinif = (sl.get('SINIF') or '').strip().upper()
        if sinif not in ('KESIN', 'EVRENSEL'):
            continue
        for yon, dizi, ad in (('ileri', c['F'], (sl.get('oligo_adi_F') or c['hedef'] + '_F')),
                              ('geri', c['R'], (sl.get('oligo_adi_R') or c['hedef'] + '_R'))):
            oligo.append({
                'oligo adi': ad.strip(), 'dizi (5→3)': dizi, 'uzunluk': len(dizi),
                'hedef': c['hedef'], 'yon': yon, 'plaka': c['plaka'], 'Ta (C)': c['ta'],
                'urun (bp)': c['urun'], 'sinif': sinif,
                'siparis sarti': (sl.get('siparis_sarti') or '-').strip()})
    ek_uyari = [c['hedef'] for c in ciftler if c.get('panelde_yok')]
    notlar = [u'Tedarikciye bu sayfa verilir. %d cift = %d oligo. Diziler 5→3 '
              u'yonundedir, hicbirinde dejenere baz yoktur.' % (len(oligo) // 2, len(oligo))]
    if ek_uyari:
        notlar.append(u'UYARI: %s icin panel tablosunda satir yok; plaka ve Ta '
                      u'bilinmiyor, "BELIRSIZ" yazildi.' % ', '.join(ek_uyari))
    sayfa(wb, '1 Siparis',
          ['oligo adi', 'dizi (5→3)', 'uzunluk', 'hedef', 'yon', 'plaka',
           'Ta (C)', 'urun (bp)', 'sinif', 'siparis sarti'],
          oligo, {'dizi (5→3)': 28, 'hedef': 34, 'oligo adi': 26}, notlar,
          renk=lambda s: KOTU if s['plaka'] == 'BELIRSIZ' else None)

    # ================= 2 PANEL =================
    pr = []
    for c in ciftler:
        sl = SL.get(c['hedef'], {})
        pm = PM.get(c['hedef'], {})
        g = GEO.get(c['hedef'], {})
        gi, gr = g.get('Ileri', {}), g.get('Geri', {})
        ih = '; '.join(x for x in ((gi.get('Ihlal') or '').strip(),
                                   (gr.get('Ihlal') or '').strip()) if x)
        tmF, tmR = _f(gi.get('Tm')), _f(gr.get('Tm'))
        pr.append({
            'hedef': c['hedef'], 'sinif': (sl.get('SINIF') or '').strip(),
            'plaka': c['plaka'], 'Ta (C)': c['ta'],
            'ileri primer': c['F'], 'uz F': len(c['F']),
            'GC% F': gi.get('GC%', ''), 'Tm F': gi.get('Tm', ''),
            'geri primer': c['R'], 'uz R': len(c['R']),
            'GC% R': gr.get('GC%', ''), 'Tm R': gr.get('Tm', ''),
            'dTm': ('%.2f' % abs(tmF - tmR)) if (tmF and tmR) else '',
            'urun (bp)': c['urun'],
            'ayrim (x)': (sl.get('ayrim_mm1') or '').strip(),
            'dCq': (sl.get('dCq_karsiligi') or '').strip(),
            'kapsam': (pm.get('ASIL_kapsam_mm1') or '').strip(),
            'kural ihlali': ih or '-',
            '4 katman hukmu': (DG.get(c['hedef'], {}).get('KARAR') or '-')[:60]})

    def panel_renk(s):
        if str(s['4 katman hukmu']).startswith('RISKLI'):
            return KOTU
        if s['kural ihlali'] != '-':
            return UYARI
        return None
    sayfa(wb, '2 Panel',
          ['hedef', 'sinif', 'plaka', 'Ta (C)', 'ileri primer', 'uz F', 'GC% F', 'Tm F',
           'geri primer', 'uz R', 'GC% R', 'Tm R', 'dTm', 'urun (bp)', 'ayrim (x)',
           'dCq', 'kapsam', 'kural ihlali', '4 katman hukmu'],
          pr, {'hedef': 34, 'ileri primer': 26, 'geri primer': 26,
               'kural ihlali': 30, '4 katman hukmu': 30},
          [u'Tm, GC ve dTm panelin kendi motoruyla (primer3; mv 50 mM, dv 1,5 mM, '
           u'dNTP 0,6 mM, oligo 50 nM) 11.08.2026 07:02\'de yeniden olculdu. '
           u'Kirmizi satir: dort katmanli dogrulama "RISKLI" diyor ama cift '
           u'siparis listesinde duruyor - "8 Bulgular" sayfasina bakin. '
           u'Sari satir: panelin GC / 3\' uc / Tm penceresi kurallarindan sapma var.'],
          renk=panel_renk)

    # ================= 3 KIMLIKLER ve 4 CINS-TUR =================
    ksat = [x.rstrip('\n').split('\t')
            for x in io.open(kaynaklar['kutu kimlikleri'], encoding='utf-8')]
    kb = None
    for i, r in enumerate(ksat):
        if r and r[0].strip() == 'kutu':
            kb = i
            break
    K = []
    if kb is not None:
        H = ksat[kb]
        for r in ksat[kb + 1:]:
            if not r or not r[0].strip():
                continue
            # SUTUN SAYISI SARTI: dosyanin sonunda 1-2 sutunluk NOT satirlari
            # var; yalniz "ilk hucre dolu mu" diye bakmak onlari da kutu sanar
            # (ilk surumde 96 yerine 102 satir uretti). Tam satir sarti konuldu.
            if len(r) < len(H) - 1:
                continue
            K.append(dict(zip(H, r)))

    kim = []
    for r in K:
        e1, e2 = _f(r.get('en_iyi_kimlik_%')), _f(r.get('ikinci_kimlik_%'))
        kim.append({
            'kutu': r.get('kutu', ''),
            'Kraken etiketi': r.get('MEVCUT_KAYITLI_KIMLIK', ''),
            'olculen kimlik': r.get('DOGRULANAN_KIMLIK', ''),
            'ad degisti mi': r.get('UYUSUYOR_MU', ''),
            'savunulabilir duzey': r.get('SAVUNULABILIR_DUZEY', ''),
            'onerilen ad': r.get('ONERILEN_AD', ''),
            'en iyi isabet': (r.get('en_iyi_isabet') or '')[:110],
            'kimlik %': r.get('en_iyi_kimlik_%', ''),
            'veritabani': r.get('en_iyi_vtb', ''),
            'ikinci %': r.get('ikinci_kimlik_%', ''),
            'marj': ('%.2f' % (e1 - e2)) if (e1 is not None and e2 is not None) else '',
            'lokus': r.get('lokus', ''),
            'uyesi oldugu hedefler': (r.get('UYESI_OLDUGU_HEDEFLER') or '')[:70]})
    sayfa(wb, '3 Kimlikler',
          ['kutu', 'Kraken etiketi', 'olculen kimlik', 'ad degisti mi',
           'savunulabilir duzey', 'onerilen ad', 'en iyi isabet', 'kimlik %',
           'veritabani', 'ikinci %', 'marj', 'lokus', 'uyesi oldugu hedefler'],
          kim, {'olculen kimlik': 40, 'en iyi isabet': 46, 'Kraken etiketi': 26,
                'onerilen ad': 30, 'uyesi oldugu hedefler': 32},
          [u'Her kutunun kimligi bagimsiz olarak yeniden olculdu. "ad degisti mi" '
           u'sutunu HAYIR ise Kraken etiketi ile olculen kimlik ayrisiyor demektir.'],
          renk=lambda s: UYARI if str(s['ad degisti mi']).strip().upper().startswith('HAYIR') else None)

    # --- 4 Cins-Tur: tur adi VERILEMEYEN kutularda en yuksek yuzdeli tur adayi
    sys.path.insert(0, T('verification'))
    try:
        import identity_verification as KD
        ad_coz, TE, CE, AP = KD.ad_coz, KD.TUR_ESIGI, KD.CINS_ESIGI, KD.AYRIM_PAYI
    except Exception:                       # panelin fonksiyonu yoksa sayfa uretilmez
        ad_coz = None
    ct = []
    if ad_coz:
        for r in K:
            duzey = (r.get('SAVUNULABILIR_DUZEY') or '').strip().upper()
            if duzey.startswith('TUR'):
                continue                    # zaten tur adi var
            e1, e2 = _f(r.get('en_iyi_kimlik_%')), _f(r.get('ikinci_kimlik_%'))
            lokus = (r.get('lokus') or 'SSU').strip().upper()
            te = TE.get(lokus, 98.7)
            ce = CE.get(lokus, 94.5)
            c1, t1, _tam = ad_coz(r.get('en_iyi_isabet', ''))
            c2, t2, _tam2 = ad_coz(r.get('ikinci_isabet', ''))
            # ad_coz ikili ad bulamazsa soyun en dar halkasina dus; hangi
            # yoldan gelindigi "ad kaynagi" sutununda yazar.
            aday1 = t1 or (c1 + ' sp.' if c1 else '') or _terminal_ad(r.get('en_iyi_isabet', ''))
            aday2 = t2 or (c2 + ' sp.' if c2 else '') or _terminal_ad(r.get('ikinci_isabet', ''))
            kaynak1 = ('ikili ad' if t1 else ('cins' if c1 else 'soyun en dar halkasi'))
            marj = (e1 - e2) if (e1 is not None and e2 is not None) else None
            if e1 is None:
                neden = u'sayisal kimlik yok'
            elif e1 < ce:
                neden = (u'kimlik %%%.2f, cins esiginin (%%%.1f) bile altinda' % (e1, ce))
            elif e1 < te:
                neden = (u'kimlik %%%.2f, tur esiginin (%%%.1f) altinda' % (e1, te))
            elif marj is not None and marj < AP:
                neden = (u'tur esigini geciyor ama ikinci isabetle arasinda yalniz '
                         u'%%%.2f var (gereken %%%.1f)' % (marj, AP))
            else:
                neden = u'tur adi cikarilamadi (kayit basliginda ikili ad yok)'
            ct.append({
                'kutu': r.get('kutu', ''),
                'olculen duzey': r.get('SAVUNULABILIR_DUZEY', ''),
                'su an yazan ad': r.get('ONERILEN_AD', ''),
                'EN YUKSEK YUZDELI TUR ADAYI': aday1 or '-',
                'ad kaynagi': kaynak1,
                'kimlik %': ('%.2f' % e1) if e1 is not None else '',
                'ikinci aday': aday2 or '-',
                'ikinci %': ('%.2f' % e2) if e2 is not None else '',
                'marj': ('%.2f' % marj) if marj is not None else '',
                'lokus': lokus,
                'tur esigi': '%.1f' % te,
                'neden tur adi verilmedi': neden})
    sayfa(wb, '4 Cins-Tur',
          ['kutu', 'olculen duzey', 'su an yazan ad', 'EN YUKSEK YUZDELI TUR ADAYI',
           'ad kaynagi', 'kimlik %', 'ikinci aday', 'ikinci %', 'marj', 'lokus',
           'tur esigi', 'neden tur adi verilmedi'],
          ct, {'su an yazan ad': 34, 'EN YUKSEK YUZDELI TUR ADAYI': 30,
               'neden tur adi verilmedi': 46, 'ikinci aday': 24},
          [u'Tur adi VERILEMEYEN kutularda en yuksek yuzdeyi veren tur adayi. '
           u'Bu sutun bir KIMLIK IDDIASI DEGILDIR; "en yakin kayit su, su yuzdeyle" '
           u'demektir. Esikler panelin kendi kurallaridir: SSU/LSU tur %98,7, '
           u'ITS %98,5; cins %94,5 / %90,0; en iyi ile ikinci arasindaki fark '
           u'%0,5\'ten kucukse tur atamasi savunulmaz. Adlar panelin kendi ad_coz() '
           u'fonksiyonuyla cikarildi, yeni bir kural konmadi.'],
          renk=lambda s: IYI if (_f(s['kimlik %']) or 0) >= 98.7 and (_f(s['marj']) or 0) >= 0.5 else None)

    # ================= 5 ESIK =================
    es = []
    for c in ciftler:
        r = ES.get(c['hedef'])
        if not r:
            continue
        es.append({'hedef': c['hedef'], 'dCq': r.get('dCq', ''),
                   'duz esik (3,00)': r.get('duz_kural', ''),
                   'R (bolluk orani)': r.get('R', ''),
                   'gerekli dCq': r.get('gerekli_dCq', ''),
                   'bolluk kurali': r.get('bolluk_kurali', ''),
                   'iki kural ayrisiyor mu': r.get('ayrisiyor_mu', ''),
                   'R bayat mi': r.get('R_bayat_mi', ''),
                   'dCq 08.08': r.get('dCq_08_08', '')})
    sayfa(wb, '5 Esik',
          ['hedef', 'dCq', 'duz esik (3,00)', 'R (bolluk orani)', 'gerekli dCq',
           'bolluk kurali', 'iki kural ayrisiyor mu', 'R bayat mi', 'dCq 08.08'],
          es, {'hedef': 34},
          [u'Iki esik kurali yan yana. Duz kural: dCq >= 3,00. Bolluga agirlikli '
           u'kural: dCq >= max(log2(R)+4,3 ; 3,32). Hangisinin uygulanacagi bir '
           u'OLCUT TERCIHIDIR, olcum degildir; bu yuzden burada hukum yazilmadi. '
           u'"R bayat mi" EVET olan satirlarda R 8 Agustos uyeligiyle hesaplandi ve '
           u'o gunden sonra cift degisti - o satirlarda bolluk hukmu guvenilmez.'],
          renk=lambda s: UYARI if str(s['iki kural ayrisiyor mu']).strip() == 'EVET' else None)

    # ================= 6 NCBI =================
    nc = []
    for c in ciftler:
        r = NC.get(c['hedef'])
        if not r:
            continue
        nc.append({'hedef': c['hedef'],
                   'adli hedef disi (siki)': r.get('siki_kural_adli', ''),
                   'adsiz cevre klonu': r.get('siki_adsiz', ''),
                   'toplam urun': r.get('toplam', ''),
                   'gevsek kural (eski sayim)': r.get('gevsek_kural_adli', ''),
                   'ornek adli hedef disi': (r.get('ornek_adli_basliklar') or '')[:110]})
    sayfa(wb, '6 NCBI 4katman',
          ['hedef', 'adli hedef disi (siki)', 'adsiz cevre klonu', 'toplam urun',
           'gevsek kural (eski sayim)', 'ornek adli hedef disi'],
          nc, {'hedef': 34, 'ornek adli hedef disi': 52},
          [u'NCBI Primer-BLAST, GenBank\'in TAMAMINA karsi olcer. Buradaki sayilar '
           u'"bu primer DUNYADA baska neyi tutar" sorusunun cevabidir; numunede ne '
           u'oldugunu 2. katman olcer ve siparis karari oradan gelir. Adsiz cevre '
           u'klonlari ("uncultured ...") hukme GIRMEZ: hicbir taksona bagli '
           u'degiller ve hedeflerimiz adlandirilmamis soylar oldugundan hedefin '
           u'kendisi olabilirler.'])

    # ================= 7 TOPLANTI =================
    tp = [{'karar': r.get('karar', ''), 'istenen': r.get('istenen', ''),
           'istenen duzey': r.get('istenen_duzey', ''),
           'paneldeki karsiligi': r.get('paneldeki_karsiligi', ''),
           'durum': r.get('durum', ''), 'dCq': r.get('dCq', ''),
           'not': (r.get('not') or '')[:150]} for r in TP]
    sayfa(wb, '7 Toplanti',
          ['karar', 'istenen', 'istenen duzey', 'paneldeki karsiligi', 'durum',
           'dCq', 'not'],
          tp, {'istenen': 30, 'paneldeki karsiligi': 34, 'not': 60},
          [u'Toplantida numarali olarak istenen her hedef ve bugunku durumu. '
           u'Hicbir durum elle yazilmadi; siparis listesinden okundu.'],
          renk=lambda s: KOTU if str(s['durum']).startswith('YAPILAMADI') else
          (IYI if 'kosulsuz' in str(s['durum']) else None))

    # ================= 8 BULGULAR =================
    bulgu_yolu = T('TEK_TUS_SONUC', 'DENETIM_RAPORU.md')
    bl = []
    if os.path.exists(bulgu_yolu):
        onem = ''
        for l in io.open(bulgu_yolu, encoding='utf-8'):
            l = l.rstrip('\n')
            m = re.match(r'^###\s+(.+?)\s*\(\d+\)', l)
            if m:
                onem = m.group(1).strip()
                continue
            m = re.match(r'^-\s+\*\*(.+?)\*\*\s+—\s+(.*)$', l)
            if m:
                bl.append({'onem': onem, 'bulgu': m.group(1),
                           'ayrinti': m.group(2)[:300]})
    sayfa(wb, '8 Bulgular',
          ['onem', 'bulgu', 'ayrinti'], bl,
          {'onem': 20, 'bulgu': 40, 'ayrinti': 90},
          [u'Denetim kapisinin son kosusundaki acik bulgular. "SIPARISI DURDURUR" '
           u'olanlar giderilmeden siparis verilmemelidir. Ayrintili anlatim ve '
           u'karar secenekleri GECE_BULGULARI.md dosyasindadir.'],
          renk=lambda s: KOTU if 'DURDUR' in str(s['onem']).upper() else
          (UYARI if 'DIKKAT' in str(s['onem']).upper() else GRI))

    cy = T('PrimerJury_PANEL_%s.xlsx' % time.strftime('%Y%m%d'))
    wb.save(cy)
    print('yazildi: %s' % cy)
    print('  sayfalar: %s' % ', '.join(wb.sheetnames))
    print('  siparis oligo: %d | panel cifti: %d | kutu: %d | cins-tur satiri: %d'
          % (len(oligo), len(pr), len(kim), len(ct)))
    return 0


if __name__ == '__main__':
    sys.exit(main())
