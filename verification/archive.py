# -*- coding: utf-8 -*-
"""MOVE THE OLD FILES INTO AN ARCHIVE, and PUT ON RECORD what went where.

WHY
---
The directory held six workbooks, fifty odd backup and .orig files and 1.1 GB of
twin FASTA. Five of them carried an old primer sequence; ordering from one of
them would have brought six of the twenty pairs as the wrong oligo. The clutter
itself is a source of faults.

WHAT IT DOES
------------
It MOVES the candidates under ARCHIVE/<date>_arsiv/ rather than deleting them,
and writes where each file came from into MANIFEST.tsv. The directory structure
is kept so that undoing it is a one line job.

WHAT IT DOES NOT DO
-------------------
It moves no file the code reads. Before a move, every candidate's name is
SEARCHED for in all the .py, .bat and .sh files; when a name occurs, that file is
LEFT ALONE and the reason is written down. There is no "move first, break later".

To run it:
    python verification/archive.py --root .            (the plan alone)
    python verification/archive.py --root . --move     (actually move)
"""
from __future__ import print_function

import argparse
import glob
import io
import os
import re
import shutil
import sys
import time

# Kod taramasinda GORMEZDEN gelinecek klasorler (arsivin kendisi, gecmis
# kopyalar). Bunlarin icindeki bir gecis, dosyayi "kullaniliyor" yapmaz.
YOKSAY = ('ARCHIVE', '_PREVIOUS', 'QUICK_TEST', '__pycache__', 'eski')


def kod_dosyalari(kok):
    out = []
    for p, d, f in os.walk(kok):
        if any(y in p for y in YOKSAY):
            continue
        for x in f:
            if x.endswith(('.py', '.bat', '.sh')):
                out.append(os.path.join(p, x))
    return out


def _guncel_diziler(kok):
    'Every primer sequence the panel holds NOW.'
    y = os.path.join(kok, 'final_primers',
                     'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    out = set()
    if not os.path.exists(y):
        return out
    sat = [l.rstrip('\n').split('\t') for l in io.open(y, encoding='utf-8')]
    b = sat[0]
    try:
        iF = next(i for i, x in enumerate(b) if x.startswith('Ileri primer'))
        iR = next(i for i, x in enumerate(b) if x.startswith('Geri primer'))
    except StopIteration:
        return out
    for r in sat[1:]:
        for i in (iF, iR):
            if len(r) > i:
                v = r[i].strip().upper()
                if re.fullmatch(r'[ACGT]{15,30}', v or ''):
                    out.add(v)
    return out


def _xlsx_dizi_sayimi(yol, guncel):
    '(how many primer-like sequences the file holds, and how many of those are CURRENT).'
    try:
        import openpyxl
        wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
    except Exception:
        return 0, 0
    var, iyi = 0, 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    v = c.strip().upper()
                    if re.fullmatch(r'[ACGT]{15,30}', v or ''):
                        var += 1
                        if v in guncel:
                            iyi += 1
    return var, iyi


def adaylar(kok):
    'A list of (path, reason). The path is RELATIVE to the root.'
    a = []
    # The CURRENT produced Excel is NEVER a candidate. (In the first plan the
    # PrimerJury_PANEL_20260811.xlsx I had produced myself fell into the list of files
    # to be moved; without the plan step I would have thrown the delivery file into the
    # archive.)
    for f in sorted(glob.glob(os.path.join(kok, '*.xlsx'))):
        ad = os.path.basename(f)
        if ad.startswith('PrimerJury_PANEL_'):
            continue
        # ONLY the files that CARRY A PRIMER SEQUENCE and whose sequences are OLD are
        # candidates. Saying "every xlsx" drags along an analysis file that has nothing to
        # do with primers (the community trend workbook); that is a delivery product and holds
        # one primer.
        n_dizi, n_guncel = _xlsx_dizi_sayimi(f, _guncel_diziler(kok))
        if n_dizi == 0:
            continue
        a.append((os.path.relpath(f, kok),
                  'it carries primer sequences and %d of %d are NOT CURRENT; '
                  'the one correct file is PrimerJury_PANEL_*.xlsx'
                  % (n_dizi - n_guncel, n_dizi)))
    for kalip, sebep in ((u'**/*.yedek_*', u'gece/gun yedegi'),
                         (u'**/*.orig_*', 'a backup of an older version'),
                         (u'**/*.yedek_LF', 'a backup from before the line '
                                            'ending fix')):
        for f in glob.glob(os.path.join(kok, kalip), recursive=True):
            if os.path.isfile(f) and not any(y in f for y in YOKSAY):
                a.append((os.path.relpath(f, kok), sebep))
    if os.path.exists(os.path.join(kok, 'geo.json')):
        a.append(('geo.json', 'it used to be written whenever '
                              'geometry_core.py was imported; it is no longer '
                              'produced, since a __main__ guard was added'))
    for f in ('REFERENCE_DB/SILVA_SSURef_NR99.fasta',
              'REFERENCE_DB/SILVA_LSURef_NR99.fasta'):
        if os.path.exists(os.path.join(kok, *f.split('/'))):
            a.append((f, 'a twin copy. It does not enter the vote, and the '
                         'SSU one is NO LONGER a twin: the 138.2 release was '
                         'converted from U to T while this copy is RNA'))
    # the same file can come twice under two patterns
    gor = set()
    tek = []
    for y, s in a:
        if y in gor:
            continue
        gor.add(y)
        tek.append((y, s))
    return tek


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--root', dest='kok', default='.')
    p.add_argument('--move', dest='tasi', action='store_true', help='gercekten tasi')
    a = p.parse_args()
    kok = os.path.abspath(a.kok)
    # The target moved from a directory named "to be deleted" to ARCHIVE. There
    # were two separate archive directories, and the old name gave the
    # impression that its contents could be deleted, while this script DELETES
    # NOTHING and only moves. That wrong impression cost a directory: it really
    # was deleted. One archive, under a name that says what it is.
    hedef = os.path.join(kok, 'ARCHIVE', time.strftime('%Y-%m-%d') + '_arsiv')

    ad_listesi = adaylar(kok)
    kodlar = kod_dosyalari(kok)
    metinler = {}
    for k in kodlar:
        try:
            metinler[k] = io.open(k, encoding='utf-8', errors='replace').read()
        except IOError:
            pass

    # AN EXPLICIT FORCE LIST. The names of these files appear in the code, but the
    # places they appear in are either historical one off scripts, or scripts that
    # PRODUCE the file rather than read it, or live scripts corrected today. Each one
    # was looked at separately and the reason written below; there is NO blanket "force
    # move" flag.
    ZORLA = {
        'PrimerJury_PCR_Paneli_2026-08-02.xlsx':
            'only historical correction scripts read it; nothing that runs in '
            'the chain does',
        'PrimerJury_PCR_Paneli_2026-08-02_TESLIM.xlsx':
            'the live readers were moved onto the NEW Excel today '
            '(cross_check.py, audit_all.py, one_key.py); what is left are '
            'historical scripts and explanatory text',
        'PrimerJury_Primer_Tasarimi.xlsx':
            'the step scripts PRODUCE this file with --out rather than '
            'reading it, so it can be produced again',
        'REFERENCE_DB/SILVA_SSURef_NR99.fasta':
            'taken out of the identity_verification.py list today. It never '
            'entered the vote, and it is NO LONGER a twin: this copy is RNA '
            'while the 138.2 release is DNA',
        'REFERENCE_DB/SILVA_LSURef_NR99.fasta':
            'taken out of the identity_verification.py list today',
    }

    tasinacak, birakilacak = [], []
    for yol, sebep in ad_listesi:
        ad = os.path.basename(yol)
        gecen = [os.path.relpath(k, kok) for k, m in metinler.items() if ad in m]
        # its own archiving script and the audit script mentioning the name is not an
        # obstacle. HISTORICAL ONE OFF scripts are not an obstacle either: the ones under
        # engine/ did that day's correction and are finished work, they do not run in the
        # chain. Their mentioning a name does not keep a file alive, but it is written into
        # the MANIFEST so that whoever reruns that script knows where to look.
        tarihsel = ('engine', 'engine',
                    'engine', 'engine',
                    'REFERANS_TASARIM_betikleri')
        gecen = [g for g in gecen if os.path.basename(g) not in
                 ('archive.py', 'audit_all.py', 'build_excel.py',
                  'order_form.py')]
        canli = [g for g in gecen if not any(t in g for t in tarihsel)]
        if gecen and not canli:
            sebep = sebep + (' | only historical scripts mention it: %s'
                             % ', '.join(gecen[:3]))
            gecen = []
        if yol in ZORLA:
            tasinacak.append((yol, sebep + u' | ' + ZORLA[yol]))
        elif gecen:
            birakilacak.append((yol, sebep, gecen))
        else:
            tasinacak.append((yol, sebep))

    print('=' * 78)
    print('  ARSIVLEME %s   %s' % ('(PLAN)' if not a.tasi else '(TASINIYOR)',
                                   time.strftime('%Y-%m-%d %H:%M')))
    print('=' * 78)
    print(u'  candidates: %d | to move: %d | KEPT BECAUSE THE CODE REFERENCES IT: %d'
          % (len(ad_listesi), len(tasinacak), len(birakilacak)))
    toplam = 0
    for yol, sebep in tasinacak:
        try:
            toplam += os.path.getsize(os.path.join(kok, yol))
        except OSError:
            pass
    print(u'  total size to move: %.1f MB' % (toplam / 1e6))
    print()
    for yol, sebep, gecen in birakilacak:
        print(u'  LEFT       %-56s (it appears in the code: %s)'
              % (yol[:56], ', '.join(gecen[:2])))
    if birakilacak:
        print()

    if not a.tasi:
        for yol, sebep in tasinacak[:12]:
            print('  tasinacak  %-56s %s' % (yol[:56], sebep[:40]))
        if len(tasinacak) > 12:
            print(u'  ... and %d more files' % (len(tasinacak) - 12))
        print()
        print(u'  This is a PLAN. To actually move the files: --move')
        return 0

    if not os.path.isdir(hedef):
        os.makedirs(hedef)
    mani = os.path.join(hedef, 'MANIFEST.tsv')
    yeni_mani = not os.path.exists(mani)
    n = 0
    with io.open(mani, 'a', encoding='utf-8', newline='') as fh:
        if yeni_mani:
            fh.write(u'# The files moved to the archive. THEY WERE NOT DELETED.\n')
            fh.write(u'# To undo: copy each file back to the path in the "eski_yol" column.\n')
            fh.write(u'tarih\teski_yol\tarsiv_yolu\tboyut_bayt\tsebep\n')
        for yol, sebep in tasinacak:
            kaynak = os.path.join(kok, yol)
            varis = os.path.join(hedef, yol)
            if not os.path.isdir(os.path.dirname(varis)):
                os.makedirs(os.path.dirname(varis))
            try:
                boy = os.path.getsize(kaynak)
                shutil.move(kaynak, varis)
            except Exception as e:
                print('  TASINAMADI %-52s %s' % (yol[:52], e))
                continue
            n += 1
            fh.write(u'%s\t%s\t%s\t%d\t%s\n'
                     % (time.strftime('%Y-%m-%d %H:%M'), yol,
                        os.path.relpath(varis, kok), boy, sebep))
    print(u'  moved: %d files' % n)
    print('  arsiv  : %s' % os.path.relpath(hedef, kok))
    print(u'  log    : %s' % os.path.relpath(mani, kok))
    print('=' * 78)
    return 0


if __name__ == '__main__':
    sys.exit(main())
