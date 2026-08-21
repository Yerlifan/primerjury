# -*- coding: utf-8 -*-
"""ESKI DOSYALARI ARSIVE TASI  -  ve neyin nereye gittigini KAYDA GEC.

NEDEN
-----
Klasorde alti xlsx, elli kusur .yedek_/.orig_ dosyasi ve 1,1 GB ikiz FASTA
duruyordu. Besinde eski primer dizisi vardi; birinden siparis verilseydi yirmi
ciftin altisi yanlis oligo gelecekti. Karisikligin kendisi bir hata kaynagi.

NE YAPAR
--------
Adaylari _SILINECEKLER/2026-08-11_arsiv/ altina TASIR (silmez) ve her dosyanin
nereden geldigini MANIFEST.tsv'ye yazar. Geri almak tek satirlik is olsun diye
klasor yapisi korunur.

NE YAPMAZ
---------
Kodun okudugu hicbir dosyayi tasimaz. Tasimadan once her adayin adi butun
.py/.bat/.sh dosyalarinda ARANIR; gecen bir ad varsa o dosya BIRAKILIR ve
sebebi yazilir. "Once tasi sonra kir" olmaz.

Kosum:
    python KURTARMA/arsivle.py --kok .            (yalniz plan)
    python KURTARMA/arsivle.py --kok . --tasi     (gercekten tasi)
"""
from __future__ import print_function

import argparse
import glob
import io
import os
import re
import shutil
import sys
import time

# Kod taramasinda GORMEZDEN gelinecek klasorler (arsivin kendisi, gecmis
# kopyalar). Bunlarin icindeki bir gecis, dosyayi "kullaniliyor" yapmaz.
YOKSAY = ('_SILINECEKLER', '_ONCEKI', 'HIZLI_TEST', '__pycache__', 'eski')


def kod_dosyalari(kok):
    out = []
    for p, d, f in os.walk(kok):
        if any(y in p for y in YOKSAY):
            continue
        for x in f:
            if x.endswith(('.py', '.bat', '.sh')):
                out.append(os.path.join(p, x))
    return out


def _guncel_diziler(kok):
    """Panelin SU ANKI butun primer dizileri."""
    y = os.path.join(kok, 'primer_final',
                     'devir_ciftleri_20260802_sonrotus_TESLIM.tsv')
    out = set()
    if not os.path.exists(y):
        return out
    sat = [l.rstrip('\n').split('\t') for l in io.open(y, encoding='utf-8')]
    b = sat[0]
    try:
        iF = next(i for i, x in enumerate(b) if x.startswith('Ileri primer'))
        iR = next(i for i, x in enumerate(b) if x.startswith('Geri primer'))
    except StopIteration:
        return out
    for r in sat[1:]:
        for i in (iF, iR):
            if len(r) > i:
                v = r[i].strip().upper()
                if re.fullmatch(r'[ACGT]{15,30}', v or ''):
                    out.add(v)
    return out


def _xlsx_dizi_sayimi(yol, guncel):
    """(dosyadaki primer benzeri dizi sayisi, bunlarin GUNCEL olani)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
    except Exception:
        return 0, 0
    var, iyi = 0, 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            for c in row:
                if isinstance(c, str):
                    v = c.strip().upper()
                    if re.fullmatch(r'[ACGT]{15,30}', v or ''):
                        var += 1
                        if v in guncel:
                            iyi += 1
    return var, iyi


def adaylar(kok):
    """(yol, sebep) listesi. Yol koke GORELIDIR."""
    a = []
    # GUNCEL uretilen Excel ASLA aday olmaz. (Ilk planda kendi urettigim
    # MicRhoBooster_PANEL_20260811.xlsx tasinacaklar listesine dusmustu -
    # plan adimi olmasa teslim dosyasini arsive atacaktim.)
    for f in sorted(glob.glob(os.path.join(kok, '*.xlsx'))):
        ad = os.path.basename(f)
        if ad.startswith('MicRhoBooster_PANEL_'):
            continue
        # SADECE PRIMER DIZISI TASIYAN ve o dizileri ESKI olan dosyalar aday.
        # "Butun xlsx'ler" demek, primerle ilgisi olmayan bir analiz dosyasini
        # (Topluluk_Trend_Analizi) da surukler - o bir teslim urunudur ve
        # icinde tek primer yoktur.
        n_dizi, n_guncel = _xlsx_dizi_sayimi(f, _guncel_diziler(kok))
        if n_dizi == 0:
            continue
        a.append((os.path.relpath(f, kok),
                  u'primer dizisi tasiyor ve %d/%d tanesi GUNCEL DEGIL; '
                  u'tek dogru dosya MicRhoBooster_PANEL_*.xlsx'
                  % (n_dizi - n_guncel, n_dizi)))
    for kalip, sebep in ((u'**/*.yedek_*', u'gece/gun yedegi'),
                         (u'**/*.orig_*', u'eski surum yedegi'),
                         (u'**/*.yedek_LF', u'satir sonu duzeltmesi oncesi yedek')):
        for f in glob.glob(os.path.join(kok, kalip), recursive=True):
            if os.path.isfile(f) and not any(y in f for y in YOKSAY):
                a.append((os.path.relpath(f, kok), sebep))
    if os.path.exists(os.path.join(kok, 'geo.json')):
        a.append(('geo.json', u'geo.py ice aktarilinca yaziliyordu; artik '
                  u'uretilmiyor (__main__ korumasi eklendi)'))
    for f in ('REFERANS_DB/SILVA_SSURef_NR99.fasta',
              'REFERANS_DB/SILVA_LSURef_NR99.fasta'):
        if os.path.exists(os.path.join(kok, *f.split('/'))):
            a.append((f, u'ikiz kopya; oylamaya girmiyor ve SSU olani artik '
                      u'ikiz DEGIL (138.2 surumu U->T cevrildi, bu kopya RNA)'))
    # ayni dosya iki kalibla iki kez gelebilir
    gor = set()
    tek = []
    for y, s in a:
        if y in gor:
            continue
        gor.add(y)
        tek.append((y, s))
    return tek


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--kok', default='.')
    p.add_argument('--tasi', action='store_true', help='gercekten tasi')
    a = p.parse_args()
    kok = os.path.abspath(a.kok)
    hedef = os.path.join(kok, '_SILINECEKLER', time.strftime('%Y-%m-%d') + '_arsiv')

    ad_listesi = adaylar(kok)
    kodlar = kod_dosyalari(kok)
    metinler = {}
    for k in kodlar:
        try:
            metinler[k] = io.open(k, encoding='utf-8', errors='replace').read()
        except IOError:
            pass

    # ACIK ZORLAMA LISTESI. Bu dosyalarin adi kodda geciyor ama gecen yerler
    # ya tarihsel tek seferlik betikler, ya dosyayi URETEN (okumayan) betikler,
    # ya da bugun duzeltilmis canli betikler. Her biri tek tek bakildi ve
    # sebebi asagiya yazildi - toplu bir "zorla tasi" bayragi YOK.
    ZORLA = {
        'MicRhoBooster_PCR_Paneli_2026-08-02.xlsx':
            u'yalniz tarihsel duzeltme betikleri okuyor (DUZELTME_/MADDE123_/'
            u'SON_ETAP_/WSL devir_2026-08-02); zincirde kosan hicbir sey okumuyor',
        'MicRhoBooster_PCR_Paneli_2026-08-02_TESLIM.xlsx':
            u'canli okuyucular bugun YENI Excel e baglandi (capraz_kontrol.py, '
            u'hepsini_denetle.py, tek_tus.py); kalan gecisler tarihsel betikler '
            u've aciklama metinleri',
        'MicRhoBooster_Primer_Tasarimi.xlsx':
            u'WSL betikleri bu dosyayi URETIYOR (--out), okumuyor; gerekirse '
            u'yeniden uretilir',
        'REFERANS_DB/SILVA_SSURef_NR99.fasta':
            u'kimlik_dogrulama.py listesinden bugun cikarildi (oylamaya zaten '
            u'girmiyordu ve artik ikiz DEGIL: bu kopya RNA, 138.2 surumu DNA)',
        'REFERANS_DB/SILVA_LSURef_NR99.fasta':
            u'kimlik_dogrulama.py listesinden bugun cikarildi',
    }

    tasinacak, birakilacak = [], []
    for yol, sebep in ad_listesi:
        ad = os.path.basename(yol)
        gecen = [os.path.relpath(k, kok) for k, m in metinler.items() if ad in m]
        # kendi arsivleme betigi ve denetim betigi adi anmasi engel degil
        # TARIHSEL TEK SEFERLIK betikler engel degildir: DUZELTME_betikleri/,
        # SON_ETAP_betikleri/ ve MADDE123_betikleri/ altindakiler o gunku
        # duzeltmeyi yapip bitmis islerdir, zincirde kosmuyorlar. Onlarin adi
        # anmasi bir dosyayi canli tutmaz - ama MANIFEST'e yazilir ki
        # birisi o betigi yeniden kosarsa nereye bakacagini bilsin.
        tarihsel = ('DUZELTME_betikleri', 'SON_ETAP_betikleri',
                    'MADDE123_betikleri', 'FARKLI_LOKUS_betikleri',
                    'REFERANS_TASARIM_betikleri')
        gecen = [g for g in gecen if os.path.basename(g) not in
                 ('arsivle.py', 'hepsini_denetle.py', 'excel_uret.py',
                  'siparis_formu.py')]
        canli = [g for g in gecen if not any(t in g for t in tarihsel)]
        if gecen and not canli:
            sebep = sebep + (u' | yalniz tarihsel betikler aniyor: %s'
                             % ', '.join(gecen[:3]))
            gecen = []
        if yol in ZORLA:
            tasinacak.append((yol, sebep + u' | ' + ZORLA[yol]))
        elif gecen:
            birakilacak.append((yol, sebep, gecen))
        else:
            tasinacak.append((yol, sebep))

    print('=' * 78)
    print('  ARSIVLEME %s   %s' % ('(PLAN)' if not a.tasi else '(TASINIYOR)',
                                   time.strftime('%Y-%m-%d %H:%M')))
    print('=' * 78)
    print('  aday: %d | tasinacak: %d | KODDA GECTIGI ICIN BIRAKILAN: %d'
          % (len(ad_listesi), len(tasinacak), len(birakilacak)))
    toplam = 0
    for yol, sebep in tasinacak:
        try:
            toplam += os.path.getsize(os.path.join(kok, yol))
        except OSError:
            pass
    print('  tasinacak toplam boyut: %.1f MB' % (toplam / 1e6))
    print()
    for yol, sebep, gecen in birakilacak:
        print('  BIRAKILDI  %-56s (kodda geciyor: %s)'
              % (yol[:56], ', '.join(gecen[:2])))
    if birakilacak:
        print()

    if not a.tasi:
        for yol, sebep in tasinacak[:12]:
            print('  tasinacak  %-56s %s' % (yol[:56], sebep[:40]))
        if len(tasinacak) > 12:
            print('  ... ve %d dosya daha' % (len(tasinacak) - 12))
        print()
        print('  Bu bir PLANDIR. Gercekten tasimak icin: --tasi')
        return 0

    if not os.path.isdir(hedef):
        os.makedirs(hedef)
    mani = os.path.join(hedef, 'MANIFEST.tsv')
    yeni_mani = not os.path.exists(mani)
    n = 0
    with io.open(mani, 'a', encoding='utf-8', newline='') as fh:
        if yeni_mani:
            fh.write(u'# Arsive tasinan dosyalar. SILINMEDILER.\n')
            fh.write(u'# Geri almak icin: bu dosyadaki "eski_yol" sutununa geri '
                     u'kopyalayin.\n')
            fh.write(u'tarih\teski_yol\tarsiv_yolu\tboyut_bayt\tsebep\n')
        for yol, sebep in tasinacak:
            kaynak = os.path.join(kok, yol)
            varis = os.path.join(hedef, yol)
            if not os.path.isdir(os.path.dirname(varis)):
                os.makedirs(os.path.dirname(varis))
            try:
                boy = os.path.getsize(kaynak)
                shutil.move(kaynak, varis)
            except Exception as e:
                print('  TASINAMADI %-52s %s' % (yol[:52], e))
                continue
            n += 1
            fh.write(u'%s\t%s\t%s\t%d\t%s\n'
                     % (time.strftime('%Y-%m-%d %H:%M'), yol,
                        os.path.relpath(varis, kok), boy, sebep))
    print('  tasinan: %d dosya' % n)
    print('  arsiv  : %s' % os.path.relpath(hedef, kok))
    print('  kayit  : %s' % os.path.relpath(mani, kok))
    print('=' * 78)
    return 0


if __name__ == '__main__':
    sys.exit(main())
